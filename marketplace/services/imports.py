from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from django.utils.text import get_valid_filename, slugify

from marketplace.models import Brand, Category, Part


class UploadLimitError(Exception):
    def __init__(self, message: str, status_code: int = 413):
        super().__init__(message)
        self.status_code = status_code


@dataclass
class ImportResult:
    mode: str
    created: int
    updated: int
    skipped_no_price: int
    skipped_invalid: int
    errors: list[dict[str, Any]]
    total_rows: int
    processed_rows: int
    failed_rows: int
    success_rate: int


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in value.lower() if ch.isalnum())


def _first_of(normalized_to_original: dict[str, str], *variants: str) -> str | None:
    for key in variants:
        if key in normalized_to_original:
            return normalized_to_original[key]
    return None


def _parse_price(raw: str) -> Decimal | None:
    cleaned = (raw or "").strip()
    cleaned = cleaned.replace("€", "").replace("$", "").replace("₽", "").strip()
    cleaned = cleaned.replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        value = Decimal(cleaned).quantize(Decimal("0.01"))
        return value if value > 0 else None
    except Exception:
        return None


def _parse_non_negative_int(raw: str, default: int) -> int:
    cleaned = (raw or "").strip()
    if not cleaned:
        return default
    try:
        return max(0, int(float(cleaned.replace(",", "."))))
    except Exception:
        return default


def _parse_non_negative_decimal(raw: str) -> Decimal | None:
    cleaned = (raw or "").strip()
    if not cleaned:
        return None
    cleaned = cleaned.replace(" ", "").replace(",", ".")
    try:
        value = Decimal(cleaned)
        return value if value >= 0 else None
    except Exception:
        return None


def _build_header_mapping(
    fieldnames: list[str],
) -> tuple[
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
    str | None,
]:
    normalized_to_original = {_normalize_header(h): h for h in fieldnames if h}
    part_col = _first_of(normalized_to_original, "partnumber", "partno", "partnum", "part#", "number", "sku", "article", "artikul")
    desc_col = _first_of(normalized_to_original, "description", "descriptionenglisch", "name", "title", "productname")
    price_exw_col = _first_of(normalized_to_original, "priceexw", "exw", "unitprice", "price")
    price_fob_sea_col = _first_of(normalized_to_original, "pricefobsea", "fobsea")
    price_fob_air_col = _first_of(normalized_to_original, "pricefobair", "fobair")
    currency_col = _first_of(normalized_to_original, "currency", "curr")
    stock_col = _first_of(normalized_to_original, "stock", "stockqty", "qty", "quantity")
    oem_col = _first_of(normalized_to_original, "oem", "oemnumber", "oemno")
    brand_col = _first_of(normalized_to_original, "brand", "make", "manufacturer")
    cross_col = _first_of(normalized_to_original, "crossnumber", "crossnum", "cross")
    condition_col = _first_of(normalized_to_original, "condition")
    warehouse_col = _first_of(normalized_to_original, "warehouseaddress", "warehouse")
    sea_port_col = _first_of(normalized_to_original, "seaport")
    air_port_col = _first_of(normalized_to_original, "airport")
    weight_col = _first_of(normalized_to_original, "weight")
    length_col = _first_of(normalized_to_original, "length")
    width_col = _first_of(normalized_to_original, "width")
    height_col = _first_of(normalized_to_original, "height")
    moq_col = _first_of(normalized_to_original, "moq")
    lead_time_col = _first_of(normalized_to_original, "leadtimedays", "leadtime", "productionleaddays")
    return (
        part_col,
        desc_col,
        price_exw_col,
        price_fob_sea_col,
        price_fob_air_col,
        currency_col,
        stock_col,
        oem_col,
        brand_col,
        cross_col,
        condition_col,
        warehouse_col,
        sea_port_col,
        air_port_col,
        weight_col,
        length_col,
        width_col,
        height_col,
        moq_col,
        lead_time_col,
    )


def _format_limit_bytes(max_bytes: int) -> str:
    mb = max_bytes / (1024 * 1024)
    return f"{mb:.0f} MB" if mb >= 10 else f"{mb:.1f} MB"


def _csv_rows(raw: bytes) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV должен быть в кодировке UTF-8.") from exc

    sample = text[:4096]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;|\t").delimiter
    except Exception:
        delimiter = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError("Файл не содержит заголовок.")

    rows: list[tuple[int, dict[str, str]]] = []
    for row_num, row in enumerate(reader, start=2):
        normalized = {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        rows.append((row_num, normalized))
    return [str(h or "").strip() for h in reader.fieldnames], rows


def _xlsx_rows(raw: bytes) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError("Для импорта XLSX установите зависимость openpyxl.") from exc

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration as exc:
        raise ValueError("Файл не содержит заголовок.") from exc

    headers = [str(cell or "").strip() for cell in header_row]
    if not any(headers):
        raise ValueError("Файл не содержит заголовок.")

    rows: list[tuple[int, dict[str, str]]] = []
    for row_num, row_values in enumerate(row_iter, start=2):
        values = [str(value).strip() if value is not None else "" for value in row_values]
        row_dict = {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))}
        rows.append((row_num, row_dict))
    return headers, rows


def process_seller_csv_upload(
    *,
    seller,
    upload,
    category_name: str,
    default_stock: int,
    import_mode: str,
    default_image_url: str = "/static/marketplace/epiroc-logo.webp",
) -> ImportResult:
    filename = get_valid_filename(upload.name.lower())
    extension = Path(filename).suffix.lower()
    if extension not in {".csv", ".xlsx"}:
        raise ValueError("Поддерживаются файлы CSV (UTF-8) и XLSX.")

    file_size = int(getattr(upload, "size", 0) or 0)
    if file_size > int(settings.MAX_IMPORT_FILE_BYTES):
        raise UploadLimitError(
            f"Файл слишком большой. Допустимо до {_format_limit_bytes(int(settings.MAX_IMPORT_FILE_BYTES))}.",
            status_code=413,
        )

    raw = upload.read()
    if extension == ".xlsx":
        fieldnames, source_rows = _xlsx_rows(raw)
    else:
        fieldnames, source_rows = _csv_rows(raw)

    (
        part_col,
        desc_col,
        price_exw_col,
        price_fob_sea_col,
        price_fob_air_col,
        currency_col,
        stock_col,
        oem_col,
        brand_col,
        cross_col,
        condition_col,
        warehouse_col,
        sea_port_col,
        air_port_col,
        weight_col,
        length_col,
        width_col,
        height_col,
        moq_col,
        lead_time_col,
    ) = _build_header_mapping(fieldnames)

    if not (part_col and warehouse_col and (price_fob_sea_col or price_fob_air_col)):
        raise ValueError(
            "Не найдены обязательные колонки. Нужны PartNumber/Part Number, WarehouseAddress и хотя бы одна цена "
            "(Price_FOB_SEA или Price_FOB_AIR)."
        )

    category_slug = slugify(category_name)[:140] or "import-category"
    category, _ = Category.objects.get_or_create(slug=category_slug, defaults={"name": category_name})
    if category.name != category_name:
        category.name = category_name
        category.save(update_fields=["name"])
    brand = Brand.objects.filter(name__iexact=category_name).first()

    created = 0
    updated = 0
    skipped_no_price = 0
    skipped_invalid = 0
    batch_size = 1000
    errors: list[dict[str, Any]] = []
    row_count = 0
    import_timestamp = timezone.now()

    existing_parts = Part.objects.filter(seller=seller).only("id", "oem_number")
    existing_by_oem = {p.oem_number: p.id for p in existing_parts if p.oem_number}
    pending_new_by_oem: dict[str, Part] = {}
    to_create: list[Part] = []
    to_update: list[Part] = []

    def flush_updates():
        nonlocal updated
        if not to_update:
            return
        with transaction.atomic():
            Part.objects.bulk_update(
                to_update,
                [
                    "title",
                    "description",
                    "price",
                    "stock_quantity",
                    "condition",
                    "category",
                    "brand",
                    "image_url",
                    "currency",
                    "data_updated_at",
                    "is_active",
                ],
                batch_size=batch_size,
            )
        updated += len(to_update)
        to_update.clear()

    def flush_creates():
        nonlocal created
        if not to_create:
            return
        with transaction.atomic():
            Part.objects.bulk_create(to_create, batch_size=batch_size)
        created += len(to_create)
        to_create.clear()
        pending_new_by_oem.clear()

    for row_num, row in source_rows:
        row_count += 1
        if row_count > int(settings.MAX_IMPORT_ROWS):
            raise UploadLimitError(
                f"Превышен лимит строк: максимум {settings.MAX_IMPORT_ROWS}.",
                status_code=413,
            )
        part_number = (row.get(part_col) or "").strip()
        description = (row.get(desc_col) or "").strip() if desc_col else ""
        price_exw_raw = (row.get(price_exw_col) or "").strip() if price_exw_col else ""
        price_fob_sea_raw = (row.get(price_fob_sea_col) or "").strip() if price_fob_sea_col else ""
        price_fob_air_raw = (row.get(price_fob_air_col) or "").strip() if price_fob_air_col else ""
        warehouse_address = (row.get(warehouse_col) or "").strip() if warehouse_col else ""
        raw_condition = (row.get(condition_col) or "").strip().upper() if condition_col else ""
        condition_map = {"ORIGINAL": "oem", "OEM": "oem", "REMAN": "reman", "AFTERMARKET": "aftermarket"}
        normalized_condition = condition_map.get(raw_condition or "OEM")

        if not part_number:
            skipped_invalid += 1
            if len(errors) < 100:
                errors.append(
                    {
                        "row": row_num,
                        "error_type": "missing_required_field",
                        "code": "missing_part_number",
                        "reason": "Не заполнен артикул (Part Number).",
                        "hint": "Заполните колонку Part Number и повторите импорт.",
                        "original_data": row,
                    }
                )
            continue

        if not warehouse_address:
            skipped_no_price += 1
            if len(errors) < 100:
                errors.append(
                    {
                        "row": row_num,
                        "error_type": "missing_required_field",
                        "code": "missing_warehouse_address",
                        "reason": "Не заполнено обязательное поле WarehouseAddress.",
                        "hint": "Укажите адрес склада в колонке WarehouseAddress.",
                        "original_data": row,
                    }
                )
            continue

        if raw_condition and not normalized_condition:
            skipped_invalid += 1
            if len(errors) < 100:
                errors.append(
                    {
                        "row": row_num,
                        "error_type": "invalid_condition",
                        "code": "invalid_condition",
                        "reason": f"Condition = {raw_condition}. Допустимо: ORIGINAL / OEM / REMAN / AFTERMARKET.",
                        "hint": "Исправьте Condition на одно из допустимых значений.",
                        "original_data": row,
                    }
                )
            continue

        price_exw = _parse_price(price_exw_raw) if price_exw_raw else None
        price_fob_sea = _parse_price(price_fob_sea_raw) if price_fob_sea_raw else None
        price_fob_air = _parse_price(price_fob_air_raw) if price_fob_air_raw else None
        if price_fob_sea is None and price_fob_air is None:
            skipped_no_price += 1
            if len(errors) < 100:
                errors.append(
                    {
                        "row": row_num,
                        "error_type": "missing_required_field",
                        "code": "missing_fob_price",
                        "reason": "Не заполнена ни одна обязательная цена: Price_FOB_SEA или Price_FOB_AIR.",
                        "hint": "Укажите цену FOB SEA или FOB AIR больше 0.",
                        "original_data": row,
                    }
                )
            continue
        if price_exw is not None and price_fob_sea is not None and price_fob_sea < price_exw:
            skipped_invalid += 1
            if len(errors) < 100:
                errors.append(
                    {
                        "row": row_num,
                        "error_type": "invalid_format",
                        "code": "fob_less_than_exw",
                        "reason": "Price_FOB_SEA не может быть ниже Price_EXW.",
                        "hint": "Проверьте логическую связку цен: FOB >= EXW.",
                        "original_data": row,
                    }
                )
            continue
        if price_exw is not None and price_fob_air is not None and price_fob_air < price_exw:
            skipped_invalid += 1
            if len(errors) < 100:
                errors.append(
                    {
                        "row": row_num,
                        "error_type": "invalid_format",
                        "code": "fob_air_less_than_exw",
                        "reason": "Price_FOB_AIR не может быть ниже Price_EXW.",
                        "hint": "Проверьте логическую связку цен: FOB >= EXW.",
                        "original_data": row,
                    }
                )
            continue

        price = price_exw or price_fob_sea or price_fob_air
        if price is None:
            skipped_no_price += 1
            if len(errors) < 100:
                errors.append(
                    {
                        "row": row_num,
                        "error_type": "invalid_format",
                        "code": "invalid_price",
                        "reason": "Некорректный формат цены.",
                        "hint": "Используйте число больше 0, например 1250.50.",
                        "original_data": row,
                    }
                )
            continue
        title = description if description else f"Part {part_number}"

        stock_value = default_stock
        if stock_col:
            raw_stock = (row.get(stock_col) or "").strip()
            if raw_stock:
                stock_value = _parse_non_negative_int(raw_stock, default_stock)

        currency_value = "USD"
        if currency_col:
            raw_currency = (row.get(currency_col) or "").strip().upper()
            if raw_currency in {"USD", "EUR", "RUB", "CNY"}:
                currency_value = raw_currency

        oem_number_value = (row.get(oem_col) or "").strip() if oem_col else part_number
        if not oem_number_value:
            oem_number_value = part_number
        row_brand = (row.get(brand_col) or "").strip() if brand_col else ""
        effective_brand = brand
        if row_brand:
            brand_slug = slugify(row_brand)[:180] or "brand"
            effective_brand, _ = Brand.objects.get_or_create(
                name=row_brand,
                defaults={"slug": brand_slug, "region": "global"},
            )
        cross_numbers = (row.get(cross_col) or "").strip() if cross_col else ""
        moq_value = _parse_non_negative_int((row.get(moq_col) or "").strip() if moq_col else "", 1)
        lead_time_value = _parse_non_negative_int((row.get(lead_time_col) or "").strip() if lead_time_col else "", 1)
        weight_value = _parse_non_negative_decimal((row.get(weight_col) or "").strip() if weight_col else "")
        length_value = _parse_non_negative_decimal((row.get(length_col) or "").strip() if length_col else "")
        width_value = _parse_non_negative_decimal((row.get(width_col) or "").strip() if width_col else "")
        height_value = _parse_non_negative_decimal((row.get(height_col) or "").strip() if height_col else "")

        match_key = oem_number_value or part_number

        if import_mode == "preview":
            if existing_by_oem.get(match_key):
                updated += 1
            else:
                created += 1
            continue

        existing_id = existing_by_oem.get(match_key)
        if existing_id:
            to_update.append(
                Part(
                    id=existing_id,
                    seller=seller,
                    title=title[:255],
                    oem_number=oem_number_value,
                    description=description,
                    price=price,
                    stock_quantity=stock_value,
                    condition=normalized_condition,
                    image_url=default_image_url,
                    category=category,
                    brand=effective_brand,
                    currency=currency_value,
                    moq=max(1, moq_value),
                    production_lead_days=max(1, lead_time_value),
                    prep_to_ship_days=max(1, lead_time_value),
                    shipping_lead_days=max(1, lead_time_value),
                    gross_weight_kg=weight_value if weight_value is not None else Decimal("0.100"),
                    length_cm=length_value if length_value is not None else Decimal("1.00"),
                    width_cm=width_value if width_value is not None else Decimal("1.00"),
                    height_cm=height_value if height_value is not None else Decimal("1.00"),
                    country_of_origin=warehouse_address[:120] or "Unknown",
                    cross_numbers=cross_numbers,
                    supplier_part_uid=warehouse_address[:80],
                    data_updated_at=import_timestamp,
                    is_active=True,
                )
            )
            if len(to_update) >= batch_size:
                flush_updates()
            continue

        if match_key in pending_new_by_oem:
            obj = pending_new_by_oem[match_key]
            obj.title = title[:255]
            obj.description = description
            obj.price = price
            obj.stock_quantity = stock_value
            obj.currency = currency_value
            obj.condition = normalized_condition
            obj.brand = effective_brand
            obj.moq = max(1, moq_value)
            obj.production_lead_days = max(1, lead_time_value)
            obj.prep_to_ship_days = max(1, lead_time_value)
            obj.shipping_lead_days = max(1, lead_time_value)
            if weight_value is not None:
                obj.gross_weight_kg = weight_value
            if length_value is not None:
                obj.length_cm = length_value
            if width_value is not None:
                obj.width_cm = width_value
            if height_value is not None:
                obj.height_cm = height_value
            obj.country_of_origin = warehouse_address[:120] or obj.country_of_origin
            obj.cross_numbers = cross_numbers
            obj.supplier_part_uid = warehouse_address[:80]
            obj.data_updated_at = import_timestamp
            continue

        base = slugify(f"{part_number}-{title}")[:220] or "part"
        new_part = Part(
            seller=seller,
            title=title[:255],
            slug=f"{base}-{uuid4().hex[:8]}",
            oem_number=oem_number_value,
            description=description,
            price=price,
            stock_quantity=stock_value,
            condition=normalized_condition,
            image_url=default_image_url,
            category=category,
            brand=effective_brand,
            currency=currency_value,
            moq=max(1, moq_value),
            production_lead_days=max(1, lead_time_value),
            prep_to_ship_days=max(1, lead_time_value),
            shipping_lead_days=max(1, lead_time_value),
            gross_weight_kg=weight_value if weight_value is not None else Decimal("0.100"),
            length_cm=length_value if length_value is not None else Decimal("1.00"),
            width_cm=width_value if width_value is not None else Decimal("1.00"),
            height_cm=height_value if height_value is not None else Decimal("1.00"),
            country_of_origin=warehouse_address[:120] or "Unknown",
            cross_numbers=cross_numbers,
            supplier_part_uid=warehouse_address[:80],
            data_updated_at=import_timestamp,
            is_active=True,
        )
        to_create.append(new_part)
        pending_new_by_oem[match_key] = new_part
        if len(to_create) >= batch_size:
            flush_creates()

    if import_mode != "preview":
        flush_updates()
        flush_creates()

    processed_rows = created + updated
    failed_rows = skipped_no_price + skipped_invalid
    total_rows = row_count
    success_rate = int(round((processed_rows / total_rows) * 100)) if total_rows else 0

    return ImportResult(
        mode=import_mode,
        created=created,
        updated=updated,
        skipped_no_price=skipped_no_price,
        skipped_invalid=skipped_invalid,
        errors=errors,
        total_rows=total_rows,
        processed_rows=processed_rows,
        failed_rows=failed_rows,
        success_rate=success_rate,
    )
