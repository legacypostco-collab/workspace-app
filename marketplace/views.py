from decimal import Decimal
from functools import wraps
import csv
import io
import json
import logging
import os
from datetime import timedelta
from uuid import uuid4
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.conf import settings
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, F, Q, Sum
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.text import slugify
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from files.models import StoredFile
from files.storage import read_stored_file_bytes, store_import_source_file
from imports.models import ImportJob, ImportPreviewSession
from imports.services import ColumnMappingResolver, ImportParser
from imports.tasks import process_import_job
from .forms import BulkPriceLookupForm, CheckoutForm, LoginForm, RegisterForm, RFQCreateForm, SellerBulkUploadForm, SellerPartForm
from .models import (
    Brand,
    Category,
    Drawing,
    Order,
    OrderClaim,
    OrderDocument,
    OrderEvent,
    OrderItem,
    Part,
    RFQ,
    RFQItem,
    SellerImportRun,
    SupplierRatingEvent,
    UserProfile,
    WebhookDeliveryLog,
)
from .rules import AutoModeInputs, decide_auto_mode
from .services.imports import UploadLimitError, process_seller_csv_upload
from .services.logistics import logistics_estimate
from .services.observability import Timer, log_api_error, metric_inc
from projections.models import DashboardProjection
from projections.services import refresh_supplier_dashboard_projection
from dashboard.services import DashboardProjectionBuilder

CART_SESSION_KEY = "cart"
COMPARE_SESSION_KEY = "compare_parts"
ORDER_TRANSITIONS = {
    "pending": {"reserve_paid", "cancelled"},
    "reserve_paid": {"pending", "confirmed", "cancelled"},
    "confirmed": {"reserve_paid", "in_production", "cancelled"},
    "in_production": {"confirmed", "ready_to_ship", "cancelled"},
    "ready_to_ship": {"in_production", "transit_abroad", "shipped", "cancelled"},
    "transit_abroad": {"ready_to_ship", "customs", "cancelled"},
    "customs": {"transit_abroad", "transit_rf", "cancelled"},
    "transit_rf": {"customs", "issuing", "cancelled"},
    "issuing": {"transit_rf", "shipped", "cancelled"},
    "shipped": {"issuing", "delivered", "cancelled"},
    "delivered": {"shipped", "completed"},
    "completed": set(),
    "cancelled": set(),
}


def _find_status_path(current: str, target: str, _max_depth: int = 15) -> list[str] | None:
    """BFS to find shortest path from current to target through ORDER_TRANSITIONS."""
    if current == target:
        return []
    from collections import deque
    queue: deque[tuple[str, list[str]]] = deque([(current, [])])
    visited = {current}
    while queue and len(visited) < _max_depth:
        node, path = queue.popleft()
        for nxt in ORDER_TRANSITIONS.get(node, set()):
            if nxt == "cancelled":
                continue
            if nxt == target:
                return path + [nxt]
            if nxt not in visited:
                visited.add(nxt)
                queue.append((nxt, path + [nxt]))
    return None

logger = logging.getLogger("marketplace")


SELLER_IMPORT_MAPPING_FIELDS: list[tuple[str, str]] = [
    ("oem", "Артикул (Part Number)"),
    ("name", "Название"),
    ("brand", "Бренд"),
    ("price_exw", "Цена EXW"),
    ("price_fob_sea", "Цена FOB Море"),
    ("price_fob_air", "Цена FOB Авиа"),
    ("quantity", "Остаток"),
    ("condition", "Состояние"),
    ("warehouse_address", "Склад"),
    ("cross_number", "Кросс-номер"),
    ("sea_port", "Морской порт"),
    ("air_port", "Авиа-порт"),
    ("weight", "Вес (кг)"),
    ("length", "Длина (см)"),
    ("width", "Ширина (см)"),
    ("height", "Высота (см)"),
]

# Auto-mapping: common header synonyms → field key
_AUTO_MAP_SYNONYMS: dict[str, list[str]] = {
    "oem": ["partnumber", "part_number", "part number", "oem", "артикул", "номер", "pn", "p/n", "каталожный"],
    "name": ["name", "title", "description", "наименование", "название", "описание"],
    "brand": ["brand", "manufacturer", "бренд", "производитель", "марка"],
    "price_exw": ["price", "price_exw", "цена", "цена exw", "price exw", "стоимость"],
    "price_fob_sea": ["price_fob_sea", "fob sea", "цена fob море"],
    "price_fob_air": ["price_fob_air", "fob air", "цена fob авиа"],
    "quantity": ["qty", "quantity", "stock", "остаток", "количество", "кол-во", "наличие"],
    "condition": ["condition", "состояние", "new/used"],
    "warehouse_address": ["warehouse", "warehouseaddress", "склад", "адрес склада", "location"],
    "cross_number": ["crossnumber", "cross", "кросс", "cross_number", "аналог"],
    "weight": ["weight", "вес", "масса"],
}


def _auto_map_columns(headers: list[str]) -> dict[str, str]:
    """Try to match spreadsheet headers to our import fields."""
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for field_key, synonyms in _AUTO_MAP_SYNONYMS.items():
        for header in headers:
            if header in used:
                continue
            h_low = header.lower().strip()
            if h_low in synonyms or any(s in h_low for s in synonyms):
                mapping[field_key] = header
                used.add(header)
                break
    return mapping


def _get_cart(request: HttpRequest) -> dict[str, int]:
    return request.session.get(CART_SESSION_KEY, {})


def _set_cart(request: HttpRequest, cart: dict[str, int]) -> None:
    request.session[CART_SESSION_KEY] = cart
    request.session.modified = True


def _get_compare_ids(request: HttpRequest) -> list[int]:
    raw = request.session.get(COMPARE_SESSION_KEY, [])
    out: list[int] = []
    for val in raw:
        try:
            out.append(int(val))
        except Exception:
            continue
    return out


def _set_compare_ids(request: HttpRequest, ids: list[int]) -> None:
    request.session[COMPARE_SESSION_KEY] = [int(x) for x in ids]
    request.session.modified = True


def _log_order_event(order: Order, event_type: str, source: str = "system", actor: User | None = None, meta: dict | None = None):
    event = OrderEvent.objects.create(
        order=order,
        event_type=event_type,
        source=source,
        actor=actor,
        meta=meta or {},
    )
    _emit_webhooks_for_order_event(event)


# SLA нормативы по этапам (в часах) из таблицы "Этапы ЛК"
SLA_STAGE_NORMS: dict[str, int] = {
    "pending": 48,          # Ожидание оплаты: ≤ 48 ч
    "reserve_paid": 48,
    "confirmed": 168,       # Формирование заказа: ≤ 7 дн (2+5)
    "in_production": 168,
    "ready_to_ship": 48,
    "transit_abroad": 240,  # Транзит (авто, КНР): ≤ 10 дн
    "customs": 48,          # Таможня: ≤ 2 рабочих дня
    "transit_rf": 24,       # Транзит РФ: ≤ 1 рабочий день
    "issuing": 24,          # Выдача: ≤ 1 рабочий день
    "shipped": 24,
    "delivered": 72,        # Приёмка: ≤ 3 рабочих дня
}


def _recalc_order_sla(order: Order):
    previous = order.sla_status
    now = timezone.now()
    status = "on_track"

    norm_hours = SLA_STAGE_NORMS.get(order.status)
    if norm_hours:
        # Определяем, когда заказ вошёл в текущий статус
        last_event = (
            OrderEvent.objects.filter(
                order=order,
                event_type="status_changed",
                meta__to=order.status,
            )
            .order_by("-created_at")
            .first()
        )
        entered_at = last_event.created_at if last_event else order.created_at
        elapsed_hours = (now - entered_at).total_seconds() / 3600

        if elapsed_hours >= norm_hours:
            status = "breached"
        elif elapsed_hours >= norm_hours * 0.75:
            status = "at_risk"

    if status != previous:
        order.sla_status = status
        if status == "breached":
            order.sla_breaches_count += 1
        order.save(update_fields=["sla_status", "sla_breaches_count"])
        _log_order_event(order, "sla_status_changed", source="system", meta={"from": previous, "to": status})


def _create_order_from_rows(
    *,
    rows,
    total: Decimal,
    customer_name: str,
    customer_email: str,
    customer_phone: str,
    delivery_address: str,
    buyer: User | None,
    source: str,
    source_id: int | None = None,
    logistics_override_cost: Decimal | None = None,
):
    reserve_percent = Decimal("10.00")
    total_weight = Decimal("0.00")
    total_volume = Decimal("0.00")
    for row in rows:
        part = row["part"]
        qty = Decimal(row["quantity"])
        total_weight += (Decimal(part.gross_weight_kg or 0) * qty)
        cm3 = Decimal(part.length_cm or 0) * Decimal(part.width_cm or 0) * Decimal(part.height_cm or 0)
        total_volume += ((cm3 / Decimal("1000000")) * qty)

    if logistics_override_cost is not None:
        logistics_result = {
            "ok": True,
            "provider": "manual_override",
            "currency": "USD",
            "cost": str(logistics_override_cost.quantize(Decimal("0.01"))),
        }
    else:
        logistics_payload = {
            "origin": settings.LOGISTICS_DEFAULT_ORIGIN,
            "destination": delivery_address or settings.LOGISTICS_DEFAULT_DESTINATION,
            "mode": settings.LOGISTICS_DEFAULT_MODE,
            "incoterm": settings.LOGISTICS_DEFAULT_INCOTERM,
            "weight_kg": str(total_weight.quantize(Decimal("0.01"))),
            "volume_m3": str(total_volume.quantize(Decimal("0000000.01"))),
            "currency": "USD",
        }
        logistics_result = logistics_estimate(logistics_payload)
        if settings.LOGISTICS_STRICT_MODE and not logistics_result.get("ok", False):
            raise ValueError(logistics_result.get("error", "Logistics calculation failed"))

    logistics_cost = Decimal("0.00")
    logistics_currency = "USD"
    logistics_provider = "internal_fallback"
    if logistics_result.get("ok"):
        try:
            logistics_cost = Decimal(str(logistics_result.get("cost", "0"))).quantize(Decimal("0.01"))
            if logistics_cost < 0:
                logistics_cost = Decimal("0.00")
        except Exception:
            logistics_cost = Decimal("0.00")
        logistics_currency = str(logistics_result.get("currency") or "USD")
        logistics_provider = str(logistics_result.get("provider") or "internal_fallback")

    grand_total = (total + logistics_cost).quantize(Decimal("0.01"))
    reserve_amount = ((grand_total * reserve_percent) / Decimal("100")).quantize(Decimal("0.01"))

    # Определяем схему оплаты
    payment_scheme = request.POST.get("payment_scheme", "simple")
    if payment_scheme not in ("simple", "staged"):
        payment_scheme = "simple"

    mid_payment_amount = Decimal("0.00")
    customs_payment_amount = Decimal("0.00")
    if payment_scheme == "staged":
        mid_payment_amount = (grand_total * Decimal("0.50")).quantize(Decimal("0.01"))
        customs_payment_amount = (grand_total * Decimal("0.40")).quantize(Decimal("0.01"))

    with transaction.atomic():
        order = Order.objects.create(
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            delivery_address=delivery_address,
            buyer=buyer,
            total_amount=grand_total,
            supplier_confirm_deadline=timezone.now() + timedelta(hours=24),
            sla_status="on_track",
            logistics_cost=logistics_cost,
            logistics_currency=logistics_currency,
            logistics_provider=logistics_provider,
            logistics_meta=logistics_result,
            reserve_percent=reserve_percent,
            reserve_amount=reserve_amount,
            payment_scheme=payment_scheme,
            mid_payment_amount=mid_payment_amount,
            customs_payment_amount=customs_payment_amount,
            payment_status="awaiting_reserve",
            )
        order.invoice_number = f"INV-{timezone.now():%Y%m%d}-{order.id}"
        order.save(update_fields=["invoice_number"])
        order_items = []
        for row in rows:
            part = row["part"]
            qty = row["quantity"]
            order_items.append(
                OrderItem(
                    order=order,
                    part=part,
                    quantity=qty,
                    unit_price=part.price,
                )
            )
            Part.objects.filter(id=part.id).update(stock_quantity=F("stock_quantity") - qty)
        OrderItem.objects.bulk_create(order_items)
        _log_order_event(
            order,
            "order_created",
            source=source,
            actor=buyer if buyer and buyer.is_authenticated else None,
            meta={
                "items_count": len(order_items),
                "base_total": str(total),
                "logistics_cost": str(logistics_cost),
                "reserve_amount": str(reserve_amount),
                "total_amount": str(grand_total),
                "source_id": source_id,
            },
        )
        return order


def _role_for(user: User | None) -> str | None:
    if not user or not user.is_authenticated:
        return None
    if user.is_superuser:
        return "seller"
    profile = getattr(user, "profile", None)
    return profile.role if profile else "buyer"


def _is_demo_user(user) -> bool:
    """Check if user is a demo account."""
    return getattr(user, "username", "").startswith("demo_")


def _tpl(user, path: str) -> str:
    """Return template path. Same templates for all users; sidebar is filtered by context_processors."""
    return path


def _profile_for(user: User | None):
    if not user or not user.is_authenticated:
        return None
    return getattr(user, "profile", None)


def _has_order_access(user: User, order: Order, role: str | None) -> bool:
    if user.is_superuser:
        return True
    if role == "seller":
        return order.items.filter(part__seller=user).exists()
    return order.buyer_id == user.id


def _allowed_regions_set(user: User) -> set[str]:
    profile = _profile_for(user)
    if not profile or not profile.allowed_regions:
        return set()
    return {x.strip().lower() for x in profile.allowed_regions.split(",") if x.strip()}


def _operator_can_access_part(user: User, part: Part) -> bool:
    if user.is_superuser:
        return True
    profile = _profile_for(user)
    if not profile or profile.role != "seller":
        return False
    if not _has_seller_permission(user, "can_manage_orders"):
        return False

    allowed_brand_ids = set(profile.allowed_brands.values_list("id", flat=True))
    if allowed_brand_ids and part.brand_id and part.brand_id not in allowed_brand_ids:
        return False
    allowed_regions = _allowed_regions_set(user)
    if allowed_regions and part.brand and (part.brand.region or "").lower() not in allowed_regions:
        return False
    return True


def _webhook_payload_for_event(event: OrderEvent) -> dict:
    return {
        "event": event.event_type,
        "source": event.source,
        "created_at": event.created_at.isoformat(),
        "order": {
            "id": event.order_id,
            "status": event.order.status,
            "payment_status": event.order.payment_status,
            "total_amount": str(event.order.total_amount),
            "logistics_cost": str(event.order.logistics_cost),
        },
        "meta": event.meta or {},
    }


def _send_webhook_attempt(*, event: OrderEvent, endpoint: str, payload: dict, attempt: int) -> bool:
    headers = {"Content-Type": "application/json"}
    secret = getattr(settings, "WEBHOOK_SECRET", "") or ""
    if secret:
        headers["X-Webhook-Secret"] = secret
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    log = WebhookDeliveryLog.objects.create(
        order_event=event,
        order=event.order,
        endpoint=endpoint,
        success=False,
        attempt=attempt,
        request_payload=payload,
    )
    try:
        req = Request(endpoint, data=body, headers=headers, method="POST")
        with urlopen(req, timeout=float(getattr(settings, "WEBHOOK_TIMEOUT_SEC", 2))) as resp:
            status_code = int(getattr(resp, "status", 200))
            response_body = resp.read().decode("utf-8", errors="ignore")[:4000]
        is_ok = 200 <= status_code < 300
        log.success = is_ok
        log.status_code = status_code
        log.response_body = response_body
        log.save(update_fields=["success", "status_code", "response_body", "updated_at"])
        return is_ok
    except HTTPError as exc:
        err_body = ""
        try:
            err_body = exc.read().decode("utf-8", errors="ignore")[:4000]
        except Exception:
            err_body = ""
        log.error = f"HTTPError: {exc}"
        log.status_code = int(getattr(exc, "code", 0) or 0)
        log.response_body = err_body
        log.save(update_fields=["error", "status_code", "response_body", "updated_at"])
        return False
    except URLError as exc:
        log.error = f"URLError: {exc}"
        log.save(update_fields=["error", "updated_at"])
        return False
    except Exception as exc:
        log.error = f"Exception: {exc}"
        log.save(update_fields=["error", "updated_at"])
        return False


def _emit_webhooks_for_order_event(event: OrderEvent) -> None:
    endpoints = [x.strip() for x in (getattr(settings, "WEBHOOK_ENDPOINTS", "") or "").split(",") if x.strip()]
    if not endpoints:
        return

    payload = _webhook_payload_for_event(event)
    max_attempts = max(1, int(getattr(settings, "WEBHOOK_RETRY_MAX_ATTEMPTS", 5) or 5))
    for endpoint in endpoints:
        for attempt in range(1, max_attempts + 1):
            if _send_webhook_attempt(event=event, endpoint=endpoint, payload=payload, attempt=attempt):
                break


def _retry_webhook_log(log: WebhookDeliveryLog) -> bool:
    max_attempts = max(1, int(getattr(settings, "WEBHOOK_RETRY_MAX_ATTEMPTS", 5) or 5))
    if log.success or int(log.attempt) >= max_attempts:
        return False

    payload = log.request_payload or {}
    if not payload and log.order_event:
        payload = _webhook_payload_for_event(log.order_event)
    if not payload:
        payload = {
            "event": "unknown",
            "source": "system",
            "created_at": timezone.now().isoformat(),
            "order": {
                "id": log.order_id,
                "status": log.order.status,
                "payment_status": log.order.payment_status,
                "total_amount": str(log.order.total_amount),
                "logistics_cost": str(log.order.logistics_cost),
            },
            "meta": {},
        }

    headers = {"Content-Type": "application/json"}
    secret = getattr(settings, "WEBHOOK_SECRET", "") or ""
    if secret:
        headers["X-Webhook-Secret"] = secret
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    attempt = int(log.attempt) + 1
    retry_log = WebhookDeliveryLog.objects.create(
        order_event=log.order_event,
        order=log.order,
        endpoint=log.endpoint,
        success=False,
        attempt=attempt,
        request_payload=payload,
    )
    try:
        req = Request(log.endpoint, data=body, headers=headers, method="POST")
        with urlopen(req, timeout=float(getattr(settings, "WEBHOOK_TIMEOUT_SEC", 2))) as resp:
            status_code = int(getattr(resp, "status", 200))
            response_body = resp.read().decode("utf-8", errors="ignore")[:4000]
        ok = 200 <= status_code < 300
        retry_log.success = ok
        retry_log.status_code = status_code
        retry_log.response_body = response_body
        retry_log.save(update_fields=["success", "status_code", "response_body", "updated_at"])
        return ok
    except HTTPError as exc:
        err_body = ""
        try:
            err_body = exc.read().decode("utf-8", errors="ignore")[:4000]
        except Exception:
            err_body = ""
        retry_log.error = f"HTTPError: {exc}"
        retry_log.status_code = int(getattr(exc, "code", 0) or 0)
        retry_log.response_body = err_body
        retry_log.save(update_fields=["error", "status_code", "response_body", "updated_at"])
        return False
    except URLError as exc:
        retry_log.error = f"URLError: {exc}"
        retry_log.save(update_fields=["error", "updated_at"])
        return False
    except Exception as exc:
        retry_log.error = f"Exception: {exc}"
        retry_log.save(update_fields=["error", "updated_at"])
        return False


def _can_upload_order_documents(user: User, role: str | None) -> bool:
    if user.is_superuser:
        return True
    if role == "seller":
        return _has_seller_permission(user, "can_manage_orders")
    return role == "buyer"


def _can_manage_claims(user: User, role: str | None) -> bool:
    if user.is_superuser:
        return True
    if role == "seller":
        return _has_seller_permission(user, "can_manage_orders")
    return role == "buyer"


def _build_payment_url(order: Order) -> tuple[str, str]:
    payment_ref = f"INV-{order.id}-{order.created_at:%Y%m%d}"
    payment_base = (settings.PAYMENT_PROVIDER_URL or "").strip()
    payment_currency = settings.PAYMENT_CURRENCY or "USD"
    payment_query = urlencode(
        {
            "merchant": settings.PAYMENT_MERCHANT_ID or "demo-merchant",
            "invoice": payment_ref,
            "order_id": order.id,
            "amount": str(order.total_amount),
            "currency": payment_currency,
            "customer_email": order.customer_email,
        }
    )
    if payment_base:
        delimiter = "&" if "?" in payment_base else "?"
        return f"{payment_base}{delimiter}{payment_query}", payment_ref
    return f"https://pay.consolidator.parts/pay?{payment_query}", payment_ref


def _has_seller_permission(user: User, permission: str) -> bool:
    if user.is_superuser:
        return True
    profile = _profile_for(user)
    if not profile or profile.role != "seller":
        return False
    return bool(getattr(profile, permission, False))


def _apply_seller_brand_scope(user: User, qs):
    profile = _profile_for(user)
    if not profile or profile.role != "seller":
        return qs.none()
    allowed_brand_ids = list(profile.allowed_brands.values_list("id", flat=True))
    if allowed_brand_ids:
        return qs.filter(brand_id__in=allowed_brand_ids)
    return qs


def _seller_rfqs_qs(user: User):
    return (
        RFQ.objects.filter(items__matched_part__seller=user)
        .distinct()
        .prefetch_related("items__matched_part__brand", "items__matched_part__category")
        .order_by("-created_at")
    )


def _part_stale_snapshot(part: Part) -> dict[str, object]:
    updated_at = part.data_updated_at or part.updated_at or timezone.now()
    age_days = max(0, (timezone.now() - updated_at).days)
    if age_days > 180:
        state = "blocked"
        label = "Blocked"
    elif age_days > 90:
        state = "limited"
        label = "Limited"
    else:
        state = "fresh"
        label = "Fresh"
    return {
        "days": age_days,
        "state": state,
        "label": label,
        "is_stale": age_days > 90,
    }


def _part_price_history(part: Part) -> list[dict[str, object]]:
    points: list[dict[str, object]] = []
    order_items = (
        OrderItem.objects.filter(part=part)
        .select_related("order")
        .order_by("order__created_at")[:24]
    )
    for item in order_items:
        points.append(
            {
                "date": item.order.created_at,
                "price": item.unit_price,
                "source": f"order#{item.order_id}",
            }
        )
    current_point = {
        "date": part.data_updated_at or part.updated_at or timezone.now(),
        "price": part.price,
        "source": "current_catalog",
    }
    if not points or points[-1]["price"] != current_point["price"]:
        points.append(current_point)
    return points


def _part_demand_stats(part: Part) -> dict[str, int]:
    rfq_items = RFQItem.objects.filter(matched_part=part)
    order_items = OrderItem.objects.filter(part=part).select_related("order")
    return {
        "rfq_count": rfq_items.count(),
        "quoted_count": rfq_items.exclude(rfq__status="new").count(),
        "orders_count": order_items.values("order_id").distinct().count(),
        "ordered_units": order_items.aggregate(total=Sum("quantity"))["total"] or 0,
        "delivered_orders": order_items.filter(order__status__in=["delivered", "completed"]).values("order_id").distinct().count(),
    }


def seller_required(view):
    @wraps(view)
    def wrapped(request: HttpRequest, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")
        if _role_for(request.user) != "seller":
            messages.error(request, "Доступно только для seller.")
            return redirect("dashboard")
        return view(request, *args, **kwargs)

    return wrapped


def operator_required(view):
    @wraps(view)
    def wrapped(request: HttpRequest, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")
        role = _role_for(request.user)
        if not (request.user.is_superuser or role == "seller"):
            messages.error(request, "Доступно только оператору.")
            return redirect("dashboard")
        return view(request, *args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(request: HttpRequest, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")
        if not request.user.is_superuser:
            messages.error(request, "Доступно только администратору.")
            return redirect("dashboard")
        return view(request, *args, **kwargs)

    return wrapped


def _cart_rows(request: HttpRequest):
    cart = _get_cart(request)
    if not cart:
        return [], Decimal("0.00")

    part_ids = [int(k) for k in cart.keys()]
    parts = Part.objects.filter(id__in=part_ids, is_active=True, price__gt=0)
    rows = []
    total = Decimal("0.00")

    for part in parts:
        qty = int(cart.get(str(part.id), 0))
        line_total = part.price * qty
        total += line_total
        rows.append({"part": part, "quantity": qty, "line_total": line_total})
    return rows, total


def _seed_if_empty() -> None:
    if Category.objects.exists() or Part.objects.exists():
        return
    engine = Category.objects.create(name="Engine", slug="engine")
    hydraulic = Category.objects.create(name="Hydraulic", slug="hydraulic")
    filters = Category.objects.create(name="Filters", slug="filters")

    Part.objects.bulk_create(
        [
            Part(
                title="Piston Assembly CAT 3306",
                slug="piston-assembly-cat-3306",
                oem_number="1234567890",
                description="Premium piston assembly for CAT 3306. Industrial-grade steel.",
                price=Decimal("15000.00"),
                stock_quantity=8,
                condition="oem",
                category=engine,
                image_url="https://images.unsplash.com/photo-1581094271901-8022df4466f9",
            ),
            Part(
                title="Hydraulic Pump Komatsu PC200",
                slug="hydraulic-pump-komatsu-pc200",
                oem_number="HP-KM-88211",
                description="Reliable hydraulic pump, tested for heavy-duty cycles.",
                price=Decimal("48900.00"),
                stock_quantity=3,
                condition="reman",
                category=hydraulic,
                image_url="https://images.unsplash.com/photo-1581092921461-eab62e97a780",
            ),
            Part(
                title="Fuel Filter John Deere 6140",
                slug="fuel-filter-jd-6140",
                oem_number="RE48786",
                description="OEM-compatible filter with high filtration efficiency.",
                price=Decimal("2950.00"),
                stock_quantity=40,
                condition="oem",
                category=filters,
                image_url="https://images.unsplash.com/photo-1613214150388-81f6b8f8246a",
            ),
        ]
    )


def _mixed_featured_parts(limit: int = 12):
    base_qs = (
        Part.objects.filter(is_active=True, price__gt=0)
        .exclude(title__icontains="pc220rock bucket1")
        .exclude(title__istartswith="Komatsu Pc")
        .select_related("category", "brand")
    )
    brands = (
        Brand.objects.annotate(parts_count=Count("parts", filter=Q(parts__is_active=True, parts__price__gt=0)))
        .filter(parts_count__gt=0)
        .order_by("-parts_count", "name")
    )

    featured = []
    used_ids = set()

    # One representative item per brand first -> guarantees brand mix.
    for brand in brands:
        part = base_qs.filter(brand=brand).order_by("-updated_at", "-id").first()
        if not part:
            continue
        featured.append(part)
        used_ids.add(part.id)
        if len(featured) >= limit:
            return featured

    # Fill remaining slots with strongest recent items.
    remainder = base_qs.exclude(id__in=used_ids).order_by("-updated_at", "-id")[: max(0, limit - len(featured))]
    featured.extend(list(remainder))
    return featured


def _parse_bulk_lookup_lines(raw_text: str, limit: int = 1000) -> list[str]:
    seen = set()
    out = []
    for raw_line in (raw_text or "").splitlines():
        normalized = raw_line.strip().strip(",;")
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        out.append(normalized)
        if len(out) >= limit:
            break
    return out


def _parse_lookup_request_line(raw_line: str) -> tuple[str, int]:
    line = (raw_line or "").strip()
    if not line:
        return "", 1
    for separator in (";", "\t", ","):
        if separator in line:
            left, right = line.rsplit(separator, 1)
            left = left.strip()
            right = right.strip()
            try:
                return left, max(1, int(right))
            except Exception:
                pass
    if " " in line:
        left, right = line.rsplit(" ", 1)
        left = left.strip()
        right = right.strip()
        try:
            return left, max(1, int(right))
        except Exception:
            pass
    return line, 1


def _normalize_article_value(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return ""
    for src, dst in {
        "–": "-",
        "—": "-",
        "−": "-",
        "‑": "-",
        "\u00a0": " ",
        "\u200b": "",
        "\u200c": "",
        "\u200d": "",
        "\ufeff": "",
    }.items():
        value = value.replace(src, dst)
    value = " ".join(value.split())
    value = value.replace(" / ", "/").replace(" - ", "-")
    return value.upper()


def _article_input_hint(raw: str, normalized: str) -> str:
    original = (raw or "").strip()
    if not original:
        return ""
    if len(normalized) < 5:
        return "Too short for exact article lookup"
    if not any(ch.isdigit() for ch in normalized):
        return "Looks like a name, not an article"
    if " " in normalized:
        return "Contains internal spaces; verify article formatting"
    if original != normalized:
        return "Normalized input before exact lookup"
    return "No exact match in catalog"


def _parse_bulk_lookup_requests(raw_text: str, limit: int = 1000) -> list[tuple[str, str, int]]:
    seen = set()
    out = []
    for raw_line in (raw_text or "").splitlines():
        query, quantity = _parse_lookup_request_line(raw_line)
        normalized_query = _normalize_article_value(query)
        if not normalized_query or normalized_query in seen:
            continue
        seen.add(normalized_query)
        out.append((query, normalized_query, quantity))
        if len(out) >= limit:
            break
    return out


def _resolve_bulk_lookup_match(base_qs, normalized_query: str, quantity: int) -> tuple[Part | None, str, int]:
    if not normalized_query:
        return None, normalized_query, quantity

    part = base_qs.filter(oem_number__iexact=normalized_query).first()
    if part:
        return part, normalized_query, quantity

    hyphen_qty_match = normalized_query.rsplit("-", 1)
    if len(hyphen_qty_match) == 2:
        candidate_article, trailing_qty = hyphen_qty_match
        try:
            inferred_qty = max(1, int(trailing_qty))
        except Exception:
            inferred_qty = None
        if inferred_qty and candidate_article:
            candidate_part = base_qs.filter(oem_number__iexact=candidate_article).first()
            if candidate_part:
                return candidate_part, candidate_article, inferred_qty

    return None, normalized_query, quantity


def _bulk_lookup_rows(queries: list[tuple[str, str, int]]) -> list[dict]:
    rows = []
    base_qs = _eligible_parts_qs().select_related("brand", "category")
    for original_query, normalized_query, quantity in queries:
        part, matched_query, resolved_quantity = _resolve_bulk_lookup_match(base_qs, normalized_query, quantity)
        match_type = "exact" if part else "missing"
        review_flag = False
        stock_label = "-"
        input_hint = ""
        if part:
            stock_qty = max(0, int(part.stock_quantity or 0))
            if stock_qty > 0:
                stock_label = f"In stock: {stock_qty}"
            else:
                stock_label = "Backorder / check"
            review_flag = part.price >= Decimal("10000.00") or stock_qty == 0
        else:
            input_hint = _article_input_hint(original_query, normalized_query)
        if part and review_flag:
            next_step_label = "Check"
            next_step_tone = "warn"
        elif part:
            next_step_label = "Ready"
            next_step_tone = "ok"
        else:
            next_step_label = "Manual quote"
            next_step_tone = "warn"
        rows.append(
            {
                "query": original_query,
                "normalized_query": matched_query,
                "quantity": resolved_quantity,
                "found": bool(part),
                "part": part,
                "match_type": match_type,
                "line_total": (part.price * resolved_quantity).quantize(Decimal("0.01")) if part else None,
                "stock_label": stock_label,
                "review_flag": review_flag,
                "input_hint": input_hint,
                "next_step_label": next_step_label,
                "next_step_tone": next_step_tone,
            }
        )
    return rows


def _bulk_lookup_csv_response(rows: list[dict]) -> HttpResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Request", "Corrected Article", "Quantity", "Match Status", "Part Number", "Part Name", "Brand", "Price", "Line Total", "Availability", "Review", "Hint"])
    for row in rows:
        part = row["part"]
        writer.writerow(
            [
                row["query"],
                row["normalized_query"] if row["normalized_query"] != row["query"] else "",
                row["quantity"],
                "EXACT" if row["found"] else "NOT_FOUND",
                part.oem_number if part else "",
                part.title if part else "",
                part.brand.name if part and part.brand else "",
                part.price if part else "",
                row["line_total"] if row["line_total"] is not None else "",
                row["stock_label"],
                "review" if row["review_flag"] else "",
                row["input_hint"],
            ]
        )
    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="bulk-price-lookup.csv"'
    return response


def _bulk_lookup_to_rfq_lines(rows: list[dict]) -> str:
    lines = []
    for row in rows:
        article = row["normalized_query"] or row["query"]
        if article:
            lines.append(f"{article};{row['quantity']}")
    return "\n".join(lines)


def home(request: HttpRequest) -> HttpResponse:
    # Always show the landing page — authenticated users can navigate to their cabinet via sidebar
    return render(request, "landing.html")


def home_marketplace(request: HttpRequest) -> HttpResponse:
    """Original marketplace home page (kept for internal use)."""
    _seed_if_empty()
    featured = _mixed_featured_parts(limit=12)
    bulk_form = BulkPriceLookupForm(request.POST or None)
    bulk_results = []
    bulk_total = 0
    bulk_found = 0
    bulk_missing = 0
    bulk_amount = Decimal("0.00")
    bulk_requested_units = 0
    bulk_found_units = 0
    bulk_missing_units = 0
    if request.method == "POST" and bulk_form.is_valid():
        requests_data = _parse_bulk_lookup_requests(bulk_form.cleaned_data["items_text"])
        bulk_results = _bulk_lookup_rows(requests_data)
        bulk_total = len(bulk_results)
        bulk_found = sum(1 for row in bulk_results if row["found"])
        bulk_missing = bulk_total - bulk_found
        bulk_amount = sum((row["line_total"] or Decimal("0.00")) for row in bulk_results)
        bulk_requested_units = sum(int(row["quantity"]) for row in bulk_results)
        bulk_found_units = sum(int(row["quantity"]) for row in bulk_results if row["found"])
        bulk_missing_units = bulk_requested_units - bulk_found_units
        action = request.POST.get("action")
        if action == "export":
            return _bulk_lookup_csv_response(bulk_results)
        if action == "cart":
            cart = _get_cart(request)
            added_positions = 0
            added_units = 0
            skipped_positions = 0
            for row in bulk_results:
                if not row["found"] or not row["part"]:
                    skipped_positions += 1
                    continue
                part = row["part"]
                current = int(cart.get(str(part.id), 0))
                qty = max(1, int(row["quantity"]))
                target_qty = current + qty
                if part.stock_quantity > 0:
                    target_qty = min(part.stock_quantity, target_qty)
                cart[str(part.id)] = target_qty
                added_positions += 1
                added_units += max(0, target_qty - current)
            _set_cart(request, cart)
            if added_positions:
                messages.success(
                    request,
                    f"Added {added_positions} positions ({added_units} units) to cart."
                    + (f" Skipped {skipped_positions} not found." if skipped_positions else ""),
                )
            else:
                messages.warning(request, "Nothing was added to cart. Exact matches were not found.")
            return redirect("cart")
        if action == "rfq":
            request.session["rfq_prefill_items_text"] = _bulk_lookup_to_rfq_lines(bulk_results)
            request.session.modified = True
            messages.success(
                request,
                f"RFQ draft prepared for {bulk_total} positions."
                + (f" {bulk_missing} will need manual quote review." if bulk_missing else ""),
            )
            return redirect("rfq_new")
    top_categories = (
        Category.objects.annotate(parts_count=Count("parts", filter=Q(parts__is_active=True, parts__price__gt=0)))
        .filter(parts_count__gt=0)
        .order_by("-parts_count", "name")[:12]
    )
    top_brands = (
        Brand.objects.annotate(parts_count=Count("parts", filter=Q(parts__is_active=True, parts__price__gt=0)))
        .filter(parts_count__gt=0)
        .order_by("-parts_count", "name")[:12]
    )
    return render(
        request,
        "marketplace/home.html",
        {
            "featured": featured,
            "top_categories": top_categories,
            "top_brands": top_brands,
            "bulk_form": bulk_form,
            "bulk_results": bulk_results,
            "bulk_total": bulk_total,
            "bulk_found": bulk_found,
            "bulk_missing": bulk_missing,
            "bulk_amount": bulk_amount,
            "bulk_requested_units": bulk_requested_units,
            "bulk_found_units": bulk_found_units,
            "bulk_missing_units": bulk_missing_units,
        },
    )


def demo_center(request: HttpRequest) -> HttpResponse:
    buyer = User.objects.filter(username="demo_buyer").first()
    seller = User.objects.filter(username="demo_seller").first()
    operator = User.objects.filter(username="demo_operator").first()

    demo_rfq = (
        RFQ.objects.filter(created_by=buyer).order_by("-id").first()
        if buyer
        else None
    )
    demo_orders = (
        Order.objects.filter(buyer=buyer).order_by("-id")[:8]
        if buyer
        else []
    )
    return render(
        request,
        "marketplace/demo_center.html",
        {
            "buyer_user": buyer,
            "seller_user": seller,
            "operator_user": operator,
            "demo_rfq": demo_rfq,
            "demo_orders": demo_orders,
        },
    )


def register_view(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.email = form.cleaned_data["email"]
            user.first_name = form.cleaned_data["first_name"]
            user.last_name = form.cleaned_data["last_name"]
            user.save()
            UserProfile.objects.create(
                user=user,
                role=form.cleaned_data["role"],
                company_name=form.cleaned_data["company_name"],
            )
            login(request, user)
            messages.success(request, "Регистрация завершена.")
            return redirect("dashboard")
    else:
        form = RegisterForm()
    return render(request, "marketplace/register.html", {"form": form})


def login_view(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        data = request.POST.copy()
        raw_login = data.get("username", "").strip()
        if "@" in raw_login:
            user = User.objects.filter(email__iexact=raw_login).first()
            if user:
                data["username"] = user.username
        form = LoginForm(request, data=data)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if user.is_superuser:
                return redirect("admin_panel_dashboard")
            role = _role_for(user)
            if role == "operator":
                return redirect("operator_dashboard")
            if role == "seller":
                return redirect("seller_dashboard")
            return redirect("buyer_dashboard")
    else:
        form = LoginForm(request)
    return render(request, "marketplace/login.html", {"form": form})


def logout_view(request: HttpRequest) -> HttpResponse:
    logout(request)
    messages.info(request, "Вы вышли из системы.")
    return redirect("home")


def demo_login(request: HttpRequest) -> HttpResponse:
    """Быстрый вход в демо-кабинет по роли."""
    from django.contrib.auth import authenticate

    DEMO_USERS = {
        "seller": ("demo_seller", "seller_dashboard"),
        "buyer": ("demo_buyer", "buyer_dashboard"),
        "operator": ("demo_operator", "operator_dashboard"),
    }
    role = request.GET.get("role", "")

    # Admin demo login — use the superuser account
    if role == "admin":
        admin_user = User.objects.filter(is_superuser=True).first()
        if admin_user:
            login(request, admin_user)
            return redirect("admin_panel_dashboard")
        return redirect("login")

    entry = DEMO_USERS.get(role)
    if not entry:
        return redirect("login")
    username, redirect_to = entry
    user = authenticate(request, username=username, password="demo12345")
    if user is None:
        return redirect("login")
    login(request, user)
    return redirect(redirect_to)


def catalog(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect("admin_panel_catalog")
    _seed_if_empty()
    query = request.GET.get("q", "").strip()
    condition = request.GET.get("condition", "").strip()
    category_slug = request.GET.get("category", "").strip()
    brand_slug = request.GET.get("brand", "").strip()
    min_price_raw = request.GET.get("min_price", "").strip()
    max_price_raw = request.GET.get("max_price", "").strip()
    in_stock = request.GET.get("in_stock", "").strip() == "1"
    sort = request.GET.get("sort", "newest").strip() or "newest"

    parts = Part.objects.filter(is_active=True, price__gt=0).select_related("category", "brand")
    if query:
        parts = parts.filter(Q(title__icontains=query) | Q(oem_number__icontains=query))
    if condition:
        parts = parts.filter(condition=condition)
    if category_slug:
        parts = parts.filter(category__slug=category_slug)
    if brand_slug:
        parts = parts.filter(brand__slug=brand_slug)
    if in_stock:
        parts = parts.filter(stock_quantity__gt=0)
    if min_price_raw:
        try:
            parts = parts.filter(price__gte=Decimal(min_price_raw))
        except Exception:
            pass
    if max_price_raw:
        try:
            parts = parts.filter(price__lte=Decimal(max_price_raw))
        except Exception:
            pass

    ordering_map = {
        "newest": "-created_at",
        "price_asc": "price",
        "price_desc": "-price",
        "stock_desc": "-stock_quantity",
    }
    ordering = ordering_map.get(sort, "-created_at")

    paginator = Paginator(parts.order_by(ordering), 48)
    page_obj = paginator.get_page(request.GET.get("page"))
    categories = Category.objects.all().order_by("name")
    brands = Brand.objects.all().order_by("name")
    compare_ids = set(_get_compare_ids(request))
    page_params = request.GET.copy()
    page_params.pop("page", None)
    pagination_query = page_params.urlencode()
    return render(
        request,
        "marketplace/catalog.html",
        {
            "parts": page_obj.object_list,
            "page_obj": page_obj,
            "categories": categories,
            "brands": brands,
            "query": query,
            "selected_condition": condition,
            "selected_category": category_slug,
            "selected_brand": brand_slug,
            "selected_min_price": min_price_raw,
            "selected_max_price": max_price_raw,
            "selected_in_stock": in_stock,
            "selected_sort": sort,
            "pagination_query": pagination_query,
            "compare_ids": compare_ids,
        },
    )


def part_detail(request: HttpRequest, slug: str) -> HttpResponse:
    _seed_if_empty()
    part = get_object_or_404(Part.objects.select_related("category", "brand"), slug=slug, is_active=True, price__gt=0)
    related = Part.objects.filter(category=part.category, is_active=True, price__gt=0).select_related("brand").exclude(id=part.id)[:4]
    return render(request, "marketplace/part_detail.html", {"part": part, "related": related})


@require_POST
def cart_add(request: HttpRequest, part_id: int) -> HttpResponse:
    part = get_object_or_404(Part, id=part_id, is_active=True, price__gt=0)
    cart = _get_cart(request)
    current = int(cart.get(str(part.id), 0))
    if current < part.stock_quantity:
        cart[str(part.id)] = current + 1
    _set_cart(request, cart)
    return redirect(request.POST.get("next") or "cart")


@require_POST
def cart_remove(request: HttpRequest, part_id: int) -> HttpResponse:
    cart = _get_cart(request)
    if str(part_id) in cart:
        del cart[str(part_id)]
        _set_cart(request, cart)
    return redirect("cart")


def cart_view(request: HttpRequest) -> HttpResponse:
    _seed_if_empty()
    rows, total = _cart_rows(request)
    return render(request, "marketplace/cart.html", {"rows": rows, "total": total})


def compare_view(request: HttpRequest) -> HttpResponse:
    _seed_if_empty()
    ids = _get_compare_ids(request)
    if not ids:
        return render(request, "marketplace/compare.html", {"parts": [], "has_items": False})
    parts = list(Part.objects.filter(id__in=ids, is_active=True, price__gt=0).select_related("brand", "category"))
    parts.sort(key=lambda p: ids.index(p.id) if p.id in ids else 10**9)
    return render(request, "marketplace/compare.html", {"parts": parts, "has_items": bool(parts)})


@require_POST
def compare_add(request: HttpRequest, part_id: int) -> HttpResponse:
    part = get_object_or_404(Part, id=part_id, is_active=True, price__gt=0)
    ids = _get_compare_ids(request)
    if part.id not in ids:
        if len(ids) >= 4:
            ids.pop(0)
        ids.append(part.id)
    _set_compare_ids(request, ids)
    return redirect(request.POST.get("next") or "compare")


@require_POST
def compare_remove(request: HttpRequest, part_id: int) -> HttpResponse:
    ids = _get_compare_ids(request)
    ids = [x for x in ids if x != int(part_id)]
    _set_compare_ids(request, ids)
    return redirect(request.POST.get("next") or "compare")


@require_POST
def compare_clear(request: HttpRequest) -> HttpResponse:
    _set_compare_ids(request, [])
    return redirect("compare")


def checkout(request: HttpRequest) -> HttpResponse:
    _seed_if_empty()
    rows, total = _cart_rows(request)
    if not rows:
        return redirect("catalog")

    initial = {}
    if request.user.is_authenticated:
        full_name = " ".join(x for x in [request.user.first_name, request.user.last_name] if x).strip()
        initial = {
            "customer_name": full_name or request.user.username,
            "customer_email": request.user.email,
        }

    if request.method == "POST":
        form = CheckoutForm(request.POST)
        if form.is_valid():
            try:
                order = _create_order_from_rows(
                    rows=rows,
                    total=total,
                    customer_name=form.cleaned_data["customer_name"],
                    customer_email=form.cleaned_data["customer_email"],
                    customer_phone=form.cleaned_data["customer_phone"],
                    delivery_address=form.cleaned_data["delivery_address"],
                    buyer=request.user if request.user.is_authenticated else None,
                    source="buyer" if request.user.is_authenticated else "system",
                )
                _set_cart(request, {})
                return redirect(f"/dashboard/?order_created={order.id}")
            except ValueError as exc:
                messages.error(request, f"Заказ не создан: {exc}")
    else:
        form = CheckoutForm(initial=initial)

    return render(request, "marketplace/checkout.html", {"rows": rows, "total": total, "form": form})


@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    if request.user.is_superuser:
        return redirect("admin_panel_dashboard")
    role = _role_for(request.user)
    if role == "seller":
        return redirect("dashboard_seller")
    if role == "operator":
        return redirect("operator_dashboard")
    return redirect("dashboard_buyer")


@login_required
def kpi_reports(request: HttpRequest) -> HttpResponse:
    if request.user.is_superuser:
        return redirect("admin_panel_analytics")
    role = _role_for(request.user)
    is_seller = role == "seller"
    if is_seller and not _has_seller_permission(request.user, "can_view_analytics"):
        messages.error(request, "Нет прав на аналитику.")
        return redirect("dashboard")

    if is_seller:
        scoped_orders = Order.objects.filter(items__part__seller=request.user).distinct()
        scoped_parts = Part.objects.filter(seller=request.user)
    else:
        scoped_orders = Order.objects.filter(buyer=request.user)
        scoped_parts = Part.objects.none()

    total_orders = scoped_orders.count()
    completed_orders = scoped_orders.filter(status="completed").count()
    cancelled_orders = scoped_orders.filter(status="cancelled").count()
    delivered_orders = scoped_orders.filter(status="delivered").count()
    breached_orders = scoped_orders.filter(sla_status="breached").count()
    total_revenue = sum((o.total_amount for o in scoped_orders[:500]), Decimal("0.00"))
    total_claims = OrderClaim.objects.filter(order__in=scoped_orders).count()
    open_claims = OrderClaim.objects.filter(order__in=scoped_orders, status__in=["open", "in_review"]).count()

    conversion = Decimal("0.00")
    if total_orders:
        conversion = (Decimal(completed_orders) / Decimal(total_orders) * Decimal("100")).quantize(Decimal("0.01"))

    cards = [
        {"label": "Orders", "value": total_orders},
        {"label": "Completed", "value": completed_orders},
        {"label": "Delivered", "value": delivered_orders},
        {"label": "Cancelled", "value": cancelled_orders},
        {"label": "SLA Breached", "value": breached_orders},
        {"label": "Open Claims", "value": open_claims},
        {"label": "Total Claims", "value": total_claims},
        {"label": "Revenue (USD)", "value": total_revenue},
        {"label": "Completion rate %", "value": conversion},
    ]

    return render(
        request,
        "marketplace/kpi_reports.html",
        {
            "role": role,
            "cards": cards,
            "parts_count": scoped_parts.count() if is_seller else 0,
            "recent_orders": scoped_orders.order_by("-id")[:20],
        },
    )


@login_required
def kpi_reports_export_csv(request: HttpRequest) -> HttpResponse:
    role = _role_for(request.user)
    is_seller = role == "seller"
    if is_seller and not _has_seller_permission(request.user, "can_view_analytics"):
        messages.error(request, "Нет прав на аналитику.")
        return redirect("dashboard")
    scoped_orders = Order.objects.filter(items__part__seller=request.user).distinct() if is_seller else Order.objects.filter(buyer=request.user)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="kpi_orders_export.csv"'
    writer = csv.writer(response)
    writer.writerow(["order_id", "status", "payment_status", "sla_status", "total_amount", "logistics_cost", "created_at"])
    for order in scoped_orders.order_by("-id")[:5000]:
        writer.writerow(
            [
                order.id,
                order.status,
                order.payment_status,
                order.sla_status,
                order.total_amount,
                order.logistics_cost,
                order.created_at.isoformat(),
            ]
        )
    return response


@login_required
def claims_export_csv(request: HttpRequest) -> HttpResponse:
    role = _role_for(request.user)
    is_seller = role == "seller"
    if is_seller and not _has_seller_permission(request.user, "can_view_analytics"):
        messages.error(request, "Нет прав на аналитику.")
        return redirect("dashboard")
    scoped_orders = Order.objects.filter(items__part__seller=request.user).distinct() if is_seller else Order.objects.filter(buyer=request.user)
    claims = OrderClaim.objects.filter(order__in=scoped_orders).select_related("order", "opened_by", "resolved_by").order_by("-id")[:5000]

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="claims_export.csv"'
    writer = csv.writer(response)
    writer.writerow(["claim_id", "order_id", "status", "title", "opened_by", "resolved_by", "created_at", "updated_at"])
    for claim in claims:
        writer.writerow(
            [
                claim.id,
                claim.order_id,
                claim.status,
                claim.title,
                claim.opened_by.username if claim.opened_by else "",
                claim.resolved_by.username if claim.resolved_by else "",
                claim.created_at.isoformat(),
                claim.updated_at.isoformat(),
            ]
        )
    return response


@login_required
def dashboard_buyer(request: HttpRequest) -> HttpResponse:
    return redirect("buyer_dashboard")


@seller_required
def dashboard_seller(request: HttpRequest) -> HttpResponse:
    return redirect("seller_dashboard")


@seller_required
def _build_seller_catalog_context(request: HttpRequest) -> dict:
    parts_qs = _apply_seller_brand_scope(
        request.user,
        Part.objects.filter(seller=request.user).select_related("category", "brand"),
    )
    successful_import_statuses = {ImportJob.Status.COMPLETED, ImportJob.Status.PARTIAL_SUCCESS}
    latest_successful_import = (
        ImportJob.objects.filter(supplier=request.user, status__in=successful_import_statuses)
        .select_related("source_file")
        .order_by("-created_at")
        .first()
    )
    query = (request.GET.get("q") or "").strip()
    recent_only = (request.GET.get("recent") or "").strip().lower() in {"1", "true", "yes", "24h"}
    last_import_only = (request.GET.get("last_import") or "").strip().lower() in {"1", "true", "yes", "latest"}
    import_run_id_raw = (request.GET.get("import_run") or "").strip()
    preview_id_raw = (request.GET.get("preview_id") or "").strip()
    selected_import_run = None
    active_preview = None
    if import_run_id_raw.isdigit():
        selected_import_run = (
            ImportJob.objects.filter(supplier=request.user, id=int(import_run_id_raw))
            .select_related("source_file")
            .first()
        )
    if preview_id_raw.isdigit():
        active_preview = ImportPreviewSession.objects.filter(supplier=request.user, id=int(preview_id_raw)).first()
    if selected_import_run:
        import_started = selected_import_run.started_at or selected_import_run.created_at
        import_finished = selected_import_run.finished_at or selected_import_run.updated_at
        import_window_start = import_started - timedelta(minutes=10)
        import_window_end = import_finished + timedelta(minutes=10)
        parts_qs = parts_qs.filter(data_updated_at__gte=import_window_start, data_updated_at__lte=import_window_end)
    if last_import_only and latest_successful_import:
        import_started = latest_successful_import.started_at or latest_successful_import.created_at
        import_finished = latest_successful_import.finished_at or latest_successful_import.updated_at
        import_window_start = import_started - timedelta(minutes=10)
        import_window_end = import_finished + timedelta(minutes=10)
        parts_qs = parts_qs.filter(data_updated_at__gte=import_window_start, data_updated_at__lte=import_window_end)
    if recent_only:
        parts_qs = parts_qs.filter(data_updated_at__gte=timezone.now() - timedelta(hours=24))
    if query:
        parts_qs = parts_qs.filter(Q(title__icontains=query) | Q(oem_number__icontains=query) | Q(brand__name__icontains=query))
    brand_filter = (request.GET.get("brand") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    if brand_filter:
        parts_qs = parts_qs.filter(brand__id=brand_filter)
    if status_filter:
        parts_qs = parts_qs.filter(availability_status=status_filter)
    # Brand list for filter dropdown
    brand_list = (
        _apply_seller_brand_scope(request.user, Part.objects.filter(seller=request.user))
        .values_list("brand__id", "brand__name")
        .distinct()
        .order_by("brand__name")
    )
    brand_list = [{"id": bid, "name": bname} for bid, bname in brand_list if bname]
    parts_qs = parts_qs.order_by("-data_updated_at", "-id")
    paginator = Paginator(parts_qs, 50)
    page_number = request.GET.get("page") or 1
    parts_page = paginator.get_page(page_number)
    for part in parts_page.object_list:
        stale_snapshot = _part_stale_snapshot(part)
        part.stale_days = stale_snapshot["days"]
        part.stale_state = stale_snapshot["state"]
        part.stale_label = stale_snapshot["label"]
        part.is_stale = stale_snapshot["is_stale"]
    bulk_form = SellerBulkUploadForm()
    profile = _profile_for(request.user)
    import_runs = (
        ImportJob.objects.filter(supplier=request.user)
        .select_related("source_file")
        .order_by("-created_at")[:10]
    )
    projection = DashboardProjection.objects.filter(supplier=request.user).first()
    if projection is None:
        projection = refresh_supplier_dashboard_projection(request.user)
    seller_rfqs_count = _seller_rfqs_qs(request.user).count()
    seller_orders_count = Order.objects.filter(items__part__seller=request.user).distinct().count()
    recent_updates_count = _apply_seller_brand_scope(
        request.user,
        Part.objects.filter(seller=request.user),
    ).filter(data_updated_at__gte=timezone.now() - timedelta(hours=24)).count()
    preview_header_options: list[str] = []
    preview_rows_matrix: list[list[str]] = []
    preview_mapping = {}
    mapping_rows: list[dict[str, object]] = []
    if active_preview:
        preview_header_options = list(
            dict.fromkeys(
                (list(active_preview.sample_rows[0].keys()) if active_preview.sample_rows else [])
                + list(active_preview.detected_columns.values())
            )
        )
        preview_rows_matrix = [[row.get(header, "") for header in preview_header_options] for row in (active_preview.sample_rows or [])]
        preview_mapping = active_preview.column_mapping or active_preview.detected_columns or {}
    for field_key, field_label in SELLER_IMPORT_MAPPING_FIELDS:
        mapping_rows.append(
            {
                "key": field_key,
                "label": field_label,
                "selected": preview_mapping.get(field_key, ""),
            }
        )
    return {
        "parts": parts_page,
        "parts_total": paginator.count,
        "query": query,
        "recent_only": recent_only,
        "last_import_only": last_import_only,
        "selected_import_run": selected_import_run,
        "active_preview": active_preview,
        "preview_header_options": preview_header_options,
        "preview_rows_matrix": preview_rows_matrix,
        "preview_mapping": preview_mapping,
        "mapping_fields": SELLER_IMPORT_MAPPING_FIELDS,
        "mapping_rows": mapping_rows,
        "bulk_form": bulk_form,
        "profile": profile,
        "upload_report": request.session.get("seller_upload_report"),
        "import_runs": import_runs,
        "latest_successful_import": latest_successful_import,
        "dashboard_projection": projection,
        "seller_rfqs_count": seller_rfqs_count,
        "seller_orders_count": seller_orders_count,
        "recent_updates_count": recent_updates_count,
        "return_qs": request.GET.urlencode(),
        "brand_list": brand_list,
        "brand_filter": brand_filter,
        "status_filter": status_filter,
    }


@seller_required
def seller_dashboard(request: HttpRequest) -> HttpResponse:
    if request.user.is_superuser:
        return redirect("admin_panel_dashboard")
    # Avoid rebuilding a heavy dashboard projection on every request.
    # Keep it consistent with the API behavior (refresh if stale).
    from dashboard.models import DashboardProjection as SupplierDashboardProjection

    stale_after = timedelta(minutes=5)
    projection = SupplierDashboardProjection.objects.filter(supplier=request.user, user=request.user).first()
    is_stale = True
    if projection and projection.updated_at:
        is_stale = projection.updated_at < (timezone.now() - stale_after)
    if projection is None or is_stale:
        projection = DashboardProjectionBuilder().build(supplier=request.user, user=request.user)
    dashboard_payload = DashboardProjectionBuilder().payload(projection)
    response = render(
        request,
        _tpl(request.user, "seller/dashboard/index.html"),
        {
            "dashboard_payload": dashboard_payload,
            "seller_page_title": "Кабинет поставщика",
            "seller_page_subtitle": "Главная рабочая панель: что требует внимания сейчас и куда перейти дальше.",
            "seller_active_nav": "dashboard",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
            ],
        },
    )
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    return response


@seller_required
def seller_product_list(request: HttpRequest) -> HttpResponse:
    return render(
        request,
        _tpl(request.user, "seller/products/catalog.html"),
        {
            **_build_seller_catalog_context(request),
            "seller_page_title": "Товары и прайсы",
            "seller_page_subtitle": "Загрузка прайсов, preview, история импортов, каталог и массовые действия в одном модуле.",
            "seller_active_nav": "products",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Товары и прайсы", "url": reverse("seller_product_list")},
            ],
        },
    )



@seller_required
def seller_orders(request: HttpRequest) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_orders"):
        messages.error(request, "Нет прав на работу с заказами.")
        return redirect("seller_dashboard")

    orders_qs = (
        Order.objects.filter(items__part__seller=request.user)
        .distinct()
        .prefetch_related("items__part", "documents", "claims")
        .order_by("-created_at")
    )
    query = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    allowed_statuses = {code for code, _ in Order.STATUS_CHOICES}
    if status and status in allowed_statuses:
        orders_qs = orders_qs.filter(status=status)
    if query:
        filter_q = (
            Q(customer_name__icontains=query)
            | Q(customer_email__icontains=query)
            | Q(items__part__oem_number__icontains=query)
            | Q(items__part__title__icontains=query)
        )
        if query.isdigit():
            filter_q = filter_q | Q(id=int(query))
        orders_qs = orders_qs.filter(filter_q).distinct()

    paginator = Paginator(orders_qs, 30)
    page_number = request.GET.get("page") or 1
    orders_page = paginator.get_page(page_number)

    rows = []
    for order in orders_page:
        seller_items = [item for item in order.items.all() if item.part and item.part.seller_id == request.user.id]
        if not seller_items:
            continue
        open_claims = sum(1 for claim in order.claims.all() if claim.status in {"open", "in_review"})
        rows.append(
            {
                "order": order,
                "items_count": len(seller_items),
                "units_total": sum(int(item.quantity) for item in seller_items),
                "documents_count": len(order.documents.all()),
                "open_claims_count": open_claims,
            }
        )

    return render(
        request,
        _tpl(request.user, "seller/orders/list.html"),
        {
            "rows": rows,
            "orders": orders_page,
            "orders_total": paginator.count,
            "query": query,
            "status": status,
            "status_choices": Order.STATUS_CHOICES,
            "seller_page_title": "Заказы",
            "seller_page_subtitle": "Список заказов по вашим товарам, фильтры, статусы и переход в карточку заказа.",
            "seller_active_nav": "orders",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Заказы", "url": reverse("seller_orders")},
            ],
        },
    )


@seller_required
def seller_sla(request: HttpRequest) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_orders"):
        messages.error(request, "Нет прав на просмотр SLA.")
        return redirect("seller_dashboard")

    orders_qs = Order.objects.filter(
        items__part__seller=request.user
    ).distinct().prefetch_related("items__part").order_by("-created_at")

    # Recalc SLA
    for order in orders_qs[:100]:
        _recalc_order_sla(order)

    # Kanban columns — новые этапы из таблицы "Этапы ЛК"
    # statuses: список DB-статусов, попадающих в эту колонку
    # drop_status: статус, выставляемый при drag-and-drop в эту колонку
    kanban_columns_cfg = [
        {
            "key": "pending",
            "label": "Ожидание оплаты",
            "statuses": ["pending"],
            "sla_hours": 48,
            "trigger": "Счёт сформирован",
            "action": "Кнопка «Отправить счёт покупателю»",
            "action_type": "Кнопка",
            "who": "Продавец / система",
        },
        {
            "key": "confirmed",
            "label": "Формирование заказа",
            "statuses": ["reserve_paid", "confirmed", "in_production", "ready_to_ship"],
            "sla_hours": 168,
            "trigger": "Предоплата поступила → Груз готов к отгрузке",
            "action": "Фиксация оплаты → Кнопка «Передано в логистику»",
            "action_type": "Автомат / Кнопка",
            "who": "Фингрид / Поставщик",
        },
        {
            "key": "transit_abroad",
            "label": "Логистика (Зарубеж)",
            "statuses": ["transit_abroad"],
            "sla_hours": 240,
            "trigger": "Фактическая передача перевозчику",
            "action": "Сканирование QR-кода отгрузки",
            "action_type": "QR-скан",
            "who": "Зарубежный логист",
        },
        {
            "key": "customs",
            "label": "Таможенное оформление",
            "statuses": ["customs"],
            "sla_hours": 48,
            "trigger": "Таможня завершена",
            "action": "Кнопка «Груз растаможен» + декларация",
            "action_type": "Кнопка + документ",
            "who": "Таможенный брокер",
        },
        {
            "key": "transit_rf",
            "label": "Логистика (РФ)",
            "statuses": ["transit_rf"],
            "sla_hours": 24,
            "trigger": "Передача в логистику РФ",
            "action": "Сканирование QR-кода передачи",
            "action_type": "QR-скан",
            "who": "РФ-логист",
        },
        {
            "key": "issuing",
            "label": "Выдача",
            "statuses": ["issuing", "shipped"],
            "sla_hours": 24,
            "trigger": "Передача на приёмку",
            "action": "Сканирование QR-кода выдачи",
            "action_type": "QR-скан",
            "who": "Оператор платформы",
        },
        {
            "key": "delivered",
            "label": "Доставлен",
            "statuses": ["delivered"],
            "sla_hours": 72,
            "trigger": "Фактическая приёмка груза",
            "action": "QR-код / документы / видеоприёмка",
            "action_type": "QR / документ / видео",
            "who": "Заказчик / оператор",
        },
        {
            "key": "completed",
            "label": "Заказ закрыт",
            "statuses": ["completed"],
            "sla_hours": 1,
            "trigger": "Документы приняты",
            "action": "Автоматическое закрытие",
            "action_type": "Автомат",
            "who": "Система",
        },
    ]
    kanban_statuses = [(col["key"], col["label"]) for col in kanban_columns_cfg]

    columns = []
    for col_cfg in kanban_columns_cfg:
        status_orders = []
        for order in orders_qs.filter(status__in=col_cfg["statuses"]):
            seller_items = [item for item in order.items.all() if item.part and item.part.seller_id == request.user.id]
            status_orders.append({
                "order": order,
                "items_count": len(seller_items),
                "units_total": sum(int(item.quantity) for item in seller_items),
            })
        columns.append({
            "key": col_cfg["key"],
            "label": col_cfg["label"],
            "orders": status_orders,
            "count": len(status_orders),
            "sla_hours": col_cfg["sla_hours"],
            "trigger": col_cfg.get("trigger", ""),
            "action": col_cfg.get("action", ""),
            "action_type": col_cfg.get("action_type", ""),
            "who": col_cfg.get("who", ""),
        })

    # Timeline data — current_step по индексу колонки канбана
    status_to_col_idx = {}
    for idx, col_cfg in enumerate(kanban_columns_cfg):
        for s in col_cfg["statuses"]:
            status_to_col_idx[s] = idx

    timeline_orders = []
    for col in columns:
        for row in col["orders"]:
            current_idx = status_to_col_idx.get(row["order"].status, -1)
            timeline_orders.append({
                "order": row["order"],
                "items_count": row["items_count"],
                "current_step": current_idx,
            })

    # SLA KPI metrics
    all_orders = list(orders_qs)
    sla_on_track = sum(1 for o in all_orders if o.sla_status == "on_track")
    sla_at_risk = sum(1 for o in all_orders if o.sla_status == "at_risk")
    sla_breached = sum(1 for o in all_orders if o.sla_status == "breached")

    # Average time per stage (in hours, from events)
    from django.utils import timezone as tz
    now = tz.now()
    avg_confirm_hours = 0
    avg_production_hours = 0
    avg_ship_hours = 0
    confirmed_orders = [o for o in all_orders if o.status not in ("pending",)]
    if confirmed_orders:
        total_h = sum(
            ((o.supplier_confirm_deadline - o.created_at).total_seconds() / 3600 if o.supplier_confirm_deadline else 24)
            for o in confirmed_orders
        )
        avg_confirm_hours = round(total_h / len(confirmed_orders))
    shipped_orders = [o for o in all_orders if o.status in ("shipped", "delivered", "completed")]
    if shipped_orders:
        total_h = sum(
            ((o.ship_deadline - o.created_at).total_seconds() / 3600 if o.ship_deadline else 72)
            for o in shipped_orders
        )
        avg_ship_hours = round(total_h / len(shipped_orders))
    production_orders = [o for o in all_orders if o.status in ("in_production", "ready_to_ship", "shipped", "delivered")]
    if production_orders:
        avg_production_hours = round((avg_ship_hours + avg_confirm_hours) / 2) if avg_ship_hours else 48

    # Stage time analytics — time each order spent at each stage
    from collections import defaultdict
    order_ids = [o.id for o in all_orders]
    stage_events = OrderEvent.objects.filter(
        order_id__in=order_ids,
        event_type="status_changed",
    ).order_by("order_id", "created_at").values("order_id", "meta", "created_at")

    events_by_order = defaultdict(list)
    for ev in stage_events:
        events_by_order[ev["order_id"]].append(ev)

    status_order_list = [s[0] for s in kanban_statuses]
    stage_analytics = []
    for order in all_orders:
        evs = events_by_order.get(order.id, [])
        # Map: status key → datetime when order entered that status
        status_start = {"pending": order.created_at}
        for ev in evs:
            to_s = (ev["meta"] or {}).get("to")
            if to_s and to_s in status_order_list:
                if to_s not in status_start:
                    status_start[to_s] = ev["created_at"]

        stage_times = []
        for i, (sk, sl) in enumerate(kanban_statuses):
            if sk not in status_start:
                stage_times.append(None)
                continue
            start = status_start[sk]
            # end = when moved to next kanban status
            end = None
            for j in range(i + 1, len(kanban_statuses)):
                next_sk = kanban_statuses[j][0]
                if next_sk in status_start:
                    end = status_start[next_sk]
                    break
            if end is None and order.status == sk:
                end = now
            if start and end:
                hours = round((end - start).total_seconds() / 3600, 1)
                stage_times.append(hours)
            else:
                stage_times.append(None)

        total_hours = sum(h for h in stage_times if h is not None)
        stage_analytics.append({
            "order": order,
            "stage_times": stage_times,
            "total_hours": round(total_hours, 1),
            "current_stage_idx": status_order_list.index(order.status) if order.status in status_order_list else -1,
        })

    return render(
        request,
        "seller/sla/list.html",
        {
            "columns": columns,
            "kanban_statuses": kanban_statuses,
            "timeline_orders": timeline_orders,
            "orders_total": orders_qs.count(),
            "sla_on_track": sla_on_track,
            "sla_at_risk": sla_at_risk,
            "sla_breached": sla_breached,
            "avg_confirm_hours": avg_confirm_hours,
            "avg_production_hours": avg_production_hours,
            "avg_ship_hours": avg_ship_hours,
            "stage_analytics": stage_analytics,
            "seller_page_title": "Контроль SLA",
            "seller_page_subtitle": "Канбан-доска поставок — перетаскивайте карточки между этапами.",
            "seller_active_nav": "sla",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Контроль SLA", "url": reverse("seller_sla")},
            ],
        },
    )


@seller_required
def seller_qr_control(request: HttpRequest) -> HttpResponse:
    """Страница QR-контроля: генерация кодов, история сканирований."""
    seller = request.user

    # Заказы продавца (активные)
    active_orders = Order.objects.filter(
        items__part__seller=seller,
    ).distinct().exclude(status__in=["cancelled", "completed"]).order_by("-created_at")

    # История сканирований QR — события с qr_code в meta
    qr_base = (
        OrderEvent.objects.filter(
            order__items__part__seller=seller,
            event_type="status_changed",
        )
        .exclude(meta__qr_code=None)
        .exclude(meta__qr_code="")
    )

    # Статистика
    total_scans = qr_base.count()
    from django.utils import timezone as tz
    today_start = tz.now().replace(hour=0, minute=0, second=0, microsecond=0)
    scans_today = qr_base.filter(created_at__gte=today_start).count()
    orders_with_qr = qr_base.values("order_id").distinct().count()

    qr_scan_events = (
        qr_base.select_related("order", "actor").order_by("-created_at")[:50]
    )

    # Статусы которые используют QR по бизнес-логике
    QR_STAGES = {
        "transit_abroad": {
            "label": "Логистика (Зарубеж)",
            "action": "Сканирование QR-кода отгрузки",
            # самолёт
            "svg": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="#64B5F6" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M22 2L11 13"/><path d="M22 2L15 22l-4-9-9-4 19-7z"/></svg>',
        },
        "transit_rf": {
            "label": "Логистика (РФ)",
            "action": "Сканирование при приёме груза",
            # грузовик
            "svg": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="#64B5F6" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><rect x="1" y="3" width="15" height="13" rx="1"/><path d="M16 8h4l3 3v5h-7V8z"/><circle cx="5.5" cy="18.5" r="2.5"/><circle cx="18.5" cy="18.5" r="2.5"/></svg>',
        },
        "issuing": {
            "label": "Выдача",
            "action": "QR-скан при получении заказа",
            # коробка с рукой / выдача
            "svg": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="#64B5F6" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M21 16V8a2 2 0 00-1-1.73l-7-4a2 2 0 00-2 0l-7 4A2 2 0 002 8v8a2 2 0 001 1.73l7 4a2 2 0 002 0l7-4A2 2 0 0021 16z"/><polyline points="3.27 6.96 12 12.01 20.73 6.96"/><line x1="12" y1="22.08" x2="12" y2="12"/></svg>',
        },
        "delivered": {
            "label": "Доставлен",
            "action": "Подтверждение доставки",
            # локация / точка назначения
            "svg": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="#64B5F6" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M21 10c0 7-9 13-9 13s-9-6-9-13a9 9 0 0118 0z"/><circle cx="12" cy="10" r="3"/></svg>',
        },
    }

    # Timeline data — те же этапы что и в канбане SLA
    kanban_columns_cfg = [
        {"key": "pending", "label": "Ожидание", "statuses": ["pending"]},
        {"key": "confirmed", "label": "Формирование", "statuses": ["reserve_paid", "confirmed", "in_production", "ready_to_ship"]},
        {"key": "transit_abroad", "label": "Зарубеж", "statuses": ["transit_abroad"]},
        {"key": "customs", "label": "Таможня", "statuses": ["customs"]},
        {"key": "transit_rf", "label": "РФ", "statuses": ["transit_rf"]},
        {"key": "issuing", "label": "Выдача", "statuses": ["issuing", "shipped"]},
        {"key": "delivered", "label": "Доставлен", "statuses": ["delivered"]},
        {"key": "completed", "label": "Закрыт", "statuses": ["completed"]},
    ]
    kanban_statuses = [(col["key"], col["label"]) for col in kanban_columns_cfg]

    status_to_col_idx = {}
    for idx, col_cfg in enumerate(kanban_columns_cfg):
        for s in col_cfg["statuses"]:
            status_to_col_idx[s] = idx

    # Все заказы продавца (включая completed для таймлайна)
    all_seller_orders = Order.objects.filter(
        items__part__seller=seller,
    ).distinct().exclude(status="cancelled").order_by("-created_at")

    timeline_orders = []
    for order in all_seller_orders:
        current_idx = status_to_col_idx.get(order.status, -1)
        timeline_orders.append({
            "order": order,
            "current_step": current_idx,
        })

    return render(
        request,
        "seller/qr/list.html",
        {
            "active_orders": active_orders,
            "qr_scan_events": qr_scan_events,
            "total_scans": total_scans,
            "scans_today": scans_today,
            "orders_with_qr": orders_with_qr,
            "active_orders_count": active_orders.count(),
            "qr_stages": QR_STAGES,
            "timeline_orders": timeline_orders,
            "kanban_statuses": kanban_statuses,
            "seller_page_title": "QR-контроль",
            "seller_page_subtitle": "Генерация QR-кодов и отслеживание заказов",
            "seller_active_nav": "qr",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "QR-контроль", "url": reverse("seller_qr_control")},
            ],
        },
    )


@seller_required
def seller_rating(request: HttpRequest) -> HttpResponse:
    """Рейтинг поставщика: разбивка, метрики, предупреждения, советы."""
    from datetime import timedelta
    from decimal import Decimal

    seller = request.user
    profile = seller.profile

    # Rating breakdown
    rating_score = float(profile.rating_score)
    external_score = float(profile.external_score)
    behavioral_score = float(profile.behavioral_score)
    supplier_status = profile.supplier_status
    status_display = profile.get_supplier_status_display()

    # Orders for metrics
    now = timezone.now()
    d30_ago = now - timedelta(days=30)
    d60_ago = now - timedelta(days=60)

    seller_orders = Order.objects.filter(items__part__seller=seller).distinct()
    orders_30d = seller_orders.filter(created_at__gte=d30_ago)
    orders_prev = seller_orders.filter(created_at__gte=d60_ago, created_at__lt=d30_ago)

    # SLA
    orders_30d_list = list(orders_30d.only("sla_status", "sla_breaches_count", "status"))
    total_30d = len(orders_30d_list)
    on_track_30d = sum(1 for o in orders_30d_list if o.sla_status == "on_track")
    at_risk_30d = sum(1 for o in orders_30d_list if o.sla_status == "at_risk")
    breached_30d = sum(1 for o in orders_30d_list if o.sla_status == "breached")
    sla_pct = round((on_track_30d / total_30d * 100) if total_30d else 100, 1)

    orders_prev_list = list(orders_prev.only("sla_status"))
    total_prev = len(orders_prev_list)
    on_track_prev = sum(1 for o in orders_prev_list if o.sla_status == "on_track")
    sla_pct_prev = round((on_track_prev / total_prev * 100) if total_prev else 100, 1)
    sla_trend = round(sla_pct - sla_pct_prev, 1)

    # Conversion (RFQ → Order)
    from marketplace.models import RFQ
    rfqs_30d = RFQ.objects.filter(created_at__gte=d30_ago).count()
    orders_created_30d = orders_30d.count()
    conversion_pct = round((orders_created_30d / rfqs_30d * 100) if rfqs_30d else 0, 1)

    rfqs_prev = RFQ.objects.filter(created_at__gte=d60_ago, created_at__lt=d30_ago).count()
    orders_created_prev = orders_prev.count()
    conv_prev = round((orders_created_prev / rfqs_prev * 100) if rfqs_prev else 0, 1)
    conv_trend = round(conversion_pct - conv_prev, 1)

    # Claims
    seller_order_ids = list(seller_orders.values_list("id", flat=True)[:500])
    open_claims = OrderClaim.objects.filter(order_id__in=seller_order_ids, status__in=["open", "in_review"]).count()
    total_claims = OrderClaim.objects.filter(order_id__in=seller_order_ids).count()

    # Cancellations
    cancelled_30d = orders_30d.filter(status="cancelled").count()

    # Total SLA breaches
    sla_breaches_total = sum(o.sla_breaches_count for o in orders_30d_list)

    # Rating events
    rating_events = list(
        SupplierRatingEvent.objects.filter(supplier=seller).order_by("-created_at")[:20]
    )

    # Warnings / action items
    warnings = []
    if breached_30d > 0:
        warnings.append({
            "level": "critical",
            "title": f"{breached_30d} нарушений SLA",
            "text": "Каждое нарушение снижает поведенческий рейтинг на ~2 балла. Ускорьте обработку заказов.",
            "icon": "alert",
        })
    if open_claims > 0:
        warnings.append({
            "level": "critical",
            "title": f"{open_claims} открытых рекламаций",
            "text": "Нерешённые рекламации снижают внешний рейтинг. Решите их как можно скорее.",
            "icon": "claim",
        })
    if at_risk_30d > 0:
        warnings.append({
            "level": "warning",
            "title": f"{at_risk_30d} заказов под угрозой SLA",
            "text": "Эти заказы скоро выйдут за пределы SLA. Примите меры сейчас.",
            "icon": "clock",
        })
    if conversion_pct < 50 and rfqs_30d > 0:
        warnings.append({
            "level": "warning",
            "title": f"Конверсия {conversion_pct}% ниже нормы",
            "text": "Проверьте цены и наличие товара. Целевой показатель — выше 50%.",
            "icon": "trend",
        })
    if rating_score < 80:
        warnings.append({
            "level": "warning",
            "title": "Рейтинг ниже порога «Надёжный»",
            "text": f"Текущий рейтинг {rating_score:.1f}. Нужно 80+ для статуса «Надёжный».",
            "icon": "star",
        })
    if cancelled_30d > 0:
        warnings.append({
            "level": "info",
            "title": f"{cancelled_30d} отмен за 30 дней",
            "text": "Отмены негативно влияют на поведенческий рейтинг.",
            "icon": "cancel",
        })
    if not warnings:
        warnings.append({
            "level": "success",
            "title": "Всё в порядке",
            "text": "Показатели в норме. Продолжайте поддерживать высокий уровень сервиса.",
            "icon": "check",
        })

    # Event type labels
    event_labels = dict(SupplierRatingEvent.EVENT_CHOICES)

    return render(
        request,
        "seller/rating/list.html",
        {
            "rating_score": rating_score,
            "external_score": external_score,
            "behavioral_score": behavioral_score,
            "supplier_status": supplier_status,
            "status_display": status_display,
            "sla_pct": sla_pct,
            "sla_trend": sla_trend,
            "conversion_pct": conversion_pct,
            "conv_trend": conv_trend,
            "open_claims": open_claims,
            "total_claims": total_claims,
            "cancelled_30d": cancelled_30d,
            "breached_30d": breached_30d,
            "at_risk_30d": at_risk_30d,
            "sla_breaches_total": sla_breaches_total,
            "rating_events": rating_events,
            "event_labels": event_labels,
            "warnings": warnings,
            "seller_page_title": "Рейтинг",
            "seller_page_subtitle": "Подробная разбивка рейтинга и рекомендации",
            "seller_active_nav": "rating",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Рейтинг", "url": reverse("seller_rating")},
            ],
        },
    )


@seller_required
@seller_required
def seller_negotiations(request: HttpRequest) -> HttpResponse:
    """Согласование: уровни скидок, лояльность, переторжка, чертежи."""
    return render(request, "seller/negotiations/list.html", {})


def seller_analytics(request: HttpRequest) -> HttpResponse:
    """Аналитика: отчёты, интеграции, рассылка, API."""
    return render(request, "seller/analytics/list.html", {})


def seller_team(request: HttpRequest) -> HttpResponse:
    """Команда: орг-схема, права, чат, задачи, активность, рейтинги."""
    return render(request, "seller/team/list.html", {})


def seller_integrations(request: HttpRequest) -> HttpResponse:
    """Интеграции: 1С, ТОИР, ERP, Битрикс24, индивидуальная."""
    return render(request, "seller/integrations/list.html", {})


def seller_logistics(request: HttpRequest) -> HttpResponse:
    """Логистика: карта, терминалы, отслеживание, калькулятор, аукцион."""
    return render(request, _tpl(request.user, "seller/logistics/list.html"), {})


def seller_reports(request: HttpRequest) -> HttpResponse:
    """Отчёты: сводные, продажи, финансовые, операционные, экспорт, расписание."""
    return render(request, "seller/reports/list.html", {
        "seller_active_nav": "reports",
        "history_reports": [],
    })


# ═══════════════════════════════════════════════════════════════════
# BUYER CABINET
# ═══════════════════════════════════════════════════════════════════

def buyer_dashboard(request: HttpRequest) -> HttpResponse:
    return render(request, _tpl(request.user, "buyer/dashboard/index.html"), {})

def buyer_catalog(request: HttpRequest) -> HttpResponse:
    return render(request, _tpl(request.user, "buyer/catalog/list.html"), {})

def buyer_rfq_list(request: HttpRequest) -> HttpResponse:
    return render(request, _tpl(request.user, "buyer/rfq/list.html"), {})

def buyer_orders(request: HttpRequest) -> HttpResponse:
    return render(request, _tpl(request.user, "buyer/orders/list.html"), {})

def buyer_shipments(request: HttpRequest) -> HttpResponse:
    return render(request, _tpl(request.user, "buyer/shipments/list.html"), {})

def buyer_claims(request: HttpRequest) -> HttpResponse:
    return render(request, "buyer/claims/list.html", {})

def buyer_suppliers(request: HttpRequest) -> HttpResponse:
    return render(request, "buyer/suppliers/list.html", {})

def buyer_negotiations(request: HttpRequest) -> HttpResponse:
    return render(request, "buyer/negotiations/list.html", {})

def buyer_finance(request: HttpRequest) -> HttpResponse:
    return render(request, "buyer/finance/list.html", {})

def buyer_analytics(request: HttpRequest) -> HttpResponse:
    return render(request, "buyer/analytics/list.html", {})


def seller_finance(request: HttpRequest) -> HttpResponse:
    """Финансовый кабинет поставщика: оплаты, документы, таймлайн."""
    from decimal import Decimal

    seller = request.user

    orders_qs = (
        Order.objects.filter(items__part__seller=seller)
        .distinct()
        .select_related("buyer")
        .prefetch_related("documents", "events", "items")
        .order_by("-created_at")
    )

    # Фильтр по статусу оплаты
    payment_filter = request.GET.get("payment", "")
    if payment_filter and payment_filter in dict(Order.PAYMENT_STATUS_CHOICES):
        orders_qs = orders_qs.filter(payment_status=payment_filter)

    # Поиск по номеру заказа
    search_q = request.GET.get("q", "").strip()
    if search_q:
        orders_qs = orders_qs.filter(id__icontains=search_q)

    orders_list = list(orders_qs[:100])

    # Метрики (считаем по всем заказам, без фильтра)
    all_orders = list(
        Order.objects.filter(items__part__seller=seller)
        .distinct()
        .only("total_amount", "reserve_amount", "payment_status", "reserve_paid_at")
    )
    total_revenue = sum((o.total_amount for o in all_orders), Decimal("0.00"))
    paid_revenue = sum(
        (o.total_amount for o in all_orders if o.payment_status == "paid"),
        Decimal("0.00"),
    )
    awaiting_revenue = sum(
        (
            o.total_amount
            for o in all_orders
            if o.payment_status in ("awaiting_reserve", "reserve_paid")
        ),
        Decimal("0.00"),
    )
    reserves_collected = sum(
        (o.reserve_amount for o in all_orders if o.reserve_paid_at),
        Decimal("0.00"),
    )

    # Канбан-этапы для таймлайна в drawer
    kanban_columns_cfg = [
        {"key": "pending", "label": "Ожидание", "statuses": ["pending"]},
        {"key": "confirmed", "label": "Формирование", "statuses": ["reserve_paid", "confirmed", "in_production", "ready_to_ship"]},
        {"key": "transit_abroad", "label": "Зарубеж", "statuses": ["transit_abroad"]},
        {"key": "customs", "label": "Таможня", "statuses": ["customs"]},
        {"key": "transit_rf", "label": "РФ", "statuses": ["transit_rf"]},
        {"key": "issuing", "label": "Выдача", "statuses": ["issuing", "shipped"]},
        {"key": "delivered", "label": "Доставлен", "statuses": ["delivered"]},
        {"key": "completed", "label": "Закрыт", "statuses": ["completed"]},
    ]
    kanban_statuses = [(col["key"], col["label"]) for col in kanban_columns_cfg]

    status_to_col_idx = {}
    for idx, col_cfg in enumerate(kanban_columns_cfg):
        for s in col_cfg["statuses"]:
            status_to_col_idx[s] = idx

    # Собираем данные по каждому заказу
    finance_rows = []
    for order in orders_list:
        docs = list(order.documents.all())
        events = list(order.events.all())
        invoice_event = next(
            (e for e in events if e.event_type == "invoice_opened"), None
        )
        reserve_event = next(
            (e for e in events if e.event_type == "reserve_paid"), None
        )
        final_event = next(
            (e for e in events if e.event_type == "final_payment_paid"), None
        )
        mid_event = next(
            (e for e in events if e.event_type == "mid_payment_paid"), None
        )
        customs_event = next(
            (e for e in events if e.event_type == "customs_payment_paid"), None
        )
        current_step = status_to_col_idx.get(order.status, -1)

        finance_rows.append(
            {
                "order": order,
                "docs": docs,
                "docs_count": len(docs),
                "invoice_event": invoice_event,
                "reserve_event": reserve_event,
                "mid_event": mid_event,
                "customs_event": customs_event,
                "final_event": final_event,
                "current_step": current_step,
            }
        )

    return render(
        request,
        "seller/finance/list.html",
        {
            "finance_rows": finance_rows,
            "kanban_statuses": kanban_statuses,
            "total_revenue": total_revenue,
            "paid_revenue": paid_revenue,
            "awaiting_revenue": awaiting_revenue,
            "reserves_collected": reserves_collected,
            "payment_filter": payment_filter,
            "search_q": search_q,
            "payment_choices": Order.PAYMENT_STATUS_CHOICES,
            "seller_page_title": "Финансы",
            "seller_page_subtitle": "Оплаты, документы и финансовый контроль",
            "seller_active_nav": "finance",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Финансы", "url": reverse("seller_finance")},
            ],
        },
    )


@seller_required
def seller_drawings(request: HttpRequest) -> HttpResponse:
    """Страница управления чертежами поставщика."""
    seller = request.user
    drawings = Drawing.objects.filter(seller=seller)

    # Статистика
    total = drawings.count()
    drafts = drawings.filter(status="draft").count()
    on_review = drawings.filter(status="on_review").count()
    approved = drawings.filter(status="approved").count()
    rejected = drawings.filter(status="rejected").count()
    archived = drawings.filter(status="archived").count()

    # Фильтрация
    status_filter = request.GET.get("status", "")
    format_filter = request.GET.get("format", "")
    search_q = request.GET.get("q", "").strip()

    qs = drawings.exclude(status="archived") if not status_filter else drawings
    if status_filter:
        qs = qs.filter(status=status_filter)
    if format_filter:
        qs = qs.filter(file_format=format_filter)
    if search_q:
        qs = qs.filter(
            models.Q(title__icontains=search_q)
            | models.Q(oem_number__icontains=search_q)
            | models.Q(description__icontains=search_q)
        )

    # Форматы для фильтра
    formats_used = (
        drawings.values_list("file_format", flat=True).distinct().order_by("file_format")
    )

    return render(
        request,
        "seller/drawings/list.html",
        {
            "drawings": qs,
            "total": total,
            "drafts": drafts,
            "on_review": on_review,
            "approved": approved,
            "rejected": rejected,
            "archived": archived,
            "formats_used": list(formats_used),
            "status_filter": status_filter,
            "format_filter": format_filter,
            "search_q": search_q,
            "seller_page_title": "Чертежи",
            "seller_page_subtitle": "Управление чертежами и CAD-файлами",
            "seller_active_nav": "drawings",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Чертежи", "url": reverse("seller_drawings")},
            ],
        },
    )


@seller_required
def seller_order_detail(request: HttpRequest, order_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_orders"):
        messages.error(request, "Нет прав на работу с заказами.")
        return redirect("seller_dashboard")

    order = get_object_or_404(
        Order.objects.prefetch_related("items__part", "events", "documents", "claims"),
        id=order_id,
    )
    has_access = order.items.filter(part__seller=request.user).exists()
    if not has_access:
        messages.error(request, "Нет доступа к этому заказу.")
        return redirect("seller_orders")

    _recalc_order_sla(order)
    seller_items = [item for item in order.items.all() if item.part and item.part.seller_id == request.user.id]
    open_claims = [claim for claim in order.claims.all() if claim.status in {"open", "in_review"}]
    events = order.events.all()[:100]
    documents = order.documents.all()[:100]
    allowed_statuses = {"confirmed", "in_production", "ready_to_ship", "shipped", "delivered", "cancelled"}
    status_choices = [(value, label) for value, label in Order.STATUS_CHOICES if value in allowed_statuses]
    return render(
        request,
        "seller/orders/detail.html",
        {
            "order": order,
            "seller_items": seller_items,
            "events": events,
            "documents": documents,
            "claims": order.claims.all()[:100],
            "open_claims": open_claims,
            "status_choices": status_choices,
            "seller_page_title": f"Заказ #{order.id}",
            "seller_page_subtitle": "Карточка заказа, события, документы и действия поставщика.",
            "seller_active_nav": "orders",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Заказы и SLA", "url": reverse("seller_orders")},
                {"label": f"Заказ #{order.id}", "url": reverse("seller_order_detail", args=[order.id])},
            ],
        },
    )


@seller_required
def seller_request_list(request: HttpRequest) -> HttpResponse:
    rfqs_qs = _seller_rfqs_qs(request.user)
    query = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    if status:
        rfqs_qs = rfqs_qs.filter(status=status)
    if query:
        rfqs_qs = rfqs_qs.filter(
            Q(customer_name__icontains=query)
            | Q(company_name__icontains=query)
            | Q(customer_email__icontains=query)
            | Q(items__query__icontains=query)
            | Q(items__matched_part__oem_number__icontains=query)
            | Q(items__matched_part__title__icontains=query)
        ).distinct()

    rfq_rows = []
    for rfq in rfqs_qs[:100]:
        seller_items = [item for item in rfq.items.all() if item.matched_part and item.matched_part.seller_id == request.user.id]
        if not seller_items:
            continue
        total_qty = sum(item.quantity for item in seller_items)
        rfq_rows.append(
            {
                "rfq": rfq,
                "seller_items_count": len(seller_items),
                "total_qty": total_qty,
                "sample_items": seller_items[:3],
                "estimated_total": sum(item.estimated_line_total for item in seller_items),
            }
        )

    return render(
        request,
        _tpl(request.user, "seller/requests/list.html"),
        {
            "rfq_rows": rfq_rows,
            "query": query,
            "status": status,
            "status_choices": RFQ.STATUS_CHOICES,
            "seller_page_title": "Запросы клиентов",
            "seller_page_subtitle": "Все RFQ, где уже найдены позиции по вашему ассортименту и требуется ответ поставщика.",
            "seller_active_nav": "requests",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Запросы клиентов", "url": reverse("seller_request_list")},
            ],
        },
    )


@seller_required
def seller_request_detail(request: HttpRequest, rfq_id: int) -> HttpResponse:
    rfq = get_object_or_404(_seller_rfqs_qs(request.user), id=rfq_id)
    seller_items = [item for item in rfq.items.all() if item.matched_part and item.matched_part.seller_id == request.user.id]
    total_qty = sum(item.quantity for item in seller_items)
    estimated_total = sum(item.estimated_line_total for item in seller_items)

    # Расчёт итоговой скидки
    from decimal import Decimal
    total_discount_amount = Decimal("0.00")
    for item in seller_items:
        line = item.estimated_line_total
        item_discount = Decimal("0.00")
        if item.discount_percent:
            item_discount += line * item.discount_percent / 100
        if item.discount_fixed:
            item_discount += item.discount_fixed * item.quantity
        item.discount_amount = item_discount.quantize(Decimal("0.01"))
        total_discount_amount += item_discount
    if rfq.discount_percent:
        total_discount_amount += estimated_total * rfq.discount_percent / 100
    total_after_discount = estimated_total - total_discount_amount

    return render(
        request,
        "seller/requests/detail.html",
        {
            "rfq": rfq,
            "seller_items": seller_items,
            "seller_items_count": len(seller_items),
            "total_qty": total_qty,
            "estimated_total": estimated_total,
            "total_discount_amount": total_discount_amount.quantize(Decimal("0.01")),
            "total_after_discount": total_after_discount.quantize(Decimal("0.01")),
            "seller_page_title": f"RFQ #{rfq.id}",
            "seller_page_subtitle": "Карточка входящего запроса по вашему ассортименту.",
            "seller_active_nav": "requests",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Запросы клиентов", "url": reverse("seller_request_list")},
                {"label": f"RFQ #{rfq.id}", "url": reverse("seller_request_detail", args=[rfq.id])},
            ],
        },
    )


@seller_required
def seller_rfq_inbox(request: HttpRequest) -> HttpResponse:
    return seller_request_list(request)


@seller_required
@require_POST
def seller_parts_bulk_action(request: HttpRequest) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на массовое управление ассортиментом.")
        return redirect("seller_product_list")

    return_qs = (request.POST.get("return_qs") or "").strip()
    action = (request.POST.get("action") or "").strip()
    selected_ids: list[int] = []
    for raw_id in request.POST.getlist("part_ids"):
        try:
            selected_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue
    selected_ids = list(dict.fromkeys(selected_ids))

    scoped_parts = _apply_seller_brand_scope(
        request.user,
        Part.objects.filter(seller=request.user),
    )
    if selected_ids:
        scoped_parts = scoped_parts.filter(id__in=selected_ids)
    else:
        messages.warning(request, "Выберите хотя бы одну позицию.")
        if return_qs:
            return redirect(f"{reverse('seller_product_list')}?{return_qs}")
        return redirect("seller_product_list")

    now = timezone.now()
    if action == "hide":
        updated_count = scoped_parts.update(is_active=False, data_updated_at=now)
        messages.success(request, f"Скрыто позиций: {updated_count}.")
    elif action == "unhide":
        updated_count = scoped_parts.update(is_active=True, data_updated_at=now)
        messages.success(request, f"Активировано позиций: {updated_count}.")
    elif action == "status":
        status_value = (request.POST.get("availability_status") or "").strip()
        allowed_statuses = {code for code, _ in Part.AVAILABILITY_STATUS_CHOICES}
        if status_value not in allowed_statuses:
            messages.error(request, "Неверный статус доступности.")
        else:
            updated_count = scoped_parts.update(availability_status=status_value, data_updated_at=now)
            messages.success(request, f"Статус обновлен у {updated_count} позиций.")
    elif action == "stock":
        if not _has_seller_permission(request.user, "can_manage_pricing"):
            messages.error(request, "Нет прав на массовое обновление остатков.")
        else:
            try:
                stock_value = int(request.POST.get("stock_quantity"))
                if stock_value < 0:
                    raise ValueError
            except (TypeError, ValueError):
                messages.error(request, "Остаток должен быть целым числом >= 0.")
            else:
                updated_count = scoped_parts.update(stock_quantity=stock_value, data_updated_at=now)
                messages.success(request, f"Остаток обновлен у {updated_count} позиций.")
    else:
        messages.error(request, "Неизвестное массовое действие.")

    if return_qs:
        return redirect(f"{reverse('seller_product_list')}?{return_qs}")
    return redirect("seller_product_list")


@seller_required
@require_POST
@csrf_exempt
def seller_part_inline_update(request: HttpRequest, part_id: int) -> JsonResponse:
    """AJAX inline-edit for a single Part field."""
    import json as _json
    try:
        body = _json.loads(request.body)
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    field = body.get("field", "")
    value = body.get("value", "")

    ALLOWED_FIELDS = {
        "price": "can_manage_pricing",
        "stock_quantity": "can_manage_pricing",
        "condition": "can_manage_assortment",
        "availability_status": "can_manage_assortment",
        "is_active": "can_manage_assortment",
    }
    if field not in ALLOWED_FIELDS:
        return JsonResponse({"ok": False, "error": f"Field '{field}' not allowed"}, status=400)

    perm = ALLOWED_FIELDS[field]
    if not _has_seller_permission(request.user, perm):
        return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

    part = Part.objects.filter(id=part_id, seller=request.user).first()
    if not part:
        return JsonResponse({"ok": False, "error": "Not found"}, status=404)

    try:
        if field == "price":
            from decimal import Decimal, InvalidOperation
            part.price = Decimal(str(value)).quantize(Decimal("0.01"))
        elif field == "stock_quantity":
            part.stock_quantity = max(0, int(value))
        elif field == "condition":
            if value in ("oem", "aftermarket", "reman"):
                part.condition = value
        elif field == "availability_status":
            if value in ("active", "limited", "made_to_order", "discontinued", "blocked"):
                part.availability_status = value
        elif field == "is_active":
            part.is_active = str(value).lower() in ("true", "1")

        part.data_updated_at = timezone.now()
        part.save()
        return JsonResponse({"ok": True, "value": str(getattr(part, field))})
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)


def brands_directory(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect("admin_panel_catalog")
    query = (request.GET.get("q") or "").strip()
    regions = [
        ("europe", "Europe"),
        ("china", "China"),
        ("components", "Component Manufacturers"),
    ]
    grouped = []
    for key, title in regions:
        brands_qs = (
            Brand.objects.filter(region=key)
            .annotate(parts_count=Count("parts", filter=Q(parts__is_active=True, parts__price__gt=0)))
            .filter(parts_count__gt=0)
        )
        if query:
            brands_qs = brands_qs.filter(name__icontains=query)
        brands = brands_qs.order_by("-parts_count", "name")
        grouped.append({"key": key, "title": title, "brands": brands, "count": brands.count()})
    return render(request, "marketplace/brands_directory.html", {"groups": grouped, "query": query})


def categories_directory(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect("admin_panel_catalog")
    categories = (
        Category.objects.all()
        .order_by("name")
        .annotate(parts_count=Count("parts", filter=Q(parts__is_active=True, parts__price__gt=0)))
    )
    total_parts = sum(c.parts_count for c in categories)
    return render(
        request,
        "marketplace/categories_directory.html",
        {"categories": categories, "total_parts": total_parts},
    )


def _match_part_for_query(query: str):
    normalized = (query or "").strip()
    if not normalized:
        return None

    base_qs = _eligible_parts_qs()

    exact = (
        base_qs.filter(oem_number__iexact=normalized)
        .select_related("brand", "category")
        .first()
    )
    if exact:
        return exact

    contains_oem = (
        base_qs.filter(oem_number__icontains=normalized)
        .select_related("brand", "category")
        .first()
    )
    if contains_oem:
        return contains_oem

    contains_title = (
        base_qs.filter(title__icontains=normalized)
        .select_related("brand", "category")
        .first()
    )
    return contains_title


def _supplier_profile_for_part(part: Part):
    seller = part.seller
    if not seller:
        return None
    return getattr(seller, "profile", None)


def _supplier_status_for_part(part: Part) -> str:
    profile = _supplier_profile_for_part(part)
    if not profile or profile.role != "seller":
        return "sandbox"
    return profile.supplier_status or "sandbox"


def _supplier_rating_for_part(part: Part) -> Decimal:
    profile = _supplier_profile_for_part(part)
    if not profile or profile.role != "seller":
        return Decimal("0.00")
    return Decimal(profile.rating_score or 0)


def _is_offer_fresh(part: Part, max_age_days: int = 30) -> bool:
    return part.updated_at >= timezone.now() - timedelta(days=max_age_days)


def _eligible_parts_qs():
    return Part.objects.filter(
        is_active=True,
        price__gt=0,
        currency__isnull=False,
        incoterm__isnull=False,
        moq__gt=0,
        gross_weight_kg__gt=0,
        length_cm__gt=0,
        width_cm__gt=0,
        height_cm__gt=0,
    ).exclude(availability_status__in=["blocked", "discontinued"]).exclude(mapping_status="needs_review")


def _match_confidence_and_pool(query: str):
    normalized = (query or "").strip()
    if not normalized:
        return Decimal("0.00"), Part.objects.none()

    base_qs = _eligible_parts_qs()

    exact = base_qs.filter(oem_number__iexact=normalized)
    if exact.exists():
        return Decimal("95.00"), exact

    by_oem = base_qs.filter(oem_number__icontains=normalized)
    if by_oem.exists():
        return Decimal("75.00"), by_oem

    by_title = base_qs.filter(title__icontains=normalized)
    if by_title.exists():
        return Decimal("65.00"), by_title

    return Decimal("0.00"), Part.objects.none()


def _split_offers_by_status(parts_qs):
    offers = {
        "trusted": [],
        "sandbox": [],
        "risky": [],
        "rejected": [],
    }
    for part in parts_qs.select_related("brand", "category", "seller__profile"):
        if not _is_offer_fresh(part):
            continue
        status = _supplier_status_for_part(part)
        offers.setdefault(status, []).append(part)
    for key in offers:
        offers[key] = sorted(offers[key], key=lambda p: (p.price, -_supplier_rating_for_part(p), p.id))
    return offers


def _select_best_part(candidates):
    if not candidates:
        return None
    return sorted(candidates, key=lambda p: (p.price, -_supplier_rating_for_part(p), p.id))[0]


def _build_rfq_item_decision(query: str, requested_mode: str):
    confidence, raw_pool = _match_confidence_and_pool(query)
    offers = _split_offers_by_status(raw_pool)
    trusted = offers["trusted"]
    sandbox = offers["sandbox"]
    risky = offers["risky"]

    auto_decision = decide_auto_mode(
        AutoModeInputs(
            part_found=bool(raw_pool.exists()),
            confidence=float(confidence),
            trusted_count=len(trusted),
            sandbox_count=len(sandbox),
            fresh_data=True,
        ),
        confidence_threshold=70.0,
    )

    if requested_mode == "manual_oem":
        return {
            "state": "oem_manual",
            "matched_part": None,
            "confidence": confidence,
            "decision_reason": "Manual OEM mode selected.",
            "recommended_supplier_status": "",
            "offers": offers,
        }

    if requested_mode == "auto" and auto_decision.eligible_auto:
        matched = _select_best_part(trusted)
        return {
            "state": "auto_matched",
            "matched_part": matched,
            "confidence": confidence,
            "decision_reason": auto_decision.reason,
            "recommended_supplier_status": "trusted",
            "offers": offers,
        }

    # AUTO fallback or explicit SEMI: operator review required.
    if requested_mode == "auto":
        reason = auto_decision.reason
    else:
        reason = "Semi mode selected, operator confirmation required."

    preferred = _select_best_part(trusted) or _select_best_part(sandbox) or None
    preferred_status = _supplier_status_for_part(preferred) if preferred else ""
    if not preferred and risky:
        preferred = _select_best_part(risky)
        preferred_status = "risky"

    return {
        "state": "needs_review",
        "matched_part": preferred,
        "confidence": confidence,
        "decision_reason": reason,
        "recommended_supplier_status": preferred_status,
        "offers": offers,
    }


def _parse_rfq_items(raw: str) -> list[tuple[str, int]]:
    items: list[tuple[str, int]] = []
    for line in (raw or "").splitlines():
        text = line.strip()
        if not text:
            continue
        left, sep, right = text.partition(";")
        query = left.strip()
        if not query:
            continue
        quantity = 1
        if sep:
            try:
                quantity = max(1, int(right.strip()))
            except Exception:
                quantity = 1
        items.append((query, quantity))
    return items


def _rfq_rows(rfq: RFQ):
    rows = []
    total = Decimal("0.00")
    for item in rfq.items.select_related("matched_part__seller__profile"):
        part = item.matched_part
        if not part or not part.is_eligible_for_matching:
            continue
        supplier_status = _supplier_status_for_part(part)
        if supplier_status in {"rejected", "risky"}:
            continue
        line_total = part.price * item.quantity
        total += line_total
        rows.append(
            {
                "item": item,
                "part": part,
                "quantity": item.quantity,
                "line_total": line_total,
                "supplier_status": supplier_status,
            }
        )
    return rows, total


@login_required
def rfq_list(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect("admin_panel_rfq")
    role = _role_for(request.user)
    if role == "seller":
        rfqs = RFQ.objects.all().prefetch_related("items")[:50]
    else:
        rfqs = RFQ.objects.filter(created_by=request.user).prefetch_related("items")[:50]
    return render(request, "marketplace/rfq_list.html", {"rfqs": rfqs, "role": role})


@login_required
def rfq_new(request: HttpRequest) -> HttpResponse:
    initial = {}
    if request.user.is_authenticated:
        full_name = " ".join(x for x in [request.user.first_name, request.user.last_name] if x).strip()
        initial = {
            "customer_name": full_name or request.user.username,
            "customer_email": request.user.email,
        }
    prefill_items_text = request.session.pop("rfq_prefill_items_text", "").strip()
    if prefill_items_text:
        initial["items_text"] = prefill_items_text

    if request.method == "POST":
        form = RFQCreateForm(request.POST)
        if form.is_valid():
            items_data = _parse_rfq_items(form.cleaned_data["items_text"])
            if not items_data:
                form.add_error("items_text", "Добавьте хотя бы одну позицию.")
            else:
                with transaction.atomic():
                    rfq = RFQ.objects.create(
                        created_by=request.user,
                        customer_name=form.cleaned_data["customer_name"],
                        customer_email=form.cleaned_data["customer_email"],
                        company_name=form.cleaned_data["company_name"],
                        mode=form.cleaned_data["mode"],
                        urgency=form.cleaned_data["urgency"],
                        notes=form.cleaned_data["notes"],
                        status="new",
                    )
                    all_auto_approved = True
                    for query, quantity in items_data:
                        decision = _build_rfq_item_decision(query, form.cleaned_data["mode"])
                        item = RFQItem.objects.create(
                            rfq=rfq,
                            query=query,
                            quantity=quantity,
                            matched_part=decision["matched_part"],
                            state=decision["state"],
                            confidence=decision["confidence"],
                            decision_reason=decision["decision_reason"],
                            recommended_supplier_status=decision["recommended_supplier_status"],
                        )
                        if decision["state"] != "auto_matched":
                            all_auto_approved = False

                        # Log risk-related decisions into rating events.
                        matched_part = decision["matched_part"]
                        if matched_part and matched_part.seller:
                            supplier_status = _supplier_status_for_part(matched_part)
                            if supplier_status == "sandbox":
                                SupplierRatingEvent.objects.create(
                                    supplier=matched_part.seller,
                                    event_type="sandbox_selected",
                                    impact_score=Decimal("0.00"),
                                    meta={"rfq_id": rfq.id, "rfq_item_id": item.id, "query": query},
                                )
                            elif supplier_status == "risky":
                                SupplierRatingEvent.objects.create(
                                    supplier=matched_part.seller,
                                    event_type="risky_selected",
                                    impact_score=Decimal("-1.00"),
                                    meta={"rfq_id": rfq.id, "rfq_item_id": item.id, "query": query},
                                )

                    rfq.status = "quoted" if all_auto_approved else "needs_review"
                    rfq.save(update_fields=["status"])

                messages.success(request, f"RFQ #{rfq.id} создан.")
                return redirect("rfq_detail", rfq_id=rfq.id)
    else:
        form = RFQCreateForm(initial=initial)

    return render(request, "marketplace/rfq_new.html", {"form": form})


@login_required
def rfq_detail(request: HttpRequest, rfq_id: int) -> HttpResponse:
    if request.user.is_superuser:
        return redirect("admin_panel_rfq_detail", rfq_id=rfq_id)
    rfq = get_object_or_404(RFQ.objects.prefetch_related("items__matched_part__brand", "items__matched_part__category"), id=rfq_id)
    role = _role_for(request.user)
    if role != "seller" and rfq.created_by_id != request.user.id:
        messages.error(request, "Нет доступа к этому RFQ.")
        return redirect("dashboard")

    rows, total = _rfq_rows(rfq)
    item_cards = []
    for item in rfq.items.all():
        decision = _build_rfq_item_decision(item.query, rfq.mode)
        item_cards.append(
            {
                "item": item,
                "trusted": decision["offers"]["trusted"][:3],
                "sandbox": decision["offers"]["sandbox"][:3],
                "risky": decision["offers"]["risky"][:3],
            }
        )
    return render(
        request,
        "marketplace/rfq_detail.html",
        {"rfq": rfq, "role": role, "rows": rows, "total": total, "item_cards": item_cards},
    )


@login_required
def rfq_proposal(request: HttpRequest, rfq_id: int) -> HttpResponse:
    rfq = get_object_or_404(RFQ.objects.prefetch_related("items__matched_part"), id=rfq_id)
    role = _role_for(request.user)
    if role == "seller" and not request.user.is_superuser:
        messages.error(request, "КП доступно для клиента.")
        return redirect("rfq_detail", rfq_id=rfq.id)
    if rfq.created_by_id and rfq.created_by_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Нет доступа к этому RFQ.")
        return redirect("rfq_list")

    rows, total = _rfq_rows(rfq)
    if not rows:
        messages.error(request, "Нет доступных позиций для формирования КП.")
        return redirect("rfq_detail", rfq_id=rfq.id)

    initial = {
        "customer_name": rfq.customer_name,
        "customer_email": rfq.customer_email,
    }
    if request.user.is_authenticated:
        initial["customer_email"] = request.user.email or rfq.customer_email

    if request.method == "POST":
        form = CheckoutForm(request.POST)
        if form.is_valid():
            logistics_cost = Decimal("0.00")
            logistics_raw = (request.POST.get("logistics_cost") or "").strip()
            if logistics_raw:
                try:
                    logistics_cost = Decimal(logistics_raw).quantize(Decimal("0.01"))
                    if logistics_cost < 0:
                        logistics_cost = Decimal("0.00")
                except Exception:
                    logistics_cost = Decimal("0.00")
            try:
                order = _create_order_from_rows(
                    rows=rows,
                    total=total,
                    customer_name=form.cleaned_data["customer_name"],
                    customer_email=form.cleaned_data["customer_email"],
                    customer_phone=form.cleaned_data["customer_phone"],
                    delivery_address=form.cleaned_data["delivery_address"],
                    buyer=request.user if request.user.is_authenticated else None,
                    source="buyer",
                    source_id=rfq.id,
                    logistics_override_cost=logistics_cost if logistics_cost > 0 else None,
                )
                _log_order_event(
                    order,
                    "status_changed",
                    source="system",
                    actor=request.user if request.user.is_authenticated else None,
                    meta={"note": "Proposal accepted", "base_total": str(total), "logistics_cost": str(logistics_cost)},
                )
                rfq.status = "quoted"
                rfq.save(update_fields=["status"])
                messages.success(request, f"КП принято. Заказ #{order.id} создан.")
                return redirect("order_invoice", order_id=order.id)
            except ValueError as exc:
                messages.error(request, f"КП не может быть принято: {exc}")
    else:
        form = CheckoutForm(initial=initial)

    return render(
        request,
        "marketplace/rfq_proposal.html",
        {"rfq": rfq, "rows": rows, "total": total, "form": form},
    )


@login_required
@require_POST
def rfq_logistics_estimate(request: HttpRequest, rfq_id: int) -> JsonResponse:
    rfq = get_object_or_404(RFQ, id=rfq_id)
    role = _role_for(request.user)
    if role == "seller" and not request.user.is_superuser:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)
    if rfq.created_by_id and rfq.created_by_id != request.user.id and not request.user.is_superuser:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    payload = {
        "origin": (request.POST.get("origin") or "").strip(),
        "destination": (request.POST.get("destination") or "").strip(),
        "mode": (request.POST.get("mode") or "sea").strip().lower(),
        "incoterm": (request.POST.get("incoterm") or "FOB").strip().upper(),
        "weight_kg": (request.POST.get("weight_kg") or "0").strip(),
        "volume_m3": (request.POST.get("volume_m3") or "0").strip(),
        "currency": (request.POST.get("currency") or "USD").strip().upper(),
    }
    result = logistics_estimate(payload)
    if not result.get("ok", False):
        return JsonResponse({"ok": False, "error": result.get("error", "logistics_calculation_failed"), "result": result}, status=502)
    return JsonResponse({"ok": True, "result": result})


@login_required
def rfq_proposal_pdf(request: HttpRequest, rfq_id: int) -> HttpResponse:
    rfq = get_object_or_404(RFQ.objects.prefetch_related("items__matched_part"), id=rfq_id)
    role = _role_for(request.user)
    if role == "seller" and not request.user.is_superuser:
        messages.error(request, "КП доступно для клиента.")
        return redirect("rfq_detail", rfq_id=rfq.id)
    if rfq.created_by_id and rfq.created_by_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Нет доступа к этому RFQ.")
        return redirect("rfq_list")

    rows, total = _rfq_rows(rfq)
    if not rows:
        messages.error(request, "Нет доступных позиций для формирования КП.")
        return redirect("rfq_detail", rfq_id=rfq.id)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except Exception:
        messages.error(request, "PDF-экспорт требует пакет reportlab. Выполните: pip install reportlab")
        return redirect("rfq_proposal", rfq_id=rfq.id)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    doc_no = f"KP-{rfq.created_at:%Y%m%d}-{rfq.id}"

    left = 15 * mm
    right = 195 * mm
    y = height - 15 * mm

    # Header band
    pdf.setFillColor(colors.HexColor("#0f2f66"))
    pdf.roundRect(left, y - 16 * mm, right - left, 16 * mm, 4 * mm, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(left + 4 * mm, y - 9 * mm, "CONSOLIDATOR PARTS")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(left + 4 * mm, y - 13 * mm, "Commercial Proposal / Коммерческое предложение")

    y -= 23 * mm
    pdf.setFillColor(colors.HexColor("#0c1530"))
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left, y, f"KP No: {doc_no}")
    pdf.setFont("Helvetica", 9)
    pdf.drawRightString(right, y, f"Date: {rfq.created_at:%d.%m.%Y}")

    y -= 8 * mm
    pdf.setFillColor(colors.HexColor("#1a2748"))
    pdf.roundRect(left, y - 20 * mm, right - left, 20 * mm, 2 * mm, fill=0, stroke=1)
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left + 3 * mm, y - 5 * mm, "Customer")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left + 3 * mm, y - 10 * mm, rfq.customer_name[:70])
    pdf.drawString(left + 3 * mm, y - 14 * mm, rfq.customer_email[:70])
    pdf.drawString(left + 3 * mm, y - 18 * mm, (rfq.company_name or "-")[:70])

    y -= 27 * mm
    # Table header
    pdf.setFillColor(colors.HexColor("#e9f0ff"))
    pdf.rect(left, y - 7 * mm, right - left, 7 * mm, fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#1b2d57"))
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(left + 2 * mm, y - 4.8 * mm, "#")
    pdf.drawString(left + 8 * mm, y - 4.8 * mm, "Part")
    pdf.drawString(left + 88 * mm, y - 4.8 * mm, "OEM")
    pdf.drawString(left + 120 * mm, y - 4.8 * mm, "Qty")
    pdf.drawString(left + 136 * mm, y - 4.8 * mm, "Lead")
    pdf.drawString(left + 154 * mm, y - 4.8 * mm, "Price")
    pdf.drawString(left + 176 * mm, y - 4.8 * mm, "Line Total")
    y -= 9 * mm

    pdf.setFont("Helvetica", 8)
    for idx, row in enumerate(rows, start=1):
        if y < 36 * mm:
            pdf.showPage()
            y = height - 20 * mm
            pdf.setFont("Helvetica", 8)
        part_title = (row["part"].title or "")[:34]
        oem = (row["part"].oem_number or "")[:20]
        lead = f"{row['part'].production_lead_days} d"
        pdf.setFillColor(colors.HexColor("#0f1f42"))
        pdf.drawString(left + 2 * mm, y, str(idx))
        pdf.drawString(left + 8 * mm, y, part_title)
        pdf.drawString(left + 88 * mm, y, oem)
        pdf.drawRightString(left + 131 * mm, y, str(row["quantity"]))
        pdf.drawRightString(left + 148 * mm, y, lead)
        pdf.drawRightString(left + 171 * mm, y, f"${row['part'].price}")
        pdf.drawRightString(right, y, f"${row['line_total']}")
        y -= 5.2 * mm

    y -= 2 * mm
    pdf.setStrokeColor(colors.HexColor("#b7c7e8"))
    pdf.line(left + 130 * mm, y, right, y)
    y -= 7 * mm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColor(colors.HexColor("#0f2f66"))
    pdf.drawRightString(right, y, f"TOTAL: ${total}")

    y -= 10 * mm
    pdf.setFont("Helvetica", 8)
    pdf.setFillColor(colors.HexColor("#253c72"))
    pdf.drawString(left, y, "Terms: Prices are valid for 3 business days. Delivery terms are confirmed at order stage.")
    y -= 4.5 * mm
    pdf.drawString(left, y, "Условия: КП действительно 3 рабочих дня. Финальные условия поставки подтверждаются при заказе.")

    y -= 12 * mm
    pdf.setStrokeColor(colors.HexColor("#7d95c6"))
    pdf.line(left, y, left + 70 * mm, y)
    pdf.line(left + 95 * mm, y, right, y)
    pdf.setFont("Helvetica", 8)
    pdf.drawString(left, y - 4 * mm, "Authorized Signature (Seller)")
    pdf.drawString(left + 95 * mm, y - 4 * mm, "Authorized Signature (Buyer)")

    pdf.showPage()
    pdf.save()
    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{doc_no}.pdf"'
    return response


@login_required
def rfq_checkout(request: HttpRequest, rfq_id: int) -> HttpResponse:
    rfq = get_object_or_404(RFQ.objects.prefetch_related("items__matched_part"), id=rfq_id)
    role = _role_for(request.user)
    if role == "seller" and not request.user.is_superuser:
        messages.error(request, "Оформление из RFQ доступно только buyer.")
        return redirect("rfq_detail", rfq_id=rfq.id)
    if rfq.created_by_id and rfq.created_by_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Нет доступа к этому RFQ.")
        return redirect("rfq_list")

    rows, total = _rfq_rows(rfq)
    if not rows:
        messages.error(request, "В RFQ нет доступных позиций для заказа.")
        return redirect("rfq_detail", rfq_id=rfq.id)

    initial = {
        "customer_name": rfq.customer_name,
        "customer_email": rfq.customer_email,
    }

    if request.method == "POST":
        form = CheckoutForm(request.POST)
        if form.is_valid():
            try:
                order = _create_order_from_rows(
                    rows=rows,
                    total=total,
                    customer_name=form.cleaned_data["customer_name"],
                    customer_email=form.cleaned_data["customer_email"],
                    customer_phone=form.cleaned_data["customer_phone"],
                    delivery_address=form.cleaned_data["delivery_address"],
                    buyer=request.user if request.user.is_authenticated else None,
                    source="buyer",
                    source_id=rfq.id,
                )
                rfq.status = "quoted"
                rfq.save(update_fields=["status"])
                messages.success(request, f"Заказ #{order.id} создан из RFQ #{rfq.id}.")
                return redirect("dashboard_buyer")
            except ValueError as exc:
                messages.error(request, f"Заказ не создан: {exc}")
    else:
        form = CheckoutForm(initial=initial)

    return render(
        request,
        "marketplace/rfq_checkout.html",
        {"rfq": rfq, "rows": rows, "total": total, "form": form},
    )


@operator_required
def operator_queue(request: HttpRequest) -> HttpResponse:
    active_tab = (request.GET.get("tab") or "queue").strip().lower()
    if active_tab not in {"queue", "manual", "risky"}:
        active_tab = "queue"

    rfq_items = (
        RFQItem.objects.filter(state="needs_review")
        .select_related("rfq", "matched_part__brand", "matched_part__category")
        .order_by("-rfq__created_at", "id")[:200]
    )
    queue_rows = []
    risky_rows = []
    for item in rfq_items:
        decision = _build_rfq_item_decision(item.query, "semi")
        trusted = [p for p in decision["offers"]["trusted"] if _operator_can_access_part(request.user, p)][:5]
        sandbox = [p for p in decision["offers"]["sandbox"] if _operator_can_access_part(request.user, p)][:5]
        risky = [p for p in decision["offers"]["risky"] if _operator_can_access_part(request.user, p)][:5]
        row = (
            {
                "item": item,
                "rfq": item.rfq,
                "trusted": trusted,
                "sandbox": sandbox,
                "risky": risky,
            }
        )
        queue_rows.append(row)
        if risky:
            risky_rows.append(row)

    manual_rows = list(
        RFQItem.objects.filter(state="oem_manual")
        .select_related("rfq")
        .order_by("-rfq__created_at", "id")[:200]
    )
    metrics = {
        "needs_review_count": len(queue_rows),
        "manual_oem_count": len(manual_rows),
        "risky_candidates_count": len(risky_rows),
    }
    return render(
        request,
        "marketplace/operator_queue.html",
        {
            "active_tab": active_tab,
            "queue_rows": queue_rows,
            "manual_rows": manual_rows,
            "risky_rows": risky_rows,
            "metrics": metrics,
        },
    )


@operator_required
def operator_webhooks(request: HttpRequest) -> HttpResponse:
    state = (request.GET.get("state") or "failed").strip().lower()
    if state not in {"all", "failed", "success"}:
        state = "failed"
    endpoint_filter = (request.GET.get("endpoint") or "").strip()

    logs_qs = WebhookDeliveryLog.objects.select_related("order", "order_event")
    if state == "failed":
        logs_qs = logs_qs.filter(success=False)
    elif state == "success":
        logs_qs = logs_qs.filter(success=True)
    if endpoint_filter:
        logs_qs = logs_qs.filter(endpoint__icontains=endpoint_filter)
    logs_qs = logs_qs.order_by("-created_at")

    paginator = Paginator(logs_qs, 40)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    max_attempts = max(1, int(getattr(settings, "WEBHOOK_RETRY_MAX_ATTEMPTS", 5) or 5))
    metrics = {
        "total": WebhookDeliveryLog.objects.count(),
        "failed": WebhookDeliveryLog.objects.filter(success=False).count(),
        "success": WebhookDeliveryLog.objects.filter(success=True).count(),
    }

    return render(
        request,
        "marketplace/operator_webhooks.html",
        {
            "state": state,
            "endpoint_filter": endpoint_filter,
            "page_obj": page_obj,
            "max_attempts": max_attempts,
            "metrics": metrics,
        },
    )


@operator_required
@require_POST
def operator_retry_webhook(request: HttpRequest, log_id: int) -> HttpResponse:
    log = get_object_or_404(WebhookDeliveryLog.objects.select_related("order", "order_event"), id=log_id)
    ok = _retry_webhook_log(log)
    if ok:
        messages.success(request, f"Webhook #{log_id} успешно доставлен при ретрае.")
    else:
        messages.error(request, f"Webhook #{log_id} не доставлен. Проверь endpoint/секрет/сеть.")
    return redirect(f"{reverse('operator_webhooks')}?state=failed")


@operator_required
@require_POST
def operator_retry_failed_webhooks(request: HttpRequest) -> HttpResponse:
    limit_raw = (request.POST.get("limit") or "30").strip()
    try:
        limit = max(1, min(200, int(limit_raw)))
    except Exception:
        limit = 30
    max_attempts = max(1, int(getattr(settings, "WEBHOOK_RETRY_MAX_ATTEMPTS", 5) or 5))
    failed_logs = list(
        WebhookDeliveryLog.objects.select_related("order", "order_event")
        .filter(success=False, attempt__lt=max_attempts)
        .order_by("created_at")[:limit]
    )
    if not failed_logs:
        messages.info(request, "Нет webhook-ошибок для ретрая.")
        return redirect(f"{reverse('operator_webhooks')}?state=failed")

    ok_count = 0
    fail_count = 0
    for log in failed_logs:
        if _retry_webhook_log(log):
            ok_count += 1
        else:
            fail_count += 1
    messages.success(request, f"Ретрай завершен: успешно {ok_count}, ошибок {fail_count}.")
    return redirect(f"{reverse('operator_webhooks')}?state=failed")


@operator_required
@require_POST
def operator_assign_supplier(request: HttpRequest, rfq_item_id: int) -> HttpResponse:
    item = get_object_or_404(RFQItem.objects.select_related("rfq"), id=rfq_item_id)
    part_id_raw = (request.POST.get("part_id") or "").strip()
    if not part_id_raw.isdigit():
        messages.error(request, "Не выбран поставщик/позиция.")
        return redirect(f"{reverse('operator_queue')}?tab=queue")

    selected_part = get_object_or_404(Part.objects.select_related("seller__profile"), id=int(part_id_raw), is_active=True, price__gt=0)
    if not _operator_can_access_part(request.user, selected_part):
        messages.error(request, "Этот бренд/регион недоступен вашей операторской роли.")
        return redirect(f"{reverse('operator_queue')}?tab=queue")
    status = _supplier_status_for_part(selected_part)
    if status == "rejected":
        messages.error(request, "Исключённый поставщик не может быть выбран.")
        return redirect(f"{reverse('operator_queue')}?tab=queue")

    sandbox_confirm = request.POST.get("sandbox_confirm") == "1"
    risky_confirm = request.POST.get("risky_confirm") == "1"
    risky_double_confirm = request.POST.get("risky_double_confirm") == "1"

    if status == "sandbox" and not sandbox_confirm:
        messages.error(request, "Для Песочницы требуется подтверждение оператора.")
        return redirect(f"{reverse('operator_queue')}?tab=queue")

    if status == "risky" and not (risky_confirm and risky_double_confirm):
        messages.error(request, "Для Рискового поставщика требуется двойное подтверждение.")
        return redirect(f"{reverse('operator_queue')}?tab=risky")

    item.matched_part = selected_part
    item.recommended_supplier_status = status
    item.state = "auto_matched" if status == "trusted" else "needs_review"
    item.decision_reason = f"Operator assigned supplier status={status}, part_id={selected_part.id}"
    item.save(update_fields=["matched_part", "recommended_supplier_status", "state", "decision_reason"])

    if selected_part.seller:
        if status == "sandbox":
            SupplierRatingEvent.objects.create(
                supplier=selected_part.seller,
                event_type="sandbox_selected",
                impact_score=Decimal("0.00"),
                meta={"rfq_id": item.rfq_id, "rfq_item_id": item.id, "actor": request.user.username},
            )
        elif status == "risky":
            SupplierRatingEvent.objects.create(
                supplier=selected_part.seller,
                event_type="risky_selected",
                impact_score=Decimal("-1.00"),
                meta={"rfq_id": item.rfq_id, "rfq_item_id": item.id, "actor": request.user.username},
            )

    # If all items resolved with matched parts, mark RFQ quoted.
    unresolved_exists = item.rfq.items.filter(matched_part__isnull=True).exists()
    if not unresolved_exists:
        item.rfq.status = "quoted"
        item.rfq.save(update_fields=["status"])

    messages.success(request, f"Поставщик назначен для позиции RFQ #{item.rfq_id}.")
    return redirect(f"{reverse('operator_queue')}?tab=queue")


@operator_required
@require_POST
def operator_escalate_manual_oem(request: HttpRequest, rfq_item_id: int) -> HttpResponse:
    item = get_object_or_404(RFQItem.objects.select_related("rfq"), id=rfq_item_id)
    reason = (request.POST.get("manual_reason") or "").strip()
    if not reason:
        messages.error(request, "Укажи причину перевода в ручной OEM-поиск.")
        return redirect(f"{reverse('operator_queue')}?tab=manual")

    item.state = "oem_manual"
    item.matched_part = None
    item.recommended_supplier_status = ""
    item.decision_reason = f"Manual OEM escalation by {request.user.username}: {reason}"
    item.save(update_fields=["state", "matched_part", "recommended_supplier_status", "decision_reason"])

    item.rfq.status = "needs_review"
    item.rfq.save(update_fields=["status"])

    # Audit event (supplier is optional in this event, so we anchor to operator user).
    SupplierRatingEvent.objects.create(
        supplier=request.user,
        event_type="manual_oem_escalation",
        impact_score=Decimal("0.00"),
        meta={"rfq_id": item.rfq_id, "rfq_item_id": item.id, "reason": reason},
    )

    messages.success(request, f"Позиция RFQ #{item.rfq_id} переведена в ручной OEM-поиск.")
    return redirect(f"{reverse('operator_queue')}?tab=manual")


@seller_required
def seller_product_detail(request: HttpRequest, part_id: int) -> HttpResponse:
    part = get_object_or_404(
        _apply_seller_brand_scope(request.user, Part.objects.select_related("brand", "category")),
        id=part_id,
        seller=request.user,
    )
    missing_fields = part.mandatory_missing_fields()
    completeness_total = 8
    completeness_done = max(0, completeness_total - len(missing_fields))
    completeness_percent = int(round((completeness_done / completeness_total) * 100)) if completeness_total else 0
    stale_snapshot = _part_stale_snapshot(part)
    price_history = _part_price_history(part)
    demand_stats = _part_demand_stats(part)
    related_parts = (
        _apply_seller_brand_scope(
            request.user,
            Part.objects.filter(seller=request.user, brand=part.brand).exclude(id=part.id).select_related("brand", "category"),
        )
        .order_by("-data_updated_at", "-id")[:6]
    )
    return render(
        request,
        "seller/products/detail.html",
        {
            "part": part,
            "missing_fields": missing_fields,
            "completeness_percent": completeness_percent,
            "stale_snapshot": stale_snapshot,
            "price_history": price_history,
            "demand_stats": demand_stats,
            "related_parts": related_parts,
            "seller_page_title": part.title,
            "seller_page_subtitle": "Карточка товара поставщика: данные, логистика, полнота и быстрые действия.",
            "seller_active_nav": "products",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Товары и прайсы", "url": reverse("seller_product_list")},
                {"label": part.title, "url": reverse("seller_product_detail", args=[part.id])},
            ],
        },
    )


@seller_required
def seller_part_create(request: HttpRequest) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на управление ассортиментом.")
        return redirect("seller_product_list")
    if not _has_seller_permission(request.user, "can_manage_pricing"):
        messages.error(request, "Нет прав на создание позиций с ценой.")
        return redirect("seller_product_list")

    if request.method == "POST":
        form = SellerPartForm(request.POST)
        if form.is_valid():
            part = form.save(commit=False)
            base = slugify(part.title)[:220] or "part"
            part.slug = f"{base}-{uuid4().hex[:8]}"
            part.seller = request.user
            part.save()
            messages.success(request, "Товар создан.")
            return redirect("seller_product_list")
    else:
        form = SellerPartForm()
    return render(
        request,
        "marketplace/seller_part_form.html",
        {
            "form": form,
            "mode": "create",
            "seller_page_title": "Новый товар",
            "seller_page_subtitle": "Создание новой позиции вручную.",
            "seller_active_nav": "products",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Товары и прайсы", "url": reverse("seller_product_list")},
                {"label": "Новый товар", "url": reverse("seller_part_create")},
            ],
        },
    )


@seller_required
def seller_part_edit(request: HttpRequest, part_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на редактирование ассортимента.")
        return redirect("seller_product_list")

    part = get_object_or_404(_apply_seller_brand_scope(request.user, Part.objects.all()), id=part_id, seller=request.user)
    if request.method == "POST":
        old_price = part.price
        form = SellerPartForm(request.POST, instance=part)
        if form.is_valid():
            updated = form.save(commit=False)
            if not _has_seller_permission(request.user, "can_manage_pricing"):
                updated.price = old_price
            updated.save()
            messages.success(request, "Товар обновлен.")
            return redirect("seller_product_list")
    else:
        form = SellerPartForm(instance=part)
    return render(
        request,
        "marketplace/seller_part_form.html",
        {
            "form": form,
            "mode": "edit",
            "part": part,
            "seller_page_title": f"Редактирование: {part.title}",
            "seller_page_subtitle": "Обновление данных позиции, цены и логистики.",
            "seller_active_nav": "products",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Товары и прайсы", "url": reverse("seller_product_list")},
                {"label": part.title, "url": reverse("seller_product_detail", args=[part.id])},
                {"label": "Редактирование", "url": reverse("seller_part_edit", args=[part.id])},
            ],
        },
    )


@seller_required
@require_POST
def seller_import_google_sheet(request: HttpRequest) -> HttpResponse:
    """Import from a public Google Sheets URL."""
    import re as _re
    import urllib.request as _urllib_request

    sheet_url = (request.POST.get("sheet_url") or "").strip()
    if "docs.google.com/spreadsheets" not in sheet_url:
        messages.error(request, "Нужна корректная ссылка Google Sheets.")
        return redirect("seller_product_list")

    # Extract sheet ID
    match = _re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_url)
    if not match:
        messages.error(request, "Не удалось извлечь ID таблицы из ссылки.")
        return redirect("seller_product_list")

    sheet_id = match.group(1)
    csv_export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"

    try:
        req = _urllib_request.Request(csv_export_url, headers={"User-Agent": "ConsolidatorParts/1.0"})
        with _urllib_request.urlopen(req, timeout=15) as resp:
            raw = resp.read()
    except Exception:
        messages.error(request, "Не удалось загрузить Google-таблицу. Убедитесь, что таблица открыта для чтения (доступ по ссылке).")
        return redirect("seller_product_list")

    try:
        from marketplace.services.imports import _csv_rows
        headers, rows = _csv_rows(raw)
    except Exception as exc:
        messages.error(request, f"Ошибка парсинга таблицы: {exc}")
        return redirect("seller_product_list")

    sample = [row_dict for _, row_dict in rows[:10]]
    detected = {h: h for h in headers}
    auto_mapping = _auto_map_columns(headers)

    # If key columns are mapped, auto-confirm and start import
    has_key_cols = auto_mapping.get("oem") and auto_mapping.get("price_exw")
    initial_status = (
        ImportPreviewSession.Status.MAPPING_CONFIRMED if has_key_cols
        else ImportPreviewSession.Status.DRAFT
    )

    preview = ImportPreviewSession.objects.create(
        supplier=request.user,
        source_type=ImportPreviewSession.SourceType.GOOGLE_SHEET,
        source_url=sheet_url,
        status=initial_status,
        detected_columns=detected,
        sample_rows=sample,
        column_mapping=auto_mapping,
    )

    total = len(rows)
    mapped = len(auto_mapping)
    return redirect(f"{reverse('seller_product_list')}?preview_id={preview.id}&gs_imported=1&gs_rows={total}&gs_cols={mapped}")


@seller_required
def seller_import_preview(request: HttpRequest, preview_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на загрузку ассортимента.")
        return redirect("seller_product_list")
    preview = get_object_or_404(ImportPreviewSession, id=preview_id, supplier=request.user)
    if preview.source_type not in (ImportPreviewSession.SourceType.CSV, ImportPreviewSession.SourceType.GOOGLE_SHEET):
        messages.error(request, "Для этого источника preview пока не поддерживается в UI.")
        return redirect("seller_product_list")
    if preview.source_type == ImportPreviewSession.SourceType.CSV and not preview.source_file_id:
        messages.error(request, "Файл источника не найден.")
        return redirect("seller_product_list")

    header_options = list(dict.fromkeys(list(preview.sample_rows[0].keys() if preview.sample_rows else []) + list(preview.detected_columns.values())))
    preview_rows_matrix = [[row.get(header, "") for header in header_options] for row in (preview.sample_rows or [])]
    current_mapping = preview.column_mapping or preview.detected_columns or {}
    return render(
        request,
        "marketplace/seller_import_preview.html",
        {
            "preview": preview,
            "header_options": header_options,
            "preview_rows_matrix": preview_rows_matrix,
            "mapping_fields": SELLER_IMPORT_MAPPING_FIELDS,
            "current_mapping": current_mapping,
        },
    )


@seller_required
@require_POST
def seller_import_preview_confirm(request: HttpRequest, preview_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на загрузку ассортимента.")
        return redirect("seller_product_list")
    preview = get_object_or_404(ImportPreviewSession, id=preview_id, supplier=request.user)
    mapping = {key: (request.POST.get(f"mapping__{key}") or "").strip() for key, _ in SELLER_IMPORT_MAPPING_FIELDS}

    header_options = list(dict.fromkeys(list(preview.sample_rows[0].keys() if preview.sample_rows else []) + list(preview.detected_columns.values())))
    ok, reason = ColumnMappingResolver().validate_mapping(mapping, header_options)
    if not ok:
        messages.error(request, reason)
        return redirect(f"{reverse('seller_product_list')}?preview_id={preview.id}")

    preview.column_mapping = mapping
    preview.status = ImportPreviewSession.Status.MAPPING_CONFIRMED
    preview.save(update_fields=["column_mapping", "status", "updated_at"])
    messages.success(request, "Маппинг колонок подтвержден. Можно запускать импорт.")
    return redirect(f"{reverse('seller_product_list')}?preview_id={preview.id}")


@seller_required
@require_POST
def seller_import_preview_start(request: HttpRequest, preview_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на загрузку ассортимента.")
        return redirect("seller_product_list")
    preview = get_object_or_404(ImportPreviewSession, id=preview_id, supplier=request.user)
    if preview.status != ImportPreviewSession.Status.MAPPING_CONFIRMED:
        messages.error(request, "Сначала подтвердите маппинг колонок.")
        return redirect(f"{reverse('seller_product_list')}?preview_id={preview.id}")

    idempotency_key = preview.source_file.checksum_sha256 if preview.source_file_id else ""
    job = ImportJob.objects.create(
        supplier=request.user,
        source_type=preview.source_type,
        source_file=preview.source_file,
        source_url=preview.source_url,
        preview_session=preview,
        column_mapping_json=preview.column_mapping or {},
        status=ImportJob.Status.QUEUED,
        idempotency_key=idempotency_key,
    )
    try:
        process_import_job.delay(job.id)
    except Exception as exc:
        logger.warning(
            "import_job_enqueue_failed_from_preview",
            extra={"job_id": job.id, "supplier_id": request.user.id, "error": str(exc)},
        )
        messages.error(request, "Не удалось поставить импорт в очередь.")
        return redirect(f"{reverse('seller_product_list')}?preview_id={preview.id}")

    messages.success(request, f"Импорт запущен (job #{job.id}). Можно следить за статусом на экране результата.")
    return redirect("seller_import_result", import_id=job.id)


@seller_required
def seller_import_result(request: HttpRequest, import_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на просмотр результатов импорта.")
        return redirect("seller_product_list")

    job = get_object_or_404(
        ImportJob.objects.select_related("source_file", "error_report__file"),
        id=import_id,
        supplier=request.user,
    )
    rows = list(
        job.rows.filter(status=ImportRow.Status.INVALID)
        .order_by("row_no")
        .values(
            "row_no",
            "part_number_raw",
            "error_code",
            "error_message",
            "error_hint",
            "raw_payload",
        )[:100]
    )
    return render(
        request,
        "seller/products/import_result.html",
        {
            "import_job": job,
            "error_rows_preview": rows,
            "seller_page_title": f"Результат импорта #{job.id}",
            "seller_page_subtitle": "Статус обработки, итоговые счетчики и ошибки по строкам.",
            "seller_active_nav": "products",
            "seller_breadcrumbs": [
                {"label": "Кабинет поставщика", "url": reverse("seller_dashboard")},
                {"label": "Товары и прайсы", "url": reverse("seller_product_list")},
                {"label": f"Импорт #{job.id}", "url": reverse("seller_import_result", args=[job.id])},
            ],
        },
    )


@seller_required
@require_POST
def seller_bulk_upload(request: HttpRequest) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на bulk upload.")
        return redirect("seller_product_list")
    if not _has_seller_permission(request.user, "can_manage_pricing"):
        messages.error(request, "Нет прав на bulk upload цен.")
        return redirect("seller_product_list")

    def _humanize_import_error(raw_message: str) -> str:
        message = (raw_message or "").strip()
        if "Превышен лимит строк" in message:
            return f"{message} Разбейте файл на несколько частей и загрузите по очереди."
        if "Файл слишком большой" in message:
            return f"{message} Сожмите файл или разделите его на несколько файлов."
        if "Файл должен содержать колонки" in message:
            return "Не найдены обязательные колонки. Нужны PartNumber/Part Number, WarehouseAddress и хотя бы одна цена Price_FOB_SEA/Price_FOB_AIR."
        return message or "Ошибка обработки импорта."

    form = SellerBulkUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Некорректная форма загрузки.")
        return redirect("seller_product_list")
    import_mode = (request.POST.get("import_mode") or "apply").strip().lower()
    if import_mode not in {"preview", "apply"}:
        import_mode = "apply"

    upload = form.cleaned_data["file"]
    upload_name = getattr(upload, "name", "") or ""
    if import_mode == "preview":
        try:
            stored = store_import_source_file(upload)
            stored_file = StoredFile.objects.create(
                supplier=request.user,
                source_type=StoredFile.SourceType.IMPORT_CSV,
                storage_key=stored.storage_key,
                original_name=stored.original_name,
                content_type=stored.content_type,
                size_bytes=stored.size_bytes,
                checksum_sha256=stored.checksum_sha256,
            )
            preview = ImportPreviewSession.objects.create(
                supplier=request.user,
                source_type=ImportPreviewSession.SourceType.CSV,
                source_file=stored_file,
                status=ImportPreviewSession.Status.DRAFT,
            )
            preview_result = ImportParser().build_preview(stored_file.storage_key)
            preview.detected_columns = preview_result.detected_columns
            preview.sample_rows = preview_result.sample_rows
            preview.column_mapping = preview_result.detected_columns
            preview.save(update_fields=["detected_columns", "sample_rows", "column_mapping", "updated_at"])
            target_url = f"{reverse('seller_product_list')}?preview_id={preview.id}"
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"redirect_url": target_url}, status=200)
            return redirect(target_url)
        except Exception:
            logger.exception("seller_import_preview_build_failed", extra={"seller_id": request.user.id, "filename": upload_name})
            return JsonResponse({"error": "Не удалось построить preview импорта."}, status=400)

    metric_inc("import_attempts_total")
    timer = Timer()
    try:
        report = process_seller_csv_upload(
            seller=request.user,
            upload=upload,
            category_name=form.cleaned_data["category"].strip() or "Epiroc",
            default_stock=form.cleaned_data["default_stock"],
            import_mode=import_mode,
        )
    except UploadLimitError as exc:
        error_message = _humanize_import_error(str(exc))
        SellerImportRun.objects.create(
            seller=request.user,
            filename=upload_name,
            mode=import_mode,
            status="failed",
            skipped_invalid_count=1,
            error_count=1,
            errors=[{"row": 0, "reason": error_message}],
        )
        refresh_supplier_dashboard_projection(request.user)
        metric_inc("import_limits_triggered_total")
        logger.warning(
            "import_limit_exceeded",
            extra={"seller_id": request.user.id, "status": exc.status_code, "reason": error_message},
        )
        return JsonResponse({"error": error_message}, status=exc.status_code)
    except ValueError as exc:
        error_message = _humanize_import_error(str(exc))
        SellerImportRun.objects.create(
            seller=request.user,
            filename=upload_name,
            mode=import_mode,
            status="failed",
            skipped_invalid_count=1,
            error_count=1,
            errors=[{"row": 0, "reason": error_message}],
        )
        refresh_supplier_dashboard_projection(request.user)
        metric_inc("import_validation_errors_total")
        logger.warning("import_validation_error", extra={"seller_id": request.user.id, "reason": error_message})
        return JsonResponse({"error": error_message}, status=400)
    except Exception:
        SellerImportRun.objects.create(
            seller=request.user,
            filename=upload_name,
            mode=import_mode,
            status="failed",
            skipped_invalid_count=1,
            error_count=1,
            errors=[{"row": 0, "reason": "Ошибка обработки импорта."}],
        )
        refresh_supplier_dashboard_projection(request.user)
        metric_inc("import_internal_errors_total")
        logger.exception("import_internal_error", extra={"seller_id": request.user.id})
        return JsonResponse({"error": "Ошибка обработки импорта."}, status=500)

    request.session["seller_upload_report"] = {
        "mode": report.mode,
        "created": report.created,
        "updated": report.updated,
        "skipped_no_price": report.skipped_no_price,
        "skipped_invalid": report.skipped_invalid,
        "total_rows": report.total_rows,
        "processed_rows": report.processed_rows,
        "failed_rows": report.failed_rows,
        "success_rate": report.success_rate,
        "errors": report.errors,
    }
    SellerImportRun.objects.create(
        seller=request.user,
        filename=upload_name,
        mode=report.mode,
        status="success",
        created_count=report.created,
        updated_count=report.updated,
        skipped_no_price_count=report.skipped_no_price,
        skipped_invalid_count=report.skipped_invalid,
        error_count=len(report.errors),
        errors=report.errors,
    )
    refresh_supplier_dashboard_projection(request.user)
    metric_inc("import_success_total")
    logger.info(
        "import_finished",
        extra={
            "seller_id": request.user.id,
            "mode": report.mode,
            "created": report.created,
            "updated": report.updated,
            "skipped_no_price": report.skipped_no_price,
            "skipped_invalid": report.skipped_invalid,
            "latency_ms": timer.elapsed_ms(),
        },
    )
    if import_mode == "preview":
        messages.info(
            request,
            f"Предпросмотр: всего строк {report.total_rows}, успешно {report.processed_rows} ({report.success_rate}%), ошибок {report.failed_rows}.",
        )
    else:
        messages.success(
            request,
            f"Импорт завершен: успешно {report.processed_rows} из {report.total_rows} строк ({report.success_rate}%).",
        )
    return redirect("seller_product_list")


@seller_required
def seller_csv_template(request: HttpRequest) -> HttpResponse:
    content = (
        "PartNumber,CrossNumber,Brand,Name,Quantity,Condition,WarehouseAddress,Price_EXW,Price_FOB_SEA,Price_FOB_AIR,Weight,Length,Width,Height,MOQ,LeadTime_days\n"
        "RE48786,RE48786A,John Deere,MAIN SWITCH,10,OEM,Shanghai CN,250.00,295.00,330.00,1.1,12,8,6,1,7\n"
    )
    response = HttpResponse(content, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="consolidator_template.csv"'
    return response


@seller_required
def seller_gsheet_template(request: HttpRequest) -> HttpResponse:
    """Generate XLSX template for Google Sheets import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Sheet 1: Data ──
    ws = wb.active
    ws.title = "Данные"

    headers = [
        "PartNumber", "Name", "Brand", "Price_EXW", "Price_FOB_SEA",
        "Price_FOB_AIR", "Quantity", "Condition", "WarehouseAddress",
        "CrossNumber", "SeaPort", "AirPort", "Weight", "Length", "Width", "Height",
    ]
    header_ru = [
        "Артикул", "Название", "Бренд", "Цена EXW ($)", "Цена FOB Море ($)",
        "Цена FOB Авиа ($)", "Остаток (шт)", "Состояние", "Адрес склада",
        "Кросс-номер", "Морской порт", "Авиа-порт", "Вес (кг)", "Длина (см)", "Ширина (см)", "Высота (см)",
    ]

    # Row 1: Russian hints
    hint_fill = PatternFill(start_color="2C2C2C", end_color="2C2C2C", fill_type="solid")
    hint_font = Font(size=9, color="919191", italic=True)
    for col_idx, label in enumerate(header_ru, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = hint_fill
        cell.font = hint_font
        cell.alignment = Alignment(horizontal="center")

    # Row 2: English headers (for import matching)
    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    header_font = Font(size=10, bold=True, color="ECECEC")
    thin_border = Border(bottom=Side(style="thin", color="3A3A3A"))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(16, len(header) + 4)

    # Sample rows
    samples = [
        ["RE48786", "Гидроцилиндр RE48786", "John Deere", 250.00, 295.00, 330.00, 10, "OEM", "Shanghai, CN", "RE48786A", "Shanghai", "PVG", 1.1, 12, 8, 6],
        ["7C-4190", "Фильтр масляный", "Caterpillar", 18.50, 22.00, 28.00, 150, "New", "Guangzhou, CN", "", "Ningbo", "CAN", 0.3, 8, 8, 10],
    ]
    data_font = Font(size=10, color="ECECEC")
    for row_idx, row_data in enumerate(samples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font

    # ── Sheet 2: Instructions ──
    wi = wb.create_sheet("Инструкция")
    wi.sheet_properties.tabColor = "64B5F6"
    wi.column_dimensions["A"].width = 80

    instructions = [
        ("Как подключить Google Sheets к Consolidator Parts", True, 14),
        ("", False, 10),
        ("1. Заполните лист \"Данные\" по образцу", False, 11),
        ("   - Строка 1 (серая) — подсказки на русском, НЕ удаляйте", False, 10),
        ("   - Строка 2 (тёмная) — названия колонок для импорта", False, 10),
        ("   - Строки 3+ — ваши данные", False, 10),
        ("", False, 10),
        ("2. Обязательные колонки:", True, 11),
        ("   PartNumber — артикул / каталожный номер детали", False, 10),
        ("   Name — название / описание", False, 10),
        ("   Price_EXW — цена EXW в долларах", False, 10),
        ("   Quantity — остаток на складе", False, 10),
        ("", False, 10),
        ("3. Загрузите файл на Google Drive", False, 11),
        ("   Google Drive → Создать → Загрузить файл → выберите этот .xlsx", False, 10),
        ("", False, 10),
        ("4. Откройте как Google Таблицу", False, 11),
        ("   Правой кнопкой → Открыть с помощью → Google Таблицы", False, 10),
        ("", False, 10),
        ("5. Откройте доступ по ссылке", False, 11),
        ("   Поделиться → Все, у кого есть ссылка → Читатель → Копировать ссылку", False, 10),
        ("", False, 10),
        ("6. Вставьте ссылку на сайте", False, 11),
        ("   Товары и прайсы → Google Sheets → Вставьте ссылку → Подключить", False, 10),
        ("", False, 10),
        ("Поддержка: support@consolidator.com", False, 10),
    ]
    for row_idx, (text, bold, size) in enumerate(instructions, 1):
        cell = wi.cell(row=row_idx, column=1, value=text)
        cell.font = Font(size=size, bold=bold, color="212647")

    # Write to response
    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="consolidator_price_template.xlsx"'
    return response


@seller_required
def seller_price_export(request: HttpRequest) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на выгрузку прайса.")
        return redirect("seller_product_list")

    parts = (
        _apply_seller_brand_scope(request.user, Part.objects.filter(seller=request.user))
        .select_related("brand", "category")
        .order_by("oem_number", "title")
    )
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="supplier_prices_{timezone.now():%Y%m%d_%H%M}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Part Number", "Description", "Unitprice", "Currency", "Stock", "OEM", "Brand", "Category", "Active"])
    for part in parts:
        writer.writerow(
            [
                part.title,
                part.description or "",
                str(part.price),
                part.currency,
                part.stock_quantity,
                part.oem_number,
                part.brand.name if part.brand else "",
                part.category.name if part.category else "",
                "1" if part.is_active else "0",
            ]
        )
    return response


@seller_required
def seller_import_errors_csv(request: HttpRequest, run_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_assortment"):
        messages.error(request, "Нет прав на просмотр ошибок импорта.")
        return redirect("seller_product_list")

    job = (
        ImportJob.objects.select_related("error_report__file")
        .filter(id=run_id, supplier=request.user)
        .first()
    )
    if job and getattr(job, "error_report", None) and job.error_report.file_id:
        content = read_stored_file_bytes(job.error_report.file.storage_key)
        response = HttpResponse(content, content_type=job.error_report.file.content_type or "text/csv")
        filename = job.error_report.file.original_name or f"import_errors_{job.id}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    run = get_object_or_404(SellerImportRun, id=run_id, seller=request.user)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="import_errors_{run.id}.csv"'
    writer = csv.writer(response)
    writer.writerow(["row_number", "original_data", "error_type", "error_message", "fix_suggestion"])
    for err in run.errors or []:
        writer.writerow(
            [
                err.get("row", ""),
                json.dumps(err.get("original_data", {}), ensure_ascii=False),
                err.get("error_type") or err.get("code", ""),
                err.get("reason", ""),
                err.get("hint", ""),
            ]
        )
    return response


@seller_required
@require_POST
def seller_order_status_update(request: HttpRequest, order_id: int) -> HttpResponse:
    if not _has_seller_permission(request.user, "can_manage_orders"):
        messages.error(request, "Нет прав на управление заказами.")
        return redirect("seller_orders")

    next_url = (request.POST.get("next") or "").strip()

    order = get_object_or_404(Order, id=order_id)
    has_access = order.items.filter(part__seller=request.user).exists()
    if not has_access:
        messages.error(request, "Вы не можете менять этот заказ.")
        return redirect("seller_orders")

    allowed = {key for key, _ in Order.STATUS_CHOICES}
    status = (request.POST.get("status") or "").strip()
    if status not in allowed:
        messages.error(request, "Неверный статус.")
        return redirect(next_url or "seller_orders")
    seller_allowed_statuses = {"pending", "reserve_paid", "confirmed", "in_production", "ready_to_ship", "transit_abroad", "customs", "transit_rf", "issuing", "shipped", "delivered", "completed", "cancelled"}
    if status not in seller_allowed_statuses:
        messages.error(request, "Этот статус может быть изменен только клиентом или системой.")
        return redirect(next_url or "seller_orders")

    current = order.status
    if status != current:
        # Build path through intermediate statuses
        path = _find_status_path(current, status)
        if path is None:
            messages.error(request, f"Недопустимый переход статуса: {current} -> {status}")
            return redirect(next_url or "seller_orders")

        # Advance through each intermediate status, logging events
        for step_status in path:
            prev = order.status
            order.status = step_status
            update_fields = ["status"]
            if step_status == "confirmed" and not order.ship_deadline:
                order.ship_deadline = timezone.now() + timedelta(days=5)
                update_fields.append("ship_deadline")
            order.save(update_fields=update_fields)
            _log_order_event(
                order,
                "status_changed",
                source="seller",
                actor=request.user,
                meta={"from": prev, "to": step_status},
            )
        _recalc_order_sla(order)

        # Handle QR code scan
        qr_code = request.POST.get("qr_code", "").strip()
        if qr_code:
            _log_order_event(
                order,
                "status_changed",
                source="seller",
                actor=request.user,
                meta={"qr_code": qr_code, "status": status},
            )

        # Handle document upload (e.g. customs declaration)
        doc_file = request.FILES.get("document")
        if doc_file:
            doc_type = "customs" if status == "customs" else "other"
            OrderDocument.objects.create(
                order=order,
                doc_type=doc_type,
                title=doc_file.name,
                file_obj=doc_file,
                uploaded_by=request.user,
            )
            _log_order_event(
                order,
                "document_uploaded",
                source="seller",
                actor=request.user,
                meta={"doc_type": doc_type, "filename": doc_file.name},
            )

    if status == "cancelled":
        seller_ids = (
            order.items.values_list("part__seller_id", flat=True)
            .exclude(part__seller_id__isnull=True)
            .distinct()
        )
        for seller_id in seller_ids:
            SupplierRatingEvent.objects.create(
                supplier_id=seller_id,
                event_type="order_cancellation",
                impact_score=Decimal("-8.00"),
                meta={"order_id": order.id},
            )
    if status == "shipped" and order.ship_deadline and timezone.now() > order.ship_deadline:
        seller_ids = (
            order.items.values_list("part__seller_id", flat=True)
            .exclude(part__seller_id__isnull=True)
            .distinct()
        )
        for seller_id in seller_ids:
            SupplierRatingEvent.objects.create(
                supplier_id=seller_id,
                event_type="delivery_delay",
                impact_score=Decimal("-5.00"),
                meta={"order_id": order.id, "deadline": order.ship_deadline.isoformat()},
            )
    messages.success(request, f"Статус заказа #{order.id} обновлен: {order.get_status_display()}")
    if next_url:
        return redirect(next_url)
    return redirect("seller_orders")


@login_required
def order_detail(request: HttpRequest, order_id: int) -> HttpResponse:
    if request.user.is_superuser:
        return redirect("admin_panel_order_detail", order_id=order_id)
    order = get_object_or_404(
        Order.objects.prefetch_related("items__part", "events", "documents", "claims"),
        id=order_id,
    )
    role = _role_for(request.user)

    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")

    _recalc_order_sla(order)
    is_buyer = order.buyer_id == request.user.id
    can_upload_docs = _can_upload_order_documents(request.user, role)
    can_manage_claims = _can_manage_claims(request.user, role)
    can_open_claim = can_manage_claims and is_buyer and order.status in {"delivered", "completed"}
    return render(
        request,
        "marketplace/order_detail.html",
        {
            "order": order,
            "events": order.events.all()[:100],
            "documents": order.documents.all()[:100],
            "claims": order.claims.all()[:100],
            "is_buyer": is_buyer,
            "can_upload_docs": can_upload_docs,
            "can_manage_claims": can_manage_claims,
            "can_open_claim": can_open_claim,
            "claim_status_choices": OrderClaim.STATUS_CHOICES,
            "document_type_choices": OrderDocument.DOC_TYPE_CHOICES,
        },
    )


@login_required
def order_invoice(request: HttpRequest, order_id: int) -> HttpResponse:
    if request.user.is_superuser:
        return redirect("admin_panel_order_detail", order_id=order_id)
    order = get_object_or_404(Order.objects.prefetch_related("items__part"), id=order_id)
    role = _role_for(request.user)

    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к инвойсу этого заказа.")
        return redirect("dashboard")

    _log_order_event(
        order,
        "invoice_opened",
        source="seller" if _role_for(request.user) == "seller" else "buyer",
        actor=request.user,
    )
    subtotal = sum((item.total_price for item in order.items.all()), Decimal("0.00"))
    reserve_due = max(Decimal("0.00"), (order.reserve_amount or Decimal("0.00")))
    final_due = max(Decimal("0.00"), (order.total_amount or Decimal("0.00")) - reserve_due)
    payment_url, payment_ref = _build_payment_url(order)
    is_buyer = order.buyer_id == request.user.id
    return render(
        request,
        "marketplace/order_invoice.html",
        {
            "order": order,
            "subtotal": subtotal,
            "reserve_due": reserve_due,
            "final_due": final_due,
            "payment_url": payment_url,
            "payment_ref": payment_ref,
            "is_buyer": is_buyer,
        },
    )


@login_required
def order_invoice_pdf(request: HttpRequest, order_id: int) -> HttpResponse:
    order = get_object_or_404(Order.objects.prefetch_related("items__part"), id=order_id)
    role = _role_for(request.user)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к инвойсу этого заказа.")
        return redirect("dashboard")

    _log_order_event(
        order,
        "invoice_opened",
        source="seller" if _role_for(request.user) == "seller" else "buyer",
        actor=request.user,
        meta={"channel": "pdf"},
    )

    subtotal = sum((item.total_price for item in order.items.all()), Decimal("0.00"))
    reserve_due = max(Decimal("0.00"), (order.reserve_amount or Decimal("0.00")))
    final_due = max(Decimal("0.00"), (order.total_amount or Decimal("0.00")) - reserve_due)
    reserve_due_date = order.created_at + timedelta(days=1)
    final_due_date = order.ship_deadline or (order.created_at + timedelta(days=7))
    payment_url, payment_ref = _build_payment_url(order)

    try:
        from reportlab.graphics.barcode.qr import QrCodeWidget
        from reportlab.graphics import renderPDF
        from reportlab.graphics.shapes import Drawing
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except Exception:
        messages.error(request, "PDF-экспорт требует пакет reportlab. Выполните: pip install reportlab")
        return redirect("order_invoice", order_id=order.id)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 15 * mm
    right = 195 * mm
    y = height - 15 * mm

    # Header
    pdf.setFillColor(colors.HexColor("#0f2f66"))
    pdf.roundRect(left, y - 16 * mm, right - left, 16 * mm, 4 * mm, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(left + 4 * mm, y - 9 * mm, "CONSOLIDATOR PARTS")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(left + 4 * mm, y - 13 * mm, "Invoice / Счет на оплату")
    y -= 23 * mm

    pdf.setFillColor(colors.HexColor("#0c1530"))
    pdf.setFont("Helvetica-Bold", 11)
    invoice_no = order.invoice_number or f"INV-{order.created_at:%Y%m%d}-{order.id}"
    pdf.drawString(left, y, f"Invoice No: {invoice_no}")
    pdf.setFont("Helvetica", 9)
    pdf.drawRightString(right, y, f"Date: {order.created_at:%d.%m.%Y}")
    y -= 7 * mm
    pdf.drawString(left, y, f"Order: #{order.id}")
    pdf.drawString(left + 48 * mm, y, f"Status: {order.get_status_display()}")
    pdf.drawString(left + 95 * mm, y, f"Payment: {order.get_payment_status_display()}")

    y -= 10 * mm
    pdf.setStrokeColor(colors.HexColor("#1a2748"))
    pdf.roundRect(left, y - 21 * mm, right - left, 21 * mm, 2 * mm, fill=0, stroke=1)
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left + 3 * mm, y - 5 * mm, "Buyer")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left + 3 * mm, y - 10 * mm, order.customer_name[:70])
    pdf.drawString(left + 3 * mm, y - 14 * mm, order.customer_email[:70])
    pdf.drawString(left + 3 * mm, y - 18 * mm, order.delivery_address[:90])

    y -= 27 * mm
    # Items table
    pdf.setFillColor(colors.HexColor("#e9f0ff"))
    pdf.rect(left, y - 7 * mm, right - left, 7 * mm, fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#1b2d57"))
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(left + 2 * mm, y - 4.8 * mm, "#")
    pdf.drawString(left + 8 * mm, y - 4.8 * mm, "Part")
    pdf.drawString(left + 96 * mm, y - 4.8 * mm, "OEM")
    pdf.drawString(left + 130 * mm, y - 4.8 * mm, "Qty")
    pdf.drawString(left + 147 * mm, y - 4.8 * mm, "Price")
    pdf.drawString(left + 176 * mm, y - 4.8 * mm, "Line")
    y -= 9 * mm
    pdf.setFont("Helvetica", 8)

    for idx, item in enumerate(order.items.all(), start=1):
        if y < 50 * mm:
            pdf.showPage()
            y = height - 20 * mm
            pdf.setFont("Helvetica", 8)
        pdf.setFillColor(colors.HexColor("#0f1f42"))
        pdf.drawString(left + 2 * mm, y, str(idx))
        pdf.drawString(left + 8 * mm, y, (item.part.title or "")[:40])
        pdf.drawString(left + 96 * mm, y, (item.part.oem_number or "")[:20])
        pdf.drawRightString(left + 143 * mm, y, str(item.quantity))
        pdf.drawRightString(left + 170 * mm, y, f"${item.unit_price}")
        pdf.drawRightString(right, y, f"${item.total_price}")
        y -= 5.1 * mm

    y -= 1 * mm
    pdf.setStrokeColor(colors.HexColor("#b7c7e8"))
    pdf.line(left + 118 * mm, y, right, y)
    y -= 6 * mm
    pdf.setFont("Helvetica", 9)
    pdf.setFillColor(colors.HexColor("#253c72"))
    pdf.drawRightString(right, y, f"Subtotal: ${subtotal}")
    y -= 5 * mm
    pdf.drawRightString(right, y, f"Logistics: ${order.logistics_cost} {order.logistics_currency} ({order.logistics_provider})")
    y -= 5 * mm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColor(colors.HexColor("#0f2f66"))
    pdf.drawRightString(right, y, f"TOTAL: ${order.total_amount}")

    # Payment schedule + QR
    y -= 11 * mm
    pdf.setFont("Helvetica-Bold", 10)
    pdf.setFillColor(colors.HexColor("#0c1530"))
    pdf.drawString(left, y, "Payment Schedule / График платежей")
    y -= 6 * mm
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left, y, f"1) Reserve {order.reserve_percent}%: ${reserve_due}  due {reserve_due_date:%d.%m.%Y}")
    y -= 5 * mm
    pdf.drawString(left, y, f"2) Final payment: ${final_due}  due {final_due_date:%d.%m.%Y}")
    y -= 5 * mm
    pdf.drawString(left, y, f"Payment reference: {payment_ref}")

    qr_size = 28 * mm
    qr = QrCodeWidget(payment_url)
    bounds = qr.getBounds()
    qr_width = bounds[2] - bounds[0]
    qr_height = bounds[3] - bounds[1]
    drawing = Drawing(qr_size, qr_size, transform=[qr_size / qr_width, 0, 0, qr_size / qr_height, 0, 0])
    drawing.add(qr)
    renderPDF.draw(drawing, pdf, right - qr_size, y - qr_size + 3 * mm)
    pdf.setFont("Helvetica", 7)
    pdf.drawString(right - qr_size, y - qr_size - 1 * mm, "Scan for payment link")

    y -= 35 * mm
    pdf.setFont("Helvetica", 8)
    pdf.setFillColor(colors.HexColor("#253c72"))
    pdf.drawString(left, y, "This invoice is generated automatically by Consolidator Parts.")
    y -= 4.5 * mm
    pdf.drawString(left, y, "Настоящий счет сформирован автоматически системой Consolidator Parts.")

    pdf.showPage()
    pdf.save()
    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{invoice_no}.pdf"'
    return response


@login_required
@require_POST
def order_mark_reserve_paid(request: HttpRequest, order_id: int) -> HttpResponse:
    order = get_object_or_404(Order, id=order_id)
    role = _role_for(request.user)
    if order.buyer_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Только клиент заказа может подтверждать оплату резерва.")
        return redirect("order_invoice", order_id=order.id)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if order.status in {"cancelled", "completed"}:
        messages.error(request, "Заказ закрыт для изменения оплаты.")
        return redirect("order_invoice", order_id=order.id)
    if order.payment_status in {"reserve_paid", "paid"}:
        messages.info(request, "Резерв уже зафиксирован.")
        return redirect("order_invoice", order_id=order.id)

    previous = order.status
    order.payment_status = "reserve_paid"
    order.reserve_paid_at = timezone.now()
    if order.status == "pending":
        order.status = "reserve_paid"
    order.save(update_fields=["payment_status", "reserve_paid_at", "status"])
    _log_order_event(order, "reserve_paid", source="buyer", actor=request.user, meta={"reserve_amount": str(order.reserve_amount)})
    if previous != order.status:
        _log_order_event(order, "status_changed", source="buyer", actor=request.user, meta={"from": previous, "to": order.status})
    messages.success(request, "Резерв 10% зафиксирован.")
    return redirect("order_invoice", order_id=order.id)


@login_required
@require_POST
def order_mark_final_paid(request: HttpRequest, order_id: int) -> HttpResponse:
    order = get_object_or_404(Order, id=order_id)
    role = _role_for(request.user)
    if order.buyer_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Только клиент заказа может подтверждать финальную оплату.")
        return redirect("order_invoice", order_id=order.id)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if order.status in {"cancelled", "completed"}:
        messages.error(request, "Заказ закрыт для изменения оплаты.")
        return redirect("order_invoice", order_id=order.id)
    if order.payment_status == "paid":
        messages.info(request, "Финальная оплата уже зафиксирована.")
        return redirect("order_invoice", order_id=order.id)
    # Для simple: reserve_paid → paid; для staged: customs_paid → paid
    if order.payment_scheme == "staged":
        if order.payment_status != "customs_paid":
            messages.error(request, "Для поэтапной схемы все промежуточные платежи должны быть зафиксированы.")
            return redirect("order_invoice", order_id=order.id)
    else:
        if order.payment_status != "reserve_paid":
            messages.error(request, "Сначала нужно зафиксировать резерв 10%.")
            return redirect("order_invoice", order_id=order.id)

    order.payment_status = "paid"
    order.final_paid_at = timezone.now()
    order.save(update_fields=["payment_status", "final_paid_at"])
    _log_order_event(order, "final_payment_paid", source="buyer", actor=request.user, meta={"total_amount": str(order.total_amount)})
    messages.success(request, "Финальная оплата зафиксирована.")
    return redirect("order_invoice", order_id=order.id)


@login_required
@require_POST
def order_mark_mid_paid(request: HttpRequest, order_id: int) -> HttpResponse:
    """Подтверждение 50% после подтверждения заказа (staged scheme)."""
    order = get_object_or_404(Order, id=order_id)
    role = _role_for(request.user)
    if order.buyer_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Только клиент заказа может подтверждать оплату.")
        return redirect("order_invoice", order_id=order.id)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if order.payment_scheme != "staged":
        messages.error(request, "Промежуточный платёж доступен только для поэтапной схемы.")
        return redirect("order_invoice", order_id=order.id)
    if order.payment_status != "reserve_paid":
        messages.error(request, "Сначала нужно зафиксировать резерв 10%.")
        return redirect("order_invoice", order_id=order.id)

    order.payment_status = "mid_paid"
    order.mid_paid_at = timezone.now()
    order.save(update_fields=["payment_status", "mid_paid_at"])
    _log_order_event(order, "mid_payment_paid", source="buyer", actor=request.user, meta={"mid_payment_amount": str(order.mid_payment_amount)})
    messages.success(request, f"Промежуточная оплата 50% (${order.mid_payment_amount}) зафиксирована.")
    return redirect("order_invoice", order_id=order.id)


@login_required
@require_POST
def order_mark_customs_paid(request: HttpRequest, order_id: int) -> HttpResponse:
    """Подтверждение 40% после прохождения таможни (staged scheme)."""
    order = get_object_or_404(Order, id=order_id)
    role = _role_for(request.user)
    if order.buyer_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Только клиент заказа может подтверждать оплату.")
        return redirect("order_invoice", order_id=order.id)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if order.payment_scheme != "staged":
        messages.error(request, "Таможенный платёж доступен только для поэтапной схемы.")
        return redirect("order_invoice", order_id=order.id)
    if order.payment_status != "mid_paid":
        messages.error(request, "Сначала нужно зафиксировать промежуточный платёж 50%.")
        return redirect("order_invoice", order_id=order.id)

    order.payment_status = "customs_paid"
    order.customs_paid_at = timezone.now()
    order.save(update_fields=["payment_status", "customs_paid_at"])
    _log_order_event(order, "customs_payment_paid", source="buyer", actor=request.user, meta={"customs_payment_amount": str(order.customs_payment_amount)})
    messages.success(request, f"Таможенная оплата 40% (${order.customs_payment_amount}) зафиксирована.")
    return redirect("order_invoice", order_id=order.id)


@login_required
@require_POST
def order_confirm_quality(request: HttpRequest, order_id: int) -> HttpResponse:
    order = get_object_or_404(Order, id=order_id)
    role = _role_for(request.user)
    if order.buyer_id != request.user.id and not request.user.is_superuser:
        messages.error(request, "Только клиент заказа может подтвердить качество.")
        return redirect("order_detail", order_id=order.id)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if order.status != "delivered":
        messages.error(request, "Качество можно подтвердить только после статуса Delivered.")
        return redirect("order_detail", order_id=order.id)
    if order.payment_status != "paid":
        messages.error(request, "Перед закрытием заказа нужна финальная оплата.")
        return redirect("order_detail", order_id=order.id)

    previous = order.status
    order.status = "completed"
    order.save(update_fields=["status"])
    _log_order_event(order, "quality_confirmed", source="buyer", actor=request.user)
    _log_order_event(order, "status_changed", source="buyer", actor=request.user, meta={"from": previous, "to": order.status})
    messages.success(request, "Качество подтверждено. Заказ закрыт.")
    return redirect("order_detail", order_id=order.id)


@login_required
@require_POST
def order_add_document(request: HttpRequest, order_id: int) -> HttpResponse:
    order = get_object_or_404(Order, id=order_id)
    role = _role_for(request.user)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if not _can_upload_order_documents(request.user, role):
        messages.error(request, "Нет прав на загрузку документов.")
        return redirect("order_detail", order_id=order.id)

    doc_type = (request.POST.get("doc_type") or "other").strip()
    title = (request.POST.get("title") or "").strip()
    file_url = (request.POST.get("file_url") or "").strip()
    file_obj = request.FILES.get("file_obj")
    allowed_doc_types = {key for key, _ in OrderDocument.DOC_TYPE_CHOICES}
    blocked_extensions = {
        ".exe",
        ".sh",
        ".bat",
        ".cmd",
        ".msi",
        ".php",
        ".js",
        ".jar",
        ".com",
        ".scr",
    }
    allowed_extensions = {
        ".pdf",
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".txt",
        ".csv",
        ".doc",
        ".docx",
        ".xls",
        ".xlsx",
    }

    if doc_type not in allowed_doc_types:
        messages.error(request, "Некорректный тип документа.")
        return redirect("order_detail", order_id=order.id)
    if not title:
        messages.error(request, "Укажите название документа.")
        return redirect("order_detail", order_id=order.id)
    if not file_url and not file_obj:
        messages.error(request, "Добавьте файл или ссылку на документ.")
        return redirect("order_detail", order_id=order.id)
    if file_obj:
        ext = os.path.splitext(file_obj.name or "")[1].lower()
        if ext in blocked_extensions or ext not in allowed_extensions:
            messages.error(request, "Тип файла не разрешен.")
            return redirect("order_detail", order_id=order.id)
        if int(file_obj.size or 0) > int(settings.MAX_ORDER_DOCUMENT_BYTES):
            messages.error(request, f"Файл слишком большой (макс. {settings.MAX_ORDER_DOCUMENT_BYTES} байт).")
            return redirect("order_detail", order_id=order.id)
        # Normalize filename for storage.
        safe_name = slugify(os.path.splitext(file_obj.name)[0]) or "document"
        file_obj.name = f"{safe_name}{ext}"

    doc = OrderDocument.objects.create(
        order=order,
        doc_type=doc_type,
        title=title,
        file_url=file_url,
        file_obj=file_obj,
        uploaded_by=request.user,
    )
    _log_order_event(
        order,
        "document_uploaded",
        source="seller" if role == "seller" else "buyer",
        actor=request.user,
        meta={"document_id": doc.id, "doc_type": doc_type, "title": title},
    )
    messages.success(request, "Документ добавлен к заказу.")
    return redirect("order_detail", order_id=order.id)


@csrf_exempt
@require_POST
def payment_callback(request: HttpRequest) -> HttpResponse:
    configured_secret = (getattr(settings, "PAYMENT_CALLBACK_SECRET", "") or "").strip()
    provided_secret = (
        request.headers.get("X-Payment-Secret")
        or request.POST.get("secret")
        or request.GET.get("secret")
        or ""
    ).strip()
    if configured_secret and configured_secret != provided_secret:
        return JsonResponse({"ok": False, "error": "invalid_secret"}, status=403)

    payload: dict = {}
    if request.content_type and "application/json" in request.content_type.lower():
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except Exception:
            payload = {}
    if not payload:
        payload = request.POST.dict()

    order_id_raw = payload.get("order_id") or payload.get("orderId")
    invoice_number = (payload.get("invoice_number") or payload.get("invoice") or "").strip()
    callback_status = (payload.get("status") or payload.get("payment_status") or "").strip().lower()
    transaction_id = (payload.get("transaction_id") or payload.get("tx_id") or "").strip()

    order = None
    if order_id_raw:
        try:
            order = Order.objects.filter(id=int(order_id_raw)).first()
        except Exception:
            order = None
    if not order and invoice_number:
        order = Order.objects.filter(invoice_number=invoice_number).first()
    if not order:
        return JsonResponse({"ok": False, "error": "order_not_found"}, status=404)

    meta = {
        "callback_status": callback_status,
        "transaction_id": transaction_id,
        "invoice_number": order.invoice_number,
    }
    changed_fields: list[str] = []
    if callback_status in {"reserve_paid", "reserve_success", "deposit_paid"}:
        if order.payment_status not in {"reserve_paid", "paid"}:
            order.payment_status = "reserve_paid"
            changed_fields.append("payment_status")
        if not order.reserve_paid_at:
            order.reserve_paid_at = timezone.now()
            changed_fields.append("reserve_paid_at")
        if order.status == "pending":
            prev_status = order.status
            order.status = "reserve_paid"
            changed_fields.append("status")
            _log_order_event(order, "status_changed", source="system", meta={"from": prev_status, "to": order.status, **meta})
        order.save(update_fields=list(set(changed_fields)))
        _log_order_event(order, "reserve_paid", source="system", meta=meta)
    elif callback_status in {"mid_paid", "mid_payment", "confirmation_paid"}:
        if order.payment_scheme == "staged" and order.payment_status not in {"mid_paid", "customs_paid", "paid"}:
            order.payment_status = "mid_paid"
            changed_fields.append("payment_status")
        if not order.mid_paid_at:
            order.mid_paid_at = timezone.now()
            changed_fields.append("mid_paid_at")
        order.save(update_fields=list(set(changed_fields)))
        _log_order_event(order, "mid_payment_paid", source="system", meta=meta)
    elif callback_status in {"customs_paid", "customs_payment"}:
        if order.payment_scheme == "staged" and order.payment_status not in {"customs_paid", "paid"}:
            order.payment_status = "customs_paid"
            changed_fields.append("payment_status")
        if not order.customs_paid_at:
            order.customs_paid_at = timezone.now()
            changed_fields.append("customs_paid_at")
        order.save(update_fields=list(set(changed_fields)))
        _log_order_event(order, "customs_payment_paid", source="system", meta=meta)
    elif callback_status in {"paid", "success", "final_paid", "full_paid"}:
        if order.payment_status != "paid":
            order.payment_status = "paid"
            changed_fields.append("payment_status")
        if not order.final_paid_at:
            order.final_paid_at = timezone.now()
            changed_fields.append("final_paid_at")
        order.save(update_fields=list(set(changed_fields)))
        _log_order_event(order, "final_payment_paid", source="system", meta=meta)
    elif callback_status in {"refunded", "refund"}:
        if order.payment_status != "refunded":
            order.payment_status = "refunded"
            order.save(update_fields=["payment_status"])
        _log_order_event(order, "status_changed", source="system", meta={"from": order.status, "to": order.status, **meta})
    else:
        return JsonResponse({"ok": False, "error": "unsupported_status", "status": callback_status}, status=400)

    return JsonResponse({"ok": True, "order_id": order.id, "payment_status": order.payment_status})


@login_required
@require_POST
def order_open_claim(request: HttpRequest, order_id: int) -> HttpResponse:
    order = get_object_or_404(Order, id=order_id)
    role = _role_for(request.user)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if not _can_manage_claims(request.user, role):
        messages.error(request, "Нет прав на работу с рекламациями.")
        return redirect("order_detail", order_id=order.id)
    if order.status not in {"delivered", "completed"}:
        messages.error(request, "Рекламация доступна после доставки.")
        return redirect("order_detail", order_id=order.id)

    title = (request.POST.get("title") or "").strip()
    description = (request.POST.get("description") or "").strip()
    if not title or not description:
        messages.error(request, "Заполните тему и описание рекламации.")
        return redirect("order_detail", order_id=order.id)

    claim = OrderClaim.objects.create(
        order=order,
        title=title,
        description=description,
        status="open",
        opened_by=request.user,
    )
    _log_order_event(
        order,
        "claim_opened",
        source="seller" if role == "seller" else "buyer",
        actor=request.user,
        meta={"claim_id": claim.id, "title": title},
    )
    messages.success(request, "Рекламация открыта.")
    return redirect("order_detail", order_id=order.id)


@login_required
@require_POST
def order_update_claim_status(request: HttpRequest, claim_id: int) -> HttpResponse:
    claim = get_object_or_404(OrderClaim.objects.select_related("order"), id=claim_id)
    order = claim.order
    role = _role_for(request.user)
    if not _has_order_access(request.user, order, role):
        messages.error(request, "Нет доступа к заказу.")
        return redirect("dashboard")
    if not _can_manage_claims(request.user, role):
        messages.error(request, "Нет прав на работу с рекламациями.")
        return redirect("order_detail", order_id=order.id)

    new_status = (request.POST.get("status") or "").strip()
    allowed_statuses = {key for key, _ in OrderClaim.STATUS_CHOICES}
    if new_status not in allowed_statuses:
        messages.error(request, "Некорректный статус рекламации.")
        return redirect("order_detail", order_id=order.id)

    prev_status = claim.status
    if prev_status == new_status:
        messages.info(request, "Статус рекламации не изменился.")
        return redirect("order_detail", order_id=order.id)

    claim.status = new_status
    claim.resolved_by = request.user if new_status in {"approved", "rejected", "closed"} else claim.resolved_by
    claim.save(update_fields=["status", "resolved_by", "updated_at"])
    _log_order_event(
        order,
        "claim_status_changed",
        source="seller" if role == "seller" else "buyer",
        actor=request.user,
        meta={"claim_id": claim.id, "from": prev_status, "to": new_status},
    )
    if new_status == "approved":
        order.payment_status = "refund_pending"
        order.save(update_fields=["payment_status"])
    if new_status in {"rejected", "closed"} and order.payment_status == "refund_pending":
        order.payment_status = "paid"
        order.save(update_fields=["payment_status"])

    messages.success(request, "Статус рекламации обновлён.")
    return redirect("order_detail", order_id=order.id)


# ═══ Operator cabinet views ═══


def operator_select_role(request):
    return render(request, "operator/select_role.html", {})


def operator_logist_dashboard(request):
    return render(request, "operator/logist/dashboard.html", {"operator_role": "logist", "operator_active_nav": "dashboard"})


def operator_logist_shipments(request):
    return render(request, "operator/logist/shipments.html", {"operator_role": "logist", "operator_active_nav": "shipments"})


def operator_logist_quotes(request):
    return render(request, "operator/logist/quotes.html", {"operator_role": "logist", "operator_active_nav": "quotes"})


def operator_logist_routes(request):
    return render(request, "operator/logist/routes.html", {"operator_role": "logist", "operator_active_nav": "routes"})


def operator_customs_dashboard(request):
    return render(request, "operator/customs/dashboard.html", {"operator_role": "customs", "operator_active_nav": "dashboard"})


def operator_customs_declarations(request):
    return render(request, "operator/customs/declarations.html", {"operator_role": "customs", "operator_active_nav": "declarations"})


def operator_customs_tariffs(request):
    return render(request, "operator/customs/tariffs.html", {"operator_role": "customs", "operator_active_nav": "tariffs"})


def operator_payments_dashboard(request):
    return render(request, "operator/payments/dashboard.html", {"operator_role": "payments", "operator_active_nav": "dashboard"})


def operator_payments_invoices(request):
    return render(request, "operator/payments/invoices.html", {"operator_role": "payments", "operator_active_nav": "invoices"})


def operator_payments_escrow(request):
    return render(request, "operator/payments/escrow.html", {"operator_role": "payments", "operator_active_nav": "escrow"})


def operator_manager_dashboard(request):
    return render(request, "operator/manager/dashboard.html", {"operator_role": "manager", "operator_active_nav": "dashboard"})


def operator_manager_orders(request):
    return render(request, "operator/manager/orders.html", {"operator_role": "manager", "operator_active_nav": "orders"})


def operator_manager_clients(request):
    return render(request, "operator/manager/clients.html", {"operator_role": "manager", "operator_active_nav": "clients"})


def operator_logist_ports(request):
    return render(request, "operator/logist/ports.html", {"operator_role": "logist", "operator_active_nav": "ports"})


def operator_logist_documents(request):
    return render(request, "operator/logist/documents.html", {"operator_role": "logist", "operator_active_nav": "documents"})


def operator_logist_analytics(request):
    return render(request, "operator/logist/analytics.html", {"operator_role": "logist", "operator_active_nav": "analytics"})


def operator_customs_documents(request):
    return render(request, "operator/customs/documents.html", {"operator_role": "customs", "operator_active_nav": "documents"})


def operator_customs_requests(request):
    return render(request, "operator/customs/requests.html", {"operator_role": "customs", "operator_active_nav": "requests"})


def operator_customs_analytics(request):
    return render(request, "operator/customs/analytics.html", {"operator_role": "customs", "operator_active_nav": "analytics"})


def operator_payments_reconciliation(request):
    return render(request, "operator/payments/reconciliation.html", {"operator_role": "payments", "operator_active_nav": "reconciliation"})


def operator_payments_analytics(request):
    return render(request, "operator/payments/analytics.html", {"operator_role": "payments", "operator_active_nav": "analytics"})


def operator_manager_shipments(request):
    return render(request, "operator/manager/shipments.html", {"operator_role": "manager", "operator_active_nav": "shipments_mgr"})


def operator_manager_negotiations(request):
    return render(request, "operator/manager/negotiations.html", {"operator_role": "manager", "operator_active_nav": "negotiations"})


def operator_manager_analytics(request):
    return render(request, "operator/manager/analytics.html", {"operator_role": "manager", "operator_active_nav": "analytics"})


# ═══ Admin panel ═══


@admin_required
def admin_panel_dashboard(request):
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    total_users = User.objects.count()
    active_orders = Order.objects.exclude(status__in=["completed", "cancelled"]).count()
    month_revenue = Order.objects.filter(created_at__gte=month_start).aggregate(s=Sum("total_amount"))["s"] or 0
    rfq_today = RFQ.objects.filter(created_at__gte=today_start).count()
    total_orders = Order.objects.count()
    avg_order = (Order.objects.aggregate(s=Sum("total_amount"))["s"] or 0) / max(total_orders, 1)

    recent_events = OrderEvent.objects.select_related("order").order_by("-created_at")[:10]

    # Alerts
    pending_suppliers = UserProfile.objects.filter(role="seller", supplier_status="sandbox").count()
    overdue_orders = Order.objects.filter(sla_status="breached").count()
    open_claims = OrderClaim.objects.exclude(status="closed").count()
    blocked_parts = Part.objects.filter(availability_status="blocked").count()

    # Today summary
    orders_today = Order.objects.filter(created_at__gte=today_start).count()
    paid_today = Order.objects.filter(reserve_paid_at__gte=today_start).aggregate(s=Sum("total_amount"))["s"] or 0
    registrations_today = User.objects.filter(date_joined__gte=today_start).count()

    # Extended KPI
    total_gmv = Order.objects.aggregate(s=Sum("total_amount"))["s"] or 0
    total_parts = Part.objects.count()
    total_rfq = RFQ.objects.count()
    month_rfq = RFQ.objects.filter(created_at__gte=month_start).count()
    quoted_rfq = RFQ.objects.filter(status="quoted").count()
    rfq_conversion = round(quoted_rfq / max(total_rfq, 1) * 100, 1)
    completed_orders = Order.objects.filter(status__in=["completed", "delivered"]).count()
    completion_rate = round(completed_orders / max(total_orders, 1) * 100, 1)
    total_sellers = UserProfile.objects.filter(role="seller").count()
    total_buyers = UserProfile.objects.filter(role="buyer").count()
    pending_payments = Order.objects.filter(payment_status="awaiting_reserve").aggregate(s=Sum("total_amount"))["s"] or 0
    sla_ok = Order.objects.filter(sla_status="on_track").exclude(status__in=["completed", "cancelled"]).count()
    sla_total = Order.objects.exclude(status__in=["completed", "cancelled"]).count()
    sla_rate = round(sla_ok / max(sla_total, 1) * 100, 1)
    recent_imports_dash = SellerImportRun.objects.select_related("seller").order_by("-created_at")[:3]

    ctx = {
        "admin_active_nav": "dashboard",
        "total_users": total_users,
        "active_orders": active_orders,
        "month_revenue": month_revenue,
        "rfq_today": rfq_today,
        "avg_order": round(avg_order, 0),
        "total_orders": total_orders,
        "recent_events": recent_events,
        "pending_suppliers": pending_suppliers,
        "overdue_orders": overdue_orders,
        "open_claims": open_claims,
        "blocked_parts": blocked_parts,
        "orders_today": orders_today,
        "paid_today": paid_today,
        "rfq_today_count": rfq_today,
        "registrations_today": registrations_today,
        "total_gmv": total_gmv,
        "total_parts": total_parts,
        "total_rfq": total_rfq,
        "month_rfq": month_rfq,
        "rfq_conversion": rfq_conversion,
        "completion_rate": completion_rate,
        "total_sellers": total_sellers,
        "total_buyers": total_buyers,
        "pending_payments": pending_payments,
        "sla_rate": sla_rate,
        "recent_imports_dash": recent_imports_dash,
    }
    return render(request, "admin_panel/dashboard.html", ctx)


@admin_required
def admin_panel_users(request):
    now = timezone.now()
    week_ago = now - timedelta(days=7)

    # POST actions
    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "toggle_user":
            uid = request.POST.get("user_id")
            target = User.objects.filter(id=uid).first()
            if target and not target.is_superuser:
                target.is_active = not target.is_active
                target.save(update_fields=["is_active"])
        elif action == "bulk_block":
            ids = request.POST.getlist("selected_users")
            User.objects.filter(id__in=ids, is_superuser=False).update(is_active=False)
        elif action == "bulk_activate":
            ids = request.POST.getlist("selected_users")
            User.objects.filter(id__in=ids, is_superuser=False).update(is_active=True)
        return redirect("admin_panel_users")

    # Filters
    search = request.GET.get("q", "").strip()
    role_filter = request.GET.get("role", "")
    status_filter = request.GET.get("status", "")
    sort = request.GET.get("sort", "-date_joined")

    qs = User.objects.select_related("profile")

    if search:
        qs = qs.filter(Q(username__icontains=search) | Q(email__icontains=search) | Q(first_name__icontains=search) | Q(last_name__icontains=search))
    if role_filter == "buyer":
        qs = qs.filter(profile__role="buyer")
    elif role_filter == "seller":
        qs = qs.filter(profile__role="seller")
    elif role_filter == "admin":
        qs = qs.filter(is_superuser=True)
    if status_filter == "active":
        qs = qs.filter(is_active=True)
    elif status_filter == "blocked":
        qs = qs.filter(is_active=False)

    valid_sorts = {
        "-date_joined": "-date_joined",
        "date_joined": "date_joined",
        "username": "username",
        "-id": "-id",
    }
    qs = qs.order_by(valid_sorts.get(sort, "-date_joined"))

    # Pagination
    page_num = int(request.GET.get("page", "1"))
    per_page = 50
    offset = (page_num - 1) * per_page
    users_list = list(qs[offset:offset + per_page + 1])
    has_next = len(users_list) > per_page
    if has_next:
        users_list = users_list[:per_page]

    # Stats
    total = User.objects.count()
    buyers = UserProfile.objects.filter(role="buyer").count()
    sellers = UserProfile.objects.filter(role="seller").count()
    blocked = User.objects.filter(is_active=False).count()
    new_this_week = User.objects.filter(date_joined__gte=week_ago).count()

    ctx = {
        "admin_active_nav": "users",
        "users": users_list,
        "total": total,
        "buyers": buyers,
        "sellers": sellers,
        "blocked": blocked,
        "new_this_week": new_this_week,
        "search": search,
        "role_filter": role_filter,
        "status_filter": status_filter,
        "current_sort": sort,
        "page_num": page_num,
        "has_next": has_next,
        "has_prev": page_num > 1,
    }
    return render(request, "admin_panel/users.html", ctx)


@admin_required
def admin_panel_orders(request):
    now = timezone.now()
    week_ago = now - timedelta(days=7)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # POST: bulk actions
    if request.method == "POST":
        action = request.POST.get("action", "")
        ids = request.POST.getlist("selected_orders")
        new_status = request.POST.get("new_status", "")
        if ids and action == "bulk_status" and new_status:
            valid = [c[0] for c in Order.STATUS_CHOICES]
            if new_status in valid:
                for order in Order.objects.filter(id__in=ids):
                    old = order.status
                    if old != new_status:
                        order.status = new_status
                        order.save(update_fields=["status"])
                        OrderEvent.objects.create(
                            order=order, event_type="status_changed", source="admin",
                            meta={"comment": f"Массовое: {old} → {new_status}"},
                        )
        elif ids and action == "bulk_cancel":
            for order in Order.objects.filter(id__in=ids).exclude(status="cancelled"):
                old = order.status
                order.status = "cancelled"
                order.save(update_fields=["status"])
                OrderEvent.objects.create(
                    order=order, event_type="status_changed", source="admin",
                    meta={"comment": f"Массовая отмена: {old} → cancelled"},
                )
        return redirect("admin_panel_orders")

    # Filters
    search = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "")
    payment_filter = request.GET.get("payment", "")
    sla_filter = request.GET.get("sla", "")
    sort = request.GET.get("sort", "-created_at")

    qs = Order.objects.select_related("buyer").order_by(sort if sort in ("-created_at", "created_at", "-total_amount", "total_amount") else "-created_at")

    if search:
        if search.isdigit():
            qs = qs.filter(id=int(search))
        else:
            qs = qs.filter(Q(customer_name__icontains=search) | Q(customer_email__icontains=search))
    if status_filter:
        qs = qs.filter(status=status_filter)
    if payment_filter:
        qs = qs.filter(payment_status=payment_filter)
    if sla_filter == "breached":
        qs = qs.filter(sla_status="breached")
    elif sla_filter == "on_track":
        qs = qs.filter(sla_status="on_track")

    # Pagination
    page_num = int(request.GET.get("page", "1"))
    per_page = 50
    offset = (page_num - 1) * per_page
    orders_list = list(qs[offset:offset + per_page + 1])
    has_next = len(orders_list) > per_page
    if has_next:
        orders_list = orders_list[:per_page]

    # Stats
    total = Order.objects.count()
    active = Order.objects.exclude(status__in=["completed", "cancelled"]).count()
    completed = Order.objects.filter(status__in=["completed", "delivered"]).count()
    cancelled = Order.objects.filter(status="cancelled").count()
    total_amount = Order.objects.aggregate(s=Sum("total_amount"))["s"] or 0
    week_orders = Order.objects.filter(created_at__gte=week_ago).count()
    week_revenue = Order.objects.filter(created_at__gte=week_ago).aggregate(s=Sum("total_amount"))["s"] or 0
    problem_count = Order.objects.filter(Q(sla_status="breached") | Q(claims__status__in=["open", "in_review"])).distinct().count()

    ctx = {
        "admin_active_nav": "orders",
        "orders": orders_list,
        "total": total,
        "active": active,
        "completed": completed,
        "cancelled": cancelled,
        "total_amount": total_amount,
        "week_orders": week_orders,
        "week_revenue": week_revenue,
        "problem_count": problem_count,
        "search": search,
        "status_filter": status_filter,
        "payment_filter": payment_filter,
        "sla_filter": sla_filter,
        "current_sort": sort,
        "page_num": page_num,
        "has_next": has_next,
        "has_prev": page_num > 1,
        "status_choices": Order.STATUS_CHOICES,
        "payment_choices": Order.PAYMENT_STATUS_CHOICES,
    }
    return render(request, "admin_panel/orders.html", ctx)


@admin_required
def admin_panel_rfq(request):
    now = timezone.now()
    week_ago = now - timedelta(days=7)

    # POST: bulk actions
    if request.method == "POST":
        action = request.POST.get("action", "")
        ids = request.POST.getlist("selected_rfqs")
        if ids and action == "bulk_cancel":
            RFQ.objects.filter(id__in=ids).exclude(status="cancelled").update(status="cancelled")
        elif ids and action == "bulk_status":
            new_status = request.POST.get("new_status", "")
            if new_status in dict(RFQ.STATUS_CHOICES):
                RFQ.objects.filter(id__in=ids).update(status=new_status)
        return redirect("admin_panel_rfq")

    # Filters
    search = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "")
    urgency_filter = request.GET.get("urgency", "")
    mode_filter = request.GET.get("mode", "")

    qs = RFQ.objects.select_related("created_by").order_by("-created_at")
    if search:
        if search.isdigit():
            qs = qs.filter(id=int(search))
        else:
            qs = qs.filter(Q(customer_name__icontains=search) | Q(customer_email__icontains=search) | Q(company_name__icontains=search))
    if status_filter:
        qs = qs.filter(status=status_filter)
    if urgency_filter:
        qs = qs.filter(urgency=urgency_filter)
    if mode_filter:
        qs = qs.filter(mode=mode_filter)

    # Pagination
    page_num = int(request.GET.get("page", "1"))
    per_page = 50
    offset = (page_num - 1) * per_page
    rfq_list = list(qs[offset:offset + per_page + 1])
    has_next = len(rfq_list) > per_page
    if has_next:
        rfq_list = rfq_list[:per_page]

    # Stats
    total = RFQ.objects.count()
    active = RFQ.objects.filter(status="new").count()
    quoted = RFQ.objects.filter(status="quoted").count()
    needs_review = RFQ.objects.filter(status="needs_review").count()
    cancelled = RFQ.objects.filter(status="cancelled").count()
    week_rfq = RFQ.objects.filter(created_at__gte=week_ago).count()
    conversion = round(quoted / max(total, 1) * 100, 1)

    ctx = {
        "admin_active_nav": "rfq",
        "rfqs": rfq_list,
        "total": total,
        "active": active,
        "quoted": quoted,
        "needs_review": needs_review,
        "cancelled": cancelled,
        "week_rfq": week_rfq,
        "conversion": conversion,
        "search": search,
        "status_filter": status_filter,
        "urgency_filter": urgency_filter,
        "mode_filter": mode_filter,
        "page_num": page_num,
        "has_next": has_next,
        "has_prev": page_num > 1,
        "status_choices": RFQ.STATUS_CHOICES,
        "urgency_choices": RFQ.URGENCY_CHOICES,
        "mode_choices": RFQ.MODE_CHOICES,
    }
    return render(request, "admin_panel/rfq.html", ctx)


@admin_required
def admin_panel_finance(request):
    now = timezone.now()
    week_ago = now - timedelta(days=7)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Filters
    payment_filter = request.GET.get("payment", "")
    search = request.GET.get("q", "").strip()

    qs = Order.objects.select_related("buyer").order_by("-created_at")
    if payment_filter:
        qs = qs.filter(payment_status=payment_filter)
    if search:
        if search.isdigit():
            qs = qs.filter(id=int(search))
        else:
            qs = qs.filter(Q(customer_name__icontains=search))

    # Pagination
    page_num = int(request.GET.get("page", "1"))
    per_page = 50
    offset = (page_num - 1) * per_page
    orders_list = list(qs[offset:offset + per_page + 1])
    has_next = len(orders_list) > per_page
    if has_next:
        orders_list = orders_list[:per_page]

    # Stats
    total_revenue = Order.objects.aggregate(s=Sum("total_amount"))["s"] or 0
    month_revenue = Order.objects.filter(created_at__gte=month_start).aggregate(s=Sum("total_amount"))["s"] or 0
    week_revenue = Order.objects.filter(created_at__gte=week_ago).aggregate(s=Sum("total_amount"))["s"] or 0
    total_logistics = Order.objects.aggregate(s=Sum("logistics_cost"))["s"] or 0
    reserve_collected = Order.objects.exclude(reserve_paid_at=None).aggregate(s=Sum("reserve_amount"))["s"] or 0
    fully_paid = Order.objects.filter(payment_status="paid").aggregate(s=Sum("total_amount"))["s"] or 0
    fully_paid_count = Order.objects.filter(payment_status="paid").count()
    pending_sum = Order.objects.filter(payment_status="awaiting_reserve").aggregate(s=Sum("total_amount"))["s"] or 0
    pending_count = Order.objects.filter(payment_status="awaiting_reserve").count()
    reserve_paid_sum = Order.objects.filter(payment_status="reserve_paid").aggregate(s=Sum("total_amount"))["s"] or 0
    reserve_paid_count = Order.objects.filter(payment_status="reserve_paid").count()
    mid_paid_sum = Order.objects.filter(payment_status="mid_paid").aggregate(s=Sum("total_amount"))["s"] or 0
    mid_paid_count = Order.objects.filter(payment_status="mid_paid").count()

    # Platform commission estimate (8% default)
    commission_rate = 8
    platform_commission = round(float(total_revenue) * commission_rate / 100, 0)

    # Stuck payments — reserve paid more than 7 days ago but not fully paid
    stuck = Order.objects.filter(payment_status="reserve_paid", reserve_paid_at__lt=week_ago).count()

    ctx = {
        "admin_active_nav": "finance",
        "orders": orders_list,
        "total_revenue": total_revenue,
        "month_revenue": month_revenue,
        "week_revenue": week_revenue,
        "total_logistics": total_logistics,
        "reserve_collected": reserve_collected,
        "fully_paid": fully_paid,
        "fully_paid_count": fully_paid_count,
        "pending_sum": pending_sum,
        "pending_count": pending_count,
        "reserve_paid_sum": reserve_paid_sum,
        "reserve_paid_count": reserve_paid_count,
        "mid_paid_sum": mid_paid_sum,
        "mid_paid_count": mid_paid_count,
        "platform_commission": platform_commission,
        "commission_rate": commission_rate,
        "stuck": stuck,
        "search": search,
        "payment_filter": payment_filter,
        "page_num": page_num,
        "has_next": has_next,
        "has_prev": page_num > 1,
        "payment_choices": Order.PAYMENT_STATUS_CHOICES,
    }
    return render(request, "admin_panel/finance.html", ctx)


@admin_required
def admin_panel_catalog(request):
    if request.method == "POST":
        action = request.POST.get("action")
        admin_note = request.POST.get("admin_note", "").strip()
        now = timezone.now()

        # Bulk actions
        selected_ids = request.POST.getlist("selected_parts")
        if selected_ids and action in ("bulk_block", "bulk_activate"):
            parts_qs = Part.objects.filter(id__in=selected_ids)
            if action == "bulk_block":
                parts_qs.update(
                    availability_status="blocked", is_active=False,
                    admin_note=admin_note, moderated_at=now, moderated_by=request.user,
                )
            elif action == "bulk_activate":
                parts_qs.update(
                    availability_status="active", is_active=True,
                    admin_note="", moderated_at=now, moderated_by=request.user,
                )
            return redirect("admin_panel_catalog")

        # Single part actions
        part_id = request.POST.get("part_id")
        part = Part.objects.filter(id=part_id).first()
        if part:
            if action == "block":
                part.availability_status = "blocked"
                part.is_active = False
                part.admin_note = admin_note
                part.moderated_at = now
                part.moderated_by = request.user
                part.save(update_fields=["availability_status", "is_active", "admin_note", "moderated_at", "moderated_by"])
            elif action == "activate":
                part.availability_status = "active"
                part.is_active = True
                part.admin_note = ""
                part.moderated_at = now
                part.moderated_by = request.user
                part.save(update_fields=["availability_status", "is_active", "admin_note", "moderated_at", "moderated_by"])
        return redirect("admin_panel_catalog")

    # Filters
    seller_filter = request.GET.get("seller", "")
    brand_filter = request.GET.get("brand", "")
    category_filter = request.GET.get("category", "")
    status_filter = request.GET.get("status", "")
    search = request.GET.get("q", "").strip()
    show = request.GET.get("show", "")  # "no_price", "no_brand", "zero_stock", "duplicates"

    parts_qs = Part.objects.select_related("brand", "category", "seller").order_by("-id")
    if seller_filter:
        parts_qs = parts_qs.filter(seller__username=seller_filter)
    if brand_filter:
        parts_qs = parts_qs.filter(brand__name=brand_filter)
    if category_filter:
        parts_qs = parts_qs.filter(category__name=category_filter)
    if status_filter == "active":
        parts_qs = parts_qs.filter(availability_status="active", is_active=True)
    elif status_filter == "blocked":
        parts_qs = parts_qs.filter(availability_status="blocked")
    elif status_filter == "inactive":
        parts_qs = parts_qs.filter(is_active=False)
    if search:
        parts_qs = parts_qs.filter(Q(oem_number__icontains=search) | Q(title__icontains=search))

    # Data quality filters
    if show == "no_price":
        parts_qs = parts_qs.filter(price__lte=0)
    elif show == "no_brand":
        parts_qs = parts_qs.filter(brand__isnull=True)
    elif show == "zero_stock":
        parts_qs = parts_qs.filter(stock_quantity=0, availability="in_stock")
    elif show == "duplicates":
        # OEM numbers that appear more than once
        from django.db.models import Count as DCount
        dup_oems = (Part.objects.values("oem_number")
                    .annotate(cnt=DCount("id")).filter(cnt__gt=1)
                    .values_list("oem_number", flat=True)[:100])
        parts_qs = parts_qs.filter(oem_number__in=list(dup_oems)).order_by("oem_number")

    # Fast pagination
    page_num = int(request.GET.get("page", "1"))
    per_page = 50
    offset = (page_num - 1) * per_page
    parts_list = list(parts_qs[offset:offset + per_page + 1])
    has_next = len(parts_list) > per_page
    if has_next:
        parts_list = parts_list[:per_page]

    # Fast approximate count via EXPLAIN (SQLite) or table stats
    # Avoid COUNT(*) on millions of rows
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("SELECT MAX(id) FROM marketplace_part")
        row = cursor.fetchone()
        parts_count_approx = row[0] or 0  # approximate, fast

    categories_count = Category.objects.count()  # small table, fast
    brands_count = Brand.objects.count()  # small table, fast
    sellers_count = UserProfile.objects.filter(role="seller").count()  # small table
    blocked_count = Part.objects.filter(availability_status="blocked").count() if Part.objects.filter(availability_status="blocked").exists() else 0
    active_count = parts_count_approx - blocked_count

    # Data quality — only compute when specifically requested, otherwise show "check" link
    no_price = 0
    no_brand = 0
    zero_stock = 0
    dup_oem_count = 0
    if show == "no_price":
        no_price = parts_qs.filter(price__lte=0).count() if not search else len(parts_list)
    elif show == "no_brand":
        no_brand = parts_qs.filter(brand__isnull=True).count() if not search else len(parts_list)
    elif show == "zero_stock":
        zero_stock = parts_qs.filter(stock_quantity=0).count() if not search else len(parts_list)
    elif show == "duplicates":
        dup_oem_count = 1

    # Light lists for filters
    categories = Category.objects.order_by("name")[:30]
    brands = Brand.objects.order_by("name")[:50]
    sellers_list = User.objects.filter(parts__isnull=False).distinct().order_by("username")[:20]
    recent_imports = SellerImportRun.objects.select_related("seller").order_by("-created_at")[:5]

    ctx = {
        "admin_active_nav": "catalog",
        "parts_count": parts_count_approx,
        "active_count": active_count,
        "blocked_count": blocked_count,
        "categories_count": categories_count,
        "brands_count": brands_count,
        "sellers_count": sellers_count,
        "no_price": no_price,
        "no_brand": no_brand,
        "zero_stock": zero_stock,
        "dup_oem_count": dup_oem_count,
        "parts": parts_list,
        "page_num": page_num,
        "has_next": has_next,
        "has_prev": page_num > 1,
        "next_page": page_num + 1,
        "prev_page": page_num - 1,
        "categories": categories,
        "brands": brands,
        "sellers_list": sellers_list,
        "current_seller": seller_filter,
        "current_brand": brand_filter,
        "current_category": category_filter,
        "status_filter": status_filter,
        "search": search,
        "show": show,
        "recent_imports": recent_imports,
    }
    return render(request, "admin_panel/catalog.html", ctx)


@admin_required
def admin_panel_moderation(request):
    now = timezone.now()

    if request.method == "POST":
        action = request.POST.get("action")
        # Supplier moderation
        if action in ("approve_supplier", "reject_supplier"):
            profile_id = request.POST.get("profile_id")
            profile = UserProfile.objects.filter(id=profile_id).first()
            if profile:
                old_status = profile.supplier_status
                profile.supplier_status = "trusted" if action == "approve_supplier" else "rejected"
                profile.save(update_fields=["supplier_status"])
                # Log
                SupplierRatingEvent.objects.create(
                    supplier=profile.user,
                    event_type="manual_override",
                    score_impact=Decimal("0"),
                    reason=f"Модерация: {old_status} → {profile.supplier_status} (админ: {request.user.username})",
                )
        # Part moderation
        elif action == "activate_part":
            part_id = request.POST.get("part_id")
            part = Part.objects.filter(id=part_id).first()
            if part:
                part.availability_status = "active"
                part.is_active = True
                part.admin_note = ""
                part.moderated_at = now
                part.moderated_by = request.user
                part.save(update_fields=["availability_status", "is_active", "admin_note", "moderated_at", "moderated_by"])
        # Claim moderation
        elif action in ("close_claim", "approve_claim", "reject_claim"):
            claim_id = request.POST.get("claim_id")
            claim = OrderClaim.objects.filter(id=claim_id).first()
            if claim:
                if action == "close_claim":
                    claim.status = "closed"
                elif action == "approve_claim":
                    claim.status = "approved"
                elif action == "reject_claim":
                    claim.status = "rejected"
                claim.resolved_by = request.user
                claim.save(update_fields=["status", "resolved_by"])
        return redirect("admin_panel_moderation")

    # Tab filter
    tab = request.GET.get("tab", "all")

    sandbox_suppliers = UserProfile.objects.filter(role="seller", supplier_status="sandbox").select_related("user")
    blocked_parts = Part.objects.filter(availability_status="blocked").select_related("brand", "seller").order_by("-updated_at")[:50]
    open_claims = OrderClaim.objects.exclude(status="closed").select_related("order").order_by("-created_at")[:50]
    sla_breached = Order.objects.filter(sla_status="breached").exclude(status__in=["completed", "cancelled"]).select_related("buyer").order_by("-created_at")[:20]

    # Recent moderation history
    recent_moderated = Part.objects.filter(moderated_at__isnull=False).select_related("moderated_by", "seller", "brand").order_by("-moderated_at")[:10]
    recent_supplier_events = SupplierRatingEvent.objects.filter(event_type="manual_override").select_related("supplier").order_by("-created_at")[:10]

    total_pending = sandbox_suppliers.count() + open_claims.count() + sla_breached.count()
    if blocked_parts.exists():
        total_pending += blocked_parts.count()

    ctx = {
        "admin_active_nav": "moderation",
        "tab": tab,
        "sandbox_suppliers": sandbox_suppliers,
        "blocked_parts": blocked_parts,
        "open_claims": open_claims,
        "sla_breached": sla_breached,
        "total_pending": total_pending,
        "supplier_count": sandbox_suppliers.count(),
        "parts_count": blocked_parts.count(),
        "claims_count": open_claims.count(),
        "sla_count": sla_breached.count(),
        "recent_moderated": recent_moderated,
        "recent_supplier_events": recent_supplier_events,
    }
    return render(request, "admin_panel/moderation.html", ctx)


@admin_required
def admin_panel_settings(request):
    settings_path = os.path.join(settings.BASE_DIR, "platform_settings.json")
    saved = False

    # Load existing settings
    platform_cfg = {
        "platform_name": "Consolidator Parts",
        "support_email": "support@consolidator.com",
        "maintenance_mode": False,
        "default_commission": 8,
        "rfq_response_hours": 24,
        "max_delivery_days": 14,
        "sla_penalty_percent": 2,
        "claim_response_hours": 48,
    }
    if os.path.exists(settings_path):
        with open(settings_path, "r") as f:
            platform_cfg.update(json.load(f))

    if request.method == "POST":
        tab = request.POST.get("tab", "general")
        if tab == "general":
            platform_cfg["platform_name"] = request.POST.get("platform_name", platform_cfg["platform_name"])
            platform_cfg["support_email"] = request.POST.get("support_email", platform_cfg["support_email"])
            platform_cfg["maintenance_mode"] = "maintenance_mode" in request.POST
        elif tab == "sla":
            platform_cfg["rfq_response_hours"] = int(request.POST.get("rfq_response_hours", 24))
            platform_cfg["max_delivery_days"] = int(request.POST.get("max_delivery_days", 14))
            platform_cfg["sla_penalty_percent"] = int(request.POST.get("sla_penalty_percent", 2))
            platform_cfg["claim_response_hours"] = int(request.POST.get("claim_response_hours", 48))
        elif tab == "currencies":
            currencies = platform_cfg.get("currencies", [
                {"code": "USD", "name": "US Dollar", "is_active": True},
                {"code": "CNY", "name": "Chinese Yuan", "is_active": True},
                {"code": "EUR", "name": "Euro", "is_active": True},
                {"code": "RUB", "name": "Russian Ruble", "is_active": True},
            ])
            for cur in currencies:
                cur["is_active"] = f"cur_{cur['code']}" in request.POST
            new_code = request.POST.get("new_code", "").strip().upper()
            new_name = request.POST.get("new_name", "").strip()
            if new_code and new_name and not any(c["code"] == new_code for c in currencies):
                currencies.append({"code": new_code, "name": new_name, "is_active": True})
            platform_cfg["currencies"] = currencies
        elif tab == "notifications":
            notif_keys = ["order_created", "new_rfq", "payment_received", "claim_opened", "sla_breach"]
            notifications = {}
            for key in notif_keys:
                notifications[key] = f"notif_{key}" in request.POST
            platform_cfg["notifications"] = notifications
        with open(settings_path, "w") as f:
            json.dump(platform_cfg, f, indent=2)
        saved = True

    currencies = platform_cfg.get("currencies", [
        {"code": "USD", "name": "US Dollar", "is_active": True},
        {"code": "CNY", "name": "Chinese Yuan", "is_active": True},
        {"code": "EUR", "name": "Euro", "is_active": True},
        {"code": "RUB", "name": "Russian Ruble", "is_active": True},
    ])
    notifications = platform_cfg.get("notifications", {
        "order_created": True, "new_rfq": True, "payment_received": True,
        "claim_opened": True, "sla_breach": True,
    })
    notif_labels = {
        "order_created": "Новый заказ", "new_rfq": "Новый RFQ",
        "payment_received": "Оплата получена", "claim_opened": "Рекламация открыта",
        "sla_breach": "SLA нарушение",
    }
    ctx = {
        "admin_active_nav": "settings",
        "cfg": platform_cfg,
        "currencies": currencies,
        "notifications": notifications,
        "notif_labels": notif_labels,
        "saved": saved,
    }
    return render(request, "admin_panel/settings.html", ctx)


@admin_required
def admin_panel_analytics(request):
    now = timezone.now()

    # Period filter
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    period = request.GET.get("period", "month")

    if date_from and date_to:
        from datetime import datetime as _dt
        period = "custom"
        period_start = timezone.make_aware(_dt.strptime(date_from, "%Y-%m-%d"))
        period_end = timezone.make_aware(_dt.strptime(date_to, "%Y-%m-%d")) + timedelta(days=1)
        duration = period_end - period_start
        prev_end = period_start
        prev_start = period_start - duration
        period_label = f"{date_from} — {date_to}"
    elif period == "week":
        period_start = now - timedelta(days=7)
        prev_start = now - timedelta(days=14)
        prev_end = period_start
        period_label = "За неделю"
    elif period == "quarter":
        period_start = now - timedelta(days=90)
        prev_start = now - timedelta(days=180)
        prev_end = period_start
        period_label = "За квартал"
    elif period == "year":
        period_start = now - timedelta(days=365)
        prev_start = now - timedelta(days=730)
        prev_end = period_start
        period_label = "За год"
    elif period == "all":
        period_start = None
        prev_start = None
        prev_end = None
        period_label = "Всё время"
    else:  # month
        period_start = now - timedelta(days=30)
        prev_start = now - timedelta(days=60)
        prev_end = period_start
        period_label = "За месяц"

    # Current period
    orders_qs = Order.objects.all()
    rfq_qs = RFQ.objects.all()
    if period_start:
        orders_qs = orders_qs.filter(created_at__gte=period_start)
        rfq_qs = rfq_qs.filter(created_at__gte=period_start)

    gmv = orders_qs.aggregate(s=Sum("total_amount"))["s"] or 0
    orders_count = orders_qs.count()
    avg_order = round(float(gmv) / max(orders_count, 1), 0)
    rfq_count = rfq_qs.count()
    rfq_quoted = rfq_qs.filter(status="quoted").count()
    rfq_conversion = round(rfq_quoted / max(rfq_count, 1) * 100, 1)
    completed = orders_qs.filter(status__in=["completed", "delivered"]).count()
    completion_rate = round(completed / max(orders_count, 1) * 100, 1)
    new_users = User.objects.filter(date_joined__gte=period_start).count() if period_start else User.objects.count()

    # Previous period for comparison
    prev_gmv = 0
    prev_orders = 0
    prev_rfq = 0
    if prev_start and prev_end:
        prev_gmv = Order.objects.filter(created_at__gte=prev_start, created_at__lt=prev_end).aggregate(s=Sum("total_amount"))["s"] or 0
        prev_orders = Order.objects.filter(created_at__gte=prev_start, created_at__lt=prev_end).count()
        prev_rfq = RFQ.objects.filter(created_at__gte=prev_start, created_at__lt=prev_end).count()

    # Deltas
    gmv_delta = round(float(gmv - prev_gmv) / max(float(prev_gmv), 1) * 100, 1) if prev_gmv else 0
    orders_delta = round((orders_count - prev_orders) / max(prev_orders, 1) * 100, 1) if prev_orders else 0

    # Top buyers
    top_buyers_qs = User.objects.filter(orders__isnull=False)
    if period_start:
        top_buyers_qs = top_buyers_qs.filter(orders__created_at__gte=period_start)
    top_buyers = (
        top_buyers_qs.annotate(order_count=Count("orders", distinct=True), total_spent=Sum("orders__total_amount"))
        .order_by("-total_spent")[:10]
    )

    # Top sellers
    top_sellers = (
        User.objects.filter(parts__isnull=False)
        .annotate(parts_count=Count("parts", distinct=True))
        .order_by("-parts_count")[:10]
    )

    # By category (light — no heavy joins)
    by_category = Category.objects.annotate(parts_count=Count("parts")).order_by("-parts_count")[:10]

    # By brand
    by_brand = Brand.objects.annotate(parts_count=Count("parts")).order_by("-parts_count")[:10]

    # Status breakdown
    by_status = list(orders_qs.values("status").annotate(count=Count("id")).order_by("-count"))
    status_labels = dict(Order.STATUS_CHOICES)
    for item in by_status:
        item["label"] = status_labels.get(item["status"], item["status"])

    by_payment = list(orders_qs.values("payment_status").annotate(
        count=Count("id"), total=Sum("total_amount")
    ).order_by("-total"))
    payment_labels = dict(Order.PAYMENT_STATUS_CHOICES)
    for item in by_payment:
        item["label"] = payment_labels.get(item["payment_status"], item["payment_status"])

    total_users = User.objects.count()

    ctx = {
        "admin_active_nav": "analytics",
        "period": period,
        "period_label": period_label,
        "date_from": date_from,
        "date_to": date_to,
        "gmv": gmv,
        "orders_count": orders_count,
        "avg_order": avg_order,
        "rfq_count": rfq_count,
        "rfq_conversion": rfq_conversion,
        "completion_rate": completion_rate,
        "new_users": new_users,
        "total_users": total_users,
        "gmv_delta": gmv_delta,
        "orders_delta": orders_delta,
        "top_buyers": top_buyers,
        "top_sellers": top_sellers,
        "by_category": by_category,
        "by_brand": by_brand,
        "by_status": by_status,
        "by_payment": by_payment,
    }
    return render(request, "admin_panel/analytics.html", ctx)


@admin_required
def admin_panel_logs(request):
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    tab = request.GET.get("tab", "events")

    if tab == "webhooks":
        wh_qs = WebhookDeliveryLog.objects.select_related("order").order_by("-created_at")
        wh_success = request.GET.get("success", "")
        wh_q = request.GET.get("q", "").strip()
        if wh_success == "1":
            wh_qs = wh_qs.filter(success=True)
        elif wh_success == "0":
            wh_qs = wh_qs.filter(success=False)
        if wh_q:
            wh_qs = wh_qs.filter(endpoint__icontains=wh_q) if not wh_q.isdigit() else wh_qs.filter(order_id=int(wh_q))
        per_page = 50
        page_num = int(request.GET.get("page", 1))
        offset = (page_num - 1) * per_page
        webhooks = list(wh_qs[offset:offset + per_page + 1])
        has_next = len(webhooks) > per_page
        if has_next:
            webhooks = webhooks[:per_page]
        total_wh = WebhookDeliveryLog.objects.count()
        failed_wh = WebhookDeliveryLog.objects.filter(success=False).count()
        ctx = {
            "admin_active_nav": "logs", "tab": "webhooks",
            "webhooks": webhooks, "total_wh": total_wh, "failed_wh": failed_wh,
            "success_rate": round((total_wh - failed_wh) / max(total_wh, 1) * 100, 1),
            "wh_success": wh_success, "wh_q": wh_q,
            "page_num": page_num, "has_next": has_next, "has_prev": page_num > 1,
        }
        return render(request, "admin_panel/logs.html", ctx)

    # Events tab
    qs = OrderEvent.objects.select_related("order").order_by("-created_at")
    event_type = request.GET.get("type", "")
    source = request.GET.get("source", "")
    q = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    if event_type:
        qs = qs.filter(event_type=event_type)
    if source:
        qs = qs.filter(source=source)
    if q:
        if q.isdigit():
            qs = qs.filter(order_id=int(q))
        else:
            qs = qs.filter(Q(order__customer_name__icontains=q) | Q(meta__comment__icontains=q))
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    per_page = 50
    page_num = int(request.GET.get("page", 1))
    offset = (page_num - 1) * per_page
    events = list(qs[offset:offset + per_page + 1])
    has_next = len(events) > per_page
    if has_next:
        events = events[:per_page]

    total_events = OrderEvent.objects.count()
    today_events = OrderEvent.objects.filter(created_at__gte=today_start).count()
    week_events = OrderEvent.objects.filter(created_at__gte=now - timedelta(days=7)).count()

    ctx = {
        "admin_active_nav": "logs", "tab": "events",
        "events": events, "total_events": total_events,
        "today_events": today_events, "week_events": week_events,
        "event_type_choices": OrderEvent.EVENT_CHOICES,
        "source_choices": OrderEvent.SOURCE_CHOICES,
        "current_type": event_type, "current_source": source,
        "q": q, "date_from": date_from, "date_to": date_to,
        "page_num": page_num, "has_next": has_next, "has_prev": page_num > 1,
    }
    return render(request, "admin_panel/logs.html", ctx)


@admin_required
def admin_panel_tariffs(request):
    settings_path = os.path.join(settings.BASE_DIR, "platform_settings.json")
    platform_cfg = {}
    if os.path.exists(settings_path):
        with open(settings_path, "r") as f:
            platform_cfg = json.load(f)

    # Defaults
    default_plans = [
        {"id": "basic", "name": "Базовый", "price": 0, "commission": 10, "max_products": 100, "is_active": True},
        {"id": "professional", "name": "Профессиональный", "price": 99, "commission": 8, "max_products": 10000, "is_active": True},
        {"id": "corporate", "name": "Корпоративный", "price": 499, "commission": 5, "max_products": 0, "is_active": True},
    ]
    tariff_plans = platform_cfg.get("tariff_plans", default_plans)
    category_commissions = platform_cfg.get("category_commissions", {})
    discounts = platform_cfg.get("discounts", [])

    saved = False
    if request.method == "POST":
        tab = request.POST.get("tab", "plans")
        if tab == "plans":
            for plan in tariff_plans:
                plan["price"] = int(request.POST.get(f"price_{plan['id']}", plan["price"]))
                plan["commission"] = int(request.POST.get(f"commission_{plan['id']}", plan["commission"]))
                plan["max_products"] = int(request.POST.get(f"max_{plan['id']}", plan["max_products"]))
                plan["is_active"] = f"active_{plan['id']}" in request.POST
        elif tab == "categories":
            for cat in Category.objects.all():
                cid = str(cat.id)
                pct = request.POST.get(f"pct_{cid}", "8")
                min_a = request.POST.get(f"min_{cid}", "0")
                max_a = request.POST.get(f"max_{cid}", "1000")
                category_commissions[cid] = {
                    "percent": float(pct), "min_amount": float(min_a), "max_amount": float(max_a),
                }
        elif tab == "discounts":
            action = request.POST.get("action", "")
            if action == "add_discount":
                discounts.append({
                    "name": request.POST.get("new_name", ""),
                    "type": request.POST.get("new_type", "commission"),
                    "value": request.POST.get("new_value", "0"),
                    "period": request.POST.get("new_period", ""),
                    "is_active": True,
                })
            elif action == "save_discounts":
                for i, d in enumerate(discounts):
                    d["is_active"] = f"active_{i}" in request.POST

        platform_cfg["tariff_plans"] = tariff_plans
        platform_cfg["category_commissions"] = category_commissions
        platform_cfg["discounts"] = discounts
        with open(settings_path, "w") as f:
            json.dump(platform_cfg, f, indent=2, ensure_ascii=False)
        saved = True

    # Count suppliers per tier
    from django.db.models import Count as _Count
    seller_parts = (
        User.objects.filter(parts__isnull=False)
        .annotate(pc=_Count("parts"))
        .values_list("pc", flat=True)
    )
    tier_counts = {"basic": 0, "professional": 0, "corporate": 0}
    for pc in seller_parts:
        if pc > 10000:
            tier_counts["corporate"] += 1
        elif pc > 100:
            tier_counts["professional"] += 1
        else:
            tier_counts["basic"] += 1

    categories = Category.objects.annotate(parts_count=Count("parts")).order_by("name")

    ctx = {
        "admin_active_nav": "tariffs",
        "tariff_plans": tariff_plans,
        "category_commissions": category_commissions,
        "discounts": discounts,
        "tier_counts": tier_counts,
        "categories": categories,
        "saved": saved,
        "total_sellers": UserProfile.objects.filter(role="seller").count(),
    }
    return render(request, "admin_panel/tariffs.html", ctx)


@admin_required
def admin_panel_support(request):
    now = timezone.now()

    # Load SLA threshold
    settings_path = os.path.join(settings.BASE_DIR, "platform_settings.json")
    sla_hours = 48
    if os.path.exists(settings_path):
        with open(settings_path, "r") as f:
            sla_hours = json.load(f).get("claim_response_hours", 48)

    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "bulk_status":
            ids = request.POST.getlist("claim_ids")
            new_status = request.POST.get("bulk_new_status", "")
            if ids and new_status in dict(OrderClaim.STATUS_CHOICES):
                claims = OrderClaim.objects.filter(id__in=ids)
                for claim in claims:
                    claim.status = new_status
                    if new_status in ("closed", "approved", "rejected"):
                        claim.resolved_by = request.user
                    claim.save(update_fields=["status", "resolved_by"])
                    note = request.POST.get("admin_note", "").strip()
                    if note:
                        OrderEvent.objects.create(
                            order=claim.order, event_type="claim_status_changed",
                            source="admin", actor=request.user,
                            meta={"comment": note, "new_status": new_status, "claim_id": claim.id},
                        )
        else:
            claim_id = request.POST.get("claim_id")
            new_status = request.POST.get("new_status")
            claim = OrderClaim.objects.filter(id=claim_id).first()
            if claim and new_status in dict(OrderClaim.STATUS_CHOICES):
                claim.status = new_status
                if new_status in ("closed", "approved", "rejected"):
                    claim.resolved_by = request.user
                claim.save(update_fields=["status", "resolved_by"])
                note = request.POST.get("admin_note", "").strip()
                if note:
                    OrderEvent.objects.create(
                        order=claim.order, event_type="claim_status_changed",
                        source="admin", actor=request.user,
                        meta={"comment": note, "new_status": new_status, "claim_id": claim.id},
                    )
        return redirect(f"/admin-panel/support/?status={request.GET.get('status', '')}&q={request.GET.get('q', '')}")

    # Filters
    status_filter = request.GET.get("status", "open")
    q = request.GET.get("q", "").strip()

    qs = OrderClaim.objects.select_related("order", "opened_by").order_by("-created_at")
    if status_filter == "open":
        qs = qs.exclude(status="closed")
    elif status_filter in ("in_review", "approved", "rejected", "closed"):
        qs = qs.filter(status=status_filter)
    # else "all"

    if q:
        if q.isdigit():
            qs = qs.filter(Q(order_id=int(q)) | Q(id=int(q)))
        else:
            qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q) | Q(order__customer_name__icontains=q))

    # Pagination
    per_page = 50
    page_num = int(request.GET.get("page", 1))
    offset = (page_num - 1) * per_page
    claims_list = list(qs[offset:offset + per_page + 1])
    has_next = len(claims_list) > per_page
    if has_next:
        claims_list = claims_list[:per_page]

    # Compute age and overdue for each claim
    sla_threshold = timedelta(hours=sla_hours)
    for claim in claims_list:
        age = now - claim.created_at
        claim.hours_open = int(age.total_seconds() / 3600)
        claim.is_overdue = claim.status not in ("closed", "approved", "rejected") and age > sla_threshold

    total_claims = OrderClaim.objects.count()
    open_count = OrderClaim.objects.exclude(status="closed").count()
    resolved_count = OrderClaim.objects.filter(status="closed").count()
    overdue_count = OrderClaim.objects.exclude(
        status__in=("closed", "approved", "rejected")
    ).filter(created_at__lt=now - sla_threshold).count()

    ctx = {
        "admin_active_nav": "support",
        "claims": claims_list,
        "total_claims": total_claims,
        "open_count": open_count,
        "resolved_count": resolved_count,
        "overdue_count": overdue_count,
        "sla_hours": sla_hours,
        "status_choices": OrderClaim.STATUS_CHOICES,
        "status_filter": status_filter,
        "q": q,
        "page_num": page_num, "has_next": has_next, "has_prev": page_num > 1,
    }
    return render(request, "admin_panel/support.html", ctx)


@admin_required
def admin_panel_order_detail(request, order_id):
    order = get_object_or_404(Order.objects.select_related("buyer"), id=order_id)
    items = order.items.select_related("part", "part__brand")
    events = OrderEvent.objects.filter(order=order).order_by("-created_at")[:20]
    documents = OrderDocument.objects.filter(order=order)
    claims = OrderClaim.objects.filter(order=order)

    if request.method == "POST":
        if "new_status" in request.POST:
            new_status = request.POST["new_status"]
            valid = [c[0] for c in Order.STATUS_CHOICES]
            if new_status in valid and new_status != order.status:
                old = order.status
                order.status = new_status
                order.save(update_fields=["status"])
                OrderEvent.objects.create(
                    order=order, event_type="status_changed", source="admin",
                    meta={"comment": f"Админ: {old} → {new_status}"},
                )
        elif "new_payment_status" in request.POST:
            new_ps = request.POST["new_payment_status"]
            valid_ps = [c[0] for c in Order.PAYMENT_STATUS_CHOICES]
            if new_ps in valid_ps and new_ps != order.payment_status:
                old_ps = order.payment_status
                order.payment_status = new_ps
                if new_ps == "reserve_paid" and not order.reserve_paid_at:
                    order.reserve_paid_at = timezone.now()
                elif new_ps == "paid" and not order.final_paid_at:
                    order.final_paid_at = timezone.now()
                order.save(update_fields=["payment_status", "reserve_paid_at", "final_paid_at"])
                OrderEvent.objects.create(
                    order=order, event_type="status_changed", source="admin",
                    meta={"comment": f"Админ оплата: {old_ps} → {new_ps}"},
                )
        return redirect("admin_panel_order_detail", order_id=order.id)

    # Timeline step
    status_order = ["pending", "reserve_paid", "confirmed", "in_production", "ready_to_ship", "transit_abroad", "customs", "transit_rf", "issuing", "shipped", "delivered", "completed"]
    current_step = status_order.index(order.status) + 1 if order.status in status_order else 0

    ctx = {
        "admin_active_nav": "orders",
        "order": order,
        "items": items,
        "events": events,
        "documents": documents,
        "claims": claims,
        "status_choices": Order.STATUS_CHOICES,
        "payment_status_choices": Order.PAYMENT_STATUS_CHOICES,
        "current_step": current_step,
        "total_steps": len(status_order),
    }
    return render(request, "admin_panel/order_detail.html", ctx)


@admin_required
def admin_panel_rfq_detail(request, rfq_id):
    rfq = get_object_or_404(RFQ.objects.select_related("created_by"), id=rfq_id)
    items = RFQItem.objects.filter(rfq=rfq).select_related("matched_part", "matched_part__brand")

    if request.method == "POST":
        new_status = request.POST.get("new_status", "")
        if new_status in dict(RFQ.STATUS_CHOICES):
            rfq.status = new_status
            rfq.save(update_fields=["status"])
        return redirect("admin_panel_rfq_detail", rfq_id=rfq.id)

    ctx = {
        "admin_active_nav": "rfq",
        "rfq": rfq,
        "items": items,
        "status_choices": RFQ.STATUS_CHOICES,
    }
    return render(request, "admin_panel/rfq_detail.html", ctx)


@admin_required
def admin_panel_user_detail(request, user_id):
    target = get_object_or_404(User.objects.select_related("profile"), id=user_id)
    user_orders = Order.objects.filter(buyer=target).order_by("-created_at")[:20]
    seller_orders = Order.objects.filter(items__part__seller=target).distinct().order_by("-created_at")[:20]
    user_parts = Part.objects.filter(seller=target).order_by("-created_at")[:20]
    user_rfqs = RFQ.objects.filter(created_by=target).order_by("-created_at")[:20]
    user_imports = SellerImportRun.objects.filter(seller=target).order_by("-created_at")[:20]
    recent_parts = Part.objects.filter(seller=target, updated_at__gte=timezone.now() - timedelta(days=7)).order_by("-updated_at")[:30]

    # Supplier metrics
    total_parts_count = Part.objects.filter(seller=target).count()
    active_parts_count = Part.objects.filter(seller=target, is_active=True).count()
    seller_revenue = Order.objects.filter(items__part__seller=target).distinct().aggregate(s=Sum("total_amount"))["s"] or 0
    seller_order_count = Order.objects.filter(items__part__seller=target).distinct().count()
    buyer_order_count = Order.objects.filter(buyer=target).count()
    buyer_spent = Order.objects.filter(buyer=target).aggregate(s=Sum("total_amount"))["s"] or 0
    rating = target.profile.rating_score if hasattr(target, "profile") else 0

    if request.method == "POST":
        action = request.POST.get("action", "toggle_active")
        if action == "toggle_active" and not target.is_superuser:
            target.is_active = not target.is_active
            target.save(update_fields=["is_active"])
        elif action == "block_new_parts":
            Part.objects.filter(seller=target, updated_at__gte=timezone.now() - timedelta(days=7)).update(availability_status="blocked", is_active=False)
        elif action == "approve_new_parts":
            Part.objects.filter(seller=target, availability_status="blocked").update(availability_status="active", is_active=True)
        elif action == "save_admin_note" and hasattr(target, "profile"):
            target.profile.admin_note = request.POST.get("admin_note", "")
            target.profile.save(update_fields=["admin_note"])
        elif action == "update_profile" and hasattr(target, "profile"):
            profile = target.profile
            new_role = request.POST.get("role")
            if new_role in ("buyer", "seller"):
                profile.role = new_role
            profile.company_name = request.POST.get("company_name", profile.company_name)
            new_dept = request.POST.get("department")
            if new_dept in dict(UserProfile.DEPARTMENT_CHOICES):
                profile.department = new_dept
            profile.can_manage_assortment = "can_manage_assortment" in request.POST
            profile.can_manage_pricing = "can_manage_pricing" in request.POST
            profile.can_manage_orders = "can_manage_orders" in request.POST
            profile.can_manage_drawings = "can_manage_drawings" in request.POST
            profile.can_view_analytics = "can_view_analytics" in request.POST
            profile.can_manage_team = "can_manage_team" in request.POST
            new_status = request.POST.get("supplier_status")
            if new_status in dict(UserProfile.SUPPLIER_STATUS_CHOICES):
                profile.supplier_status = new_status
            profile.save()
        return redirect("admin_panel_user_detail", user_id=target.id)

    ctx = {
        "admin_active_nav": "users",
        "target_user": target,
        "user_orders": user_orders,
        "seller_orders": seller_orders,
        "user_parts": user_parts,
        "user_rfqs": user_rfqs,
        "user_imports": user_imports,
        "recent_parts": recent_parts,
        "total_parts_count": total_parts_count,
        "active_parts_count": active_parts_count,
        "seller_revenue": seller_revenue,
        "seller_order_count": seller_order_count,
        "buyer_order_count": buyer_order_count,
        "buyer_spent": buyer_spent,
        "rating": rating,
    }
    return render(request, "admin_panel/user_detail.html", ctx)


@admin_required
def admin_panel_imports(request):
    seller_filter = request.GET.get("seller", "")
    qs = SellerImportRun.objects.select_related("seller").order_by("-created_at")
    if seller_filter:
        qs = qs.filter(seller__username=seller_filter)
    imports = qs[:200]
    sellers_list = User.objects.filter(import_runs__isnull=False).distinct().order_by("username")
    total_imports = SellerImportRun.objects.count()
    total_created = sum(i.created_count for i in imports)
    total_updated = sum(i.updated_count for i in imports)
    total_errors = sum(i.error_count for i in imports)

    ctx = {
        "admin_active_nav": "imports",
        "imports": imports,
        "sellers_list": sellers_list,
        "current_seller": seller_filter,
        "total_imports": total_imports,
        "total_created": total_created,
        "total_updated": total_updated,
        "total_errors": total_errors,
    }
    return render(request, "admin_panel/imports.html", ctx)


@admin_required
def admin_panel_import_detail(request, import_id):
    imp = get_object_or_404(SellerImportRun.objects.select_related("seller"), id=import_id)

    # Find parts updated around the time of this import (±2 minutes)
    time_from = imp.created_at - timedelta(minutes=2)
    time_to = imp.created_at + timedelta(minutes=2)
    parts = Part.objects.filter(
        seller=imp.seller,
        data_updated_at__gte=time_from,
        data_updated_at__lte=time_to,
    ).select_related("brand", "category").order_by("-updated_at")

    if request.method == "POST":
        action = request.POST.get("action")
        admin_note = request.POST.get("admin_note", "").strip()
        now = timezone.now()
        selected_ids = request.POST.getlist("selected_parts")

        if selected_ids:
            parts_qs = Part.objects.filter(id__in=selected_ids)
            if action == "bulk_activate":
                parts_qs.update(availability_status="active", is_active=True, admin_note="", moderated_at=now, moderated_by=request.user)
            elif action == "bulk_block":
                parts_qs.update(availability_status="blocked", is_active=False, admin_note=admin_note, moderated_at=now, moderated_by=request.user)
        elif action == "activate_all":
            parts.update(availability_status="active", is_active=True, admin_note="", moderated_at=now, moderated_by=request.user)
        elif action == "block_all":
            parts.update(availability_status="blocked", is_active=False, admin_note=admin_note, moderated_at=now, moderated_by=request.user)

        return redirect("admin_panel_import_detail", import_id=imp.id)

    ctx = {
        "admin_active_nav": "imports",
        "imp": imp,
        "parts": parts,
        "parts_count": parts.count(),
    }
    return render(request, "admin_panel/import_detail.html", ctx)


@admin_required
def admin_panel_export_csv(request, report_type):
    """Universal admin CSV export."""
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response.write("\ufeff")  # BOM for Excel
    writer = csv.writer(response)

    if report_type == "users":
        response["Content-Disposition"] = 'attachment; filename="users_export.csv"'
        writer.writerow(["id", "username", "email", "role", "company", "is_active", "date_joined"])
        for u in User.objects.select_related("profile").order_by("-date_joined"):
            role = u.profile.role if hasattr(u, "profile") else "buyer"
            company = u.profile.company_name if hasattr(u, "profile") else ""
            writer.writerow([u.id, u.username, u.email, role, company, u.is_active, u.date_joined.strftime("%Y-%m-%d")])
    elif report_type == "orders":
        response["Content-Disposition"] = 'attachment; filename="orders_export.csv"'
        writer.writerow(["id", "customer_name", "buyer", "status", "payment_status", "total_amount", "logistics_cost", "sla_status", "created_at"])
        for o in Order.objects.select_related("buyer").order_by("-created_at"):
            writer.writerow([o.id, o.customer_name, o.buyer.username if o.buyer else "", o.status, o.payment_status, o.total_amount, o.logistics_cost, o.sla_status, o.created_at.strftime("%Y-%m-%d")])
    elif report_type == "rfq":
        response["Content-Disposition"] = 'attachment; filename="rfq_export.csv"'
        writer.writerow(["id", "customer_name", "company", "mode", "urgency", "status", "created_at"])
        for r in RFQ.objects.order_by("-created_at"):
            writer.writerow([r.id, r.customer_name, r.company_name, r.mode, r.urgency, r.status, r.created_at.strftime("%Y-%m-%d")])
    elif report_type == "parts":
        response["Content-Disposition"] = 'attachment; filename="catalog_export.csv"'
        writer.writerow(["id", "oem_number", "title", "brand", "category", "seller", "price", "currency", "stock", "status"])
        for p in Part.objects.select_related("brand", "category", "seller").order_by("-created_at")[:5000]:
            writer.writerow([p.id, p.oem_number, p.title, p.brand.name if p.brand else "", p.category.name if p.category else "", p.seller.username if p.seller else "", p.price, p.currency, p.stock_quantity, p.availability_status])
    elif report_type == "finance":
        response["Content-Disposition"] = 'attachment; filename="finance_export.csv"'
        writer.writerow(["order_id", "customer_name", "total_amount", "reserve_amount", "logistics_cost", "payment_status", "payment_scheme", "created_at"])
        for o in Order.objects.order_by("-created_at"):
            writer.writerow([o.id, o.customer_name, o.total_amount, o.reserve_amount, o.logistics_cost, o.payment_status, o.payment_scheme, o.created_at.strftime("%Y-%m-%d")])
    elif report_type == "events":
        response["Content-Disposition"] = 'attachment; filename="events_export.csv"'
        writer.writerow(["id", "order_id", "event_type", "source", "meta", "created_at"])
        for e in OrderEvent.objects.order_by("-created_at")[:5000]:
            writer.writerow([e.id, e.order_id, e.event_type, e.source, json.dumps(e.meta) if e.meta else "", e.created_at.strftime("%Y-%m-%d %H:%M")])
    else:
        return HttpResponse("Unknown report type", status=404)

    return response


@seller_required
def seller_report_export_csv(request, report_type):
    """Universal seller CSV export."""
    user = request.user
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response.write("\ufeff")
    writer = csv.writer(response)

    if report_type == "orders":
        response["Content-Disposition"] = 'attachment; filename="seller_orders.csv"'
        writer.writerow(["order_id", "customer", "status", "payment_status", "total_amount", "sla_status", "created_at"])
        orders = Order.objects.filter(items__part__seller=user).distinct().order_by("-created_at")
        for o in orders:
            writer.writerow([o.id, o.customer_name, o.status, o.payment_status, o.total_amount, o.sla_status, o.created_at.strftime("%Y-%m-%d")])
    elif report_type == "parts":
        response["Content-Disposition"] = 'attachment; filename="seller_parts.csv"'
        writer.writerow(["oem_number", "title", "brand", "category", "price", "currency", "stock", "condition", "status"])
        for p in Part.objects.filter(seller=user).select_related("brand", "category").order_by("-created_at"):
            writer.writerow([p.oem_number, p.title, p.brand.name if p.brand else "", p.category.name if p.category else "", p.price, p.currency, p.stock_quantity, p.condition, p.availability_status])
    elif report_type == "sla":
        response["Content-Disposition"] = 'attachment; filename="seller_sla.csv"'
        writer.writerow(["order_id", "customer", "status", "sla_status", "breaches", "confirm_deadline", "ship_deadline", "created_at"])
        orders = Order.objects.filter(items__part__seller=user).distinct().order_by("-created_at")
        for o in orders:
            writer.writerow([o.id, o.customer_name, o.status, o.sla_status, o.sla_breaches_count, o.supplier_confirm_deadline or "", o.ship_deadline or "", o.created_at.strftime("%Y-%m-%d")])
    elif report_type == "finance":
        response["Content-Disposition"] = 'attachment; filename="seller_finance.csv"'
        writer.writerow(["order_id", "customer", "total_amount", "reserve_amount", "logistics_cost", "payment_status", "payment_scheme", "created_at"])
        orders = Order.objects.filter(items__part__seller=user).distinct().order_by("-created_at")
        for o in orders:
            writer.writerow([o.id, o.customer_name, o.total_amount, o.reserve_amount, o.logistics_cost, o.payment_status, o.payment_scheme, o.created_at.strftime("%Y-%m-%d")])
    elif report_type == "stock":
        response["Content-Disposition"] = 'attachment; filename="seller_stock.csv"'
        writer.writerow(["oem_number", "title", "brand", "stock_quantity", "price", "availability", "availability_status", "updated_at"])
        for p in Part.objects.filter(seller=user).select_related("brand").order_by("-updated_at"):
            writer.writerow([p.oem_number, p.title, p.brand.name if p.brand else "", p.stock_quantity, p.price, p.availability, p.availability_status, p.updated_at.strftime("%Y-%m-%d")])
    elif report_type == "analytics":
        response["Content-Disposition"] = 'attachment; filename="seller_analytics.csv"'
        writer.writerow(["order_id", "customer", "total_amount", "status", "payment_status", "sla_status", "created_at"])
        orders = Order.objects.filter(items__part__seller=user).distinct().order_by("-created_at")
        for o in orders:
            writer.writerow([o.id, o.customer_name, o.total_amount, o.status, o.payment_status, o.sla_status, o.created_at.strftime("%Y-%m-%d")])
    else:
        return HttpResponse("Unknown report type", status=404)

    return response
