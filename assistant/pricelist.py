"""ТЗ: загрузка прайс-листа через чат с AI-маппингом колонок.

Сценарий:
  1. POST /api/assistant/upload-pricelist/ (multipart, file)
     • Читаем заголовки + первые 3 строки.
     • AI предлагает маппинг колонок на стандартные поля платформы.
     • Возвращаем PricelistImport(status='preview') + карточку с формой.

  2. POST /api/assistant/upload-pricelist/<import_id>/commit/
     body: {"mapping": {std_field: source_column}}
     • Детерминированный парсер обходит весь файл по mapping.
     • Upsert по (seller, oem_number) — повторная загрузка не плодит дубли.
     • Сохраняем mapping в PricelistMapping для следующего раза.
     • Возвращаем итог: imported / failed + errors.

  3. POST /api/assistant/upload-pricelist/<import_id>/cancel/
     Отменить превью без импорта.

Стандартные поля платформы:
  oem_number  (обязательное)
  title       (обязательное)
  price       (обязательное)
  currency    (опц., default USD)
  brand       (опц.)
  stock       (опц., default 0)
  moq         (опц., default 1)
  incoterm    (опц., FOB/CIF/DDP, default FOB)
  weight_kg   (опц.)
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
from decimal import Decimal, InvalidOperation
from typing import Any

from django.conf import settings
from django.core.files.base import ContentFile
from django.db import IntegrityError, transaction
from django.utils import timezone
from django.utils.text import slugify
from django.utils.translation import gettext as _
from django.utils.translation import gettext_lazy as _lazy
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .actions import ActionResult, register

logger = logging.getLogger(__name__)


# ── Standard fields ──────────────────────────────────────────────

STD_FIELDS = [
    # (key, label, required, enum_values_or_None, default_value_or_None)
    # required=True — поле ОБЯЗАТЕЛЬНО должно быть в файле или указано явно.
    # Остальные поля заполняются дефолтами автоматически если не замаплены.
    # oem_number необязателен: если в прайсе нет настоящего артикула (только
    # название — напр. буровые коронки «PDC bit 152.4 mm»), он генерится из
    # названия (см. _gen_oem_from_title), чтобы каждая позиция была уникальной.
    ("oem_number",        _lazy("Артикул (PartNumber)"),   False, None, None),
    ("cross_number",      _lazy("Кросс-номер (CrossNumber)"), False, None, ""),
    ("brand",             _lazy("Бренд"),                   False, [
        "Caterpillar", "Komatsu", "Hitachi", "Liebherr", "TEREX",
        "New Holland", "Wirtgen", "Iveco", "HBM-Nobas", "John Deere",
        "Volvo", "JCB", "Bobcat", "BOMAG",
        "Cummins", "Deutz", "Bosch",
        "Atlas Copco", "Epiroc", "Sandvik", "Ingersoll Rand",
        "Hyundai", "Doosan", "Kobelco", "Kubota",
        "XCMG", "FAW", "LiuGong", "Shantui", "Shacman", "SDLG",
        "Weichai", "Sinotruk", "HOWO", "Zoomlion", "Sany",
        "Bosch Rexroth", "Perkins", "Dana", "Carraro", "Denso",
        "Lincoln", "Berco", "ITR", "ETP",
        "Generic",
    ], "Generic"),
    ("title",             _lazy("Название"),                True,  None, None),
    ("stock",             _lazy("Остаток (Quantity)"),      False, None, "1"),
    ("condition",         _lazy("Тип товара"),              False, ["OEM", "AFTERMARKET", "REMAN"], "OEM"),
    ("availability",      _lazy("Наличие"),                 False, ["IN_STOCK", "BACKORDER"], "IN_STOCK"),
    ("manufacturer",      _lazy("Завод-производитель"),     False, None, ""),
    ("manufacturer_visible", _lazy("Показывать завод"),     False, ["Да", "Нет"], "Да"),
    ("price_exw",         _lazy("Цена EXW"),                True,  None, None),
    ("warehouse_address", _lazy("Адрес склада EXW"),        False, None, ""),
    ("price_fob_sea",     _lazy("Цена FOB SEA"),            False, None, "0"),
    ("price_fob_air",     _lazy("Цена FOB AIR"),            False, None, "0"),
    ("sea_port",          _lazy("Морпорт отправления"),     False, None, ""),
    ("air_port",          _lazy("Аэропорт отправления"),    False, None, ""),
    ("weight_kg",         _lazy("Вес, кг"),                 False, None, "0.5"),
    ("length_cm",         _lazy("Длина, см"),               False, None, "1"),
    ("width_cm",          _lazy("Ширина, см"),              False, None, "1"),
    ("height_cm",         _lazy("Высота, см"),              False, None, "1"),
    ("currency",          _lazy("Валюта"),                  False, ["USD", "EUR", "RUB", "CNY"], "USD"),
]

REQUIRED_FIELDS = [k for k, _, req, _, _ in STD_FIELDS if req]

# Дефолты для незамапленных полей
FIELD_DEFAULTS = {k: d for k, _, _, _, d in STD_FIELDS if d is not None}


def _gen_oem_from_title(title: str) -> str:
    """Стабильный синтетический артикул из названия — для прайсов без настоящего
    OEM (товар идентифицируется названием, напр. «PDC bit 152.4 mm 4 wings»).

    Детерминирован: повторная загрузка того же названия → тот же артикул (апсерт
    не плодит дубли). Префикс GEN- помечает синтетику. Хвост-хэш разводит
    названия, схлопывающиеся в одинаковый slug.
    """
    import hashlib
    base = slugify(title or "").upper()[:48].strip("-")
    h = hashlib.md5((title or "").strip().lower().encode("utf-8")).hexdigest()[:6].upper()
    return (f"GEN-{base}-{h}" if base else f"GEN-{h}")[:100]

# Известные бренды-аналоги — реальные изготовители деталей. Поле
# `manufacturer` = бренд-аналога. Видимость для покупателя считается
# автоматически: значение из этого whitelist → ПОКАЗЫВАЕМ (узнаваемый бренд
# повышает доверие и рейтинг в выдаче); любой другой текст трактуем как
# частную торговую марку поставщика → СКРЫВАЕМ от покупателя.
KNOWN_ANALOG_BRANDS = [
    "ITR", "Berco", "Carraro", "Dana", "Denso", "Bosch", "Bosch Rexroth",
    "Rexroth", "Perkins", "ETP", "Cummins", "Deutz", "Donaldson",
    "Fleetguard", "SKF", "NOK", "NTN", "NSK", "Parker", "KYB", "Kayaba",
    "Mahle", "Mann-Filter", "Mann", "Garrett", "Holset", "WABCO", "Sachs",
    "Hengst", "Eaton", "Vickers", "Lincoln", "FleetPride", "Federal-Mogul",
]
_KNOWN_ANALOG_NORM = {b.strip().lower() for b in KNOWN_ANALOG_BRANDS}

# Чипы быстрого выбора в мастере (самые частые в спецтехнике). Любой бренд
# из KNOWN_ANALOG_BRANDS, вписанный вручную, тоже считается «известным».
MANUFACTURER_CHIP_OPTIONS = [
    "ITR", "Berco", "Carraro", "Dana", "Denso", "Bosch",
    "Perkins", "Cummins", "Donaldson", "SKF", "Parker", "ETP",
]
# Полный список для выпадающего списка в мастере — узнаваемые бренды-аналоги
# из whitelist, по алфавиту. Дубли-варианты схлопываем.
MANUFACTURER_SELECT_OPTIONS = sorted(
    {b for b in KNOWN_ANALOG_BRANDS}, key=str.lower
)


# ── EN-translation для названий запчастей ──────────────────────
# Каталог маркетплейса международный → выходной XLSX всегда на английском.
# Известные термины переводим, остальное оставляем (лучше частичный перевод
# чем сломанный текст).
_EN_TERMS: dict[str, str] = {
    # Chinese (упрощённый) — экскаватор/спецтехника
    "岩石型斗齿": "Rock-type bucket tooth",
    "斗齿": "Bucket tooth", "齿座": "Tooth holder",
    "铲斗": "Bucket", "岩石铲斗": "Rock bucket",
    "立方": "m³",
    "履带": "Track", "履带板": "Track shoe",
    "拖链轮": "Carrier roller", "拖轮": "Carrier roller",
    "链轨": "Track link", "链节": "Track link",
    "驱动轮": "Drive sprocket", "支重轮": "Track roller",
    "引导轮": "Front idler", "托链轮": "Carrier roller",
    "回转支承": "Slewing bearing", "回转马达": "Swing motor",
    "行走马达": "Travel motor",
    "动臂": "Boom", "斗杆": "Stick", "斗杆缸": "Stick cylinder",
    "动臂缸": "Boom cylinder", "铲斗缸": "Bucket cylinder",
    "液压泵": "Hydraulic pump", "主泵": "Main pump",
    # Chinese (упрощённый) — крепёж и базовые запчасти
    "螺栓": "Bolt", "螺钉": "Screw", "螺丝": "Screw", "螺母": "Nut",
    "垫圈": "Washer", "弹垫": "Spring washer", "平垫": "Flat washer",
    "销": "Pin", "卡簧": "Snap ring", "开口销": "Cotter pin",
    "轴承": "Bearing", "密封": "Seal", "油封": "Oil seal",
    "滤芯": "Filter", "滤清器": "Filter", "机油滤芯": "Oil filter",
    "空气滤芯": "Air filter", "燃油滤芯": "Fuel filter",
    "皮带": "Belt", "链条": "Chain", "齿轮": "Gear",
    "活塞": "Piston", "活塞环": "Piston ring", "气缸": "Cylinder",
    "曲轴": "Crankshaft", "凸轮轴": "Camshaft", "连杆": "Connecting rod",
    "水泵": "Water pump", "油泵": "Oil pump", "燃油泵": "Fuel pump",
    "喷油器": "Injector", "喷嘴": "Nozzle", "火花塞": "Spark plug",
    "传感器": "Sensor", "继电器": "Relay", "保险丝": "Fuse",
    "电池": "Battery", "起动机": "Starter", "发电机": "Alternator",
    "散热器": "Radiator", "水箱": "Radiator", "风扇": "Fan",
    "刹车片": "Brake pad", "刹车盘": "Brake disc", "离合器": "Clutch",
    "轮胎": "Tire", "钢圈": "Rim", "车轮": "Wheel",
    "软管": "Hose", "管": "Pipe", "接头": "Fitting", "法兰": "Flange",
    "支架": "Bracket", "盖": "Cover", "板": "Plate", "壳体": "Housing",
    "弹簧": "Spring", "阀": "Valve", "阀门": "Valve", "电磁阀": "Solenoid valve",
    "套": "Bushing", "衬套": "Bushing", "护套": "Sleeve",
    # Russian
    "болт": "Bolt", "винт": "Screw", "гайка": "Nut",
    "шайба": "Washer", "штифт": "Pin", "шплинт": "Cotter pin",
    "подшипник": "Bearing", "сальник": "Oil seal", "уплотнение": "Seal",
    "фильтр": "Filter", "масляный фильтр": "Oil filter",
    "воздушный фильтр": "Air filter", "топливный фильтр": "Fuel filter",
    "ремень": "Belt", "цепь": "Chain", "шестерня": "Gear",
    "поршень": "Piston", "кольцо": "Ring", "цилиндр": "Cylinder",
    "коленвал": "Crankshaft", "распредвал": "Camshaft", "шатун": "Connecting rod",
    "помпа": "Pump", "насос": "Pump", "форсунка": "Injector",
    "свеча": "Spark plug", "датчик": "Sensor", "реле": "Relay",
    "предохранитель": "Fuse", "аккумулятор": "Battery",
    "стартер": "Starter", "генератор": "Alternator",
    "радиатор": "Radiator", "вентилятор": "Fan",
    "тормозная колодка": "Brake pad", "колодка": "Brake pad",
    "тормозной диск": "Brake disc", "сцепление": "Clutch",
    "шланг": "Hose", "трубка": "Pipe", "трубa": "Pipe",
    "кронштейн": "Bracket", "крышка": "Cover", "пластина": "Plate",
    "корпус": "Housing", "пружина": "Spring", "клапан": "Valve",
    "втулка": "Bushing", "уплотнитель": "Seal",
}


def _looks_non_ascii(text: str) -> bool:
    """True если в строке есть символы кириллицы или CJK."""
    for ch in text:
        c = ord(ch)
        # Cyrillic + CJK Unified Ideographs ranges
        if 0x0400 <= c <= 0x04FF or 0x4E00 <= c <= 0x9FFF or 0x3400 <= c <= 0x4DBF:
            return True
    return False


def _translate_to_en(text: str) -> str:
    """Переводит CJK/Cyrillic токены на английский по словарю.
    Неизвестные слова оставляет как есть. Если в строке нет non-ASCII —
    возвращает без изменений (fast path)."""
    if not text or not _looks_non_ascii(text):
        return text
    out = text
    # сортируем по длине: длинные термины сначала, чтобы "масляный фильтр"
    # сматчился раньше чем "фильтр".
    for src in sorted(_EN_TERMS.keys(), key=len, reverse=True):
        if src in out:
            out = out.replace(src, _EN_TERMS[src])
        # Capitalized вариант для русских терминов
        if src and src[0].isalpha() and src != src.capitalize():
            cap = src.capitalize()
            if cap in out:
                out = out.replace(cap, _EN_TERMS[src])
    return out


# ── Formula whitelist engine ────────────────────────────────────

import ast
import operator as _op
from datetime import UTC

_SAFE_OPS = {
    ast.Add: _op.add,
    ast.Sub: _op.sub,
    ast.Mult: _op.mul,
    ast.Div: _op.truediv,
    ast.FloorDiv: _op.floordiv,
    ast.Mod: _op.mod,
    ast.USub: _op.neg,
}

def _safe_round(*args):
    if len(args) == 1:
        return Decimal(str(round(float(args[0]))))
    return Decimal(str(round(float(args[0]), int(args[1]))))

_SAFE_FUNCS = {
    "round": _safe_round,
    "min": lambda *a: min(*a),
    "max": lambda *a: max(*a),
    "abs": lambda x: abs(x),
    "int": lambda x: Decimal(int(x)),
    "float": lambda x: Decimal(str(float(x))),
}


def _eval_node(node, variables: dict[str, Any]) -> Any:
    """Рекурсивный evaluator AST-узлов — только арифметика + whitelist функций."""
    if isinstance(node, ast.Expression):
        return _eval_node(node.body, variables)
    if isinstance(node, ast.Constant):
        if isinstance(node.value, (int, float, Decimal)):
            return Decimal(str(node.value))
        return node.value
    if isinstance(node, ast.Name):
        if node.id in variables:
            return variables[node.id]
        raise ValueError(f"Unknown variable: {node.id}")
    if isinstance(node, ast.BinOp):
        left = _eval_node(node.left, variables)
        right = _eval_node(node.right, variables)
        op_fn = _SAFE_OPS.get(type(node.op))
        if not op_fn:
            raise ValueError(f"Unsupported operator: {type(node.op).__name__}")
        return op_fn(Decimal(str(left)), Decimal(str(right)))
    if isinstance(node, ast.UnaryOp):
        operand = _eval_node(node.operand, variables)
        op_fn = _SAFE_OPS.get(type(node.op))
        if not op_fn:
            raise ValueError(f"Unsupported unary: {type(node.op).__name__}")
        return op_fn(Decimal(str(operand)))
    if isinstance(node, ast.Call):
        if not isinstance(node.func, ast.Name):
            raise ValueError("Only simple function calls allowed")
        fn = _SAFE_FUNCS.get(node.func.id)
        if not fn:
            raise ValueError(f"Function not allowed: {node.func.id}")
        args = [_eval_node(a, variables) for a in node.args]
        return fn(*args)
    if isinstance(node, ast.IfExp):
        test = _eval_node(node.test, variables)
        return _eval_node(node.body if test else node.orelse, variables)
    if isinstance(node, ast.Compare):
        left = _eval_node(node.left, variables)
        for cmp_op, comparator in zip(node.ops, node.comparators):
            right = _eval_node(comparator, variables)
            if isinstance(cmp_op, ast.Gt):
                if not (left > right):
                    return False
            elif isinstance(cmp_op, ast.Lt):
                if not (left < right):
                    return False
            elif isinstance(cmp_op, ast.GtE):
                if not (left >= right):
                    return False
            elif isinstance(cmp_op, ast.LtE):
                if not (left <= right):
                    return False
            elif isinstance(cmp_op, ast.Eq):
                if not (left == right):
                    return False
            else:
                raise ValueError(f"Unsupported comparison: {type(cmp_op).__name__}")
            left = right
        return True
    raise ValueError(f"Unsupported AST node: {type(node).__name__}")


def eval_formula(expression: str, variables: dict[str, Any]) -> Any:
    """Безопасное вычисление формулы. Только арифметика + whitelist функций.

    Примеры:
      eval_formula("price * 1.15", {"price": Decimal("100")})  → 115.00
      eval_formula("round(weight * 2.2, 1)", {"weight": Decimal("5")})  → 11.0
      eval_formula("price if price > 0 else 1", {"price": Decimal("0")})  → 1
    """
    try:
        tree = ast.parse(expression.strip(), mode="eval")
    except SyntaxError as e:
        raise ValueError(f"Invalid formula syntax: {e}")
    return _eval_node(tree, variables)


def validate_formula(expression: str, available_vars: list[str]) -> tuple[bool, str]:
    """Проверяет формулу на синтаксис и использование доступных переменных."""
    try:
        tree = ast.parse(expression.strip(), mode="eval")
    except SyntaxError as e:
        return False, f"Syntax error: {e}"
    for node in ast.walk(tree):
        if isinstance(node, ast.Name) and node.id not in available_vars and node.id not in _SAFE_FUNCS:
            return False, f"Unknown variable: {node.id}"
        if isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name) and node.func.id not in _SAFE_FUNCS:
                return False, f"Function not allowed: {node.func.id}"
    return True, ""


# ── File reading ─────────────────────────────────────────────────

MAX_FILE_BYTES = 50 * 1024 * 1024  # 50 МБ (ТЗ)
MIN_COLUMNS = 2          # ТЗ: минимум 2 колонки
MAX_COLUMNS = 100        # ТЗ: максимум 100 колонок
MAX_AI_HEADERS = 20      # ТЗ: AI получает максимум 20 заголовков за раз
MAX_AI_CALLS_PER_DAY = 50  # увеличил с 3 — словарь + value-detection
                            # покрывают большинство кейсов, AI как fallback


def _read_xlsx_rows(blob: bytes, max_rows: int | None = None):
    """Iter всех строк xlsx как кортежи. Максимум первый лист.

    Hybrid стратегия:
    • max_rows ≤ 100 (preview) → openpyxl read_only (быстрый старт)
    • max_rows > 100 или None (full read) → python-calamine Rust-based
      (на 616k+ ~6× быстрее openpyxl). Fallback на openpyxl.
    """
    use_calamine = max_rows is None or max_rows > 100
    if use_calamine:
        try:
            from python_calamine import CalamineWorkbook
            wb = CalamineWorkbook.from_filelike(io.BytesIO(blob))
            sheet = wb.get_sheet_by_index(0)
            rows = sheet.to_python(nrows=max_rows) if max_rows else sheet.to_python()
            for row in rows:
                yield tuple("" if v is None else str(v).strip() for v in row)
            return
        except ImportError:
            pass
        except Exception as e:
            logger.warning("calamine failed, falling back to openpyxl: %s", e)

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    n = 0
    for row in sheet.iter_rows(values_only=True):
        yield tuple("" if v is None else str(v).strip() for v in row)
        n += 1
        if max_rows is not None and n >= max_rows:
            break


def _read_csv_rows(blob: bytes, max_rows: int | None = None):
    """Iter csv/tsv с авто-определением разделителя.

    Терпит mixed line endings (\\r, \\r\\n, \\n) — csv.reader падает с
    «new-line character seen in unquoted field» когда в данных смешаны
    разные newlines. Нормализуем все newlines к \\n перед парсингом.
    Также отрезаем UTF-8 BOM если есть.
    """
    # UTF-8 BOM можно встретить в файлах из Excel
    if blob[:3] == b"\xef\xbb\xbf":
        blob = blob[3:]
    text = blob.decode("utf-8", errors="replace")
    # Нормализация newlines: \\r\\n или \\r → \\n
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        class _D(csv.excel):
            delimiter = ";"
        dialect = _D()
    reader = csv.reader(io.StringIO(text, newline=""), dialect=dialect)
    n = 0
    for row in reader:
        yield tuple((v or "").strip() for v in row)
        n += 1
        if max_rows is not None and n >= max_rows:
            break


def _detect_format(filename: str, blob: bytes) -> str:
    """Определяет формат файла. Magic-байты приоритетнее расширения —
    бывает что seller сохраняет xlsx с расширением .csv (или наоборот).
    """
    # Magic: zip-архив (xlsx) — приоритет над расширением
    if blob[:4] == b"PK\x03\x04":
        return "xlsx"
    # Magic: BIFF-Excel (старый .xls) — у нас не поддерживается, но
    # пометим как xlsx чтобы openpyxl выдал понятную ошибку.
    if blob[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xlsx"
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xls"):
        return "xlsx"
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return "csv"
    return "csv"


def _find_header_idx(rows: list) -> int:
    """Находит строку с заголовками таблицы, пропуская «шапку» прайса
    (название компании, дата, телефон — обычно в одной merged-ячейке).
    Заголовки = первая строка с ≥2 непустыми ячейками, где их число близко
    к максимуму среди первых 15 строк. Возвращает индекс (0 если не нашли).
    """
    scan = rows[:15]
    if not scan:
        return 0

    def nonempty(r):
        return sum(1 for c in r if str(c).strip())

    counts = [nonempty(r) for r in scan]
    mx = max(counts) if counts else 0
    if mx < 2:
        return 0
    # первая строка где непустых >= 2 и >= 60% от макс (это уже таблица,
    # а не однострочная шапка).
    for i, c in enumerate(counts):
        if c >= 2 and c >= mx * 0.6:
            return i
    return 0


def _read_preview(filename: str, blob: bytes) -> tuple[list[str], list[list[str]]]:
    """Возвращает (headers, sample_rows[3]). Автоматически пропускает
    шапку прайса и находит реальную строку заголовков."""
    fmt = _detect_format(filename, blob)
    # читаем чуть больше строк, чтобы найти заголовки под шапкой
    rows = list(
        _read_xlsx_rows(blob, max_rows=19) if fmt == "xlsx"
        else _read_csv_rows(blob, max_rows=19)
    )
    if not rows:
        raise ValueError("File is empty")
    hidx = _find_header_idx(rows)
    headers = list(rows[hidx])
    while headers and not str(headers[-1]).strip():
        headers.pop()
    n_cols = len(headers)
    sample = [list(r)[:n_cols] for r in rows[hidx + 1:hidx + 4]]
    return headers, sample


def _read_all(filename: str, blob: bytes):
    """Iter всех строк данных — пропускает шапку + строку заголовков."""
    fmt = _detect_format(filename, blob)
    rows_iter = _read_xlsx_rows(blob) if fmt == "xlsx" else _read_csv_rows(blob)
    # определяем header-индекс по первым 15 строкам (буферизуем их)
    buf = []
    it = iter(rows_iter)
    for _ in range(15):
        try:
            buf.append(next(it))
        except StopIteration:
            break
    hidx = _find_header_idx(buf)
    # отдаём строки ПОСЛЕ заголовков: остаток буфера + хвост итератора
    for row in buf[hidx + 1:]:
        yield row
    for row in it:
        yield row


# ── AI mapping ───────────────────────────────────────────────────

def _ai_calls_used_today(seller) -> int:
    """Сколько раз seller вызывал AI за последние 24 часа.

    Считаем по PricelistImport.ai_called=True.
    """
    from datetime import timedelta

    from marketplace.models import PricelistImport
    since = timezone.now() - timedelta(hours=24)
    return PricelistImport.objects.filter(
        seller=seller, ai_called=True, created_at__gte=since,
    ).count()


def _detect_by_values(headers: list[str], sample_rows: list[list[str]],
                       already_mapped: set[str]) -> dict[str, str]:
    """Детекция канонических полей по СОДЕРЖИМОМУ колонок (sample rows).

    Работает когда заголовки непонятные (PRO_ID, ABCD123, etc):
    смотрит на значения, угадывает что это:
      - price_exw — числа с decimal > 1
      - stock — небольшие целые
      - oem_number — alphanumeric codes (дефисы, точки)
      - title — длинные строки с пробелами
      - brand — известные бренды
      - condition — OEM/AFTERMARKET/REMAN/NEW
      - currency — USD/EUR/RUB/CNY
      - weight_kg — small decimals < 100

    already_mapped: каноники которые уже сматчились через словарь —
                    пропускаем (не дублируем).

    Returns: {header: canonical_key}
    """
    if not sample_rows:
        return {}

    import re as _re
    KNOWN_BRANDS = {
        "epiroc", "caterpillar", "cat", "komatsu", "hitachi", "liebherr",
        "sandvik", "atlas copco", "volvo", "kobelco", "doosan", "hyundai",
        "cummins", "deutz", "perkins", "bosch", "bobcat", "jcb",
    }
    CONDITION_VALUES = {"oem", "original", "aftermarket", "reman", "new",
                         "used", "remanufactured"}
    CURRENCY_VALUES = {"usd", "eur", "rub", "cny", "rmb", "jpy", "kzt"}

    n_cols = len(headers)
    n_rows = len(sample_rows)
    if n_rows == 0:
        return {}

    # Собираем колонки (по-столбцово)
    cols = [[] for _ in range(n_cols)]
    for row in sample_rows:
        for i in range(n_cols):
            if i < len(row):
                v = row[i]
                cols[i].append("" if v is None else str(v).strip())
            else:
                cols[i].append("")

    detected: dict[str, str] = {}
    used_canonical = set(already_mapped)

    for i, vals in enumerate(cols):
        if not vals:
            continue
        non_empty = [v for v in vals if v]
        if not non_empty:
            continue
        header = headers[i] if i < len(headers) else f"col{i}"

        # Стрипуем единицы измерения для числовых проверок
        decimals = []
        for v in non_empty:
            m = _re.search(r"-?\d+[.,]?\d*", v.replace(" ", ""))
            if m:
                try:
                    decimals.append(float(m.group(0).replace(",", ".")))
                except ValueError:
                    pass

        # Считаем характеристики
        all_decimal_ratio = len(decimals) / len(non_empty)
        avg_len = sum(len(v) for v in non_empty) / len(non_empty)
        has_spaces = sum(1 for v in non_empty if " " in v) / len(non_empty)
        has_dashes = sum(1 for v in non_empty if "-" in v) / len(non_empty)

        # ── currency
        if "currency" not in used_canonical:
            if all(v.lower() in CURRENCY_VALUES for v in non_empty):
                detected[header] = "currency"; used_canonical.add("currency"); continue

        # ── condition
        if "condition" not in used_canonical:
            if all(v.lower() in CONDITION_VALUES for v in non_empty):
                detected[header] = "condition"; used_canonical.add("condition"); continue

        # ── brand
        if "brand" not in used_canonical:
            if all(v.lower() in KNOWN_BRANDS for v in non_empty):
                detected[header] = "brand"; used_canonical.add("brand"); continue

        # ── price_exw — все числа, среднее > 1
        if "price" not in used_canonical and all_decimal_ratio >= 0.9:
            avg_dec = sum(decimals) / max(len(decimals), 1)
            if avg_dec > 1 and max(decimals) > 1:
                detected[header] = "price"
                used_canonical.add("price"); continue

        # ── weight (decimals < 100 in average)
        if "weight" not in used_canonical and all_decimal_ratio >= 0.9:
            avg_dec = sum(decimals) / max(len(decimals), 1)
            if avg_dec < 100:
                detected[header] = "weight"
                used_canonical.add("weight"); continue

        # ── stock — целые числа
        if "stock" not in used_canonical and all_decimal_ratio >= 0.9:
            if all(d == int(d) and 0 <= d < 10000 for d in decimals):
                detected[header] = "stock"; used_canonical.add("stock"); continue

        # ── part_number — alphanumeric codes (дефисы/точки, короткие)
        if "part_number" not in used_canonical:
            if has_dashes >= 0.5 and avg_len <= 25 and has_spaces < 0.3:
                detected[header] = "part_number"
                used_canonical.add("part_number"); continue
            # Числовые коды (Komatsu OEM)
            if all_decimal_ratio >= 0.8 and avg_len >= 6 and has_spaces < 0.2:
                detected[header] = "part_number"
                used_canonical.add("part_number"); continue

        # ── description/title — длинная строка с пробелами
        if "description" not in used_canonical:
            if has_spaces >= 0.4 or avg_len > 15:
                detected[header] = "description"
                used_canonical.add("description"); continue

    return detected


def _smart_mapping(headers: list[str], sample_rows: list[list[str]],
                    seller=None
                    ) -> tuple[dict[str, str], list[str], bool, str]:
    """Умная автоматическая загрузка прайса.

      1. Словарь COLUMN_MAP + LearnedColumnSynonym (БД)
      2. Value-based детекция по содержимому колонок (для unknown headers)
      3. Для оставшихся — AI-запрос (max 20 заголовков)
      4. AI-ответы → learn_synonym() в БД
      5. Возвращает (mapping_std, unknown, ai_called, status)
    """
    from .price_mappings import (
        CANONICAL_TO_STD,
        COLUMN_MAP,
        learn_synonym,
        load_learned_lookup,
        match_headers,
    )

    learned = load_learned_lookup()
    canonical_map, unknown_headers = match_headers(headers, learned=learned)

    # Value-based fallback для нераспознанных через словарь
    if unknown_headers and sample_rows:
        already = set(canonical_map.keys())
        by_values = _detect_by_values(headers, sample_rows, already)
        for header, canonical in by_values.items():
            if canonical not in canonical_map and header in unknown_headers:
                canonical_map[canonical] = header
                unknown_headers.remove(header)
                # Learn для будущих импортов
                try:
                    learn_synonym(canonical, header, source="value-detect")
                except Exception:
                    pass

    ai_called = False
    status = "ok"
    if unknown_headers:
        # ТЗ: AI получает максимум 20 заголовков за раз
        if len(unknown_headers) > MAX_AI_HEADERS:
            unknown_headers = unknown_headers[:MAX_AI_HEADERS]
        # ТЗ: лимит 3 AI-вызова в день на seller
        if seller is not None:
            used = _ai_calls_used_today(seller)
            if used >= MAX_AI_CALLS_PER_DAY:
                status = "quota_exceeded"
                # AI не вызываем, unknowns остаются как есть
                _build_std_mapping = lambda cmap: {  # noqa: E731
                    CANONICAL_TO_STD[c]: h for c, h in cmap.items()
                    if c in CANONICAL_TO_STD
                }
                return _build_std_mapping(canonical_map), unknown_headers, False, status

        ai_canonical = _ai_resolve_unknowns(unknown_headers, sample_rows,
                                              list(COLUMN_MAP.keys()))
        ai_called = bool(ai_canonical)
        for header, canonical in (ai_canonical or {}).items():
            if canonical not in canonical_map and canonical in COLUMN_MAP:
                canonical_map[canonical] = header
                learn_synonym(canonical, header, source="ai")
                if header in unknown_headers:
                    unknown_headers.remove(header)

    # canonical_map → std_field → header (для STD_FIELDS)
    mapping_std: dict[str, str] = {}
    for canonical, header in canonical_map.items():
        std = CANONICAL_TO_STD.get(canonical)
        if std and std not in mapping_std:
            mapping_std[std] = header

    return mapping_std, unknown_headers, ai_called, status


def _ai_resolve_unknowns(unknown_headers: list[str], sample_rows: list[list[str]],
                          allowed_canonicals: list[str]) -> dict[str, str] | None:
    """Claude tool use для структурированного маппинга колонок.

    Вместо парсинга свободного текста JSON — Claude возвращает
    структурированный ответ через tool_use (propose_mapping).
    Fallback на текстовый JSON если tool use не сработал.
    """
    api_key = getattr(settings, "ANTHROPIC_API_KEY", "")
    if not api_key or not unknown_headers:
        return {}
    try:
        from anthropic import Anthropic
        client = Anthropic(api_key=api_key)
    except Exception:
        return {}

    sample_text = "\n".join(
        " | ".join(str(c)[:30] for c in row) for row in (sample_rows or [])[:3]
    )

    propose_mapping_tool = {
        "name": "propose_mapping",
        "description": (
            "Предложить маппинг колонок прайс-листа поставщика "
            "на канонические поля платформы."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "mappings": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "source_header": {
                                "type": "string",
                                "description": _("Оригинальный заголовок колонки из файла"),
                            },
                            "canonical_key": {
                                "type": "string",
                                "enum": allowed_canonicals,
                                "description": _("Каноническое поле платформы"),
                            },
                            "confidence": {
                                "type": "number",
                                "minimum": 0,
                                "maximum": 1,
                                "description": _("Уверенность в маппинге (0-1)"),
                            },
                        },
                        "required": ["source_header", "canonical_key", "confidence"],
                    },
                },
                "unmapped": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": _("Заголовки, которые не подходят ни под один ключ"),
                },
                "questions": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "field": {"type": "string"},
                            "question": {"type": "string"},
                            "options": {
                                "type": "array",
                                "items": {"type": "string"},
                            },
                            "default": {"type": "string"},
                        },
                        "required": ["field", "question"],
                    },
                    "description": _("Вопросы оператору для уточнения"),
                },
            },
            "required": ["mappings"],
        },
    }

    prompt = (
        "Определи назначение колонок прайс-листа поставщика запчастей "
        "для спецтехники (Caterpillar, Komatsu, etc.). "
        "Используй tool propose_mapping для ответа.\n\n"
        "Правила:\n"
        "- Маппь только те колонки, в которых уверен (confidence ≥ 0.7)\n"
        "- Колонки, которые не подходят ни под один ключ — в unmapped\n"
        "- Если колонка неоднозначна — задай вопрос в questions\n\n"
        f"Нераспознанные заголовки: {', '.join(repr(h) for h in unknown_headers)}\n\n"
        f"Примеры строк:\n{sample_text}"
    )

    try:
        msg = client.messages.create(
            model=getattr(settings, "ANTHROPIC_FAST_MODEL", "claude-haiku-4-5-20251001"),
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
            tools=[propose_mapping_tool],
            tool_choice={"type": "tool", "name": "propose_mapping"},
        )
        for block in msg.content:
            if block.type == "tool_use" and block.name == "propose_mapping":
                result = {}
                for m in block.input.get("mappings", []):
                    h = m.get("source_header", "")
                    c = m.get("canonical_key", "")
                    conf = m.get("confidence", 0)
                    if (h in unknown_headers and c in allowed_canonicals
                            and conf >= 0.7):
                        result[h] = c
                return result
        return {}
    except Exception:
        logger.exception("AI tool-use mapping failed")
        return {}


def _build_mapped_preview(headers: list[str], sample_rows: list[list[str]],
                            mapping: dict[str, str]) -> dict:
    """Строит «как ляжет в базу» превью первых N строк.

    Возвращает {std_columns: […], rows: [[v, …], …]} — для UI рендеринга
    как table_preview. Применяет mapping, но без коэрсии (Decimal/int) —
    показываем как есть из ячеек.
    """
    # std_field → header → idx
    col_idx: dict[str, int] = {}
    for std, src in (mapping or {}).items():
        if not src or src.startswith("fix:") or src not in headers:
            continue
        col_idx[std] = headers.index(src)
    # Берём std-поля в фикс-порядке STD_FIELDS чтобы UI колонки
    # были в осмысленном порядке.
    std_keys = [k for k, _, _, _, _ in STD_FIELDS if k in col_idx]
    rows = []
    for row in (sample_rows or [])[:5]:
        rows.append([
            (str(row[col_idx[k]]) if col_idx[k] < len(row) else "")
            for k in std_keys
        ])
    labels = {k: lbl for k, lbl, _, _, _ in STD_FIELDS}
    return {
        "headers": [labels.get(k, k) for k in std_keys],
        "std_keys": std_keys,
        "rows": rows,
    }


# legacy-обёртка, оставлена ради совместимости с тестами
def _heuristic_mapping(headers: list[str]) -> dict[str, str]:
    """Минимальный legacy-маппер. В новом коде используется _smart_mapping."""
    mapping_std, _, _, _ = _smart_mapping(headers, sample_rows=[])
    # _smart_mapping возвращает только что нашёл по словарю — это может
    # включать любые поля, не только title+weight. По legacy-контракту
    # тут возвращали именно minimal-set. Оставлю как есть — словарь
    # достаточно консервативен.
    return mapping_std


def _ai_mapping(headers: list[str], sample_rows: list[list[str]]) -> dict[str, str]:
    """Спрашивает Claude про маппинг колонок. Если API недоступен — heuristic."""
    api_key = getattr(settings, "ANTHROPIC_API_KEY", "")
    if not api_key:
        return _heuristic_mapping(headers)

    try:
        import anthropic  # noqa: F401
    except ImportError:
        return _heuristic_mapping(headers)

    fields_doc = "\n".join(
        f"  - {k} ({label}){' [REQUIRED]' if req else ''}"
        for k, label, req, enum_v, _ in STD_FIELDS
    )
    sample_text = "\n".join(
        " | ".join(str(c)[:40] for c in row) for row in sample_rows
    )
    prompt = (
        f"Определи маппинг колонок прайс-листа поставщика на стандартные поля "
        f"платформы B2B-маркетплейса запчастей. Верни ТОЛЬКО JSON, без "
        f"объяснений, в формате {{\"std_field\": \"source_column_header\"}}.\n\n"
        f"Стандартные поля:\n{fields_doc}\n\n"
        f"Заголовки прайса:\n{', '.join(repr(h) for h in headers)}\n\n"
        f"Примеры строк:\n{sample_text}\n\n"
        f"JSON:"
    )
    try:
        from anthropic import Anthropic
        client = Anthropic(api_key=api_key)
        msg = client.messages.create(
            model=getattr(settings, "ANTHROPIC_FAST_MODEL", "claude-haiku-4-5-20251001"),
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        text = msg.content[0].text.strip()
        # Достаём JSON из ответа
        if "```" in text:
            text = text.split("```", 2)[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.split("```", 1)[0]
        text = text.strip()
        result = json.loads(text)
        # Проверяем что все значения — реальные заголовки
        return {k: v for k, v in result.items()
                 if k in {f for f, _, _, _ in STD_FIELDS} and v in headers}
    except Exception:
        logger.exception("AI mapping failed, falling back to heuristic")
        return _heuristic_mapping(headers)


# ── Deterministic parse + import ────────────────────────────────

def _coerce_decimal(raw: Any) -> Decimal | None:
    """Извлекает первое число из текста.

    Терпим к: «94,0841350448234 €», «$1,234.56», «1 234,56 RUB»,
    нескольким числам в одной ячейке («94,08 €  9,65 €» → берёт 94.08).
    """
    if raw is None or raw == "":
        return None
    import re as _re
    s = str(raw).strip()
    # Первый числовой кусок (с разделителями тысяч и десятичными)
    m = _re.search(r"-?[\d.,'\s ]+", s)
    if not m:
        return None
    raw_num = m.group(0).strip()
    # Удаляем апострофы (швейц. 1'234) и любые пробелы (включая NBSP)
    raw_num = _re.sub(r"[\s' ]", "", raw_num)
    if not raw_num or raw_num in ("-",):
        return None
    # Эвристика разделителей "," vs "."
    if "," in raw_num and "." in raw_num:
        # Оба знака: последний из них = десятичный, другой = тысячи
        if raw_num.rfind(",") > raw_num.rfind("."):
            cleaned = raw_num.replace(".", "").replace(",", ".")
        else:
            cleaned = raw_num.replace(",", "")
    elif "," in raw_num:
        # Одна запятая, после неё ровно 3 цифры, до неё ≤3 цифр → тысячи
        last = raw_num.rfind(",")
        tail = raw_num[last + 1:]
        head = raw_num[:last].replace(",", "")
        if (raw_num.count(",") == 1 and len(tail) == 3 and tail.isdigit()
                and head.isdigit() and len(head) <= 3):
            cleaned = raw_num.replace(",", "")
        else:
            cleaned = raw_num.replace(",", ".")
    else:
        # Только точка(и) — последняя десятичная, остальные тысячи
        if raw_num.count(".") > 1:
            parts = raw_num.split(".")
            cleaned = "".join(parts[:-1]) + "." + parts[-1]
        else:
            cleaned = raw_num
    if not cleaned or cleaned in ("-", ".", "-."):
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _coerce_int(raw: Any) -> int | None:
    if raw is None or raw == "":
        return None
    s = str(raw).strip()
    import re as _re
    s = _re.sub(r"[^\d]", "", s)
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _normalize_currency(raw: str) -> str:
    s = (raw or "").strip().upper().replace("$", "USD").replace("€", "EUR").replace("₽", "RUB")
    if s in {"USD", "EUR", "RUB", "CNY"}:
        return s
    return "USD"


def _normalize_incoterm(raw: str) -> str:
    s = (raw or "").strip().upper()
    if s in {"FOB", "CIF", "DDP"}:
        return s
    return "FOB"


def _apply_transform(field: str, raw_value: str, transform_rules: dict,
                      row_vars: dict[str, Any]) -> str:
    """Применяет правило трансформации к значению поля если есть."""
    rule = transform_rules.get(field)
    if not rule:
        return raw_value
    rule_type = rule.get("type", "")
    if rule_type == "formula":
        formula = rule.get("formula", "")
        if not formula:
            return raw_value
        try:
            result = eval_formula(formula, row_vars)
            return str(result)
        except Exception:
            return raw_value
    if rule_type == "map":
        value_map = rule.get("map", {})
        return value_map.get(raw_value.strip(), raw_value)
    if rule_type == "concat":
        parts = rule.get("parts", [])
        return " ".join(str(row_vars.get(p, p)) for p in parts).strip()
    return raw_value


def _lookup_reference_data(oems: list[str], brand: str | None = None) -> dict[str, dict]:
    """Поиск точных данных по OEM-номеру в нашей эталонной базе
    (таможня, дилеры, OEM-каталоги).

    Используется как enrichment layer ПЕРЕД AI оценкой:
    точные данные из реестров > AI guess > defaults.

    Source priority: customs > dealer > oem > manual > ai.

    Returns: {oem_number: {weight_kg, length_cm, ..., source, confidence}}
    """
    from marketplace.models import PartReference
    if not oems:
        return {}
    # SQLite ограничивает WHERE IN до 999 параметров — режем на чанки.
    refs = []
    CHUNK = 900
    for i in range(0, len(oems), CHUNK):
        batch = oems[i:i + CHUNK]
        refs.extend(PartReference.objects.filter(oem_number__in=batch))
    qs = sorted(refs, key=lambda r: (r.oem_number, -float(r.confidence or 0)))
    if brand:
        # Сначала с точным брендом, потом без — fallback
        pass
    SOURCE_PRIORITY = {"customs": 5, "dealer": 4, "oem": 3, "manual": 2, "ai": 1}
    best: dict[str, dict] = {}
    for ref in qs:
        cur = best.get(ref.oem_number)
        score = (
            SOURCE_PRIORITY.get(ref.source, 0),
            ref.confidence,
        )
        if cur is None or score > cur["_score"]:
            best[ref.oem_number] = {
                "weight_kg": float(ref.weight_kg) if ref.weight_kg else None,
                "length_cm": float(ref.length_cm) if ref.length_cm else None,
                "width_cm":  float(ref.width_cm)  if ref.width_cm  else None,
                "height_cm": float(ref.height_cm) if ref.height_cm else None,
                "hs_code":   ref.hs_code or None,
                "cross_numbers": ref.cross_numbers or None,
                "country_of_origin": ref.country_of_origin or None,
                "source":    ref.source,
                "confidence": ref.confidence,
                "_score":    score,
            }
    for v in best.values():
        v.pop("_score", None)
    return best


def _supplier_wide_value(key: str, mapping: dict, constants: dict | None) -> str:
    """Возвращает supplier-wide константное значение для поля:
    приоритет constants → mapping['fix:...'], '' если нигде нет."""
    c = (constants or {}).get(key)
    if c not in (None, "", 0):
        return str(c).strip()
    m = (mapping or {}).get(key, "")
    if isinstance(m, str) and m.startswith("fix:"):
        return m[4:].strip()
    return ""


def _resolve_warehouse_for_import(seller, mapping: dict, constants: dict | None,
                                    filename: str = ""):
    """Создаёт НОВЫЙ виртуальный склад на каждую загрузку прайса.

    Правило: «один импорт — один склад». Даже если ports+address совпадают
    с существующим складом — это отдельная загрузка, отдельная папка.
    Продавец сам может потом переименовать или объединить.

    Возвращает SellerWarehouse (никогда None — всегда создаёт папку).
    """
    from marketplace.models import SellerWarehouse
    sea = _supplier_wide_value("sea_port", mapping, constants)
    air = _supplier_wide_value("air_port", mapping, constants)
    addr = _supplier_wide_value("warehouse_address", mapping, constants)
    currency = _supplier_wide_value("currency", mapping, constants) or "USD"

    # Страна выводится из ISO-кода в начале портового кода (TRMER → TR)
    cc = ""
    for src in (sea, air):
        if src and len(src) >= 2:
            head = src.split()[0]
            if len(head) >= 2 and head[:2].isalpha():
                cc = head[:2].upper()
                break

    # Автоимя: «Склад #N · Страна · Город» либо «Склад #N · <filename>»
    idx = SellerWarehouse.objects.filter(seller=seller).count() + 1
    country_label = {
        "TR": "Турция", "CN": "Китай", "RU": "Россия",
        "AE": "ОАЭ", "NL": "Нидерланды", "KZ": "Казахстан",
    }.get(cc, cc or "")
    # Город — третий токен в портовой строке (Анкара, Мерсин, ...)
    city = ""
    for src in (sea, air):
        parts = [p.strip() for p in src.split("·")]
        if len(parts) >= 3:
            city = parts[2]
            break
    name_parts = [f"Склад #{idx}"]
    if country_label:
        name_parts.append(country_label)
    if city:
        name_parts.append(city)
    if len(name_parts) == 1 and filename:
        # Если логистики нет — используем имя файла как fallback
        short = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0][:50]
        name_parts.append(short)
    name = " · ".join(name_parts)

    return SellerWarehouse.objects.create(
        seller=seller, name=name[:120],
        country_code=cc[:2], sea_port=sea[:120], air_port=air[:120],
        address=addr[:1000], currency=currency[:3] or "USD",
    )


def _import_file(import_obj, mapping: dict[str, str], blob: bytes,
                 transform_rules: dict | None = None,
                 constants: dict | None = None,
                 warehouse=None):
    """Полный детерминированный импорт по подтверждённому маппингу.

    transform_rules: {std_field: {type: formula|map|concat, ...}}
    constants: {std_field: fixed_value} — дополнительные константы

    Enrichment chain (для отсутствующих в файле полей):
      1. PartReference (customs/dealer/OEM) — точные данные
      2. ai_estimates (если seller вызвал AI) — оценки
      3. fix-defaults / NULL
    """
    from marketplace.models import Brand, Category, Part

    seller = import_obj.seller
    headers = import_obj.headers
    transform_rules = transform_rules or {}
    constants = constants or {}
    ai_estimates = getattr(import_obj, "ai_estimates", None) or {}

    col_idx: dict[str, int] = {}
    fixed_vals: dict[str, str] = {}
    # Сначала дефолты, потом mapping (mapping перезаписывает)
    for fld, dflt in FIELD_DEFAULTS.items():
        fixed_vals[fld] = str(dflt)
    for fld, val in mapping.items():
        if not val:
            continue
        if isinstance(val, str) and val.startswith("fix:"):
            fixed_vals[fld] = val[4:]
        elif val in headers:
            col_idx[fld] = headers.index(val)
            fixed_vals.pop(fld, None)
    # Constants имеют приоритет над любыми дефолтами (свежий ответ юзера
    # должен перезаписывать значения из сохранённого профиля).
    # Для supplier-wide логистики (порты, склад) constants ВСЕГДА выигрывают
    # даже над замапленной колонкой — потому что эти поля обязаны быть одинаковы
    # для всей загрузки, и колонка в маркетплейс-XLSX часто бывает пустой.
    SUPPLIER_WIDE_OVERRIDE = {"sea_port", "air_port", "warehouse_address"}
    for fld, val in constants.items():
        if val is None or val == "":
            continue
        if fld in col_idx and fld not in SUPPLIER_WIDE_OVERRIDE:
            continue  # колонка из файла — приоритет
        # Если supplier-wide constant есть — удаляем col_idx чтобы fixed_vals выиграл
        if fld in col_idx and fld in SUPPLIER_WIDE_OVERRIDE:
            col_idx.pop(fld, None)
        fixed_vals[fld] = str(val)

    missing_required = [
        f for f in REQUIRED_FIELDS
        if f not in col_idx and f not in fixed_vals
    ]
    if missing_required:
        return 0, 0, 0, 0, [{"row": 0, "reason": f"missing required mapping: {missing_required}"}]

    cat = Category.objects.filter(slug="parts").first() or Category.objects.first()
    if not cat:
        cat = Category.objects.create(name="Запчасти", slug="parts")
    generic_brand = Brand.objects.filter(name__iexact="Generic").first() or \
                     Brand.objects.create(name="Generic", slug="generic")

    imported = 0
    created = 0
    updated = 0
    failed = 0
    errors: list[dict] = []

    # Кэш брендов: на каждый прайс одни и те же названия → не лезть
    # в БД n раз. seed'им существующими + Generic.
    brand_cache: dict[str, Brand] = {
        b.name.lower(): b for b in Brand.objects.all()
    }
    brand_cache.setdefault("generic", generic_brand)
    # Снапшот «известных» брендов для расчёта видимости завода: whitelist
    # аналогов + ВСЕ бренды каталога платформы (OEM и aftermarket). Берём
    # ДО цикла, чтобы новые бренды, созданные по ходу из колонки Brand, не
    # засчитывались как «известные» автоматически. Generic — не показываем.
    _known_for_visibility = set(_KNOWN_ANALOG_NORM)
    _known_for_visibility.update(
        k for k in brand_cache.keys() if k != "generic"
    )

    # Fallback-бренд: если Brand-колонка пустая, пробуем определить по имени
    # файла (например komatsu*.xlsx → Komatsu, sandvik*.xlsx → Sandvik).
    filename_brand = None
    known_brands_l = ["epiroc", "caterpillar", "komatsu", "hitachi", "sandvik",
                       "liebherr", "atlas copco"]
    fname_lc = (getattr(import_obj, "filename", "") or "").lower()
    fname_clean = fname_lc.replace("_", "").replace(" ", "").replace("-", "")
    for b in known_brands_l:
        if b.replace(" ", "") in fname_clean:
            key = b
            filename_brand = brand_cache.get(key)
            if not filename_brand:
                filename_brand = Brand.objects.filter(name__iexact=b).first()
                if filename_brand:
                    brand_cache[key] = filename_brand
            break

    # Прогресс импорта для polling из UI. Пишем в общий Redis-кэш "pricelist":
    # коммит исполняется в Celery-воркере (отдельный процесс), а поллинг идёт в
    # daphne — LocMem (per-process) их бы не связал. В тестах алиас = LocMem.
    from django.core.cache import caches
    cache = caches["pricelist"]
    progress_key = f"import_progress_{import_obj.id}"
    cache.set(progress_key, {"current": 0, "total": 0, "running": True}, 600)

    # ───────── Pass 1: парсим все строки в payload'ы ─────────
    payloads: list[dict] = []
    for row_n, row in enumerate(_read_all(import_obj.filename, blob), start=2):
            # Обновление прогресса каждые 500 строк во время парсинга:
            # без этого UI висел бы на "0 строк" пока разбираем 160k XLSX.
            if row_n % 500 == 0:
                cache.set(progress_key, {
                    "current": len(payloads),
                    "total": row_n,
                    "phase": "parsing",
                    "running": True,
                }, 600)
            if not any(c.strip() for c in row):
                continue
            # Строка-разделитель секции бренда в мультибрендовом прайсе
            # («▸ KOMATSU (оригинал)», «=== CATERPILLAR ===»): заполнена
            # только одна ячейка, часто с маркером. Пропускаем молча —
            # это не товар и не ошибка.
            _nonempty = [c for c in row if str(c).strip()]
            if len(_nonempty) <= 1:
                _first = str(_nonempty[0]).strip() if _nonempty else ""
                # Заголовок секции — маркер (▸/===) ИЛИ чисто буквенное название
                # бренда заглавными. Артикул (GOOD-2, SKF-100) содержит цифры —
                # его НЕ скипаем: одиночная ячейка-артикул без названия/цены
                # должна попасть в «битые строки», а не исчезнуть молча.
                _is_divider = (
                    (not _first)
                    or _first[0] in "▸►▶•—–-=*#"
                    or (_first.isupper() and not any(c.isdigit() for c in _first))
                )
                if _is_divider:
                    continue

            def get_raw(field):
                if field in fixed_vals:
                    return fixed_vals[field]
                idx = col_idx.get(field)
                if idx is None or idx >= len(row):
                    return ""
                return str(row[idx]).strip()

            row_vars: dict[str, Any] = {}
            for fld in col_idx:
                idx = col_idx[fld]
                raw = str(row[idx]).strip() if idx < len(row) else ""
                dec = _coerce_decimal(raw)
                row_vars[fld] = dec if dec is not None else raw

            def get(field):
                raw = get_raw(field)
                return _apply_transform(field, raw, transform_rules, row_vars)

            oem = get("oem_number")
            title = get("title")
            price_exw = _coerce_decimal(get("price_exw"))
            # Fallback: если в файле пустое название — берём OEM как title
            # (всё равно лучше чем пропустить позицию)
            if oem and not title:
                title = oem
            # Переводим title на английский для единообразия каталога
            # и сопоставления с конкурентами. Если уже ASCII — fast-path.
            if title:
                title = _translate_to_en(title)
            # OEM необязателен: нет настоящего артикула → генерим из названия,
            # чтобы каждая позиция была уникальной (товар опознаётся названием).
            if not oem and title:
                oem = _gen_oem_from_title(title)
            if not title or price_exw is None or price_exw <= 0:
                failed += 1
                if len(errors) < 50:
                    # Человекочитаемый комментарий: что не так + название
                    # колонки-источника + сырое значение из ячейки.
                    def _src_col(field):
                        i = col_idx.get(field)
                        if i is not None and headers and i < len(headers):
                            return str(headers[i]).strip()
                        return ""
                    if not title:
                        reason = "no title"
                        col = _src_col("title") or "Description"
                        val = ""
                        problem = "Название пустое — заполните (артикул необязателен)"
                    else:
                        reason = "bad price_exw"
                        col = _src_col("price_exw") or "Unitprice"
                        raw_cell = get_raw("price_exw").strip()
                        val = raw_cell or (str(price_exw) if price_exw is not None else "")
                        if price_exw is None:
                            problem = ("Пустая — впишите цену" if not raw_cell
                                       else "Не число — уберите текст/символы")
                        else:  # <= 0
                            problem = "Ноль или меньше — продавать нельзя"
                    comment = (f"Колонка «{col}»"
                               + (f": значение «{val[:24]}»" if val else " пустая")
                               + f" — {problem.lower()}.")
                    errors.append({
                        "row": row_n, "oem": oem[:60], "reason": reason,
                        "column": col, "value": val[:30], "problem": problem,
                        "comment": comment,
                    })
                continue

            brand_name = _translate_to_en((get("brand") or "").strip())
            if brand_name:
                key = brand_name.lower()
                brand = brand_cache.get(key)
                if not brand:
                    # SLUG — unique-поле, а slugify нормализует сильнее, чем
                    # .lower() (убирает ®, точки, спецсимволы), поэтому разные
                    # имена («Cummins», «Cummins®», «CUMMINS.») дают один slug.
                    # Раньше это падало на duplicate key marketplace_brand_slug.
                    # get_or_create по slug берёт существующий бренд; fallback
                    # ловит и прочие коллизии (напр. unique по name).
                    bslug = slugify(brand_name)[:200] or generic_brand.slug
                    try:
                        brand, _ = Brand.objects.get_or_create(
                            slug=bslug, defaults={"name": brand_name[:200]},
                        )
                    except IntegrityError:
                        brand = (
                            Brand.objects.filter(slug=bslug).first()
                            or Brand.objects.filter(name__iexact=brand_name).first()
                            or generic_brand
                        )
                    brand_cache[key] = brand
            elif filename_brand:
                # Колонка Brand пустая, но имя файла подсказывает бренд
                brand = filename_brand
            else:
                brand = generic_brand

            stock = _coerce_int(get("stock")) or 0
            currency = _normalize_currency(get("currency"))
            cond_raw = (get("condition") or "").strip().lower()
            condition = ("oem" if "oem" in cond_raw else
                          "reman" if "reman" in cond_raw else
                          "aftermarket" if "after" in cond_raw else
                          "oem")
            # availability: IN_STOCK / BACKORDER → in_stock / backorder
            avail_raw = (get("availability") or "").strip().lower()
            availability = "backorder" if "backorder" in avail_raw or "back" in avail_raw else "in_stock"
            # manufacturer + visible flag — переводим на английский.
            # Per-row fallback: если колонка Manufacturer есть, но в этой
            # ячейке пусто (типичный мульти-брендовый прайс) — подставляем
            # значение из мастера (constants). Это даёт column-wins-with-
            # fallback семантику: разные бренды получают свои заводы из
            # файла, а пустые ячейки — общий fallback от поставщика.
            _mfg_raw = (get("manufacturer") or "").strip()
            if not _mfg_raw:
                _mfg_raw = str(constants.get("manufacturer") or "").strip()
            manufacturer = _translate_to_en(_mfg_raw[:200])
            # Видимость завода для покупателя — автоматически по ЗНАЧЕНИЮ
            # (не по источнику: вписал руками или выбрал чип — неважно).
            # Известный бренд (whitelist аналогов ИЛИ любой бренд каталога
            # платформы) → показываем (доверие/рейтинг). Неизвестная строка
            # = частная торговая марка поставщика → скрываем. Сравниваем и
            # сырое, и переведённое значение (бренды латиницей не переводятся).
            manufacturer_visible = bool(manufacturer) and (
                manufacturer.strip().lower() in _known_for_visibility
                or _mfg_raw.strip().lower() in _known_for_visibility
            )
            cross_number = (get("cross_number") or "")[:500]
            price_fob_sea = _coerce_decimal(get("price_fob_sea")) or Decimal("0")
            price_fob_air = _coerce_decimal(get("price_fob_air")) or Decimal("0")
            warehouse_addr = (get("warehouse_address") or "")[:255]
            sea_port = (get("sea_port") or "")[:120]
            air_port = (get("air_port") or "")[:120]
            # AI-оценки per-part — fast-path skip если их нет (типичный случай).
            # Когда есть — приоритет: файл > AI > fix-дефолт.
            ai_est = ai_estimates.get(oem) if ai_estimates else None

            def _dim_fast(field, default_val):
                raw = _coerce_decimal(get(field))
                return raw if raw and raw > 0 else default_val

            def _dim_with_ai(field, default_val, ai_key):
                if field in col_idx:
                    raw = _coerce_decimal(get(field))
                    if raw and raw > 0:
                        return raw
                if ai_est and ai_est.get(ai_key):
                    try:
                        v = Decimal(str(ai_est[ai_key]))
                        if v > 0:
                            return v
                    except Exception:
                        pass
                raw = _coerce_decimal(get(field))
                return raw if raw and raw > 0 else default_val

            _dim = _dim_with_ai if ai_est else _dim_fast
            if ai_est:
                weight = _dim_with_ai("weight_kg", Decimal("0.5"), "weight_kg")
                length = _dim_with_ai("length_cm", Decimal("1.0"), "length_cm")
                width = _dim_with_ai("width_cm", Decimal("1.0"), "width_cm")
                height = _dim_with_ai("height_cm", Decimal("1.0"), "height_cm")
            else:
                weight = _dim_fast("weight_kg", Decimal("0.5"))
                length = _dim_fast("length_cm", Decimal("1.0"))
                width = _dim_fast("width_cm", Decimal("1.0"))
                height = _dim_fast("height_cm", Decimal("1.0"))

            payloads.append({
                "row_n": row_n,
                "oem": oem[:100],
                "fields": {
                    "title": title[:255],
                    "oem_number": oem[:100],
                    "slug": slugify(f"{oem}-{seller.username}")[:280],
                    "price": price_exw,
                    "currency": currency,
                    "stock_quantity": stock,
                    "condition": condition,
                    "availability": availability,
                    "manufacturer": manufacturer,
                    "manufacturer_visible": manufacturer_visible,
                    "cross_numbers": cross_number,
                    "price_fob_sea": price_fob_sea,
                    "price_fob_air": price_fob_air,
                    "warehouse_address": warehouse_addr,
                    "sea_port": sea_port,
                    "air_port": air_port,
                    "gross_weight_kg": weight,
                    "length_cm": length,
                    "width_cm": width,
                    "height_cm": height,
                    "category": cat,
                    "brand": brand,
                    "is_active": True,
                },
            })

    # ───────── Pass 1b: дедупликация внутри файла ─────────────────────
    # Поставщики (Sandvik, Atlas Copco) часто пишут одну позицию N раз
    # с Quantity=1 вместо одной строки с Quantity=N. Если цена совпадает —
    # суммируем количества в одну позицию. Если цены различаются — берём
    # первую и логируем как ошибку (это или опечатка, или две модификации).
    # Ключ дедупа — (oem, price): одинаковая цена → одна позиция (Qty=MAX);
    # РАЗНАЯ цена у того же артикула → берём ОБЕ как отдельные позиции
    # (варианты прайса). Второму+ варианту даём уникальный slug, иначе
    # БД отвергнет по unique(slug). oem остаётся тот же.
    merged_payloads: dict[str, dict] = {}
    merged_count = 0
    price_variants = 0
    oem_variant_n: dict[str, int] = {}
    for pl in payloads:
        oem = pl["oem"].lower()
        price = pl["fields"].get("price")
        key = oem + "|" + str(price)
        if key in merged_payloads:
            # Тот же oem И та же цена → одна позиция, Qty=MAX.
            ex = merged_payloads[key]
            ex["fields"]["stock_quantity"] = max(
                ex["fields"].get("stock_quantity", 0) or 0,
                pl["fields"].get("stock_quantity", 0) or 0,
            )
            merged_count += 1
            continue
        n = oem_variant_n.get(oem, 0)
        if n > 0:
            # 2-й+ вариант цены для этого артикула → уникализируем slug
            # и помечаем как отдельную позицию (всегда create, не update).
            base = pl["fields"].get("slug", "")
            pl["fields"]["slug"] = (base + "-v" + str(n + 1))[:280]
            pl["_force_create"] = True
            price_variants += 1
        oem_variant_n[oem] = n + 1
        merged_payloads[key] = pl
    payloads = list(merged_payloads.values())
    if merged_count:
        logger.info("In-file dedupe: merged %d duplicate rows by Qty MAX", merged_count)
    if price_variants:
        logger.info("In-file dedupe: %d price variants kept as separate positions", price_variants)
    # Сохраним статистику для отчёта в UI
    import_obj._dedupe_stats = {
        "merged": merged_count, "price_conflicts": price_variants,
        "unique_oems": len(payloads),
    }

    # ───────── Pass 2a: только {oem_lower: part_id} — не полный Part ─────
    # Раньше загружали полные Part объекты (десятки полей × 600k = много
    # RAM и Python overhead). Для классификации create/update нужен только
    # ID существующей записи. values_list в 5-10× быстрее .objects.filter().
    # С составным индексом (seller_id, oem_number) чанк в 900 — почти O(1).
    oems = [p["oem"] for p in payloads]
    existing_id_by_oem: dict[str, int] = {}
    CHUNK = 900
    if oems:
        cache.set(progress_key, {
            "current": 0, "total": len(oems), "phase": "matching",
            "running": True,
        }, 600)
        seller_id = seller.id
        last_progress = 0
        for i in range(0, len(oems), CHUNK):
            batch = oems[i:i + CHUNK]
            for pid, oem in (Part.objects
                              .filter(seller_id=seller_id, oem_number__in=batch)
                              .values_list("id", "oem_number")):
                existing_id_by_oem[oem.lower()] = pid
            # Прогресс каждый ~10k OEM — чаще чем раньше (раз в 9k → 5k),
            # пользователь видит непрерывное движение, нет «зависов».
            if (i - last_progress) >= 5000:
                cache.set(progress_key, {
                    "current": min(i + CHUNK, len(oems)), "total": len(oems),
                    "phase": "matching", "running": True,
                }, 600)
                last_progress = i

    # ───────── Pass 2b: enrichment из PartReference — fast-path skip ────
    # Если эталонная база пуста для данного seller'а — пропускаем целиком.
    # Эта таблица почти всегда пустая в demo/прод, не тратим время.
    from marketplace.models import PartReference
    has_refs = PartReference.objects.filter(oem_number__in=oems[:CHUNK]).exists()
    reference_hits = 0
    if has_refs:
        reference_data = _lookup_reference_data(oems)
        for pl in payloads:
            ref = reference_data.get(pl["oem"])
            if not ref:
                continue
            fields = pl["fields"]
            applied = False
            for key, target in (("weight_kg", "gross_weight_kg"),
                                  ("length_cm", "length_cm"),
                                  ("width_cm", "width_cm"),
                                  ("height_cm", "height_cm")):
                if key in col_idx:
                    continue
                if ref.get(key) is not None:
                    fields[target] = Decimal(str(ref[key]))
                    applied = True
            if "cross_number" not in col_idx and ref.get("cross_numbers"):
                if not (fields.get("cross_numbers") or "").strip():
                    fields["cross_numbers"] = ref["cross_numbers"][:500]
                    applied = True
            if applied:
                reference_hits += 1
    import_obj._enrichment_stats = {"reference_hits": reference_hits}

    # ───────── Pass 3: split create/update без построения Part объектов ──
    # Раньше создавали Part(seller=...) и setattr — Python overhead.
    # Теперь только разделяем payloads на 2 списка по наличию ID, всё
    # дальнейшее — кортежи прямо из pl["fields"].
    to_create_payloads: list = []
    to_update_payloads: list = []  # (part_id, fields_dict)
    for pl in payloads:
        pid = existing_id_by_oem.get(pl["oem"].lower())
        # Ценовые варианты (2-й+ с уникальным slug) всегда создаём отдельно,
        # иначе все варианты одного oem перезапишут одну запись.
        if pid and not pl.get("_force_create"):
            to_update_payloads.append((pid, pl["fields"]))
        else:
            to_create_payloads.append(pl["fields"])

    # ───────── Pass 4: BULK INSERT через raw SQL (5-10x быстрее ORM) ─────
    # ORM bulk_create делает много overhead'а на signal'ах, валидации,
    # auto-now полях. Для массового импорта используем raw INSERT
    # через executemany с явной транзакцией.
    from datetime import datetime

    from django.db import connection

    from assistant.part_naming import translate_title
    now = datetime.now(tz=UTC)

    insert_cols = [
        "title", "title_ru", "slug", "oem_number", "description", "price", "stock_quantity",
        "condition", "image_url", "seller_id", "brand_id", "category_id",
        "availability", "availability_status", "currency", "incoterm", "moq",
        "production_lead_days", "prep_to_ship_days", "shipping_lead_days",
        "gross_weight_kg", "length_cm", "width_cm", "height_cm",
        "country_of_origin", "cross_numbers",
        "price_fob_sea", "price_fob_air", "warehouse_address", "sea_port", "air_port",
        "hs_code", "backorder_allowed", "mapping_status", "supplier_part_uid",
        "data_updated_at", "is_active", "admin_note",
        "manufacturer", "manufacturer_visible",
        "warehouse_id",
        "created_at", "updated_at",
    ]
    table = Part._meta.db_table
    placeholders = ", ".join(["%s"] * len(insert_cols))
    insert_sql = (f"INSERT INTO {table} ({', '.join(insert_cols)}) "
                   f"VALUES ({placeholders})")

    def _val(obj, attr, default):
        v = getattr(obj, attr, None)
        return v if v is not None else default

    wh_id = warehouse.id if warehouse else None
    # Прогресс между matching и INSERT/UPDATE: пользователь видит что
    # сопоставление закончилось и теперь идёт запись в БД.
    total_pending = len(to_create_payloads) + len(to_update_payloads)
    cache.set(progress_key, {
        "current": 0, "total": total_pending, "phase": "writing",
        "running": True,
    }, 600)
    with transaction.atomic():
        if to_create_payloads:
            rows = []
            for f in to_create_payloads:
                rows.append((
                    f.get("title") or "", translate_title(f.get("title") or ""),
                    f.get("slug") or "",
                    f.get("oem_number"), "",
                    f.get("price"), f.get("stock_quantity") or 0,
                    f.get("condition") or "oem", "",
                    seller.id, f["brand"].id, f["category"].id,
                    f.get("availability") or "in_stock",
                    "active", f.get("currency") or "USD",
                    "FOB", 1, 1, 1, 1,
                    f.get("gross_weight_kg"), f.get("length_cm"),
                    f.get("width_cm"), f.get("height_cm"),
                    "Unknown", f.get("cross_numbers") or "",
                    f.get("price_fob_sea"), f.get("price_fob_air"),
                    f.get("warehouse_address") or "",
                    f.get("sea_port") or "", f.get("air_port") or "",
                    "", False, "auto", "",
                    now, True, "",
                    f.get("manufacturer") or "",
                    f.get("manufacturer_visible", True),
                    wh_id,
                    now, now,  # created_at, updated_at
                ))
            INS_BATCH = 50000
            try:
                with connection.cursor() as cur:
                    for batch_start in range(0, len(rows), INS_BATCH):
                        batch = rows[batch_start:batch_start + INS_BATCH]
                        cur.executemany(insert_sql, batch)
                        created += len(batch)
                        cache.set(progress_key, {
                            "current": created, "total": total_pending,
                            "phase": "writing", "running": True,
                        }, 600)
            except Exception as e:
                logger.exception("bulk INSERT failed for %d rows (batch %d)",
                                  len(rows), batch_start)
                failed += len(rows) - created
                err_entry = {"row": 0, "reason": f"raw insert: {type(e).__name__}: {str(e)[:200]}"}
                errors.insert(0, err_entry)
            cache.set(progress_key, {
                "current": created, "total": len(payloads), "running": True,
            }, 600)

        # Raw SQL bulk UPDATE через executemany — намного быстрее
        # Django bulk_update который генерит CASE WHEN на каждое поле.
        # 160k UPDATE: было 110с (bulk_update), стало ~5с (executemany).
        if to_update_payloads:
            # slug НЕ обновляем — он уникален и генерится из OEM, который
            # совпадает (мы матчим по seller+oem_number). Обновление slug
            # ловит UNIQUE constraint если он отличается от исходного.
            upd_cols = [
                "title", "title_ru", "price", "currency", "stock_quantity", "condition",
                "availability", "manufacturer", "manufacturer_visible",
                "cross_numbers", "price_fob_sea", "price_fob_air",
                "warehouse_address", "sea_port", "air_port",
                "gross_weight_kg", "length_cm", "width_cm", "height_cm",
                "brand_id", "category_id", "is_active",
                "warehouse_id",
                "data_updated_at", "updated_at",
            ]
            update_sql = (f"UPDATE {table} SET "
                           + ", ".join(f"{c} = %s" for c in upd_cols)
                           + " WHERE id = %s")
            upd_rows = []
            for pid, f in to_update_payloads:
                upd_rows.append((
                    f.get("title") or "", translate_title(f.get("title") or ""),
                    f.get("price"),
                    f.get("currency") or "USD",
                    f.get("stock_quantity") or 0, f.get("condition") or "oem",
                    f.get("availability") or "in_stock",
                    f.get("manufacturer") or "",
                    f.get("manufacturer_visible", True),
                    f.get("cross_numbers") or "",
                    f.get("price_fob_sea"), f.get("price_fob_air"),
                    f.get("warehouse_address") or "",
                    f.get("sea_port") or "", f.get("air_port") or "",
                    f.get("gross_weight_kg"), f.get("length_cm"),
                    f.get("width_cm"), f.get("height_cm"),
                    f["brand"].id, f["category"].id, True,
                    wh_id,
                    now, now,
                    pid,
                ))
            # Чанкуем UPDATE на пачки по 50k — каждая пачка ~3-5с в SQLite.
            # Между пачками cache.set обновляет прогресс, чтобы UI не висел
            # на «616473/616473» по 30+ секунд.
            UPD_BATCH = 50000
            try:
                with connection.cursor() as cur:
                    for batch_start in range(0, len(upd_rows), UPD_BATCH):
                        batch = upd_rows[batch_start:batch_start + UPD_BATCH]
                        cur.executemany(update_sql, batch)
                        updated += len(batch)
                        cache.set(progress_key, {
                            "current": created + updated,
                            "total": total_pending,
                            "phase": "writing", "running": True,
                        }, 600)
            except Exception as e:
                logger.exception("bulk UPDATE failed for %d rows (batch %d)",
                                  len(upd_rows), batch_start)
                # Уже успешно записанные UPDATE остаются — failed только
                # для строк которые не дошли до записи.
                failed += len(upd_rows) - updated
                err_entry = {"row": 0, "reason": f"raw update: {type(e).__name__}: {str(e)[:200]}"}
                errors.insert(0, err_entry)
            cache.set(progress_key, {
                "current": created + updated, "total": len(payloads), "running": True,
            }, 600)

    imported = created + updated
    cache.set(progress_key, {
        "current": imported, "total": imported + failed, "running": False,
    }, 600)
    return imported, created, updated, failed, errors


# ── HTTP views ───────────────────────────────────────────────────

class PricelistUploadView(APIView):
    """POST /api/assistant/upload-pricelist/  (multipart, field 'file')

    1) читаем headers + 3 строки
    2) AI/heuristic предлагает mapping
    3) сохраняем PricelistImport(status='preview') + ChatMessage с формой
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        from marketplace.models import (
            PricelistImport,
            PricelistMapping,
            SupplierImportProfile,
        )

        f = request.FILES.get("file")
        if not f:
            return Response({"error": _("Файл не передан.")}, status=400)
        if f.size > MAX_FILE_BYTES:
            mb = MAX_FILE_BYTES // 1024 // 1024
            return Response(
                {"error": _("Файл слишком большой (%(size)s МБ). Максимум %(max)s МБ.") % {
                    "size": f.size // 1024 // 1024, "max": mb}},
                status=400,
            )
        blob = f.read()
        try:
            headers, sample = _read_preview(f.name, blob)
        except Exception as e:
            return Response({"error": _("Не удалось прочитать файл: %(err)s") % {"err": e}}, status=400)
        if not headers or not any(str(h).strip() for h in headers):
            return Response({"error": _("Первая строка пустая — нет заголовков.")}, status=400)
        non_empty = [h for h in headers if str(h).strip()]
        if len(non_empty) < MIN_COLUMNS:
            return Response({"error": (
                _("В файле слишком мало колонок (%(n)s). Минимум %(min)s.") % {
                    "n": len(non_empty), "min": MIN_COLUMNS}
            )}, status=400)
        if len(headers) > MAX_COLUMNS:
            return Response({"error": (
                _("В файле слишком много колонок (%(n)s). Максимум %(max)s.") % {
                    "n": len(headers), "max": MAX_COLUMNS}
            )}, status=400)

        # 1. Ищем сохранённый профиль по fingerprint заголовков
        fingerprint = SupplierImportProfile.compute_fingerprint(headers)
        profile = SupplierImportProfile.objects.filter(
            seller=request.user, headers_fingerprint=fingerprint, is_active=True,
        ).first()

        ai_called = False
        unknown: list = []
        smart_status: str = "ok"  # default; will be set if _smart_mapping called
        transform_rules: dict = {}
        constants: dict = {}
        from_profile = False

        def _file_mapped_count(s):
            return sum(1 for v in (s or {}).values()
                        if v and not (isinstance(v, str) and v.startswith("fix:")))

        def _required_satisfied(s):
            # Профиль засчитывается только если purchase-критичные поля
            # есть среди замапленных колонок (а не один случайный «Quantity»).
            for f in REQUIRED_FIELDS:
                v = (s or {}).get(f)
                if not v or (isinstance(v, str) and v.startswith("fix:")):
                    return False
            return True

        # Supplier-wide поля (порты, склад) НЕ восстанавливаем из профиля —
        # юзер должен выбрать их заново для каждой загрузки. Профиль
        # переиспользуется только для маппинга колонок и formula-правил.
        SUPPLIER_WIDE_RESET = {"sea_port", "air_port", "warehouse_address"}
        suggested = {}
        if profile:
            suggested = {
                k: v for k, v in profile.column_mapping.items()
                if v and (v.startswith("fix:") or v in headers)
                and k not in SUPPLIER_WIDE_RESET
            }
            transform_rules = profile.transform_rules or {}
            constants = {
                k: v for k, v in (profile.constants or {}).items()
                if k not in SUPPLIER_WIDE_RESET
            }
            # Однако supplier-wide колонки В ФАЙЛЕ — это другое: если
            # SeaPort/AirPort/WarehouseAddress есть как заголовки, мапим
            # их детерминированно через словарь (НЕ через профиль, т.е.
            # значения там per-row, а не supplier-wide константа).
            from assistant.price_mappings import CANONICAL_TO_STD, _build_lookup, normalize
            _lookup_local = _build_lookup()
            for h in headers:
                canonical = _lookup_local.get(normalize(h))
                std = CANONICAL_TO_STD.get(canonical) if canonical else None
                if std in SUPPLIER_WIDE_RESET and std not in suggested:
                    suggested[std] = h
            from_profile = True
        # Если профиль не покрывает required-поля файла (oem, title, price) —
        # его маппинг бесполезен (он от другого поставщика). Делаем fresh
        # smart-mapping вместо того чтобы сохранять одну случайную колонку.
        if not profile or not _required_satisfied(suggested):
            prev = PricelistMapping.objects.filter(seller=request.user).first()
            prev_mapped = {}
            if prev and prev.mapping:
                prev_mapped = {k: v for k, v in prev.mapping.items()
                                if v and (v.startswith("fix:") or v in headers)}
            if _required_satisfied(prev_mapped):
                suggested = prev_mapped
                from_profile = False
            else:
                # Свежий smart_mapping (словарь + AI для unknown headers)
                fresh, unknown, ai_called, smart_status = _smart_mapping(
                    headers, sample, seller=request.user,
                )
                suggested = fresh
                from_profile = False
                transform_rules = {}
                constants = {}
        # AI quota exceeded — НЕ фатально. У нас есть результат словаря +
        # value detection. Просто отмечаем для UI чтобы юзер знал что
        # уникальные колонки не дораспознаны AI'ом — пусть поправит руками.
        ai_quota_exceeded = (smart_status == "quota_exceeded")

        # Сканируем колонку бренда: собираем (1) список всех брендов в файле
        # и (2) по одной показательной строке на каждый бренд — чтобы превью
        # «как ляжет в базу» демонстрировало мультибренд, а не три строки
        # из первой секции файла.
        detected_brands = []
        diverse_sample = []
        brand_src = suggested.get("brand")
        if brand_src and isinstance(brand_src, str) and not brand_src.startswith("fix:") \
                and brand_src in headers:
            try:
                bidx = headers.index(brand_src)
                # Повторные строки-заголовки внутри секций мультибренд-файла
                # ("Бренд", "Brand"…) не должны попасть в список брендов.
                _hdr_low = {str(h).strip().lower() for h in headers if h}
                seen_b = set()
                for i, r in enumerate(_read_all(f.name, blob)):
                    if i > 5000:
                        break
                    v = str(r[bidx]).strip() if bidx < len(r) else ""
                    vl = v.lower()
                    if v and vl not in seen_b and vl not in _hdr_low:
                        seen_b.add(vl)
                        detected_brands.append(v)
                        if len(diverse_sample) < 5:
                            diverse_sample.append(list(r))
                        if len(detected_brands) >= 60:
                            break
            except Exception:
                detected_brands = []
                diverse_sample = []

        # Если в файле ≥2 брендов — показываем превью по одной строке на бренд,
        # иначе обычные первые строки.
        preview_sample = diverse_sample if len(detected_brands) >= 2 else sample
        mapped_preview = _build_mapped_preview(headers, preview_sample, suggested)

        # Незамапленные опциональные поля → автозаполнение дефолтами
        auto_defaults = {}
        for key, label, req, enum_v, dflt in STD_FIELDS:
            if key not in suggested and key not in constants and dflt is not None:
                auto_defaults[key] = dflt
                suggested[key] = f"fix:{dflt}"

        # Вопросы оператору — только required без маппинга (oem, title, price)
        questions = []
        for key, label, req, enum_v, dflt in STD_FIELDS:
            if req and key not in suggested and key not in constants:
                q: dict[str, Any] = {
                    "field": key,
                    "question": _("Укажите значение для «%(label)s»") % {"label": label},
                    "type": "select" if enum_v else "text",
                }
                if enum_v:
                    q["options"] = enum_v
                    q["default"] = enum_v[0]
                questions.append(q)

        imp = PricelistImport.objects.create(
            seller=request.user,
            filename=f.name[:255],
            headers=headers,
            sample_rows=sample,
            suggested_mapping=suggested,
            ai_called=ai_called,
            status="preview",
        )
        imp.file_obj.save(f.name, ContentFile(blob), save=True)

        # total_rows НЕ считаем здесь — это требует прохода всего файла
        # (для 616k Komatsu = +2с). Используем размер файла как proxy.
        # Если нужно точное число — клиент дозапросит отдельно.
        total_rows = None

        # AI questionnaire НЕ генерируется здесь синхронно
        # (это ~4 секунды задержки). Юзер запросит его отдельно
        # через /smart-questions/ — фронт сразу показывает форму.

        # Если бренд берётся из колонки файла — собираем уникальные бренды
        # (скан файла, до ~5000 строк), чтобы показать поставщику «вижу N
        # брендов: Komatsu, Caterpillar…» — подтверждение мультибренд-прайса.
        return Response({
            "import_id": imp.id,
            "filename": f.name,
            "headers": headers,
            "sample_rows": sample,
            "total_rows": total_rows,
            "detected_brands": detected_brands,
            "mapped_preview": mapped_preview,
            "suggested_mapping": suggested,
            "ai_called": ai_called,
            "ai_quota_exceeded": ai_quota_exceeded,
            "ai_intro": "",         # подтянется async
            "smart_questions": [],  # подтянется async
            "smart_questions_pending": True,
            "unknown_headers": unknown,
            "from_saved_mapping": bool(not ai_called and not from_profile),
            "from_profile": from_profile,
            "profile_name": profile.name if profile else None,
            "transform_rules": transform_rules,
            "constants": constants,
            "questions": questions,
            "std_fields": [
                {"key": k, "label": label, "required": req,
                 "enum_values": enum_v or [], "default": dflt or ""}
                for k, label, req, enum_v, dflt in STD_FIELDS
            ],
        })


def _execute_import_job(import_id, mapping, transform_rules, constants,
                          save_profile, warehouse_id):
    """Тяжёлая часть коммита прайса в каталог (запись + статистика).

    Вызывается:
      • Celery-задачей run_pricelist_import (прод — отдельный процесс воркера,
        не грузит daphne/AI: для файла 100K+ строк это критично);
      • инлайн в тестах / PRICELIST_IMPORT_SYNC (Celery-воркер не видит
        транзакцию тестовой БД, а тесты ждут синхронный результат).

    Прогресс и результат пишутся в общий Redis-кэш caches["pricelist"], чтобы
    их видел поллинг из daphne. Возвращает result-dict (для inline-режима).
    Соединение БД НЕ закрывает — этим управляет вызывающий (Celery / request).
    """
    import logging

    from django.core.cache import caches
    from django.utils import timezone as _tz

    from marketplace.models import (
        Part,
        PricelistImport,
        PricelistMapping,
        SellerWarehouse,
        SupplierImportProfile,
    )
    plc = caches["pricelist"]
    imp = PricelistImport.objects.get(id=import_id)
    seller = imp.seller
    warehouse = None
    if warehouse_id:
        try:
            warehouse = SellerWarehouse.objects.get(id=warehouse_id)
        except SellerWarehouse.DoesNotExist:
            warehouse = None

    def _drop_empty_warehouse():
        # Склад создаётся ДО импорта (_resolve_warehouse_for_import). Если
        # импорт упал или не дал ни одной позиции — пустой склад не должен
        # оставаться (бизнес-правило: пустых складов быть не может). Удаляем
        # ТОЛЬКО если он реально пуст (мог быть переиспользован — тогда в нём
        # есть чужие позиции, его не трогаем).
        try:
            if warehouse and not Part.objects.filter(warehouse=warehouse).exists():
                warehouse.delete()
        except Exception:
            pass

    try:
        with imp.file_obj.open("rb") as fh:
            blob = fh.read()

        imported, created, updated, failed, errors = _import_file(
            imp, mapping, blob,
            transform_rules=transform_rules,
            constants=constants,
            warehouse=warehouse,
        )

        imp.final_mapping = mapping
        imp.imported_rows = imported
        imp.created_rows = created
        imp.updated_rows = updated
        imp.failed_rows = failed
        imp.error_details = errors
        imp.status = "imported"
        imp.completed_at = _tz.now()
        imp.save(update_fields=[
            "final_mapping", "imported_rows", "created_rows", "updated_rows",
            "failed_rows", "error_details", "status", "completed_at",
        ])

        PricelistMapping.objects.update_or_create(
            seller=seller, defaults={"mapping": mapping},
        )

        # Профиль поставщика для повторных загрузок
        if save_profile and imp.headers:
            fp = SupplierImportProfile.compute_fingerprint(imp.headers)
            SupplierImportProfile.objects.update_or_create(
                seller=seller,
                headers_fingerprint=fp,
                defaults={
                    "source_headers": imp.headers,
                    "column_mapping": mapping,
                    "transform_rules": transform_rules,
                    "constants": constants,
                    "name": f"Auto · {imp.filename[:60]}",
                    "use_count": 1,
                },
            )

        # Категоризация незаполненных полей (mandatory / rating_bonus / optional)
        FIELD_LABELS = {
            "weight_kg": _("вес"), "length_cm": _("длина"), "width_cm": _("ширина"),
            "height_cm": _("высота"), "stock": _("остаток"),
            "price_fob_sea": "FOB SEA", "price_fob_air": "FOB AIR",
            "warehouse_address": _("адрес склада"),
            "sea_port": _("морпорт"), "air_port": _("аэропорт"),
            "cross_number": _("кросс-номер"),
        }
        MANDATORY = {"warehouse_address", "sea_port", "air_port"}
        RATING_BONUS = {"length_cm", "width_cm", "height_cm", "cross_number"}

        def _filled(key):
            val = mapping.get(key, "")
            if val and not (isinstance(val, str) and val.startswith("fix:")):
                return True
            cval = (constants or {}).get(key)
            if cval not in (None, "", 0):
                return True
            return False

        missing_mandatory, missing_rating_bonus, missing_optional = [], [], []
        for key, label in FIELD_LABELS.items():
            if _filled(key):
                continue
            entry = {"key": key, "label": label}
            if key in MANDATORY:
                missing_mandatory.append(entry)
            elif key in RATING_BONUS:
                missing_rating_bonus.append(entry)
            else:
                missing_optional.append(entry)
        missing_from_file = missing_mandatory + missing_rating_bonus + missing_optional

        enrichment_stats = getattr(imp, "_enrichment_stats", {}) or {}
        dedupe_stats = getattr(imp, "_dedupe_stats", {}) or {}
        result = {
            "ok": True,
            "import_id": imp.id,
            "imported": imported,
            "created": created,
            "updated": updated,
            "failed": failed,
            "errors_preview": errors[:10],
            "missing_from_file": missing_from_file,
            "missing_mandatory": missing_mandatory,
            "missing_rating_bonus": missing_rating_bonus,
            "missing_optional": missing_optional,
            "ai_estimated_count": len(imp.ai_estimates or {}),
            "reference_enriched": enrichment_stats.get("reference_hits", 0),
            "merged_duplicates": dedupe_stats.get("merged", 0),
            "price_conflicts": dedupe_stats.get("price_conflicts", 0),
            "unique_oems": dedupe_stats.get("unique_oems", 0),
        }
        # Если импорт не дал ни одной позиции — не оставляем пустой склад.
        if not imported:
            _drop_empty_warehouse()
        plc.set(f"import_result_{imp.id}", result, 3600)
        plc.set(f"import_progress_{imp.id}", {
            "current": imported, "total": imported,
            "phase": "done", "running": False,
        }, 600)
        return result
    except Exception as e:  # noqa: BLE001
        logging.getLogger("pricelist").exception(
            "Import job failed for import_id=%s", import_id)
        try:
            imp.status = "failed"
            imp.error_details = [{"row": 0, "oem": "", "reason": str(e)[:300]}]
            imp.save(update_fields=["status", "error_details"])
        except Exception:
            pass
        # Импорт упал → пустой склад-husk удаляем (пустых складов быть не должно).
        _drop_empty_warehouse()
        err = {"ok": False, "import_id": import_id, "error": str(e)[:300]}
        plc.set(f"import_result_{import_id}", err, 3600)
        plc.set(f"import_progress_{import_id}", {
            "current": 0, "total": 0, "phase": "failed",
            "running": False, "error": str(e)[:300],
        }, 600)
        return err


class PricelistCommitView(APIView):
    """POST /api/assistant/upload-pricelist/<id>/commit/

    body: {"mapping": {std_field: source_column}}
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, import_id):
        from marketplace.models import (
            PricelistImport,
            PricelistMapping,
            SupplierImportProfile,
        )
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)
        # Разрешаем повтор после неудачного импорта (failed) — напр. после фикса
        # бага: импорт идемпотентен (upsert по seller+oem), повтор безопасен.
        if imp.status not in ("preview", "failed"):
            return Response({"error": f"already in status {imp.status}"}, status=400)

        mapping = request.data.get("mapping") or {}
        transform_rules = request.data.get("transform_rules") or {}
        constants = request.data.get("constants") or {}
        save_profile = request.data.get("save_profile", True)

        # Supplier-wide логистика (склад/порты) — если seller заполнил в KYB-
        # профиле и в текущей загрузке не передал ни колонку, ни константу —
        # подставляем из профиля. Без этого валидатор отклонит загрузку
        # даже когда значение уже известно платформе.
        try:
            kyb = getattr(request.user, "kyb_profile", None)
        except Exception:
            kyb = None
        for _sw_fld in ("warehouse_address", "sea_port", "air_port"):
            if constants.get(_sw_fld) or mapping.get(_sw_fld):
                continue
            if kyb:
                v = getattr(kyb, _sw_fld, "") or ""
                if v:
                    constants[_sw_fld] = v
        # Юзерские правки AI-оценок: {oem: {weight_kg, length_cm, ...}}
        # Применяются поверх ai_estimates как human-in-the-loop корректировка
        ai_overrides = request.data.get("ai_estimates_override") or {}
        if isinstance(ai_overrides, dict) and ai_overrides:
            current = imp.ai_estimates or {}
            for oem, fields in ai_overrides.items():
                if not isinstance(fields, dict):
                    continue
                cur = current.get(oem, {})
                for k in ("weight_kg", "length_cm", "width_cm", "height_cm"):
                    if k in fields:
                        try:
                            cur[k] = float(fields[k])
                        except (TypeError, ValueError):
                            pass
                cur["confidence"] = 1.0  # юзер подтвердил → 100%
                current[oem] = cur
            imp.ai_estimates = current
            imp.save(update_fields=["ai_estimates"])

        if not isinstance(mapping, dict):
            return Response({"error": "mapping must be object"}, status=400)

        # Применяем ответы на вопросы: constants ПРИОРИТЕТ над дефолтами.
        # Раньше скипались если fld уже в mapping — но mapping мог содержать
        # пустой default (fix:""), который блокировал юзерский ответ.
        for fld, val in (constants or {}).items():
            if val is None or val == "":
                continue
            cur = mapping.get(fld, "")
            # Перезатираем если: (а) нет в mapping, (б) пустой default fix:""
            #                   (в) колонка не замаплена на файл
            if (not cur
                or (isinstance(cur, str) and cur.startswith("fix:") and
                    cur[4:].strip() in ("", str(FIELD_DEFAULTS.get(fld, ""))))):
                mapping[fld] = f"fix:{val}"

        for f_key in REQUIRED_FIELDS:
            if not mapping.get(f_key):
                return Response(
                    {"error": f"required field '{f_key}' not mapped"}, status=400,
                )

        # Жёсткий чек обязательных supplier-wide полей.
        # Адрес склада — обязателен. Без него рассчитать логистику нельзя.
        SUPPLIER_WIDE_KEYS = {"warehouse_address", "sea_port", "air_port"}

        def _has_value(key: str) -> bool:
            """Проверяет что для supplier-wide поля есть данные:
            - constants из формы (приоритет)
            - fix:VALUE в mapping
            - mapped file column. Для supplier-wide колонок: ВСЕ строки sample
              должны быть заполнены одним значением (иначе констант важнее —
              иначе разольёмся пустотой по 160k строк маркетплейс-формата).
              Для не-supplier-wide: хотя бы одна непустая ячейка.
            """
            c = (constants or {}).get(key)
            if c not in (None, "", 0):
                return True
            m = mapping.get(key, "")
            if isinstance(m, str) and m.startswith("fix:") and m[4:].strip():
                return True
            if m and isinstance(m, str) and m in (imp.headers or []):
                idx = imp.headers.index(m)
                rows = imp.sample_rows or []
                if key in SUPPLIER_WIDE_KEYS:
                    seen = set()
                    for row in rows:
                        cell = str(row[idx]).strip() if idx < len(row) else ""
                        if not cell:
                            return False
                        seen.add(cell)
                    return bool(rows) and len(seen) == 1
                for row in rows:
                    if idx < len(row) and str(row[idx]).strip():
                        return True
            return False

        if not _has_value("warehouse_address"):
            return Response({
                "error": "warehouse_address_required",
                "message": _(
                    "Укажите адрес склада отгрузки — обязательное поле. "
                    "В колонке WarehouseAddress нет данных или она не заполнена — "
                    "впишите адрес в форме «📎 общих полей поставщика»."
                ),
            }, status=400)
        if not _has_value("sea_port"):
            return Response({
                "error": "sea_port_required",
                "message": _(
                    "Укажите ближайший к складу морпорт отгрузки — обязательное "
                    "поле для расчёта FOB SEA. Впишите в вопросе мастера или в "
                    "форме «📎 общих полей поставщика»."
                ),
            }, status=400)
        if not _has_value("air_port"):
            return Response({
                "error": "air_port_required",
                "message": _(
                    "Укажите ближайший к складу аэропорт отгрузки — обязательное "
                    "поле для расчёта FOB AIR. Впишите в вопросе мастера или в "
                    "форме «📎 общих полей поставщика»."
                ),
            }, status=400)

        # Бренд — обязателен. Должен быть в колонке файла (с данными),
        # в constants/fix:VALUE, ИЛИ детектится из имени файла.
        # Без бренда позиции попадут в "Generic" и потеряют сопоставление
        # с конкурентами — нам важны точные данные.
        if not _has_value("brand"):
            # Fallback: имя файла содержит известный бренд?
            fname_lc = (imp.filename or "").lower()
            fname_clean = fname_lc.replace("_", "").replace(" ", "").replace("-", "")
            known = ["epiroc", "caterpillar", "komatsu", "hitachi", "sandvik",
                     "liebherr", "atlascopco"]
            if not any(b in fname_clean for b in known):
                return Response({
                    "error": "brand_required",
                    "message": _(
                        "Не удалось определить бренд: колонка Brand пуста, "
                        "имя файла не содержит известного бренда. "
                        "Укажите бренд в смарт-вопросе или назовите файл "
                        "включив бренд (напр. komatsu_pricelist.xlsx)."
                    ),
                }, status=400)
        for fld, val in mapping.items():
            if not val:
                continue
            if isinstance(val, str) and val.startswith("fix:"):
                continue
            if val not in imp.headers:
                return Response(
                    {"error": f"unknown column '{val}' for {fld}"}, status=400,
                )

        # Валидация формул
        available_vars = list(mapping.keys())
        for fld, rule in transform_rules.items():
            if rule.get("type") == "formula" and rule.get("formula"):
                ok, reason = validate_formula(rule["formula"], available_vars)
                if not ok:
                    return Response(
                        {"error": f"Invalid formula for {fld}: {reason}"},
                        status=400,
                    )

        # Файл существует? (сам blob читает job — не тащим 15 МБ через Celery.)
        if not imp.file_obj:
            return Response({"error": "file unavailable"}, status=500)

        # Виртуальный склад для этой загрузки (создаётся на каждую загрузку).
        warehouse = _resolve_warehouse_for_import(request.user, mapping, constants, imp.filename)
        warehouse_id = warehouse.id if warehouse else None

        # ── Тяжёлый импорт уходит в Celery-воркер (отдельный процесс) ─────
        # На больших прайсах (100K+ строк, напр. полный каталог Epiroc ~160K)
        # запись длится минуты. Синхронно в запросе — Cloudflare рвёт на ~100s
        # («Failed to fetch»); в потоке daphne — CPU-bound Python (GIL) тормозит
        # чат и AI. Celery исполняет импорт в своём процессе: daphne мгновенно
        # отдаёт 202, прогресс/результат идут через общий Redis-кэш "pricelist",
        # поллинг /import-progress/ их видит.
        import logging

        from django.core.cache import caches
        plc = caches["pricelist"]
        plc.delete(f"import_result_{imp.id}")
        plc.set(f"import_progress_{imp.id}", {
            "current": 0, "total": 0, "phase": "writing", "running": True,
        }, 600)

        # Лента важных событий админа: загрузка прайса продавцом + IP/кабинет.
        try:
            from assistant.actions import _log_activity
            _xff = request.META.get("HTTP_X_FORWARDED_FOR", "")
            _ip = (_xff.split(",")[0].strip() if _xff
                   else request.META.get("REMOTE_ADDR", ""))[:64]
            _log_activity(
                "pricelist", actor=request.user, ip=_ip,
                title=_("Загрузка прайса #%(id)s · %(rows)s строк · %(file)s") % {
                    "id": imp.id, "rows": imp.total_rows,
                    "file": imp.filename or _("файл")},
                meta={"import_id": imp.id, "n_rows": imp.total_rows,
                      "filename": imp.filename or "",
                      "seller": request.user.username})
        except Exception:
            logging.getLogger("pricelist").exception("activity log (pricelist) failed")

        # В тестах / sync-режиме импорт идёт ИНЛАЙН (Celery-воркер не видит
        # транзакцию тестовой БД; существующие тесты ждут синхронный 200).
        if getattr(settings, "TESTING", False) or getattr(settings, "PRICELIST_IMPORT_SYNC", False):
            result = _execute_import_job(
                imp.id, mapping, transform_rules, constants, save_profile, warehouse_id,
            )
            if result.get("ok"):
                return Response(result)
            return Response(
                {"error": result.get("error") or "import failed"}, status=500)

        # Прод — фоновая Celery-задача. Если брокер недоступен — fallback инлайн
        # (хуже для UX на больших файлах, но импорт не теряется).
        try:
            from assistant.tasks import run_pricelist_import
            run_pricelist_import.delay(
                imp.id, mapping, transform_rules, constants, save_profile, warehouse_id,
            )
        except Exception:
            logging.getLogger("pricelist").exception(
                "Celery dispatch failed — inline fallback for import_id=%s", imp.id)
            result = _execute_import_job(
                imp.id, mapping, transform_rules, constants, save_profile, warehouse_id,
            )
            if result.get("ok"):
                return Response(result)
            return Response(
                {"error": result.get("error") or "import failed"}, status=500)

        return Response({
            "ok": True, "status": "running", "import_id": imp.id,
        }, status=202)


def _generate_marketplace_xlsx(import_obj, mapping: dict, transform_rules: dict,
                                  constants: dict) -> bytes:
    """Генерирует выходной XLSX в формате маркетплейса (как у claude.ai).

    Использует openpyxl WriteOnlyWorkbook — в 5-10 раз быстрее обычного
    Workbook потому что не создаёт Cell objects в памяти, пишет
    напрямую в zip stream.

    Прогресс пишется в Django cache → виден в UI через polling
    /generate-output-progress/ endpoint.
    """
    from django.core.cache import cache

    OUTPUT_COLS = [
        ("PartNumber",       "oem_number"),
        ("CrossNumber",      "cross_number"),
        ("Brand",            "brand"),
        ("Name",             "title"),
        ("Manufacturer",     "manufacturer"),
        ("Quantity",         "stock"),
        ("Condition",        "condition"),
        ("Availability",     "availability"),
        ("Price_EXW",        "price_exw"),
        ("Currency",         "currency"),
        ("WarehouseAddress", "warehouse_address"),
        ("Price_FOB_SEA",    "price_fob_sea"),
        ("Price_FOB_AIR",    "price_fob_air"),
        ("SeaPort",          "sea_port"),
        ("AirPort",          "air_port"),
        ("Weight",           "weight_kg"),
        ("Length",           "length_cm"),
        ("Width",            "width_cm"),
        ("Height",           "height_cm"),
    ]

    headers = import_obj.headers
    ai_estimates = getattr(import_obj, "ai_estimates", None) or {}

    # col_idx + fixed_vals (как в _import_file)
    col_idx: dict[str, int] = {}
    fixed_vals: dict[str, str] = {}
    for fld, dflt in FIELD_DEFAULTS.items():
        fixed_vals[fld] = str(dflt)
    for fld, val in (mapping or {}).items():
        if not val:
            continue
        if isinstance(val, str) and val.startswith("fix:"):
            fixed_vals[fld] = val[4:]
        elif val in headers:
            col_idx[fld] = headers.index(val)
            fixed_vals.pop(fld, None)
    # Constants имеют приоритет — даже если mapping уже содержит fix:default,
    # юзерский ответ должен перебить (свежий ответ юзера > старый профиль).
    for fld, val in (constants or {}).items():
        if val is None or val == "":
            continue
        if fld in col_idx:
            continue  # колонка замаплена на файл — это приоритет
        fixed_vals[fld] = str(val)

    # xlsxwriter — в 2-3 раза быстрее openpyxl для массовой записи.
    # constant_memory=True пишет напрямую в файл (не накапливает в RAM).
    import xlsxwriter
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"constant_memory": True, "in_memory": True})
    ws = wb.add_worksheet("Pricelist")

    hdr_fmt = wb.add_format({
        "bold": True, "font_color": "#FFFFFF", "bg_color": "#2D7A3E",
        "align": "center", "valign": "vcenter",
    })

    # Header row
    for col_n, (label, _key) in enumerate(OUTPUT_COLS):
        ws.write(0, col_n, label, hdr_fmt)
        ws.set_column(col_n, col_n, max(12, len(label) + 2))
    ws.set_row(0, 22)

    # Прогресс — для polling в UI
    progress_key = f"generate_output_progress_{import_obj.id}"
    cache.set(progress_key, {"current": 0, "total": 0, "running": True}, 600)

    try:
        with import_obj.file_obj.open("rb") as fh:
            blob = fh.read()
    except Exception:
        cache.set(progress_key, {"current": 0, "total": 0, "running": False}, 60)
        wb.close()
        raise

    written = 0
    row_idx = 1  # после header
    DIM_KEYS = ("weight_kg", "length_cm", "width_cm", "height_cm")
    PREVIEW_LIMIT = 100  # сохраняем HTML preview первых N строк
    preview_rows_html = []
    for row in _read_all(import_obj.filename, blob):
        if not any(str(c).strip() for c in row):
            continue

        # row_vars для формул
        row_vars = {}
        for fld in col_idx:
            idx = col_idx[fld]
            raw = str(row[idx]).strip() if idx < len(row) else ""
            dec = _coerce_decimal(raw)
            row_vars[fld] = dec if dec is not None else raw

        def get(field):
            if field in col_idx:
                idx = col_idx[field]
                raw = str(row[idx]).strip() if idx < len(row) else ""
            else:
                raw = fixed_vals.get(field, "")
            return _apply_transform(field, raw, transform_rules or {}, row_vars)

        oem = get("oem_number")
        if not oem:
            continue
        ai_est = ai_estimates.get(oem) if ai_estimates else None

        def _dim(field, ai_key):
            if field in col_idx:
                raw = _coerce_decimal(get(field))
                if raw and raw > 0:
                    return str(raw)
            if ai_est and ai_est.get(ai_key):
                return str(ai_est[ai_key])
            raw = _coerce_decimal(get(field))
            return str(raw) if raw and raw > 0 else ""

        # write_row — быстрее чем write по одной cell
        # Текстовые поля переводим на английский (каталог международный).
        EN_FIELDS = {"title", "manufacturer", "brand", "warehouse_address",
                     "sea_port", "air_port", "condition", "availability"}
        row_values = []
        for _label, key in OUTPUT_COLS:
            if key in DIM_KEYS:
                row_values.append(_dim(key, key))
            else:
                val = get(key) or ""
                if key in EN_FIELDS and isinstance(val, str):
                    val = _translate_to_en(val)
                row_values.append(val)
        ws.write_row(row_idx, 0, row_values)
        # Накопляем HTML preview первых строк параллельно с записью в XLSX
        if written < PREVIEW_LIMIT:
            from html import escape as _h
            cells = "".join(
                f"<td>{_h(str(v)[:50])}</td>" for v in row_values
            )
            preview_rows_html.append(f"<tr>{cells}</tr>")
        row_idx += 1
        written += 1
        if written % 1000 == 0:
            cache.set(progress_key, {
                "current": written, "total": 0, "running": True,
            }, 600)

    wb.close()
    cache.set(progress_key, {"current": written, "total": written, "running": False}, 60)

    # Кэшируем preview HTML в БД — мгновенный показ без openpyxl re-parse
    from html import escape as _h
    headers_html = "".join(f"<th>{_h(label)}</th>" for label, _ in OUTPUT_COLS)
    truncated = written > PREVIEW_LIMIT
    note = (_('<div class="opx-note">Показаны первые %(limit)s из %(total)s строк</div>') % {
        "limit": PREVIEW_LIMIT, "total": written}) if truncated else ""
    preview_html = (
        '<style>'
        'body{margin:0;padding:14px;font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:#fff;}'
        '.opx-note{font-size:11.5px;color:rgba(0,0,0,0.55);margin-bottom:10px;font-style:italic;}'
        'table{border-collapse:collapse;width:100%;font-size:11.5px;font-family:"SF Mono",Menlo,monospace;}'
        'th{background:#2D7A3E;color:#fff;padding:6px 8px;text-align:left;position:sticky;top:0;font-weight:600;letter-spacing:0.02em;}'
        'td{padding:4px 8px;border-bottom:1px solid rgba(0,0,0,0.05);color:#1a1a1a;white-space:nowrap;}'
        'tr:nth-child(even) td{background:rgba(45,122,62,0.03);}'
        '</style>'
        + note
        + f'<table><thead><tr>{headers_html}</tr></thead>'
        + '<tbody>' + ''.join(preview_rows_html) + '</tbody></table>'
    )
    try:
        import_obj.output_preview_html = preview_html
        import_obj.output_total_rows = written
        import_obj.save(update_fields=["output_preview_html", "output_total_rows"])
    except Exception:
        pass

    return buf.getvalue()


HARDCODED_QUESTIONS = [
    {
        "field": "brand",
        "question": _lazy("Бренд / производитель?"),
        "options": ["Epiroc", "Caterpillar", "Komatsu", "Hitachi", "Sandvik"],
        "default": "",
        "placeholder": _lazy("Или впишите свой бренд"),
        "apply_as": "constant",
    },
    {
        "field": "condition",
        "question": _lazy("Тип товара (Condition)?"),
        "options": ["OEM", "AFTERMARKET", "REMAN"],
        "default": "OEM",
        "apply_as": "constant",
    },
    {
        "field": "availability",
        "question": _lazy("Наличие?"),
        "options": ["IN_STOCK", "BACKORDER"],
        "default": "IN_STOCK",
        "apply_as": "constant",
    },
    {
        "field": "manufacturer",
        "question": _lazy("Кто производитель детали?"),
        "hint": _lazy("Узнаваемый бренд повышает рейтинг в выдаче и снимает вопросы покупателей. Нет в списке — впишите свой (частную марку скроем)."),
        "options": MANUFACTURER_SELECT_OPTIONS,
        "render": "select",
        "default": "",
        "placeholder": _lazy("Кликните или начните вводить — появится список брендов"),
        "apply_as": "constant",
        "skippable": True,
    },
    # Внимание: warehouse_address / sea_port / air_port — supplier-wide
    # логистика. Задаётся в отдельном блоке "📎 общих полей поставщика"
    # сразу после маппинга колонок (см. pl-defaults-card в chat-first.js),
    # с селектором страны и фильтрацией портов по стране. В мастере их нет
    # чтобы не дублировать UX. Бекенд-валидация (_has_value, MANDATORY,
    # KYB-fallback) продолжает работать — поля приходят через constants
    # из формы.
    {
        "field": "price_fob_sea",
        "question": _lazy("Наценка FOB SEA (доставка до морпорта) к цене EXW?"),
        "options": ["+0%", "+2%", "+4%"],
        "default": "+2%",
        "apply_as": "formula",
    },
    {
        "field": "price_fob_air",
        "question": _lazy("Наценка FOB AIR (доставка до аэропорта) к цене EXW?"),
        "options": ["+1%", "+3%", "+5%"],
        "default": "+3%",
        "apply_as": "formula",
    },
]


def _ai_smart_questions(headers: list[str], sample_rows: list[list[str]],
                         total_rows: int | None, mapping: dict[str, str],
                         filename: str = "", seller=None) -> dict:
    """Возвращает захардкоженный список вопросов + краткое intro.

    AI-генерация заменена на детерминированный список — у нас фикс-набор
    статусов товара (OEM/AFTERMARKET/REMAN), наличия (IN_STOCK/BACKORDER),
    наценок (FOB SEA: +5/+10/+15%, FOB AIR: +7/+15/+20%).
    """
    # Бренд по умолчанию подставляем из имени файла если узнаём (Epiroc, ...)
    known_brands = ["Epiroc", "Caterpillar", "Komatsu", "Hitachi", "Sandvik",
                     "Liebherr", "Atlas Copco"]
    detected_brand = ""
    fname_lower = (filename or "").lower()
    for b in known_brands:
        if b.lower().replace(" ", "") in fname_lower.replace("_", "").replace(" ", ""):
            detected_brand = b
            break

    # Если поле уже замаплено на колонку файла — вопрос задавать не нужно.
    # mapping[field] = "Column" — из файла; "fix:..." — дефолт, спрашиваем.
    SUPPLIER_WIDE_KEYS = {"warehouse_address", "sea_port", "air_port"}

    def _from_file(field: str) -> bool:
        """True если поле замаплено на колонку файла И значение в порядке.
        Для обычных полей: достаточно хотя бы одной непустой ячейки.
        Для supplier-wide (склад/порты): колонка должна быть полностью
        заполнена одним и тем же значением — иначе нужен один константный
        ответ от юзера, чтобы не разлились пустые/смешанные адреса на 160k
        строк (типичный кейс маркетплейс-формата с пустым WarehouseAddress).
        """
        v = (mapping or {}).get(field)
        if not v or (isinstance(v, str) and v.startswith("fix:")):
            return False
        if v not in (headers or []):
            return False
        idx = headers.index(v)
        rows = sample_rows or []
        if field in SUPPLIER_WIDE_KEYS:
            seen = set()
            for row in rows:
                cell = str(row[idx]).strip() if idx < len(row) else ""
                if not cell:
                    return False  # хоть одна пустая — не подходит
                seen.add(cell)
            return bool(rows) and len(seen) == 1
        for row in rows:
            if idx < len(row) and str(row[idx]).strip():
                return True
        return False

    # Supplier-wide поля (склад, порты) подтягиваем из KYB-профиля если есть —
    # это глобальные настройки логистики, не должны спрашиваться каждый раз.
    profile_supplier_wide = {}
    if seller is not None:
        try:
            kyb = getattr(seller, "kyb_profile", None)
            if kyb:
                for fld in ("warehouse_address", "sea_port", "air_port"):
                    v = getattr(kyb, fld, "") or ""
                    if v:
                        profile_supplier_wide[fld] = v
        except Exception:
            pass

    questions = []
    for q in HARDCODED_QUESTIONS:
        field = q["field"]
        if _from_file(field):
            continue  # значение придёт из колонки файла
        # supplier-wide поле уже есть в KYB-профиле — не спрашиваем повторно.
        if field in SUPPLIER_WIDE_KEYS and profile_supplier_wide.get(field):
            continue
        q_copy = dict(q)
        if q_copy["field"] == "brand" and detected_brand:
            q_copy["default"] = detected_brand
        questions.append(q_copy)

    # Базовое intro
    rows_label = (_("%(n)s позиций") % {"n": total_rows} if total_rows
                  else _("%(n)s колонок") % {"n": len(headers)})
    if questions:
        _nq = len(questions)
        _detail_word = (_("деталь") if _nq == 1
                        else _("детали") if 2 <= _nq <= 4 else _("деталей"))
        intro = (_("📋 Я распознал в файле %(rows)s. "
                   "Уточню %(n)s %(word)s"
                   " чтобы корректно заполнить карточки товара:") % {
                       "rows": rows_label, "n": _nq, "word": _detail_word})
    else:
        intro = (_("📋 Файл распознан полностью (%(rows)s). "
                   "Уточнять нечего — все колонки на месте.") % {"rows": rows_label})

    return {"intro": intro, "questions": questions}


def _ai_smart_questions_legacy_AI(headers: list[str], sample_rows: list[list[str]],
                         total_rows: int | None, mapping: dict[str, str],
                         filename: str = "") -> dict:
    """[legacy] AI-генерация вопросов через Claude tool use.

    Оставлено для возврата если понадобится динамика. Заменено на
    _ai_smart_questions с захардкоженным списком.
    """
    api_key = getattr(settings, "ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"intro": "", "questions": []}
    try:
        from anthropic import Anthropic
        client = Anthropic(api_key=api_key)
    except Exception:
        return {"intro": "", "questions": []}

    mapped_from_file = [k for k, v in mapping.items()
                         if v and not v.startswith("fix:")]
    missing_keys = []
    for k, label, req, _a, _b in STD_FIELDS:
        if k not in mapping or mapping[k].startswith("fix:"):
            if k not in mapped_from_file:
                missing_keys.append(k)

    sample_text = "\n".join(
        " | ".join(str(c)[:30] for c in row) for row in (sample_rows or [])[:3]
    )
    rows_info = f"{total_rows} позиций" if total_rows else f"{len(headers)} колонок"

    questionnaire_tool = {
        "name": "generate_questionnaire",
        "description": (
            "Сгенерировать дружелюбный intro + 2-4 умных вопроса "
            "для уточнения данных прайс-листа. Каждый вопрос имеет "
            "варианты ответа (chip-кнопки) или текстовое поле."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "intro": {
                    "type": "string",
                    "description": ("1-2 строки приветствия: что увидел в файле "
                                     "(объём, распознанные поля). Эмодзи ok."),
                },
                "questions": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "field": {
                                "type": "string",
                                "description": ("std-поле платформы (brand, condition, "
                                                 "warehouse_address, sea_port, air_port, "
                                                 "currency) ИЛИ price_fob_sea / "
                                                 "price_fob_air для наценок"),
                            },
                            "question": {"type": "string"},
                            "options": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": _("Chip-кнопки (3-5 вариантов)"),
                            },
                            "default": {"type": "string"},
                            "placeholder": {"type": "string"},
                            "apply_as": {
                                "type": "string",
                                "enum": ["constant", "formula"],
                                "description": ("constant — прямое значение в "
                                                 "constants. formula — для наценок типа "
                                                 "FOB SEA +15%, применяется как "
                                                 "price_exw * (1 + value/100)"),
                            },
                        },
                        "required": ["field", "question", "apply_as"],
                    },
                },
            },
            "required": ["intro", "questions"],
        },
    }

    prompt = (
        f"Юзер загрузил прайс-лист в B2B-маркетплейс запчастей.\n\n"
        f"ФАЙЛ: {filename}\n"
        f"КОЛОНКИ ({len(headers)}): {', '.join(headers)}\n"
        f"ОБЪЁМ: {rows_info}\n"
        f"ПЕРВЫЕ СТРОКИ:\n{sample_text}\n\n"
        f"РАСПОЗНАНО ИЗ ФАЙЛА: {', '.join(mapped_from_file) or '—'}\n"
        f"НЕТ В ФАЙЛЕ: {', '.join(missing_keys[:10]) or '—'}\n\n"
        "Сгенерируй questionnaire через tool generate_questionnaire.\n"
        "ПРАВИЛА:\n"
        "- intro: 1-2 строки дружелюбно, как Claude.ai\n"
        "- Задавай 2-4 вопроса (не больше!) про САМОЕ ВАЖНОЕ:\n"
        "  бренд, состояние, наценка FOB SEA, наценка FOB AIR, склад\n"
        "- НЕ спрашивай про вес/габариты/остаток (это per-part)\n"
        "- options — 3-5 chip-вариантов плюс возможность ввести своё\n"
        "- Для наценок FOB используй apply_as=formula, options как «+10%, +15%, +20%»\n"
        "- Если бренд явный (Epiroc в имени файла) — default=«Epiroc»\n"
    )
    try:
        msg = client.messages.create(
            model=getattr(settings, "ANTHROPIC_FAST_MODEL", "claude-haiku-4-5-20251001"),
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
            tools=[questionnaire_tool],
            tool_choice={"type": "tool", "name": "generate_questionnaire"},
        )
        for block in msg.content:
            if block.type == "tool_use" and block.name == "generate_questionnaire":
                inp = block.input or {}
                return {
                    "intro": inp.get("intro", ""),
                    "questions": inp.get("questions", []),
                }
        return {"intro": "", "questions": []}
    except Exception:
        logger.exception("AI questionnaire failed")
        return {"intro": "", "questions": []}


def _ai_estimate_per_part(items: list[dict],
                            on_partial=None) -> dict[str, dict]:
    """Claude tool use streaming: оценить вес/габариты per-part.

    items: [{oem, title, description?}] — до 50 штук за раз
    on_partial: callback(estimates_dict) вызывается когда из стрима
                распарсилась новая полная позиция. Это позволяет UI
                показывать прогресс ITEM-BY-ITEM как у claude.ai,
                а не batch-by-batch.
    returns: {oem: {weight_kg, length_cm, width_cm, height_cm}}
    """
    api_key = getattr(settings, "ANTHROPIC_API_KEY", "")
    if not api_key or not items:
        return {}
    try:
        from anthropic import Anthropic
        client = Anthropic(api_key=api_key)
    except Exception:
        return {}

    estimate_tool = {
        "name": "estimate_dimensions",
        "description": (
            "Оценить вес и габариты каждой запчасти на основе её "
            "названия, OEM-номера и описания. Используй типичные "
            "значения для подобных деталей спецтехники "
            "(Caterpillar, Komatsu, Hitachi, Liebherr и др.). "
            "Если не уверен — давай средние/консервативные оценки."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "estimates": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "oem": {"type": "string"},
                            "weight_kg": {"type": "number", "minimum": 0.01, "maximum": 5000},
                            "length_cm": {"type": "number", "minimum": 1, "maximum": 500},
                            "width_cm":  {"type": "number", "minimum": 1, "maximum": 500},
                            "height_cm": {"type": "number", "minimum": 1, "maximum": 500},
                            "confidence": {"type": "number", "minimum": 0, "maximum": 1},
                        },
                        "required": ["oem", "weight_kg", "length_cm",
                                      "width_cm", "height_cm"],
                    },
                },
            },
            "required": ["estimates"],
        },
        # Prompt caching: tool definition не меняется между batch'ами →
        # 5x дешевле + быстрее на повторных вызовах.
        "cache_control": {"type": "ephemeral"},
    }

    items_text = "\n".join(
        f"- OEM:{it.get('oem','')[:30]} | {it.get('title','')[:80]}"
        + (f" | {it.get('description','')[:60]}" if it.get('description') else "")
        for it in items
    )
    prompt = (
        "Оцени вес (кг) и габариты (см) каждой запчасти спецтехники "
        "по её названию и OEM-номеру. Используй здравый смысл и "
        "типичные значения для подобных деталей.\n\n"
        f"Запчасти ({len(items)}):\n{items_text}\n\n"
        "Используй tool estimate_dimensions для ответа. "
        "Верни оценку для КАЖДОЙ позиции."
    )

    try:
        result: dict[str, dict] = {}
        accumulated = ""
        last_seen_oems: set = set()

        with client.messages.stream(
            model=getattr(settings, "ANTHROPIC_FAST_MODEL", "claude-haiku-4-5-20251001"),
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}],
            tools=[estimate_tool],
            tool_choice={"type": "tool", "name": "estimate_dimensions"},
        ) as stream:
            for event in stream:
                # input_json_delta — частичный JSON tool input, приходит
                # по мере генерации Claude'ом. Аккумулируем и парсим.
                if event.type == "content_block_delta" and \
                   getattr(event.delta, "type", None) == "input_json_delta":
                    accumulated += event.delta.partial_json
                    # Эвристика: ищем закрытые объекты "{...}" с полями
                    # weight_kg/length_cm и парсим их по одному.
                    new_count = 0
                    while True:
                        obj_str, next_pos = _try_extract_next_estimate(accumulated)
                        if obj_str is None:
                            break
                        accumulated = accumulated[next_pos:]
                        try:
                            est = json.loads(obj_str)
                            oem = str(est.get("oem", "")).strip()
                            if oem and oem not in last_seen_oems:
                                result[oem] = {
                                    "weight_kg": float(est.get("weight_kg", 0) or 0),
                                    "length_cm": float(est.get("length_cm", 0) or 0),
                                    "width_cm":  float(est.get("width_cm", 0) or 0),
                                    "height_cm": float(est.get("height_cm", 0) or 0),
                                    "confidence": float(est.get("confidence", 0.8) or 0.8),
                                }
                                last_seen_oems.add(oem)
                                new_count += 1
                        except (json.JSONDecodeError, TypeError, ValueError):
                            pass
                    if new_count > 0 and on_partial:
                        try:
                            on_partial(dict(result))
                        except Exception:
                            pass

            # Финальный fallback — полный input из tool_use блока
            final_msg = stream.get_final_message()
            for block in final_msg.content:
                if block.type == "tool_use" and block.name == "estimate_dimensions":
                    for est in block.input.get("estimates", []):
                        oem = str(est.get("oem", "")).strip()
                        if oem and oem not in result:
                            result[oem] = {
                                "weight_kg": float(est.get("weight_kg", 0) or 0),
                                "length_cm": float(est.get("length_cm", 0) or 0),
                                "width_cm":  float(est.get("width_cm", 0) or 0),
                                "height_cm": float(est.get("height_cm", 0) or 0),
                                "confidence": float(est.get("confidence", 0.8) or 0.8),
                            }
        return result
    except Exception:
        logger.exception("AI per-part estimation failed")
        return {}


def _try_extract_next_estimate(buf: str) -> tuple[str | None, int]:
    """Ищет в buf первый закрытый JSON объект {..."oem":...,...}.

    Возвращает (obj_str, end_pos) где end_pos — позиция после `}`.
    Если объект не нашли — (None, 0).

    Простой brace-counter, игнорирует строки в кавычках с escape.
    """
    start = buf.find("{")
    if start < 0:
        return None, 0
    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(buf)):
        ch = buf[i]
        if escape:
            escape = False
            continue
        if ch == "\\":
            escape = True
            continue
        if ch == '"':
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return buf[start:i + 1], i + 1
    return None, 0


class PricelistAiEstimateView(APIView):
    """POST /api/assistant/upload-pricelist/<id>/ai-estimate/

    Запускает AI-оценку per-part вес/габариты для всех строк прайса
    батчами параллельно (5 workers). Сохраняет в imp.ai_estimates.
    При commit'е применяются как per-part overrides.
    """
    permission_classes = [IsAuthenticated]

    MAX_ROWS = 1000       # покрываем большинство реальных прайсов
    BATCH_SIZE = 10       # компромисс: плавный прогресс vs много API calls
    PARALLEL_WORKERS = 20  # параллельные Claude calls (rate limit 50 RPM)

    def post(self, request, import_id):
        from . import ai_credits as _aic
        if not _aic.rate_ok(request.user, "ai_estimate", 12, 3600):
            return Response({"error": _("Слишком частые AI-оценки прайса. Подождите немного.")}, status=429)
        from marketplace.models import PricelistImport
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)
        if imp.status != "preview":
            return Response({"error": f"already in status {imp.status}"}, status=400)

        mapping = imp.suggested_mapping or {}
        oem_col = mapping.get("oem_number")
        title_col = mapping.get("title")
        desc_col = mapping.get("description")
        if not oem_col or not title_col:
            return Response(
                {"error": _("Нужен маппинг oem_number и title для AI-оценки.")},
                status=400,
            )
        if oem_col.startswith("fix:") or title_col.startswith("fix:"):
            return Response(
                {"error": _("oem_number и title должны быть из файла, не fix.")},
                status=400,
            )
        if oem_col not in imp.headers or title_col not in imp.headers:
            return Response({"error": _("колонки не найдены в headers")}, status=400)
        oem_idx = imp.headers.index(oem_col)
        title_idx = imp.headers.index(title_col)
        desc_idx = (imp.headers.index(desc_col) if desc_col
                     and not desc_col.startswith("fix:")
                     and desc_col in imp.headers else None)

        try:
            with imp.file_obj.open("rb") as fh:
                blob = fh.read()
        except Exception as e:
            return Response({"error": f"file unavailable: {e}"}, status=500)

        items: list[dict] = []
        for row in _read_all(imp.filename, blob):
            if len(items) >= self.MAX_ROWS:
                break
            row_list = list(row)
            if oem_idx >= len(row_list) or title_idx >= len(row_list):
                continue
            oem = str(row_list[oem_idx] or "").strip()
            title = str(row_list[title_idx] or "").strip()
            if not oem or not title:
                continue
            item = {"oem": oem, "title": title}
            if desc_idx is not None and desc_idx < len(row_list):
                d = str(row_list[desc_idx] or "").strip()
                if d:
                    item["description"] = d
            items.append(item)

        if not items:
            return Response({"error": _("Не нашёл валидных строк для оценки.")}, status=400)

        if not getattr(settings, "ANTHROPIC_API_KEY", ""):
            return Response({
                "error": _("AI-оценка недоступна (ANTHROPIC_API_KEY не настроен на сервере)."),
            }, status=503)

        chunks = [items[i:i + self.BATCH_SIZE]
                   for i in range(0, len(items), self.BATCH_SIZE)]
        total = len(items)
        truncated = len(items) >= self.MAX_ROWS

        from concurrent.futures import ThreadPoolExecutor, as_completed

        from django.core.cache import cache
        progress_key = f"ai_estimate_progress_{imp.id}"
        cache.set(progress_key, {"current": 0, "total": total, "running": True}, 300)

        # Глобальный shared dict — обновляется streaming callback'ами
        # из разных потоков. OEM уникальны между batch'ами.
        from threading import Lock
        live_estimates: dict[str, dict] = {}
        lock = Lock()

        def on_partial(batch_partial):
            with lock:
                live_estimates.update(batch_partial)
                cache.set(progress_key, {
                    "current": len(live_estimates),
                    "total": total, "running": True,
                }, 300)

        estimates: dict[str, dict] = {}
        try:
            with ThreadPoolExecutor(max_workers=self.PARALLEL_WORKERS) as ex:
                futures = [ex.submit(_ai_estimate_per_part, c, on_partial)
                            for c in chunks]
                for fut in as_completed(futures):
                    try:
                        with lock:
                            estimates.update(fut.result() or {})
                            live_estimates.update(estimates)
                    except Exception:
                        logger.exception("AI estimate batch failed")
                    cache.set(progress_key, {
                        "current": len(live_estimates),
                        "total": total, "running": True,
                    }, 300)
                    imp.ai_estimates = dict(estimates)
                    imp.save(update_fields=["ai_estimates"])
        finally:
            cache.set(progress_key, {
                "current": len(estimates), "total": total, "running": False,
            }, 300)

        # Sample для review — приоритет позициям с низкой confidence,
        # чтобы юзер мог проверить именно их (Claude-style human-in-loop)
        items_by_oem = {it["oem"]: it for it in items}
        scored = []
        for oem, est in estimates.items():
            conf = est.get("confidence", 0.8)
            title = items_by_oem.get(oem, {}).get("title", "")
            scored.append((conf, oem, title, est))
        # Сортируем: сначала с низкой confidence — туда смотреть в первую очередь
        scored.sort(key=lambda x: x[0])
        review_sample = []
        for conf, oem, title, est in scored[:8]:
            review_sample.append({
                "oem": oem, "title": title,
                "weight_kg": est["weight_kg"],
                "length_cm": est["length_cm"],
                "width_cm": est["width_cm"],
                "height_cm": est["height_cm"],
                "confidence": conf,
            })
        low_conf_count = sum(1 for c, _, _, _ in scored if c < 0.6)

        return Response({
            "ok": True,
            "estimated": len(estimates),
            "total": total,
            "truncated": truncated,
            "review_sample": review_sample,
            "low_confidence_count": low_conf_count,
        })


class PricelistImportProgressView(APIView):
    """GET /api/assistant/upload-pricelist/<id>/import-progress/

    Возвращает текущий прогресс импорта (commit) для polling из UI.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, import_id):
        # Общий Redis-кэш: коммит исполняется в Celery-воркере (др. процесс),
        # прогресс/результат там же — LocMem (per-process) их бы не связал.
        from django.core.cache import caches

        from marketplace.models import PricelistImport
        cache = caches["pricelist"]
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)
        progress = cache.get(f"import_progress_{imp.id}") or {}
        out = {
            "current": progress.get("current", 0),
            "total": progress.get("total", 0),
            "phase": progress.get("phase", ""),
            "running": progress.get("running", False),
            "status": imp.status,
        }
        # Финальный результат фонового импорта (поток коммита кладёт его в кэш
        # по завершении). Фронт ждёт именно его, т.к. сам commit отвечает 202.
        result = cache.get(f"import_result_{imp.id}")
        if result is not None:
            out["done"] = True
            out["result"] = result
        elif imp.status in ("imported", "failed"):
            # Кэш мог истечь (TTL/рестарт) — отдаём результат из БД (fallback).
            out["done"] = True
            out["result"] = {
                "ok": imp.status == "imported",
                "import_id": imp.id,
                "imported": imp.imported_rows,
                "created": imp.created_rows,
                "updated": imp.updated_rows,
                "failed": imp.failed_rows,
                "errors_preview": (imp.error_details or [])[:10],
            }
        return Response(out)


class PricelistGenerateOutputProgressView(APIView):
    """GET /api/assistant/upload-pricelist/<id>/generate-output-progress/

    Polling endpoint для прогресс-бара генерации output XLSX.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, import_id):
        from django.core.cache import cache
        progress = cache.get(f"generate_output_progress_{import_id}") or {}
        return Response({
            "current": progress.get("current", 0),
            "total": progress.get("total", 0),
            "running": progress.get("running", False),
        })


class PricelistOutputPreviewView(APIView):
    """GET /api/assistant/upload-pricelist/<id>/output-preview/

    Возвращает HTML preview сгенерированного output_file (XLSX).
    Используется для side panel в чате (как у claude.ai).
    """
    permission_classes = [IsAuthenticated]

    PREVIEW_ROWS = 100  # ограничиваем превью

    def get(self, request, import_id):
        from django.http import HttpResponse

        from marketplace.models import PricelistImport
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)
        if not imp.output_file:
            return Response({"error": "output file not generated"}, status=404)

        # Если есть кэш HTML preview — отдаём моментально (10-100 мс
        # вместо 15+ сек openpyxl re-parse 36MB файла).
        if imp.output_preview_html:
            # Перевод применяется и к старому кэшу (генерация была до фикса
            # _translate_to_en). _looks_non_ascii fast-path делает это
            # бесплатным для уже-английского HTML.
            body = _translate_to_en(imp.output_preview_html)
            html = (
                '<!doctype html><html><head><meta charset="utf-8"></head>'
                '<body>' + body + '</body></html>'
            )
            return HttpResponse(html, content_type="text/html; charset=utf-8")

        from openpyxl import load_workbook
        try:
            with imp.output_file.open("rb") as fh:
                wb = load_workbook(io.BytesIO(fh.read()), read_only=True, data_only=True)
            ws = wb.worksheets[0]
        except Exception as e:
            return Response({"error": f"cannot read XLSX: {e}"}, status=500)

        rows_html = []
        total = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > self.PREVIEW_ROWS + 1:
                # просто считаем общее количество
                total = row_idx - 1
                continue
            cells = []
            for col_idx, val in enumerate(row):
                v = "" if val is None else str(val)
                if len(v) > 50:
                    v = v[:50] + "…"
                tag = "th" if row_idx == 1 else "td"
                cls = ' class="opx-num"' if row_idx == 1 else ''
                cells.append(f"<{tag}{cls}>{v}</{tag}>")
            rows_html.append(f"<tr>{''.join(cells)}</tr>")
        if total == 0:
            total = ws.max_row - 1 if ws.max_row else 0
        try:
            wb.close()
        except Exception:
            pass

        truncated = total > self.PREVIEW_ROWS
        more_note = (_('<div class="opx-note">Показаны первые %(rows)s из %(total)s строк</div>') % {
            "rows": self.PREVIEW_ROWS, "total": total}) if truncated else ""

        html = (
            '<!doctype html><html><head><meta charset="utf-8">'
            '<style>'
            'body{margin:0;padding:14px;font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:#fff;}'
            '.opx-note{font-size:11.5px;color:rgba(0,0,0,0.55);margin-bottom:10px;font-style:italic;}'
            'table{border-collapse:collapse;width:100%;font-size:11.5px;font-family:"SF Mono",Menlo,monospace;}'
            'th{background:#2D7A3E;color:#fff;padding:6px 8px;text-align:left;position:sticky;top:0;font-weight:600;letter-spacing:0.02em;}'
            'td{padding:4px 8px;border-bottom:1px solid rgba(0,0,0,0.05);color:#1a1a1a;white-space:nowrap;}'
            'tr:nth-child(even) td{background:rgba(45,122,62,0.03);}'
            '</style></head><body>'
            + more_note
            + '<table><thead>' + (rows_html[0] if rows_html else "") + '</thead>'
            + '<tbody>' + ''.join(rows_html[1:]) + '</tbody>'
            + '</table></body></html>'
        )
        return HttpResponse(html, content_type="text/html; charset=utf-8")


class PricelistGenerateOutputView(APIView):
    """POST /api/assistant/upload-pricelist/<id>/generate-output/

    Генерирует XLSX в формате маркетплейса (как у claude.ai).
    Body: {mapping, transform_rules, constants, ai_estimates_override}
    Returns: {filename, size, download_url}
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, import_id):
        from django.core.files.base import ContentFile

        from marketplace.models import PricelistImport
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)

        mapping = request.data.get("mapping") or imp.suggested_mapping or {}
        transform_rules = request.data.get("transform_rules") or {}
        constants = request.data.get("constants") or {}
        ai_overrides = request.data.get("ai_estimates_override") or {}

        # Применяем ai_overrides поверх ai_estimates
        if isinstance(ai_overrides, dict) and ai_overrides:
            current = imp.ai_estimates or {}
            for oem, fields in ai_overrides.items():
                if not isinstance(fields, dict):
                    continue
                cur = current.get(oem, {})
                for k in ("weight_kg", "length_cm", "width_cm", "height_cm"):
                    if k in fields:
                        try:
                            cur[k] = float(fields[k])
                        except (TypeError, ValueError):
                            pass
                cur["confidence"] = 1.0
                current[oem] = cur
            imp.ai_estimates = current

        try:
            xlsx_bytes = _generate_marketplace_xlsx(
                imp, mapping, transform_rules, constants,
            )
        except Exception as e:
            logger.exception("XLSX generation failed")
            return Response({"error": f"Не удалось сгенерировать XLSX: {e}"}, status=500)

        # Сохраняем в FileField
        base_name = (imp.filename or "pricelist").rsplit(".", 1)[0]
        out_name = f"{base_name}_marketplace.xlsx"
        imp.output_file.save(out_name, ContentFile(xlsx_bytes), save=True)

        return Response({
            "filename": out_name,
            "size": len(xlsx_bytes),
            "download_url": imp.output_file.url if imp.output_file else "",
            "rows_generated": xlsx_bytes.count(b"<row") if False else None,
        })


class PricelistSmartQuestionsView(APIView):
    """GET /api/assistant/upload-pricelist/<id>/smart-questions/

    Async-генерация Claude.ai-style questionnaire. Вызывается фронтом
    сразу после upload параллельно с показом базовой формы — чтобы
    upload-ответ не блокировался на 4 секундах Claude API.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, import_id):
        from . import ai_credits as _aic
        if not _aic.rate_ok(request.user, "smart_questions", 40, 3600):
            return Response({"intro": "", "questions": []}, status=200)
        from marketplace.models import PricelistImport
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)

        smart = _ai_smart_questions(
            imp.headers, imp.sample_rows, None,
            imp.suggested_mapping or {}, imp.filename,
            seller=request.user,
        )
        return Response({
            "intro": smart.get("intro", ""),
            "questions": smart.get("questions", []),
        })


class PricelistAiEstimateProgressView(APIView):
    """GET /api/assistant/upload-pricelist/<id>/ai-estimate-progress/

    Возвращает текущий прогресс AI-оценки для polling из UI.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, import_id):
        from django.core.cache import cache

        from marketplace.models import PricelistImport
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)
        progress = cache.get(f"ai_estimate_progress_{imp.id}") or {}
        return Response({
            "current": progress.get("current", 0),
            "total": progress.get("total", 0),
            "running": progress.get("running", False),
            "estimated_in_db": len(imp.ai_estimates or {}),
        })


class PricelistTemplateXlsxView(APIView):
    """GET /api/assistant/pricelist-template.xlsx — Excel-шаблон с
    инструкцией на отдельном листе + 16 колонок и 3 примера строк.
    """
    from rest_framework.permissions import AllowAny
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        from django.conf import settings
        from django.http import FileResponse, Http404
        path = os.path.join(
            str(settings.BASE_DIR),
            "static", "templates", "consolidator_pricelist_template.xlsx",
        )
        if not os.path.exists(path):
            raise Http404("Template not found")
        resp = FileResponse(
            open(path, "rb"),
            content_type=("application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet"),
        )
        resp["Content-Disposition"] = (
            'attachment; filename="consolidator_pricelist_template.xlsx"'
        )
        return resp


class PricelistTemplateView(APIView):
    """GET /api/assistant/pricelist-template.csv — шаблон 16 колонок."""
    from rest_framework.permissions import AllowAny
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        from django.http import HttpResponse
        rows = [
            ["PartNumber", "CrossNumber", "Brand", "Name", "Quantity",
             "Condition", "Price_EXW", "WarehouseAddress",
             "Price_FOB_SEA", "Price_FOB_AIR", "SeaPort", "AirPort",
             "Weight", "Length", "Width", "Height"],
            ["561-50-82311", "5615082311", "Komatsu", "BUSHING", "8",
             "ORIGINAL", "100.00", "Shanghai CN",
             "120.00", "150.00", "Yangshan Port", "Pudong Airport",
             "0.5", "10", "5", "5"],
            ["585-33-21240", "5853321240", "Komatsu", "DISC", "14",
             "OEM", "120.00", "Shanghai CN",
             "140.00", "170.00", "Yangshan Port", "Pudong Airport",
             "2.0", "15", "10", "3"],
        ]
        out = io.StringIO()
        writer = csv.writer(out, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)
        resp = HttpResponse(out.getvalue().encode("utf-8-sig"),  # BOM для Excel
                              content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = (
            'attachment; filename="consolidator_pricelist_template.csv"'
        )
        return resp


class PricelistCancelView(APIView):
    """POST /api/assistant/upload-pricelist/<id>/cancel/  — отменить превью."""
    permission_classes = [IsAuthenticated]

    def post(self, request, import_id):
        from marketplace.models import PricelistImport
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "import not found"}, status=404)
        if imp.status == "preview":
            imp.status = "cancelled"
            imp.completed_at = timezone.now()
            imp.save(update_fields=["status", "completed_at"])
        return Response({"ok": True, "status": imp.status})


# ── Action для отображения карточки маппинга в чате ───────────────

@register("pricelist_show_errors")
def pricelist_show_errors(params, user, role):
    """Показать список ошибок последнего импорта."""
    from marketplace.models import PricelistImport
    try:
        imp = PricelistImport.objects.get(
            id=int(params.get("import_id") or 0), seller=user,
        )
    except (PricelistImport.DoesNotExist, ValueError, TypeError):
        return ActionResult(text=_("Импорт не найден."))
    if not imp.error_details:
        return ActionResult(text=_("Ошибок нет 🎉"))
    # Понятный комментарий вместо технического reason (bad price_exw и т.п.).
    # comment пишется при импорте; для старых записей — fallback по reason.
    _REASON_HINT = {
        "bad price_exw": _("Цена пустая, 0 или не число — позиция пропущена"),
        "no oem": _("Не указан артикул (PartNumber) — позиция пропущена"),
        "no title": _("Не указано название (Description) — позиция пропущена"),
        "price_conflict": _("Артикул-дубль с разными ценами"),
        "duplicate": _("Дубль строки"),
        "no price": _("Не указана цена — позиция пропущена"),
    }
    def _reason_text(e):
        if e.get("comment"):
            return e["comment"]
        r = e.get("reason", "")
        return _REASON_HINT.get(r) or (
            _("Строка пропущена: %(reason)s") % {"reason": r} if r else _("Строка пропущена"))
    # Таблица: Строка | Артикул | Колонка | Значение | Проблема.
    # Для старых записей без структурных полей — fallback из comment.
    table_rows = []
    for e in imp.error_details:
        col = e.get("column", "")
        val = e.get("value", "")
        prob = e.get("problem") or _reason_text(e)
        if not col and not prob:  # совсем старый формат
            prob = _reason_text(e)
        table_rows.append([
            str(e.get("row", "?")),
            (e.get("oem") or "—")[:24],
            col or "—",
            (str(val) if val != "" else "—"),
            prob,
        ])
    # Группируем по типу для краткой сводки «что делать».
    from collections import Counter
    by_reason = Counter(e.get("reason", "?") for e in imp.error_details)
    hint_lines = []
    if by_reason.get("bad price_exw"):
        hint_lines.append(f"• {by_reason['bad price_exw']} с проблемной ценой → впишите цену (или удалите строку, если позиции нет)")
    if by_reason.get("no oem"):
        hint_lines.append(f"• {by_reason['no oem']} без артикула → заполните PartNumber")
    if by_reason.get("no title"):
        hint_lines.append(f"• {by_reason['no title']} без названия → заполните Description")
    how_to = ("Как исправить:\n" + "\n".join(hint_lines)
              + "\n\nИсправьте эти строки в своём файле и загрузите заново — "
              + "обновление не задвоит уже загруженные позиции.") if hint_lines else ""
    return ActionResult(
        text=_('❌ Ошибки импорта #%(p0)s: %(p1)s строк не загружено (остальные — в каталоге).\n\n%(p2)s') % {"p0": f'{imp.id}', "p1": f'{imp.failed_rows}', "p2": f'{how_to}'},
        cards=[{"type": "table_preview", "data": {
            "title": _('Ошибки импорта (показаны первые %(p0)s)') % {"p0": f'{len(table_rows)}'},
            "headers": ["Строка", "Артикул", "Колонка", "Значение", "Проблема"],
            "rows": table_rows,
        }}],
        actions=[
            {"action": "__download_url", "label": _("📥 Скачать список ошибок (.csv)"),
             "params": {"url": f"/api/assistant/upload-pricelist/{imp.id}/errors.csv"}},
            {"action": "upload_pricelist", "label": _("📤 Загрузить исправленный файл"),
             "params": {}},
            {"action": "seller_warehouses", "label": _("📦 Мои товары"),
             "params": {}},
        ],
        # Экран исправления ошибок — не место для рекомендаций «Создать RFQ /
        # Аналитика»: подавляем дефолтные подсказки фронта.
        no_suggestions=True,
    )


class PricelistErrorsCsvView(APIView):
    """GET /api/assistant/upload-pricelist/<id>/errors.csv

    Скачать список проблемных строк (строка, артикул, что не так) — чтобы
    поставщик исправил их в своём файле и загрузил заново.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, import_id):
        from django.http import HttpResponse
        from marketplace.models import PricelistImport
        try:
            imp = PricelistImport.objects.get(id=import_id, seller=request.user)
        except PricelistImport.DoesNotExist:
            return Response({"error": "not found"}, status=404)
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["Строка", "Артикул", "Проблема"])
        for e in (imp.error_details or []):
            w.writerow([e.get("row", ""), e.get("oem", ""),
                        e.get("comment") or e.get("reason", "")])
        resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="errors_{imp.id}.csv"'
        return resp


@register("pricelist_history")
def pricelist_history(params, user, role):
    """История загрузок прайса этого seller'а."""
    from marketplace.models import PricelistImport
    items = PricelistImport.objects.filter(seller=user).order_by("-created_at")[:10]
    if not items:
        return ActionResult(
            text=_("Прайс ещё не загружали."),
            actions=[{"action": "upload_pricelist", "label": _("📤 Загрузить"),
                      "params": {}}],
        )
    rows = [{
        "label": f"#{i.id} · {i.filename[:40]} · {i.created_at.strftime('%d.%m %H:%M')}",
        "value": (
            f"{i.imported_rows}↑ / {i.failed_rows}✗"
            if i.status == "imported" else i.get_status_display()
        ),
    } for i in items]
    return ActionResult(
        text=_('📋 История загрузок (%(p0)s)') % {"p0": f'{len(items)}'},
        cards=[{"type": "draft", "data": {
            "title": _("История импортов прайса"), "rows": rows,
        }}],
        actions=[{"action": "upload_pricelist", "label": _("📤 Новая загрузка"),
                  "params": {}}],
    )


# ─────────────────────────────────────────────────────────────────────
# Geo-поиск городов для автокомплита склада отгрузки.
# База — GeoNames cities15000 (≈32k городов, население >15000), сгруппирована
# по ISO2-коду страны. Каждая запись: [display_latin, search_blob]. search_blob
# содержит name + asciiname + все латинские/кириллические alternatenames в
# нижнем регистре → поиск работает и на латинице, и на кириллице ("пек"→Beijing).
# Грузится в память один раз (lazy), отдаётся через endpoint топ-N по подстроке.
# ─────────────────────────────────────────────────────────────────────
_CITIES_DB = None


def _load_cities_db():
    global _CITIES_DB
    if _CITIES_DB is None:
        path = os.path.join(os.path.dirname(__file__), "data", "cities_db.json")
        try:
            with open(path, encoding="utf-8") as f:
                _CITIES_DB = json.load(f)
        except Exception:
            logger.exception("cities_db.json load failed")
            _CITIES_DB = {}
    return _CITIES_DB


# Реальные карго-порты по странам (морпорты UN/LOCODE func=port + аэропорты
# OurAirports large/medium с IATA). Формат: {cc: {sea:[[code,name,lat,lng]],
# air:[[iata,name,lat,lng]]}}. Endpoint сортирует по близости к складу.
_PORTS_DB = None


def _load_ports_db():
    global _PORTS_DB
    if _PORTS_DB is None:
        path = os.path.join(os.path.dirname(__file__), "data", "ports_db.json")
        try:
            with open(path, encoding="utf-8") as f:
                _PORTS_DB = json.load(f)
        except Exception:
            logger.exception("ports_db.json load failed")
            _PORTS_DB = {}
    return _PORTS_DB


def _haversine_km(lat1, lon1, lat2, lon2):
    import math
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
         * math.sin(dlon / 2) ** 2)
    return 2 * R * math.asin(min(1.0, math.sqrt(a)))


# Транслитерация кириллица→латиница (BGN/PCGN-подобная). Нужна потому что
# в GeoNames у многих городов нет русских названий, только латиница
# («Ürgüp»: urgup). Юзер вводит «ургюп» → транслитерируем в «urgyup» →
# fuzzy-матч с «urgup». Закрывает города без русского варианта в базе.
_TRANSLIT = {
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e','ж':'zh','з':'z',
    'и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r',
    'с':'s','т':'t','у':'u','ф':'f','х':'kh','ц':'ts','ч':'ch','ш':'sh',
    'щ':'sch','ъ':'','ы':'y','ь':'','э':'e','ю':'yu','я':'ya',
}


def _translit_ru(s):
    """Кириллица → латиница. Не-кириллические символы оставляем как есть."""
    return ''.join(_TRANSLIT.get(ch, ch) for ch in s.lower())


def _fuzzy_city_match(q, rows, limit=12):
    """Fuzzy-поиск опечаток (как у крупных сервисов): «масква»→Moscow,
    «ургюп»→Ürgüp. Сравнивает запрос И его транслитерацию с токенами
    search-индекса через SequenceMatcher, берёт лучший ratio на город.
    Offline, надёжно, без внешних API. Только в пределах страны (быстро).
    """
    from difflib import SequenceMatcher
    qt = _translit_ru(q)               # латинская транслитерация запроса
    variants = [q] if qt == q else [q, qt]
    scored = []
    for entry in rows:
        disp, search = entry[0], entry[1]
        best = 0.0
        for tok in search.split():
            if not tok:
                continue
            for v in variants:
                r = SequenceMatcher(None, v, tok).ratio()
                if r > best:
                    best = r
            if best >= 0.95:
                break
        if best >= 0.7:
            scored.append((best, disp))
    scored.sort(key=lambda x: -x[0])
    return [d for _, d in scored[:limit]]


class GeoCitiesView(APIView):
    """GET /api/assistant/geo-cities/?q=<query>&cc=<ISO2>

    Подсказки городов для поля «Город» склада отгрузки (≈32k городов мира,
    GeoNames). Поиск двуязычный (рус+лат) по подстроке. Если точных мало —
    добавляем fuzzy-поиск опечаток в пределах страны. Кэш 10 мин.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.core.cache import cache
        q = (request.GET.get("q") or "").strip().lower()
        cc = (request.GET.get("cc") or "").strip().upper()
        ckey = f"geocities:{cc}:{q}"
        cached = cache.get(ckey)
        if cached is not None:
            return Response({"cities": cached})

        db = _load_cities_db()
        if cc and cc in db:
            pool = [(cc, db[cc])]
        elif cc:
            pool = []
        else:
            pool = list(db.items())

        # Каждая запись базы: [display, search, lat, lng]. Возвращаем
        # объекты {n, lat, lng} — координаты нужны фронту для сортировки
        # портов по реальному гео-расстоянию от склада.
        qt = _translit_ru(q) if q else q
        results = []
        seen = set()

        def _coords(entry):
            la = entry[2] if len(entry) > 2 else None
            lo = entry[3] if len(entry) > 3 else None
            return {"n": entry[0], "lat": la, "lng": lo}

        # 1) substring — точные совпадения (по запросу или транслитерации)
        for _cc, rows in pool:
            for entry in rows:
                disp, search = entry[0], entry[1]
                if not q or q in search or (qt != q and qt in search):
                    if disp not in seen:
                        seen.add(disp); results.append(_coords(entry))
                        if len(results) >= 25:
                            break
            if len(results) >= 25:
                break

        # 2) fuzzy — опечатки, если точных мало и страна задана
        if q and len(results) < 8 and cc and cc in db:
            coord_map = {e[0]: e for e in db[cc]}
            for d in _fuzzy_city_match(q, db[cc]):
                if d not in seen:
                    seen.add(d)
                    results.append(_coords(coord_map.get(d, [d])))
                    if len(results) >= 25:
                        break

        cache.set(ckey, results, 600)
        return Response({"cities": results})


class GeoPortsView(APIView):
    """GET /api/assistant/geo-ports/?cc=<ISO2>&kind=sea|air&lat=&lng=&q=

    Реальные карго-порты страны (морпорты UN/LOCODE / аэропорты-хабы
    OurAirports). Если переданы координаты склада (lat,lng) — сортируем по
    реальному расстоянию (ближайший сверху: Шэньян→Далянь). Опц. q —
    подстрочный фильтр по коду/названию. Отдаём до 40 ближайших.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cc = (request.GET.get("cc") or "").strip().upper()
        kind = (request.GET.get("kind") or "sea").strip().lower()
        if kind not in ("sea", "air"):
            kind = "sea"
        q = (request.GET.get("q") or "").strip().lower()
        try:
            wlat = float(request.GET.get("lat"))
            wlng = float(request.GET.get("lng"))
        except (TypeError, ValueError):
            wlat = wlng = None

        db = _load_ports_db()
        rows = (db.get(cc, {}) or {}).get(kind, []) if cc else []
        # каждая запись: [code, name, lat, lng]
        items = []
        for code, name, la, lo in rows:
            label = code + " · " + name
            if q and q not in code.lower() and q not in name.lower():
                continue
            dist = None
            if wlat is not None and la is not None:
                dist = _haversine_km(wlat, wlng, la, lo)
            items.append({"code": code, "name": name, "label": label, "dist": dist})

        if wlat is not None:
            items.sort(key=lambda x: (x["dist"] is None, x["dist"] or 1e9))
        return Response({"ports": items[:40]})
