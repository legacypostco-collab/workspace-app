"""File upload endpoint: спецификации (Excel/CSV/PDF) → список артикулов → search_parts.

Использует openpyxl (Excel), стандартный csv, и опциональный pypdf для PDF.
Если pypdf не установлен — PDF не парсится, возвращается понятная ошибка.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
import os
import re
import zipfile

from django.core.cache import cache
from django.shortcuts import get_object_or_404
from django.utils.translation import gettext as _
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Conversation
from .permissions import detect_user_role
from .rag import execute_action

logger = logging.getLogger(__name__)

MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_ENTRIES = 2_000


def _verify_magic_bytes(kind: str, blob: bytes) -> bool:
    """Проверка signature в первых байтах файла.

    Защищает от polyglot-атак: переименованный EXE/HTML с расширением xlsx/csv/pdf.
    Возвращает True если первые байты соответствуют ожидаемому типу.

    Сигнатуры:
      xlsx (modern Office ZIP):           PK\\x03\\x04
      pdf:                                %PDF
      csv: пропускаем (это plain text, нет уникальной сигнатуры) — но
           проверяем что нет MZ/ELF/PK сигнатур в начале (anti-EXE).
    """
    if not blob or len(blob) < 4:
        return False
    head = blob[:8]
    if kind == "xlsx":
        return head[:4] == b"PK\x03\x04"
    if kind == "pdf":
        return head[:4] == b"%PDF"
    if kind == "csv":
        # CSV это text/plain — нет magic-bytes. Но запрещаем известные binary-сигнатуры
        # (anti-EXE/dll/elf/zip/pdf под маской csv).
        forbidden_signatures = (
            b"MZ",                  # PE/EXE/DLL
            b"\x7fELF",             # ELF (Linux)
            b"\xCA\xFE\xBA\xBE",    # Mach-O fat / Java class
            b"\xCE\xFA\xED\xFE",    # Mach-O 32
            b"\xCF\xFA\xED\xFE",    # Mach-O 64
            b"PK\x03\x04",          # ZIP/XLSX/DOCX
            b"%PDF",                # PDF
            b"\xD0\xCF\x11\xE0",    # OLE2 (legacy office)
        )
        return not any(head.startswith(sig) for sig in forbidden_signatures)
    return False
MAX_ARTICLES = 200

_OEM_RE = re.compile(r"^[A-ZА-Я0-9][A-ZА-Я0-9\-./]{2,}$", re.IGNORECASE)


def _looks_like_article(token: str) -> bool:
    token = (token or "").strip().strip(".").strip()
    if not token or len(token) < 3 or len(token) > 40:
        return False
    if not _OEM_RE.match(token):
        return False
    if not any(ch.isdigit() for ch in token):
        return False
    return True


def _extract_from_text(text: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in re.split(r"[\s,;|]+", text or ""):
        token = raw.strip().strip(".").strip()
        if _looks_like_article(token) and token.upper() not in seen:
            seen.add(token.upper())
            out.append(token)
            if len(out) >= MAX_ARTICLES:
                break
    return out


def _extract_from_xlsx(blob: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    if blob.startswith(b"PK\x03\x04"):
        try:
            with zipfile.ZipFile(io.BytesIO(blob)) as archive:
                entries = archive.infolist()
                if len(entries) > MAX_XLSX_ENTRIES:
                    return []
                expanded_size = sum(entry.file_size for entry in entries)
                if expanded_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                    return []
        except (OSError, zipfile.BadZipFile):
            return []
    try:
        wb = load_workbook(
            io.BytesIO(blob),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception:
        return []
    try:
        out: list[str] = []
        seen: set[str] = set()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(
                max_row=min(sheet.max_row or 0, 5000),
                max_col=min(sheet.max_column or 0, 100),
                values_only=True,
            ):
                for cell in row:
                    if cell is None:
                        continue
                    token = str(cell).strip()
                    if _looks_like_article(token) and token.upper() not in seen:
                        seen.add(token.upper())
                        out.append(token)
                        if len(out) >= MAX_ARTICLES:
                            return out
        return out
    finally:
        wb.close()


def _extract_from_csv(blob: bytes) -> list[str]:
    try:
        text = blob.decode("utf-8-sig", errors="replace")
    except Exception:
        text = blob.decode("latin-1", errors="replace")
    out: list[str] = []
    seen: set[str] = set()
    for delim in (",", ";", "\t"):
        try:
            reader = csv.reader(io.StringIO(text), delimiter=delim)
            has_columns = False
            candidate_rows = []
            for row_no, row in enumerate(reader):
                if row_no < 5:
                    candidate_rows.append(row)
                    has_columns = has_columns or len(row) > 1
                    continue
                if not has_columns:
                    break
                for cell in row:
                    token = (cell or "").strip()
                    if _looks_like_article(token) and token.upper() not in seen:
                        seen.add(token.upper())
                        out.append(token)
                        if len(out) >= MAX_ARTICLES:
                            return out
            if has_columns:
                for row in candidate_rows:
                    for cell in row:
                        token = (cell or "").strip()
                        if _looks_like_article(token) and token.upper() not in seen:
                            seen.add(token.upper())
                            out.append(token)
                            if len(out) >= MAX_ARTICLES:
                                return out
                if out:
                    return out
        except csv.Error:
            continue
    # Fallback: token-by-token
    return _extract_from_text(text)


def _extract_from_pdf(blob: bytes) -> list[str]:
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except ImportError:
            return []
    try:
        reader = PdfReader(io.BytesIO(blob))
    except Exception:
        return []
    chunks: list[str] = []
    for page in reader.pages[:50]:
        try:
            chunks.append(page.extract_text() or "")
        except Exception:
            continue
    return _extract_from_text("\n".join(chunks))


class RecognizePhotoView(APIView):
    """POST /api/assistant/recognize-photo/   (multipart, field "photo")

    Принимает фото шильды/детали → пытается извлечь артикул, бренд, модель.
    Использует Anthropic Claude vision (claude-haiku-4-5 поддерживает images).
    Без API-ключа возвращает понятный fallback.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        from . import ai_credits as _aic
        if not _aic.rate_ok(request.user, "recognize_photo", 40, 3600):
            return Response({"error": _("Слишком частое распознавание фото. Подождите немного.")}, status=429)
        photo = request.FILES.get("photo")
        if not photo:
            return Response({"error": "photo is required"}, status=400)
        if photo.size > 10 * 1024 * 1024:
            return Response({"error": _("файл > 10 МБ")}, status=400)
        try:
            from marketplace.upload_security import validate_uploaded_file

            ext = validate_uploaded_file(
                photo,
                allowed_ext={".jpg", ".jpeg", ".png", ".webp"},
                max_bytes=10 * 1024 * 1024,
            )
        except ValueError as exc:
            return Response({"error": str(exc)}, status=400)

        api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
        if not api_key:
            return Response({
                "text": "",
                "error": _("Распознавание фото не настроено (нет ANTHROPIC_API_KEY). "
                           "Опишите деталь словами или загрузите Excel/CSV со списком."),
            }, status=200)
        allowed, _remaining = _aic.try_consume(
            request.user,
            detect_user_role(request.user, request=request),
        )
        if not allowed:
            return Response(
                {"error": _("Лимит запросов к распознаванию исчерпан.")},
                status=429,
            )

        try:
            import base64

            import anthropic
            blob = photo.read()
            b64 = base64.b64encode(blob).decode("ascii")
            media_type = {
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".png": "image/png",
                ".webp": "image/webp",
            }[ext]
            client = anthropic.Anthropic(api_key=api_key)
            resp = client.messages.create(
                model=os.getenv("ANTHROPIC_VISION_MODEL", "claude-haiku-4-5-20251001"),
                max_tokens=600,
                system=(
                    "Ты — AI-снабженец. Получаешь фото шильды или детали техники. "
                    "Извлеки: бренд (Komatsu/CAT/Volvo/...), модель техники, "
                    "артикул/part number, серийный номер, любые видимые "
                    "технические параметры. Отвечай в формате JSON: "
                    '{"brand":"...","model":"...","part_number":"...",'
                    '"serial":"...","notes":"кратко"}. '
                    "Если ничего не видно — верни {} с notes о причине."
                ),
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {
                            "type": "base64", "media_type": media_type, "data": b64,
                        }},
                        {"type": "text", "text": _("Распознай шильду / деталь.")},
                    ],
                }],
            )
            text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()
            return Response({"text": text})
        except Exception:
            logger.exception("photo recognition failed user_id=%s", request.user.pk)
            return Response(
                {"error": _("Не удалось распознать изображение. Повторите позже.")},
                status=502,
            )


class TranscribeAudioView(APIView):
    """POST /api/assistant/transcribe-audio/   (multipart, field "audio")

    Принимает аудио-blob от MediaRecorder. Возвращает текст расшифровки.
    Если есть OPENAI_API_KEY — использует Whisper API. Иначе — fallback
    через Web Speech API на клиенте (этот endpoint просто скажет,
    что серверная расшифровка не настроена).
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        from . import ai_credits as _aic

        if not _aic.rate_ok(request.user, "transcribe_audio", 20, 3600):
            return Response(
                {"error": _("Слишком частые запросы на расшифровку. Повторите позже.")},
                status=429,
            )
        audio = request.FILES.get("audio")
        if not audio:
            return Response({"error": "audio is required"}, status=400)
        if audio.size > 20 * 1024 * 1024:
            return Response({"error": _("файл > 20 МБ")}, status=400)
        try:
            from marketplace.upload_security import validate_uploaded_file

            validate_uploaded_file(
                audio,
                allowed_ext={".mp3", ".mp4", ".mpeg", ".mpga", ".m4a", ".wav", ".webm"},
                max_bytes=20 * 1024 * 1024,
            )
        except ValueError as exc:
            return Response({"error": str(exc)}, status=400)

        api_key = os.getenv("OPENAI_API_KEY", "").strip()
        if not api_key:
            return Response({
                "text": "",
                "error": _("Серверная расшифровка не настроена (нет OPENAI_API_KEY). Используется встроенный Web Speech API в браузере."),
            }, status=200)
        allowed, _remaining = _aic.try_consume(
            request.user,
            detect_user_role(request.user, request=request),
        )
        if not allowed:
            return Response(
                {"error": _("Лимит запросов к расшифровке исчерпан.")},
                status=429,
            )

        # Реальный вызов Whisper
        try:
            import requests
            files = {"file": (audio.name or "audio.webm", audio.read(), audio.content_type or "audio/webm")}
            data = {"model": "whisper-1", "language": "ru"}
            r = requests.post(
                "https://api.openai.com/v1/audio/transcriptions",
                headers={"Authorization": f"Bearer {api_key}"},
                files=files, data=data, timeout=60, allow_redirects=False,
            )
            if not r.ok:
                logger.warning(
                    "audio transcription provider error status=%s user_id=%s",
                    r.status_code,
                    request.user.pk,
                )
                return Response(
                    {"error": _("Не удалось расшифровать аудио. Повторите позже.")},
                    status=502,
                )
            return Response({"text": r.json().get("text", "")})
        except Exception:
            logger.exception("audio transcription failed user_id=%s", request.user.pk)
            return Response(
                {"error": _("Не удалось расшифровать аудио. Повторите позже.")},
                status=502,
            )


class UploadSpecView(APIView):
    """POST /api/assistant/upload-spec/   (multipart, field "file")

    Парсит файл, извлекает артикулы, вызывает search_parts.
    Возвращает тот же формат, что и обычный chat-ответ (text, cards, actions).
    """

    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser]

    def post(self, request):
        is_authenticated = bool(request.user and request.user.is_authenticated)
        if not is_authenticated:
            from .security import client_ip

            remote = client_ip(request)
            fingerprint = hashlib.sha256((remote or "unknown").encode("utf-8")).hexdigest()[:24]
            rate_key = f"anon_spec_upload:{fingerprint}"
            if cache.add(rate_key, 1, timeout=3600):
                upload_count = 1
            else:
                try:
                    upload_count = cache.incr(rate_key)
                except ValueError:
                    cache.set(rate_key, 1, timeout=3600)
                    upload_count = 1
            if upload_count > 10:
                return Response(
                    {"error": _("Лимит гостевых загрузок исчерпан. Войдите или повторите позже.")},
                    status=429,
                )

        upload = request.FILES.get("file")
        if not upload:
            return Response({"error": "file is required"}, status=400)
        if upload.size > MAX_FILE_BYTES:
            return Response(
                {"error": _("файл слишком большой (>%(mb)s МБ)") % {"mb": MAX_FILE_BYTES // (1024*1024)}},
                status=400,
            )
        try:
            from marketplace.upload_security import validate_uploaded_file

            ext = validate_uploaded_file(
                upload,
                allowed_ext={".xlsx", ".csv", ".pdf"},
                max_bytes=MAX_FILE_BYTES,
            )
        except ValueError as exc:
            return Response({"error": str(exc)}, status=400)

        blob = upload.read()
        kind = ext.removeprefix(".")

        # ── Magic-bytes защита от polyglot-атак ─────────────────
        # Проверяем содержимое повторно после общей проверки загрузки.
        if not _verify_magic_bytes(kind, blob):
            return Response(
                {"error": _("Содержимое файла не соответствует его расширению.")},
                status=400,
            )
        if kind == "csv" and b"\x00" in blob:
            return Response(
                {"error": _("Текстовый файл содержит бинарные данные.")},
                status=400,
            )

        if kind == "xlsx":
            articles = _extract_from_xlsx(blob)
        elif kind == "csv":
            articles = _extract_from_csv(blob)
        elif kind == "pdf":
            articles = _extract_from_pdf(blob)
            if not articles and blob[:4] == b"%PDF":
                return Response({
                    "error": "Не удалось извлечь текст из PDF. Возможно, скан без OCR.",
                    "filename": upload.name,
                }, status=200)
        else:
            return Response(
                {"error": _("Поддерживаются Excel (.xlsx), CSV и PDF.")},
                status=400,
            )

        conv_id = request.data.get("conversation_id") if is_authenticated else None
        if conv_id:
            conv = get_object_or_404(
                Conversation, id=conv_id, user=request.user, is_active=True
            )
        elif is_authenticated:
            conv = Conversation.objects.create(
                user=request.user, role=detect_user_role(request.user)
            )
        else:
            conv = None

        if not articles:
            return Response({
                "conversation_id": str(conv.id) if conv else None,
                "filename": upload.name,
                "articles_found": 0,
                "text": (
                    _('В файле «%(p0)s» не нашлось артикулов в распознаваемом формате.\nПроверьте, что артикулы в отдельной колонке и содержат цифры (например, AB-1234, 12345-XY).') % {"p0": f'{upload.name}'}
                ),
                "cards": [],
                "actions": [
                    {"label": _("Загрузить другой файл"), "action": "upload_spec", "params": {}},
                ],
                "suggestions": ["Создать RFQ вручную", "Поиск по бренду"],
            })

        # Прокидываем найденные артикулы в search_parts (multi-article path)
        try:
            result = execute_action(
                conv,
                "search_parts",
                {"articles": articles},
                request.user,
                role=(detect_user_role(request.user) if is_authenticated else "buyer"),
            )
        except Exception:  # pragma: no cover
            logger.exception(
                "uploaded specification processing failed user_id=%s",
                request.user.pk if is_authenticated else None,
            )
            return Response(
                {"error": _("Не удалось обработать спецификацию. Повторите позже.")},
                status=500,
            )

        # Префиксная подпись о загрузке
        prefix = (
            f"Из «{upload.name}» извлёк {len(articles)} артикулов. "
        )
        result_text = result.get("text") or ""
        result["text"] = prefix + result_text

        return Response({
            "conversation_id": str(conv.id) if conv else None,
            "filename": upload.name,
            "articles_found": len(articles),
            "articles": articles[:50],
            **result,
        })
